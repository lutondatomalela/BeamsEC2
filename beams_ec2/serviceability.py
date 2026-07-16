# -*- coding: utf-8 -*-
"""Estados-limite de serviço."""

from . import geometry as _previous
globals().update({k: v for k, v in vars(_previous).items() if not k.startswith("__")})
APP_VERSION = "v0.1"

# Estados-limite de serviço
# ============================================================
# ============================================================
APP_TABLE_DESCRIPTION = (
    "Workbook técnico de cálculo de vigas de betão armado com ELU, ELS por curvaturas, "
    "fendilhação directa, fluência, retracção, integração do diagrama de momentos, "
    "pormenorização, validação e auditoria completa."
)


@dataclass
class SLSParametersServiceability:
    regime: str = "Quase-permanente / longo prazo"
    rh_pct: float = 70.0
    t0_days: float = 28.0
    t_days: float = 18250.0
    ts_days: float = 7.0
    cement_class: str = "N"
    h0_override_mm: float = 0.0
    include_shrinkage: bool = True
    deflection_limit_denominator: float = 250.0
    crack_limit_mm: float = 0.30

    @property
    def long_term(self) -> bool:
        txt = normalize_text(self.regime)
        return "longo" in txt or "quase" in txt or "repet" in txt or "frequ" in txt

    @property
    def beta_tension_stiffening(self) -> float:
        return 0.5 if self.long_term else 1.0

    @property
    def kt_crack(self) -> float:
        return 0.4 if self.long_term else 0.6


def _cement_class_serviceability(value: str) -> str:
    c = str(value or "N").strip().upper()
    return c if c in {"S", "N", "R"} else "N"


def _ec2_creep_coefficient_serviceability(fcm: float, rh_pct: float, h0_mm: float, t0_days: float, t_days: float, cement_class: str = "N") -> float:
    """NP EN 1992-1-1, Anexo B: phi(t,t0)."""
    fcm = max(float(fcm), 12.0)
    rh = min(max(float(rh_pct), 40.0), 100.0)
    h0 = max(float(h0_mm), 20.0)
    t0 = max(float(t0_days), 0.5)
    t = max(float(t_days), t0 + 1e-6)
    cement = _cement_class_serviceability(cement_class)
    alpha_cement = {"S": -1.0, "N": 0.0, "R": 1.0}[cement]
    t0_adj = t0 * (9.0 / (2.0 + t0 ** 1.2) + 1.0) ** alpha_cement
    t0_adj = max(t0_adj, 0.5)
    alpha1 = (35.0 / fcm) ** 0.7
    alpha2 = (35.0 / fcm) ** 0.2
    alpha3 = (35.0 / fcm) ** 0.5
    humidity_term = (1.0 - rh / 100.0) / max(0.1 * h0 ** (1.0 / 3.0), 1e-9)
    if fcm <= 35.0:
        phi_rh = 1.0 + humidity_term
        beta_h = 1.5 * (1.0 + (0.012 * rh) ** 18.0) * h0 + 250.0
        beta_h = min(beta_h, 1500.0)
    else:
        phi_rh = (1.0 + humidity_term * alpha1) * alpha2
        beta_h = 1.5 * (1.0 + (0.012 * rh) ** 18.0) * h0 + 250.0 * alpha3
        beta_h = min(beta_h, 1500.0 * alpha3)
    beta_fcm = 16.8 / math.sqrt(fcm)
    beta_t0 = 1.0 / (0.1 + t0_adj ** 0.2)
    phi0 = phi_rh * beta_fcm * beta_t0
    dt = max(t - t0, 0.0)
    beta_c = (dt / max(beta_h + dt, 1e-9)) ** 0.3
    return max(phi0 * beta_c, 0.0)


def _interp_piecewise_serviceability(x: float, xs: List[float], ys: List[float]) -> float:
    if x <= xs[0]:
        return ys[0]
    if x >= xs[-1]:
        return ys[-1]
    for i in range(len(xs) - 1):
        if xs[i] <= x <= xs[i + 1]:
            r = (x - xs[i]) / max(xs[i + 1] - xs[i], 1e-9)
            return ys[i] + r * (ys[i + 1] - ys[i])
    return ys[-1]


def _ec2_shrinkage_strain_serviceability(fck: float, fcm: float, rh_pct: float, h0_mm: float, t_days: float, ts_days: float, cement_class: str = "N") -> Dict[str, float]:
    """NP EN 1992-1-1, 3.1.4 e Anexo B: eps_cs(t). Valores devolvidos em deformação absoluta."""
    cement = _cement_class_serviceability(cement_class)
    ads1 = {"S": 3.0, "N": 4.0, "R": 6.0}[cement]
    ads2 = {"S": 0.13, "N": 0.12, "R": 0.11}[cement]
    rh = min(max(float(rh_pct), 40.0), 100.0)
    h0 = max(float(h0_mm), 20.0)
    t = max(float(t_days), 0.0)
    ts = max(float(ts_days), 0.0)
    beta_rh = 1.55 * (1.0 - (rh / 100.0) ** 3.0)
    eps_cd0 = 0.85 * (220.0 + 110.0 * ads1) * math.exp(-ads2 * float(fcm) / 10.0) * 1e-6 * beta_rh
    kh = _interp_piecewise_serviceability(h0, [100.0, 200.0, 300.0, 500.0], [1.0, 0.85, 0.75, 0.70])
    dt = max(t - ts, 0.0)
    beta_ds = dt / max(dt + 0.04 * h0 ** 1.5, 1e-9)
    eps_cd = beta_ds * kh * eps_cd0
    eps_ca_inf = 2.5 * max(float(fck) - 10.0, 0.0) * 1e-6
    beta_as = 1.0 - math.exp(-0.2 * math.sqrt(max(t, 0.0)))
    eps_ca = beta_as * eps_ca_inf
    return {"eps_cd": eps_cd, "eps_ca": eps_ca, "eps_cs": eps_cd + eps_ca, "eps_cd0": eps_cd0, "kh": kh, "beta_ds": beta_ds}


def _profile_from_result_serviceability(r: pd.Series) -> SectionProfileGeometry:
    h = finite(r.get("h_cm"), 0.0) * 10.0
    bw = finite(r.get("bw_cm"), 0.0) * 10.0
    kind = str(r.get("section_type", "Rectangular") or "Rectangular")
    if kind.startswith("I"):
        return SectionProfileGeometry(
            h, bw,
            finite(r.get("b_top_cm"), 0.0) * 10.0,
            finite(r.get("tf_top_cm"), 0.0) * 10.0,
            finite(r.get("b_bottom_cm"), 0.0) * 10.0,
            finite(r.get("tf_bottom_cm"), 0.0) * 10.0,
            kind,
            str(r.get("section_geometry_source", "resultado")),
            str(r.get("section_geometry_confidence", "Alta")),
            finite(r.get("section_fit_error_pct"), 0.0),
            str(r.get("section_orientation_note", "")),
        )
    if kind.startswith("T"):
        return SectionProfileGeometry(
            h, bw,
            finite(r.get("b_top_cm", r.get("bf_cm", 0.0)), 0.0) * 10.0,
            finite(r.get("tf_top_cm", r.get("hf_cm", 0.0)), 0.0) * 10.0,
            bw, 0.0, "T", "resultado", "Alta", 0.0, "",
        )
    return SectionProfileGeometry(h, bw, bw, 0.0, bw, 0.0, "Rectangular", "resultado", "Alta", 0.0, "")


def _profile_exposed_perimeter_serviceability(p: SectionProfileGeometry) -> float:
    if p.kind == "Rectangular":
        return 2.0 * (p.web_mm + p.h_mm)
    return max(2.0 * p.top_width_mm + 2.0 * p.bottom_width_mm - 2.0 * p.web_mm + 2.0 * p.h_mm, 1.0)


def _profile_interval_props_serviceability(p: SectionProfileGeometry, z0: float, z1: float) -> Tuple[float, float, float]:
    """A, Q(top), I0(top) do betão entre z0 e z1."""
    a0 = max(min(float(z0), p.h_mm), 0.0)
    a1 = max(min(float(z1), p.h_mm), 0.0)
    if a1 < a0:
        a0, a1 = a1, a0
    A = Q = I0 = 0.0
    for s0, s1, b in p.segments:
        lo = max(a0, s0)
        hi = min(a1, s1)
        if hi <= lo:
            continue
        A += b * (hi - lo)
        Q += b * (hi ** 2 - lo ** 2) / 2.0
        I0 += b * (hi ** 3 - lo ** 3) / 3.0
    return A, Q, I0


def _uncracked_state_serviceability(p: SectionProfileGeometry, As_t: float, d_t: float, As_c: float, d_c: float, alpha_e: float) -> Dict[str, float]:
    Ac, Qc, I0c = _profile_interval_props_serviceability(p, 0.0, p.h_mm)
    nadd = max(alpha_e - 1.0, 0.0)
    At = Ac + nadd * (As_t + As_c)
    Qt = Qc + nadd * (As_t * d_t + As_c * d_c)
    zbar = Qt / max(At, 1e-9)
    I0 = I0c + nadd * (As_t * d_t ** 2 + As_c * d_c ** 2)
    I = max(I0 - At * zbar ** 2, 1.0)
    S = As_t * (d_t - zbar) + As_c * (d_c - zbar)
    return {"x_mm": zbar, "I_mm4": I, "S_mm3": S, "A_mm2": At}


def _cracked_state_serviceability(p: SectionProfileGeometry, As_t: float, d_t: float, As_c: float, d_c: float, alpha_e: float) -> Dict[str, float]:
    def f(x: float) -> float:
        Ac, Qc, _ = _profile_interval_props_serviceability(p, 0.0, x)
        At = Ac + alpha_e * (As_t + As_c)
        Qt = Qc + alpha_e * (As_t * d_t + As_c * d_c)
        return Qt / max(At, 1e-9) - x
    lo = 0.1
    hi = min(max(d_t - 0.1, 0.2), p.h_mm - 0.1)
    flo, fhi = f(lo), f(hi)
    if flo * fhi > 0:
        xs = [0.1 + i * (hi - 0.1) / 200.0 for i in range(201)]
        x = min(xs, key=lambda xx: abs(f(xx)))
    else:
        for _ in range(100):
            mid = 0.5 * (lo + hi)
            fm = f(mid)
            if flo * fm <= 0:
                hi = mid
                fhi = fm
            else:
                lo = mid
                flo = fm
        x = 0.5 * (lo + hi)
    Ac, Qc, I0c = _profile_interval_props_serviceability(p, 0.0, x)
    I0 = I0c + alpha_e * (As_t * d_t ** 2 + As_c * d_c ** 2)
    At = Ac + alpha_e * (As_t + As_c)
    I = max(I0 - At * x ** 2, 1.0)
    S = As_t * (d_t - x) + As_c * (d_c - x)
    return {"x_mm": x, "I_mm4": I, "S_mm3": S, "A_mm2": At}


def _rebar_phi_serviceability(label: str, default: float = 12.0) -> float:
    m = re.search(r"[Øφ]\s*(\d+(?:[.,]\d+)?)", str(label or ""))
    return float(m.group(1).replace(",", ".")) if m else float(default)


def _section_sls_states_serviceability(p: SectionProfileGeometry, As_bot: float, As_top: float, d_bot: float, d_top: float, Ecm: float, Ec_eff: float, fctm: float, eps_cs: float) -> Dict[str, Dict[str, Dict[str, float]]]:
    h = p.h_mm
    z_bot = d_bot
    z_top = h - d_top
    states = {}
    for name, pp, As_t, dt, As_c, dc, orient in [
        ("pos", p, As_bot, z_bot, As_top, z_top, 1.0),
        ("neg", p.flipped(), As_top, d_top, As_bot, h - d_bot, -1.0),
    ]:
        state_by_time = {}
        for time_name, Ec in [("short", Ecm), ("long", Ec_eff)]:
            alpha = 200000.0 / max(Ec, 1e-9)
            un = _uncracked_state_serviceability(pp, As_t, dt, As_c, dc, alpha)
            cr = _cracked_state_serviceability(pp, As_t, dt, As_c, dc, alpha)
            A, zg, Ig, _ = pp.gross_properties()
            yt = max(pp.h_mm - zg, 1e-9)
            Mcr = fctm * Ig / yt / 1e6
            kcs_un = eps_cs * alpha * un["S_mm3"] / max(un["I_mm4"], 1e-9) if time_name == "long" else 0.0
            kcs_cr = eps_cs * alpha * cr["S_mm3"] / max(cr["I_mm4"], 1e-9) if time_name == "long" else 0.0
            state_by_time[time_name] = {
                "alpha_e": alpha, "Mcr_kNm": Mcr, "orientation": orient,
                "As_t": As_t, "As_c": As_c, "d_t": dt, "d_c": dc,
                "x_un_mm": un["x_mm"], "I_un_mm4": un["I_mm4"], "S_un_mm3": un["S_mm3"],
                "x_cr_mm": cr["x_mm"], "I_cr_mm4": cr["I_mm4"], "S_cr_mm3": cr["S_mm3"],
                "kcs_un": kcs_un, "kcs_cr": kcs_cr,
            }
        states[name] = state_by_time
    return states


def _local_sls_response_serviceability(M_kNm: float, states: Dict, time_name: str, beta: float, fctm: float, Ecm: float, Es: float,
                              p: SectionProfileGeometry, phi_t: float, clear_spacing: float, cover_mm: float, phi_st: float,
                              kt: float, eps_cs: float) -> Dict[str, float | str]:
    sign = 1.0 if M_kNm >= 0 else -1.0
    key = "pos" if sign >= 0 else "neg"
    st = states[key][time_name]
    M = abs(float(M_kNm))
    Mcr = max(st["Mcr_kNm"], 1e-12)
    zeta = 0.0 if M <= Mcr else max(0.0, min(1.0, 1.0 - beta * (Mcr / M) ** 2))
    M_nmm = M * 1e6
    Ec = Es / max(st["alpha_e"], 1e-9)
    k_un = M_nmm / max(Ec * st["I_un_mm4"], 1e-9)
    k_cr = M_nmm / max(Ec * st["I_cr_mm4"], 1e-9)
    k_load = sign * (zeta * k_cr + (1.0 - zeta) * k_un)
    k_sh = st["orientation"] * (zeta * st["kcs_cr"] + (1.0 - zeta) * st["kcs_un"])
    sigma_s = st["alpha_e"] * M_nmm * max(st["d_t"] - st["x_cr_mm"], 0.0) / max(st["I_cr_mm4"], 1e-9)
    sigma_c = M_nmm * st["x_cr_mm"] / max(st["I_cr_mm4"], 1e-9) if zeta > 0 else M_nmm * st["x_un_mm"] / max(st["I_un_mm4"], 1e-9)
    wk = 0.0
    srmax = 0.0
    rho = 0.0
    if M > Mcr and st["As_t"] > 0:
        hceff = min(2.5 * max(p.h_mm - st["d_t"], 0.0), max((p.h_mm - st["x_cr_mm"]) / 3.0, 0.0), p.h_mm / 2.0)
        pp = p if key == "pos" else p.flipped()
        Ac_eff, _, _ = _profile_interval_props_serviceability(pp, pp.h_mm - hceff, pp.h_mm)
        rho = st["As_t"] / max(Ac_eff, 1e-9)
        c_eff = max(cover_mm + phi_st, 0.0)
        ctc = max(clear_spacing + phi_t, phi_t)
        if ctc <= 5.0 * (c_eff + phi_t / 2.0) + 1e-9:
            srmax = 3.4 * c_eff + 0.8 * 0.5 * 0.425 * phi_t / max(rho, 1e-9)
        else:
            srmax = 1.3 * max(pp.h_mm - st["x_cr_mm"], 0.0)
        alpha_short = Es / max(Ecm, 1e-9)
        eps_diff = (sigma_s - kt * fctm / max(rho, 1e-9) * (1.0 + alpha_short * rho)) / Es
        eps_diff = max(eps_diff, 0.6 * sigma_s / Es, 0.0)
        wk = srmax * eps_diff
    return {
        "curvature_load_1_per_mm": k_load,
        "curvature_shrinkage_1_per_mm": k_sh,
        "curvature_total_1_per_mm": k_load + k_sh,
        "zeta": zeta, "Mcr_kNm": Mcr, "x_cr_mm": st["x_cr_mm"],
        "I_un_mm4": st["I_un_mm4"], "I_cr_mm4": st["I_cr_mm4"],
        "sigma_s_MPa": sigma_s, "sigma_c_MPa": sigma_c,
        "wk_mm": wk, "sr_max_mm": srmax, "rho_p_eff": rho,
        "section_state": "Fendilhada/intermédia" if zeta > 0 else "Não fendilhada",
    }


def _integrate_curvature_serviceability(x_mm, kappa):
    import numpy as np
    x = np.asarray(x_mm, dtype=float)
    k = np.asarray(kappa, dtype=float)
    theta = np.zeros_like(x)
    y = np.zeros_like(x)
    for i in range(1, len(x)):
        dx = x[i] - x[i - 1]
        theta[i] = theta[i - 1] + 0.5 * (k[i] + k[i - 1]) * dx
        y[i] = y[i - 1] + 0.5 * (theta[i] + theta[i - 1]) * dx
    if len(x) > 1 and x[-1] > x[0]:
        y = y - (x - x[0]) * y[-1] / (x[-1] - x[0])
    return theta, y


def _case_is_els_serviceability(case_value) -> bool:
    s = str(case_value or "").upper()
    return any(k in s for k in ["ELS", "SLS", "SERV", "RARA", "FREQ", "QUASE", "QP", "CQC", "Q.P."])


def _service_group_serviceability(app, r: pd.Series, target: str):
    df = app.df_clean if getattr(app, "df_clean", None) is not None else pd.DataFrame()
    if df.empty:
        return pd.DataFrame(), "sem dados importados", ""
    cand = df[df["member"].astype(str) == str(r.get("member", ""))].copy()
    story = str(r.get("story", "") or "")
    if story and "story" in cand.columns:
        same = cand[cand["story"].astype(str) == story]
        if not same.empty:
            cand = same
    if cand.empty:
        return pd.DataFrame(), "viga não encontrada", ""
    if target:
        mask = cand["case"].astype(str).eq(target) | cand["case"].map(extract_combination_number).astype(str).eq(str(target))
        sel = cand[mask]
        return sel, ("combinação indicada" if not sel.empty else "combinação não encontrada"), str(target)
    els = cand[cand["case"].map(_case_is_els_serviceability)]
    if els.empty:
        return pd.DataFrame(), "sem combinação ELS identificada", ""
    axis_var = getattr(app, "var_moment_axis", None)
    axis = str(axis_var.get() if axis_var is not None else "MY").lower()
    best_case = None
    best_val = -1.0
    for case, grp in els.groupby("case", dropna=False):
        val = grp.get(axis, pd.Series(dtype=float)).map(lambda v: abs(finite(v, 0.0))).max() if axis in grp.columns else 0.0
        if val > best_val:
            best_val = val
            best_case = case
    sel = els[els["case"].astype(str) == str(best_case)]
    return sel, "combinação ELS seleccionada automaticamente", str(best_case)


def _moment_diagram_serviceability(group: pd.DataFrame, length_m: float, moment_col: str):
    import numpy as np
    if group is None or group.empty:
        return np.array([]), np.array([]), "Insuficiente", ""
    g = group.sort_values("__row_order").copy()
    L = max(float(length_m), 1e-6)
    m = np.array([finite(v, 0.0) for v in g.get(moment_col, pd.Series([0.0] * len(g)))], dtype=float)
    st = np.array([safe_float(v, float("nan")) for v in g.get("station", pd.Series([float("nan")] * len(g)))], dtype=float)
    source = "estações importadas"
    finite_mask = np.isfinite(st)
    if finite_mask.sum() >= 2:
        stf = st[finite_mask]
        mf = m[finite_mask]
        order = np.argsort(stf)
        stf, mf = stf[order], mf[order]
        if stf.max() <= 1.001 and L > 1.5:
            stf = stf * L
            source = "estações normalizadas 0-1"
        else:
            stf = stf - stf.min()
        if stf.max() > L * 1.10 or stf.max() <= 1e-9:
            stf = np.linspace(0.0, L, len(mf))
            source = "ordem das linhas distribuída no vão"
        else:
            if stf.max() < L * 0.95:
                stf = stf * (L / max(stf.max(), 1e-9))
                source += "; ajustadas ao comprimento"
        x, M = stf, mf
    else:
        x = np.linspace(0.0, L, len(m))
        M = m
        source = "ordem das linhas distribuída no vão"
    if len(x) == 0:
        return x, M, "Insuficiente", source
    unique = {}
    for xx, mm in zip(x, M):
        unique.setdefault(round(float(xx), 9), []).append(float(mm))
    x = np.array(sorted(unique.keys()), dtype=float)
    M = np.array([sum(unique[xx]) / len(unique[xx]) for xx in x], dtype=float)
    quality = "Adequado" if len(x) >= 5 else ("Aceitável" if len(x) >= 3 else "Insuficiente")
    return x, M, quality, source


def _calculate_member_sls_serviceability(app, r: pd.Series, group: pd.DataFrame, case_label: str, source_note: str, params: SLSParametersServiceability):
    import numpy as np
    p = _profile_from_result_serviceability(r)
    A, zbar, Ig, _ = p.gross_properties()
    u = _profile_exposed_perimeter_serviceability(p)
    h0_auto = 2.0 * A / max(u, 1e-9)
    h0 = params.h0_override_mm if params.h0_override_mm > 0 else h0_auto
    fck = finite(r.get("fck_MPa"), parse_concrete_strength(r.get("material", DEFAULT_CONCRETE_CLASS)))
    cp = concrete_props(fck)
    Ecm = cp["Ecm"]
    phi = _ec2_creep_coefficient_serviceability(cp["fcm"], params.rh_pct, h0, params.t0_days, params.t_days, params.cement_class) if params.long_term else 0.0
    Ec_eff = Ecm / (1.0 + phi)
    shrink = _ec2_shrinkage_strain_serviceability(fck, cp["fcm"], params.rh_pct, h0, params.t_days, params.ts_days, params.cement_class)
    eps_cs = shrink["eps_cs"] if params.long_term and params.include_shrinkage else 0.0
    As_bot = finite(r.get("as_prov_bot_mm2"), 0.0)
    As_top = finite(r.get("as_prov_top_mm2"), 0.0)
    d_bot = finite(r.get("d_bot_mm"), p.h_mm - 50.0)
    d_top = finite(r.get("d_top_mm"), p.h_mm - 50.0)
    states = _section_sls_states_serviceability(p, As_bot, As_top, d_bot, d_top, Ecm, Ec_eff, cp["fctm"], eps_cs)
    axis_var = getattr(app, "var_moment_axis", None)
    moment_col = str(axis_var.get() if axis_var is not None else "MY").lower()
    x_m, M_pts, quality, diagram_source = _moment_diagram_serviceability(group, finite(r.get("length_m"), 0.0), moment_col)
    if len(x_m) < 2:
        return {"service_status": "Dados insuficientes", "service_note": "Flecha não verificada: diagrama de momentos ELS indisponível.", "service_diagram_quality": "Insuficiente"}, []
    L_m = max(finite(r.get("length_m"), 0.0), float(x_m[-1]))
    x_grid_m = np.linspace(0.0, L_m, max(201, 40 * (len(x_m) - 1) + 1))
    M_grid = np.interp(x_grid_m, x_m, M_pts)
    phi_bot = _rebar_phi_serviceability(r.get("bot_rebar", ""), 12.0)
    phi_top = _rebar_phi_serviceability(r.get("top_rebar", ""), 12.0)
    clear_bot = finite(r.get("bot_clear_spacing_mm"), 999.0)
    clear_top = finite(r.get("top_clear_spacing_mm"), 999.0)
    phi_st = finite(r.get("phi_st_mm"), 8.0)
    point_rows = []
    k_short = []
    k_final = []
    wk_vals = []
    sigs_vals = []
    sigc_vals = []
    for xx, mm in zip(x_grid_m, M_grid):
        is_pos = mm >= 0
        phi_t = phi_bot if is_pos else phi_top
        clear = clear_bot if is_pos else clear_top
        rs = _local_sls_response_serviceability(mm, states, "short", 1.0, cp["fctm"], Ecm, 200000.0, p, phi_t, clear, finite(r.get("cover_mm"), 35.0), phi_st, 0.6, 0.0)
        rl = _local_sls_response_serviceability(mm, states, "long", params.beta_tension_stiffening, cp["fctm"], Ecm, 200000.0, p, phi_t, clear, finite(r.get("cover_mm"), 35.0), phi_st, params.kt_crack, eps_cs)
        k_short.append(rs["curvature_total_1_per_mm"])
        k_final.append(rl["curvature_total_1_per_mm"] if params.long_term else rs["curvature_total_1_per_mm"])
        active = rl if params.long_term else rs
        wk_vals.append(active["wk_mm"])
        sigs_vals.append(active["sigma_s_MPa"])
        sigc_vals.append(active["sigma_c_MPa"])
        point_rows.append({
            "viga": _beam_label_labels(r), "member": r.get("member", ""), "name": r.get("name", ""), "story": r.get("story", ""),
            "service_case": case_label, "x_m": float(xx), "M_service_kNm": float(mm),
            "section_state": active["section_state"], "zeta": active["zeta"], "Mcr_kNm": active["Mcr_kNm"],
            "x_cr_mm": active["x_cr_mm"], "I_un_mm4": active["I_un_mm4"], "I_cr_mm4": active["I_cr_mm4"],
            "sigma_s_MPa": active["sigma_s_MPa"], "sigma_c_MPa": active["sigma_c_MPa"], "wk_mm": active["wk_mm"],
            "sr_max_mm": active["sr_max_mm"], "rho_p_eff": active["rho_p_eff"],
            "curvature_load_1_per_m": active["curvature_load_1_per_mm"] * 1000.0,
            "curvature_shrinkage_1_per_m": active["curvature_shrinkage_1_per_mm"] * 1000.0,
            "curvature_total_1_per_m": active["curvature_total_1_per_mm"] * 1000.0,
        })
    x_mm = x_grid_m * 1000.0
    _, y_short = _integrate_curvature_serviceability(x_mm, k_short)
    _, y_final = _integrate_curvature_serviceability(x_mm, k_final)
    for pr, yi, yf in zip(point_rows, y_short, y_final):
        pr["deflection_instant_mm"] = float(yi)
        pr["deflection_final_mm"] = float(yf)
    delta_inst = float(np.max(np.abs(y_short)))
    delta_final = float(np.max(np.abs(y_final)))
    delta_used = delta_final if params.long_term else delta_inst
    limit = L_m * 1000.0 / max(params.deflection_limit_denominator, 1.0)
    wk_max = float(max(wk_vals) if wk_vals else 0.0)
    sigs_max = float(max(sigs_vals) if sigs_vals else 0.0)
    sigc_max = float(max(sigc_vals) if sigc_vals else 0.0)
    crack_status = "OK" if wk_max <= params.crack_limit_mm + 1e-9 else "Não verifica"
    defl_status = "OK" if delta_used <= limit + 1e-9 else "Não verifica"
    steel_status = "OK" if sigs_max <= 0.8 * finite(r.get("fyk_MPa"), 500.0) + 1e-9 else "Não verifica"
    concrete_lim = 0.45 * fck if params.long_term else 0.60 * fck
    concrete_status = "OK" if sigc_max <= concrete_lim + 1e-9 else "Não verifica"
    axial_service = max(abs(finite(v, 0.0)) for v in group.get("fx", pd.Series([0.0]))) if "fx" in group.columns else 0.0
    axial_ratio = axial_service * 1000.0 / max(A * cp["fctm"], 1e-9)
    axial_status = "OK" if axial_ratio <= 0.10 else "Verificar"
    checks = [crack_status, defl_status, steel_status, concrete_status]
    status = "OK" if all(v == "OK" for v in checks) else "Não verifica"
    note = "Cálculo por curvaturas EC2 7.4.3, com secções fendilhada/não fendilhada, fluência e retracção."
    if quality == "Insuficiente":
        status = "Dados insuficientes"
        note = "Flecha não validada: importar pelo menos 3 estações do diagrama de momentos da combinação ELS."
    elif axial_status != "OK":
        status = "Dados insuficientes"
        note = "Esforço axial de serviço relevante; o módulo ELS flexional requer verificação N-M dedicada."
    return {
        "service_combination": case_label,
        "service_case_source": source_note,
        "service_regime": params.regime,
        "service_points_imported": len(x_m),
        "service_points_integrated": len(x_grid_m),
        "service_diagram_quality": quality,
        "service_diagram_source": diagram_source,
        "service_RH_pct": params.rh_pct,
        "service_t0_days": params.t0_days,
        "service_t_days": params.t_days,
        "service_ts_days": params.ts_days,
        "service_cement_class": params.cement_class,
        "service_h0_mm": h0,
        "service_h0_source": "utilizador" if params.h0_override_mm > 0 else "2Ac/u automático",
        "service_phi_creep": phi,
        "service_Ecm_MPa": Ecm,
        "service_Ec_eff_MPa": Ec_eff,
        "service_eps_cd_permille": shrink["eps_cd"] * 1000.0,
        "service_eps_ca_permille": shrink["eps_ca"] * 1000.0,
        "service_eps_cs_permille": eps_cs * 1000.0,
        "service_sigma_s_MPa": sigs_max,
        "service_sigma_s_lim_MPa": 0.8 * finite(r.get("fyk_MPa"), 500.0),
        "service_sigma_c_MPa": sigc_max,
        "service_sigma_c_lim_MPa": concrete_lim,
        "service_wk_est_mm": wk_max,
        "service_wk_lim_mm": params.crack_limit_mm,
        "service_deflection_instant_mm": delta_inst,
        "service_deflection_final_mm": delta_final,
        "service_deflection_est_mm": delta_used,
        "service_deflection_lim_mm": limit,
        "service_deflection_limit": f"L/{params.deflection_limit_denominator:.0f}",
        "service_crack_status": crack_status,
        "service_deflection_status": defl_status,
        "service_stress_status": steel_status,
        "service_concrete_status": concrete_status,
        "service_axial_ratio": axial_ratio,
        "service_status": status,
        "service_note": note,
        "sls_method": "NP EN 1992-1-1:2010, 7.2, 7.3.4 e 7.4.3; Anexo B para fluência/retracção",
    }, point_rows


def _sls_params_from_app_serviceability(app) -> SLSParametersServiceability:
    return SLSParametersServiceability(
        regime=app.var_service_regime.get() if hasattr(app, "var_service_regime") else "Quase-permanente / longo prazo",
        rh_pct=finite(app.var_service_rh.get(), 70.0) if hasattr(app, "var_service_rh") else 70.0,
        t0_days=finite(app.var_service_t0.get(), 28.0) if hasattr(app, "var_service_t0") else 28.0,
        t_days=finite(app.var_service_t.get(), 18250.0) if hasattr(app, "var_service_t") else 18250.0,
        ts_days=finite(app.var_service_ts.get(), 7.0) if hasattr(app, "var_service_ts") else 7.0,
        cement_class=app.var_service_cement.get() if hasattr(app, "var_service_cement") else "N",
        h0_override_mm=finite(app.var_service_h0.get(), 0.0) if hasattr(app, "var_service_h0") else 0.0,
        include_shrinkage=bool(app.var_service_shrinkage.get()) if hasattr(app, "var_service_shrinkage") else True,
        deflection_limit_denominator=finite(app.var_ld_limit.get(), 250.0),
        crack_limit_mm=finite(app.var_crack_limit.get(), 0.30),
    )


def _apply_serviceability_serviceability(app, results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        app.df_sls_points = pd.DataFrame()
        app.df_sls_parameters = pd.DataFrame()
        return results
    out = results.copy()
    target = app.var_service_case.get().strip() if hasattr(app, "var_service_case") else ""
    params = _sls_params_from_app_serviceability(app)
    all_points = []
    for idx, r in out.iterrows():
        group, source_note, case_label = _service_group_serviceability(app, r, target)
        if group.empty:
            out.at[idx, "service_status"] = "Dados insuficientes"
            out.at[idx, "service_note"] = "Sem combinação ELS correspondente; resultado de serviço não validado."
            out.at[idx, "service_case_source"] = source_note
            out.at[idx, "service_diagram_quality"] = "Insuficiente"
            continue
        sls, points = _calculate_member_sls_serviceability(app, r, group, case_label, source_note, params)
        for k, v in sls.items():
            out.at[idx, k] = v
        all_points.extend(points)
        current = str(out.at[idx, "status"] if "status" in out.columns else "OK")
        if sls.get("service_status") == "Não verifica":
            out.at[idx, "status"] = "Falha"
            existing = str(out.at[idx, "failure_reason"] if "failure_reason" in out.columns else "").strip()
            out.at[idx, "failure_reason"] = "; ".join(x for x in [existing, "ELS não verifica"] if x)
            out.at[idx, "failure_type"] = "els"
        elif sls.get("service_status") == "Dados insuficientes" and current == "OK":
            out.at[idx, "status"] = "Verificar"
    app.df_sls_points = pd.DataFrame(all_points)
    app.df_sls_parameters = pd.DataFrame([
        ["Regime", params.regime], ["Combinação ELS", target or "automática"], ["RH [%]", params.rh_pct],
        ["t0 [dias]", params.t0_days], ["t [dias]", params.t_days], ["ts [dias]", params.ts_days],
        ["Classe de cimento", params.cement_class], ["h0 [mm]", params.h0_override_mm or "automático 2Ac/u"],
        ["Retracção", "Sim" if params.include_shrinkage else "Não"],
        ["Limite de flecha", f"L/{params.deflection_limit_denominator:.0f}"], ["wk,lim [mm]", params.crack_limit_mm],
    ], columns=["Parâmetro", "Valor"])
    return out


def _sls_audit_df_serviceability(results: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "viga", "member", "name", "story", "case", "service_combination", "service_regime",
        "service_points_imported", "service_diagram_quality", "service_phi_creep", "service_eps_cs_permille",
        "service_Ec_eff_MPa", "service_sigma_c_MPa", "service_sigma_c_lim_MPa",
        "service_sigma_s_MPa", "service_sigma_s_lim_MPa", "service_wk_est_mm", "service_wk_lim_mm",
        "service_deflection_instant_mm", "service_deflection_final_mm", "service_deflection_est_mm",
        "service_deflection_lim_mm", "service_deflection_limit", "service_crack_status",
        "service_deflection_status", "service_concrete_status", "service_stress_status", "service_status", "service_note",
    ]
    if results is None or results.empty:
        return pd.DataFrame()
    df = results.copy()
    if "viga" not in df.columns:
        df["viga"] = df.apply(_beam_label_labels, axis=1)
    return df[[c for c in cols if c in df.columns]].copy()

# As rotinas PDF existentes resolvem este nome em tempo de execução.
globals()["_sls_audit_df_calculation"] = _sls_audit_df_serviceability
globals()["_sls_audit_df_report_base"] = _sls_audit_df_serviceability


_calc_memory_geometry_base = _calc_memory_df_geometry

def _calc_memory_df_serviceability(results: pd.DataFrame) -> pd.DataFrame:
    base = _calc_memory_geometry_base(results)
    rows = []
    if results is not None and not results.empty:
        for _, r in results.iterrows():
            common = {
                "Viga": _beam_label_labels(r), "Caso": r.get("case", ""), "Piso": r.get("story", ""),
                "Secção": r.get("section_geometry_summary", r.get("section_type", "")),
            }
            def add(item, value, unit, ref, note=""):
                rows.append({**common, "Etapa": "ELS definitivo", "Item": item, "Valor": value, "Unidade": unit, "Critério/Referência": ref, "Estado/Nota": note})
            add("Combinação / regime", f"{r.get('service_combination','')} / {r.get('service_regime','')}", "-", "NP EN 1992-1-1, Secção 7", r.get("service_diagram_quality", ""))
            add("Fluência phi(t,t0)", _fmt_report_base(r.get("service_phi_creep"), 3), "-", "NP EN 1992-1-1, Anexo B")
            add("Retracção eps_cs", _fmt_report_base(r.get("service_eps_cs_permille"), 4), "‰", "NP EN 1992-1-1, 3.1.4 / 7.4.3")
            add("Tensões betão / aço", f"{_fmt_report_base(r.get('service_sigma_c_MPa'),2)} / {_fmt_report_base(r.get('service_sigma_s_MPa'),2)}", "MPa", "NP EN 1992-1-1, 7.2", f"{r.get('service_concrete_status','')} / {r.get('service_stress_status','')}")
            add("wk / limite", f"{_fmt_report_base(r.get('service_wk_est_mm'),3)} / {_fmt_report_base(r.get('service_wk_lim_mm'),3)}", "mm", "NP EN 1992-1-1, 7.3.4", r.get("service_crack_status", ""))
            add("Flecha instantânea / final", f"{_fmt_report_base(r.get('service_deflection_instant_mm'),2)} / {_fmt_report_base(r.get('service_deflection_final_mm'),2)}", "mm", "NP EN 1992-1-1, 7.4.3", r.get("service_deflection_status", ""))
            add("Flecha usada / limite", f"{_fmt_report_base(r.get('service_deflection_est_mm'),2)} / {_fmt_report_base(r.get('service_deflection_lim_mm'),2)}", "mm", r.get("service_deflection_limit", "L/250"), r.get("service_status", ""))
    return pd.concat([base, pd.DataFrame(rows)], ignore_index=True) if rows else base

globals()["_calc_memory_df_geometry"] = _calc_memory_df_serviceability
globals()["_calc_memory_df_labels"] = _calc_memory_df_serviceability
globals()["_calc_memory_df_reporting"] = _calc_memory_df_serviceability
globals()["_calc_memory_df_report_base"] = _calc_memory_df_serviceability


# --------------------------- GUI  ---------------------------
def _ensure_sls_vars_serviceability(self):
    defaults = {
        "var_service_case": (tk.StringVar, ""),
        "var_service_regime": (tk.StringVar, "Quase-permanente / longo prazo"),
        "var_service_rh": (tk.StringVar, "70"),
        "var_service_t0": (tk.StringVar, "28"),
        "var_service_t": (tk.StringVar, "18250"),
        "var_service_ts": (tk.StringVar, "7"),
        "var_service_cement": (tk.StringVar, "N"),
        "var_service_h0": (tk.StringVar, "0"),
        "var_service_shrinkage": (tk.BooleanVar, True),
        "var_pdf_scope": (tk.StringVar, "Relatório técnico"),
    }
    for name, (cls, value) in defaults.items():
        if not hasattr(self, name):
            setattr(self, name, cls(value=value))


def _build_sidebar_serviceability(self, parent):
    _ensure_sls_vars_serviceability(self)
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    if finite(self.var_ld_limit.get(), 20.0) < 100:
        self.var_ld_limit.set("250")
    hero = ttk.LabelFrame(parent, text="BeamsEC2")
    hero.pack(fill="x", pady=(0, 8))
    link = ttk.Label(hero, text=APP_NAME, style="Header.TLabel", cursor="hand2")
    link.pack(anchor="w"); link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
    ttk.Label(hero, text="Dimensionamento de vigas de betão armado", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(hero, text="ELU, ELS por curvaturas, fluência, retracção e relatórios auditáveis segundo a NP EN 1992-1-1.", style="Subtle.TLabel", wraplength=340, justify="left").pack(anchor="w", pady=(2, 0))

    data = ttk.LabelFrame(parent, text="1. Entrada")
    data.pack(fill="x", pady=(0, 8))
    ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

    params = ttk.LabelFrame(parent, text="2. Parâmetros de cálculo")
    params.pack(fill="x", pady=(0, 8))
    self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
    ttk.Label(params, text="Aço fyk [MPa]").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 2)
    self._add_label_entry(params, "Limite de flecha L/", self.var_ld_limit, 3)
    ttk.Label(params, text="Momento principal").grid(row=4, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_moment_axis, values=["MY", "MZ"], state="readonly").grid(row=4, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Corte vertical").grid(row=5, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_shear_axis, values=["FZ", "FY"], state="readonly").grid(row=5, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Torção").grid(row=6, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_torsion_axis, values=["MX", "MY", "MZ", "Nenhuma"], state="readonly").grid(row=6, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    params.columnconfigure(1, weight=1)

    sls = ttk.LabelFrame(parent, text="3. ELS por cálculo")
    sls.pack(fill="x", pady=(0, 8))
    ttk.Label(sls, text="Combinação ELS").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(sls, textvariable=self.var_service_case).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(sls, text="Regime").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_service_regime, values=["Quase-permanente / longo prazo", "Frequente / repetida", "Característica / curta duração"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(sls, "Humidade RH [%]", self.var_service_rh, 2)
    self._add_label_entry(sls, "Idade carregamento t0 [d]", self.var_service_t0, 3)
    self._add_label_entry(sls, "Idade final t [d]", self.var_service_t, 4)
    self._add_label_entry(sls, "Início secagem ts [d]", self.var_service_ts, 5)
    ttk.Label(sls, text="Classe de cimento").grid(row=6, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_service_cement, values=["S", "N", "R"], state="readonly", width=8).grid(row=6, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(sls, "h0 [mm] (0=auto)", self.var_service_h0, 7)
    ttk.Checkbutton(sls, text="Considerar retracção", variable=self.var_service_shrinkage).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Label(sls, text="Para a flecha final, importar ≥3 estações do diagrama ELS por viga.", style="Subtle.TLabel", wraplength=320, justify="left").grid(row=9, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 4))
    sls.columnconfigure(1, weight=1)

    filters = ttk.LabelFrame(parent, text="4. Filtros")
    filters.pack(fill="x", pady=(0, 8))
    ttk.Label(filters, text="Viga/Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha", "Verificar"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "corte", "torcao", "pormenorizacao", "els", "dados", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    filters.columnconfigure(1, weight=1)

    actions = ttk.LabelFrame(parent, text="5. Cálculo e exportação")
    actions.pack(fill="x", pady=(0, 8))
    ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ttk.Label(actions, text="PDF").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(actions, textvariable=self.var_pdf_scope, values=PDF_SCOPES_Reporting, state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(actions, text="Exportar .csv", command=self.export_csv).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

    status_box = ttk.LabelFrame(parent, text="6. Estado")
    status_box.pack(fill="x", pady=(0, 8))
    ttk.Label(status_box, textvariable=self.status_var, wraplength=340, justify="left").pack(fill="x", padx=6, pady=(4, 2))
    ttk.Progressbar(status_box, variable=self.progress_var, maximum=100).pack(fill="x", padx=6, pady=(2, 2))
    ttk.Label(status_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

    notes = ttk.LabelFrame(parent, text="7. Notas rápidas")
    notes.pack(fill="x", pady=(0, 8))
    ttk.Label(notes, text=(
        f"• Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}.\n"
        "• ELS: NP EN 1992-1-1, 7.2, 7.3.4 e 7.4.3; fluência/retracção pelo Anexo B.\n"
        "• A flecha é obtida por integração numérica das curvaturas ao longo da viga.\n"
        "• Com menos de 3 estações ELS, a flecha é assinalada como não validada."
    ), wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)

BeamsEC2App._build_sidebar = _build_sidebar_serviceability


def _build_instructions_tab_serviceability(self, parent):
    outer = ttk.Frame(parent, padding=10); outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1); outer.columnconfigure(0, weight=1)
    ttk.Label(outer, text="Instruções de utilização", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    host = ttk.Frame(outer); host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(host)
    content = (
        "OBJECTIVO\n"
        "BeamsEC2 dimensiona e verifica vigas de betão armado segundo a NP EN 1992-1-1. O módulo ELS calcula tensões, largura de fendas e flechas por curvaturas, considerando os estados não fendilhado e fendilhado, a contribuição do betão traccionado, a fluência e a retracção.\n\n"
        "DADOS DE IMPORTAÇÃO\n"
        + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
        "UNIDADES\n"
        "Forças em kN; momentos em kNm; Station e Length em m; dimensões em cm; áreas em cm² e inércias em cm⁴.\n\n"
        "DIAGRAMA ELS\n"
        "Para uma verificação completa da flecha, exporte a combinação de serviço com pelo menos três estações ao longo de cada viga, incluindo as zonas de momento máximo e de mudança de sinal. Com duas extremidades apenas, não é possível reconstruir de forma inequívoca o diagrama de curvaturas.\n\n"
        "PARÂMETROS DIFERIDOS\n"
        "RH é a humidade relativa média; t0 é a idade do primeiro carregamento; t é a idade de avaliação; ts é o início da secagem; a classe de cimento é S, N ou R. h0=0 activa o cálculo automático 2Ac/u.\n\n"
        "RESULTADOS\n"
        "O memória de cálculo inclui o resumo ELS, os parâmetros de fluência/retracção e os resultados ponto a ponto. Os PDFs mantêm o resumo executivo, o relatório técnico e a memória de cálculo em ficheiros separados."
    )
    txt.insert("1.0", content); txt.config(state="disabled")

BeamsEC2App._build_instructions_tab = _build_instructions_tab_serviceability


_validate_inputs_serviceability_base = BeamsEC2App.validate_inputs

def _validate_inputs_serviceability(self):
    _ensure_sls_vars_serviceability(self)
    err = _validate_inputs_serviceability_base(self)
    if err:
        return err
    rh = finite(self.var_service_rh.get(), -1.0)
    t0 = finite(self.var_service_t0.get(), -1.0)
    t = finite(self.var_service_t.get(), -1.0)
    ts = finite(self.var_service_ts.get(), -1.0)
    h0 = finite(self.var_service_h0.get(), -1.0)
    if not 40.0 <= rh <= 100.0:
        return "Humidade relativa RH inválida; utilizar um valor entre 40 % e 100 %."
    if t0 <= 0 or t <= t0:
        return "Idades ELS inválidas: deve verificar-se t > t0 > 0."
    if ts < 0 or h0 < 0:
        return "Valores ts ou h0 inválidos."
    if finite(self.var_ld_limit.get(), 0.0) <= 0:
        return "Limite de flecha L/n inválido."
    return None

BeamsEC2App.validate_inputs = _validate_inputs_serviceability


def _run_design_serviceability(self):
    _ensure_sls_vars_serviceability(self)
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err); return
    designer = BeamDesigner(
        cover_mm=finite(self.var_cover.get(), 35.0), agg_mm=DEFAULT_AGGREGATE_MM_Reporting,
        fyk=finite(self.var_fyk.get(), 500.0), gamma_c=DEFAULT_GAMMA_C_Reporting, gamma_s=DEFAULT_GAMMA_S_Reporting,
        cot_theta=2.0, crack_limit_mm=finite(self.var_crack_limit.get(), 0.30),
        deflection_ld_limit=finite(self.var_ld_limit.get(), 250.0), calc_mode=self.var_calc_mode.get(),
    )
    elu_env = self.df_env[~self.df_env["case"].map(_case_is_els_serviceability)].copy() if "case" in self.df_env.columns else self.df_env.copy()
    if elu_env.empty:
        messagebox.showwarning("Aviso", "Não foram identificadas combinações ELU para o dimensionamento. As combinações ELS são usadas exclusivamente nas verificações de serviço.")
        return
    input_df = reduce_to_governing_cases(elu_env) if self.var_reduce_cases.get() else elu_env
    self.df_calc_input = input_df.copy(); self.progress_var.set(0.0); self.status_var.set("Análise ELU/ELS em curso...")
    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular ELU... {d}/{t} envelopes"))
    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            self.after(0, lambda: self.status_var.set("A calcular ELS por curvaturas..."))
            results = _apply_serviceability_serviceability(self, results)
            summary = build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
            failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
            ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
            validation = build_data_validation(self.df_clean, self.df_env, results)
            def finish():
                self.df_results=results; self.df_summary=summary; self.df_failures=failures; self.df_ok=ok
                self.df_validation=validation; self.df_notes=build_normative_notes(); self.df_filtered=pd.DataFrame()
                self.show_df(self.tree_results,self.df_results); self.show_df(self.tree_summary,self.df_summary)
                self.show_df(self.tree_failures,self.df_failures); self.show_df(self.tree_shortlists,self.build_shortlists_df())
                self.show_df(self.tree_validation,self.df_validation); self.show_df(self.tree_notes,self.df_notes)
                self.update_report(); self.progress_var.set(100.0)
                n_ver=int((results["status"]=="Verificar").sum()) if "status" in results.columns else 0
                self.status_var.set(f"Cálculo concluído: {len(results)} envelopes; {len(failures)} falhas; {n_ver} casos a verificar.")
            self.after(0, finish)
        except Exception as exc:
            msg=str(exc); self.after(0,lambda m=msg:messagebox.showerror("Erro",m)); self.after(0,lambda:self.status_var.set("Falha na análise.")); self.after(0,lambda:self.progress_var.set(0.0))
    threading.Thread(target=worker,daemon=True).start()

BeamsEC2App.run_design = _run_design_serviceability


_metadata_df_serviceability_base = BeamsEC2App._metadata_df
_parameters_df_serviceability_base = BeamsEC2App._parameters_df

def _metadata_df_serviceability(self):
    df = _metadata_df_serviceability_base(self)
    if "Campo" in df.columns:
        mask = df["Campo"].astype(str).str.lower().eq("âmbito")
        df.loc[mask, "Valor"] = "Dimensionamento ELU e ELS por cálculo de vigas de betão armado"
    return df


def _parameters_df_serviceability(self):
    base = _parameters_df_serviceability_base(self)
    p = _sls_params_from_app_serviceability(self)
    extra = pd.DataFrame([
        ["Regime ELS", p.regime], ["RH [%]", p.rh_pct], ["t0 [dias]", p.t0_days], ["t [dias]", p.t_days],
        ["ts [dias]", p.ts_days], ["Classe de cimento", p.cement_class], ["h0 [mm]", p.h0_override_mm or "automático 2Ac/u"],
        ["Retracção", "Sim" if p.include_shrinkage else "Não"], ["Método de flecha", "integração numérica de curvaturas - EC2 7.4.3"],
    ], columns=["Parâmetro", "Valor"])
    return pd.concat([base, extra], ignore_index=True)

BeamsEC2App._metadata_df = _metadata_df_serviceability
BeamsEC2App._parameters_df = _parameters_df_serviceability


_write_excel_serviceability_base = BeamsEC2App._write_excel

def _write_excel_serviceability(self, path: str):
    _write_excel_serviceability_base(self, path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path)
        for name in ["11_ELS", "11A_ELS_Pontos", "11B_ELS_Parametros"]:
            if name in wb.sheetnames:
                del wb[name]
        insert_at = min(12, len(wb.sheetnames))
        datasets = [
            ("11_ELS", _sls_audit_df_serviceability(self.df_results)),
            ("11A_ELS_Pontos", getattr(self, "df_sls_points", pd.DataFrame())),
            ("11B_ELS_Parametros", getattr(self, "df_sls_parameters", pd.DataFrame())),
        ]
        fill=PatternFill("solid",fgColor="1F4E5F"); thin=Side(style="thin",color="D9E2E7"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
        for offset,(name,df) in enumerate(datasets):
            ws=wb.create_sheet(name, insert_at+offset)
            if df is None: df=pd.DataFrame()
            ws.append(list(df.columns))
            for _,rr in df.iterrows(): ws.append([None if pd.isna(v) else v for v in rr.tolist()])
            for cell in ws[1]: cell.fill=fill; cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
            for row in ws.iter_rows(min_row=2):
                for cell in row: cell.border=border; cell.alignment=Alignment(vertical="top",wrap_text=True)
            ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
            for j in range(1,ws.max_column+1):
                vals=[str(ws.cell(i,j).value) for i in range(1,min(ws.max_row,250)+1) if ws.cell(i,j).value is not None]
                ws.column_dimensions[get_column_letter(j)].width=min(max([len(v) for v in vals]+[10])+2,42)
        wb.properties.title = APP_NAME
        wb.properties.creator = ""
        wb.properties.lastModifiedBy = ""
        wb.save(path)
    except Exception:
        pass

BeamsEC2App._write_excel = _write_excel_serviceability


_write_pdf_serviceability_base = BeamsEC2App._write_pdf

def _write_pdf_serviceability(self, path: str):
    if not hasattr(self,"var_pdf_scope") or self.var_pdf_scope.get() != "Relatório técnico":
        return _write_pdf_serviceability_base(self,path)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles=_pdf_styles_report_base(); results=self.df_results if self.df_results is not None else pd.DataFrame()
    summary=self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures=self.df_failures if self.df_failures is not None else pd.DataFrame()
    doc=SimpleDocTemplate(path,pagesize=landscape(A4),rightMargin=12*mm,leftMargin=12*mm,topMargin=12*mm,bottomMargin=12*mm)
    doc.title=APP_NAME; doc.author=""; doc.subject=APP_SUBJECT
    story=[_program_link_pdf_reporting(styles)]
    n_total=len(results); n_ok=int((results.get("status",pd.Series(dtype=str))=="OK").sum()); n_fail=int((results.get("status",pd.Series(dtype=str))=="Falha").sum())
    meta=[["Data",datetime.now().strftime("%Y-%m-%d %H:%M"),"Suporte normativo","NP EN 1992-1-1:2010"],["Envelopes",str(n_total),"OK/Falhas",f"{n_ok}/{n_fail}"]]
    t=Table(meta,colWidths=[38*mm,90*mm,38*mm,105*mm]); t.setStyle(self._pdf_table_style(header=False)); story += [t,Spacer(1,5*mm),Paragraph("Relatório técnico",styles["Section"])]
    geom=summary.copy(); geom=geom.assign(viga=geom.apply(_beam_label_labels,axis=1)) if not geom.empty else geom
    env=self.df_env.copy() if self.df_env is not None else pd.DataFrame(); env=env.assign(viga=env.apply(_beam_label_labels,axis=1)) if not env.empty else env
    story += [Paragraph("Geometria identificada",styles["BodyCourier"]),self._pdf_df_table(geom,["viga","story","section_type","section_geometry_summary","section_geometry_source","section_geometry_confidence","section_fit_error_pct"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Critérios de cálculo",styles["BodyCourier"]),self._pdf_df_table(self._parameters_df(),["Parâmetro","Valor"],max_rows=None,widths=[90,180]),Spacer(1,5*mm),
              Paragraph("Envelopes de esforços",styles["BodyCourier"]),self._pdf_df_table(env,["viga","story","case","section_type","n_points_found","length","material","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Flexão",styles["BodyCourier"]),self._pdf_df_table(_flexure_audit_df_calculation(summary),["viga","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Esforço transverso e torção",styles["BodyCourier"]),self._pdf_df_table(_vt_audit_df_calculation(summary),["viga","case","v_ed_kN","VRd_c_kN","VRd_max_kN","cot_theta_shear","t_ed_kNm","TRd_max_kNm","torsion_considered","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"],max_rows=None),Spacer(1,5*mm),
              Paragraph("ELS por cálculo",styles["BodyCourier"]),self._pdf_df_table(_sls_audit_df_serviceability(summary),["viga","service_combination","service_regime","service_points_imported","service_diagram_quality","service_phi_creep","service_eps_cs_permille","service_sigma_c_MPa","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_deflection_instant_mm","service_deflection_final_mm","service_deflection_lim_mm","service_status","service_note"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Pormenorização",styles["BodyCourier"]),self._pdf_df_table(_detailing_audit_df_calculation(summary),["viga","case","bot_rebar","top_rebar","bot_clear_spacing_mm","top_clear_spacing_mm","phi_st_mm","stirrup_legs","s_st_mm","skin_rebar","detailing_status","detailing_issues"],max_rows=None)]
    if failures is not None and not failures.empty:
        story += [PageBreak(),Paragraph("Falhas",styles["Section"]),self._pdf_df_table(failures.assign(viga=failures.apply(_beam_label_labels,axis=1)),["viga","story","case","failure_type","failure_reason"],max_rows=None)]
    footer_date=datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas,doc_obj):
        canvas.saveState(); canvas.setTitle(APP_NAME); canvas.setAuthor(""); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm,7*mm,f"{APP_NAME} | {footer_date}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_serviceability


_build_notes_serviceability_base = build_normative_notes

def build_normative_notes_serviceability() -> pd.DataFrame:
    notes = _build_notes_serviceability_base()
    if notes is not None and not notes.empty and "Tema" in notes.columns:
        notes = notes[~notes["Tema"].astype(str).str.strip().eq("ELS")].copy()
    extra = pd.DataFrame([
        ("ELS - tensões", "NP EN 1992-1-1, 7.2", "Tensões calculadas na secção transformada; limites 0,60 fck ou 0,45 fck conforme o regime e 0,80 fyk no aço."),
        ("ELS - fendilhação", "NP EN 1992-1-1, 7.3.4", "Cálculo directo de wk por sr,max e diferença média de extensões, com área efectiva de betão traccionado."),
        ("ELS - deformação", "NP EN 1992-1-1, 7.4.3", "Curvaturas nos estados não fendilhado e fendilhado, interpolação por zeta e integração numérica ao longo da viga."),
        ("Fluência e retracção", "NP EN 1992-1-1, 3.1.4 e Anexo B", "phi(t,t0), Ec,eff e eps_cs calculados com RH, h0, idades e classe de cimento."),
    ], columns=["Tema","Referência","Nota"])
    return pd.concat([notes,extra],ignore_index=True)

globals()["build_normative_notes"] = build_normative_notes_serviceability


