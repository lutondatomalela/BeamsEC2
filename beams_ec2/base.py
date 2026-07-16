# -*- coding: utf-8 -*-

# Núcleo base: dados, cálculo e interface
"""
BeamsEC2 — dimensionamento e verificação de vigas de betão armado.

A aplicação recebe tabelas provenientes de programas de análise/cálculo
estrutural, cria envolventes de esforços e executa verificações ELU e ELS
segundo a NP EN 1992-1-1.
"""

from __future__ import annotations

import io
import math
import os
import re
import threading
import webbrowser
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_TITLE = "BeamsEC2 - Dimensionamento de Vigas Segundo o EC2"
APP_NAME = "BeamsEC2"
APP_VERSION = "v0.1"
APP_AUTHOR = ""
APP_SUBJECT = "Dimensionamento e verificação de vigas de betão armado segundo o Eurocódigo 2"
APP_KEYWORDS = "BeamsEC2, Eurocódigo 2, EC2, NP EN 1992-1-1, vigas, betão armado, flexão, esforço transverso, torção, ELS"
APP_CATEGORY = "Structural Engineering / Reinforced Concrete Design"
APP_TABLE_DESCRIPTION = "Memória de cálculo de vigas de betão armado, com dados, envolventes, verificações ELU/ELS, pormenorização e validação."
GITHUB_URL = "https://github.com/lutondatomalela/BeamsEC2"
MAX_PREVIEW_ROWS = 60000
DEFAULT_CONCRETE_CLASS = "C30/37"

# ============================================================
# Utilidades gerais
# ============================================================
def normalize_text(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("º", "o").replace("ª", "a")
    return s


def safe_float(value, default=float("nan")) -> float:
    try:
        if pd.isna(value):
            return default
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s:
            return default
        s = s.replace("\u00a0", " ").replace(" ", "")
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+,\d+", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return default


def finite(value, default=0.0) -> float:
    v = safe_float(value, default)
    return v if math.isfinite(v) else default


def parse_concrete_strength(material: str) -> float:
    s = str(material or "")
    m = re.search(r"C\s*(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", s, re.I)
    if m:
        return float(m.group(1).replace(",", "."))
    return 30.0


def concrete_props(fck: float, alpha_cc: float = 1.0, gamma_c: float = 1.5) -> Dict[str, float]:
    fcm = fck + 8.0
    fcd = alpha_cc * fck / gamma_c
    fctm = 0.30 * fck ** (2.0 / 3.0) if fck <= 50 else 2.12 * math.log(1 + fcm / 10.0)
    ecm = 22.0 * (fcm / 10.0) ** 0.3 * 1000.0
    return {"fck": fck, "fcm": fcm, "fcd": fcd, "fctm": fctm, "Ecm": ecm}


def steel_props(fyk: float = 500.0, gamma_s: float = 1.15) -> Dict[str, float]:
    return {"fyk": fyk, "fyd": fyk / gamma_s, "Es": 200000.0}


def bar_area_mm2(phi_mm: float) -> float:
    return math.pi * phi_mm * phi_mm / 4.0


def cm_to_mm(x) -> float:
    return finite(x, 0.0) * 10.0


def m_to_mm(x) -> float:
    return finite(x, 0.0) * 1000.0


def mm2_to_cm2(x: float) -> float:
    return x / 100.0


def split_member_case(text: str) -> Tuple[str, str, str]:
    s = str(text)
    parts = [p.strip() for p in s.split("/")]
    member = parts[0] if len(parts) > 0 else ""
    node = parts[1] if len(parts) > 1 else ""
    case = parts[2] if len(parts) > 2 else ""
    return member, node, case


def extract_combination_number(case_value) -> str:
    s = str(case_value or "").strip()
    m = re.search(r"\b(\d+)\b", s)
    return m.group(1) if m else s


def classify_limit_state(case_value) -> str:
    s = str(case_value or "").upper()
    if any(k in s for k in ["ELS", "SLS", "SERV", "SERVICE", "RARA", "FREQ", "QUASE", "QP", "CQC", "Q.P.", "(S)"]):
        return "ELS"
    if any(k in s for k in ["ELU", "ULS", "STR", "EQU", "GEO", "(U)", "(C)"]):
        return "ELU"
    return "ELU"


# ============================================================
# Leitura / limpeza de tabelas
# ============================================================
def parse_pasted_table(text: str) -> pd.DataFrame:
    text = text.strip()
    if not text:
        return pd.DataFrame()

    for sep in ("\t", ";", ","):
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep, engine="python", dtype=str)
            if len(df.columns) > 1:
                return df
        except Exception:
            pass

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return pd.DataFrame()

    rows = [re.split(r"\s{2,}", line) for line in lines]
    header = rows[0]
    body = rows[1:]
    width = len(header)
    body = [r[:width] + [""] * max(0, width - len(r)) for r in body]
    return pd.DataFrame(body, columns=header)


COLUMN_ALIASES = {
    "member_case": ["member/node/case", "member/n", "member node case", "barra/no/caso", "membro/no/caso"],
    "member": ["member", "bar", "barra", "membro", "element", "elemento"],
    "node": ["node", "no", "nó", "joint", "junta"],
    "case": ["case", "caso", "combination", "combinacao", "combinação", "load case", "loadcase"],
    "station": ["station", "x", "x (m)", "position", "posição", "posicao", "dist", "distance", "station (m)"],
    "fx": ["fx (kn)", "fx", "n (kn)", "n"],
    "fy": ["fy (kn)", "fy", "vy (kn)", "vy force"],
    "fz": ["fz (kn)", "fz", "vz (kn)", "vz force"],
    "mx": ["mx (knm)", "mx", "tx (knm)", "torsion", "torcao", "torção"],
    "my": ["my (knm)", "my", "m y", "m_y"],
    "mz": ["mz (knm)", "mz", "m z", "m_z"],
    "length": ["length (m)", "length(m)", "length", "l (m)", "comprimento", "comprimento (m)"],
    "material": ["material", "betao", "betão", "concrete"],
    "hy": ["hy (cm)", "hy", "b (cm)", "b", "bw (cm)", "bw", "base (cm)"],
    "hz": ["hz (cm)", "hz", "h (cm)", "h", "altura (cm)"],
    "bf": ["bf (cm)", "bf", "beff (cm)", "beff", "b_eff", "largura banzo", "largura do banzo"],
    "hf": ["hf (cm)", "hf", "tf (cm)", "tf", "espessura banzo", "espessura do banzo"],
    "name": ["name", "nome", "label", "grupo", "beam", "viga"],
    "story": ["story", "piso", "floor", "andar", "nivel", "nível"],
}


def rename_known_columns(df: pd.DataFrame) -> pd.DataFrame:
    norm_to_original = {normalize_text(c): c for c in df.columns}
    rename_map = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if normalize_text(alias) in norm_to_original:
                rename_map[norm_to_original[normalize_text(alias)]] = target
                break
    return df.rename(columns=rename_map).copy()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = rename_known_columns(df)
    df["__row_order"] = range(len(df))

    numeric_cols = ["fx", "fy", "fz", "mx", "my", "mz", "length", "hy", "hz", "bf", "hf", "station"]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = df[c].map(safe_float)

    if "member_case" in df.columns:
        split_vals = df["member_case"].map(split_member_case)
        if "member" not in df.columns:
            df["member"] = split_vals.map(lambda x: x[0])
        if "node" not in df.columns:
            df["node"] = split_vals.map(lambda x: x[1])
        if "case" not in df.columns:
            df["case"] = split_vals.map(lambda x: x[2])
    else:
        for c in ["member", "node", "case"]:
            if c not in df.columns:
                df[c] = ""

    for c in ["name", "story", "material"]:
        if c not in df.columns:
            df[c] = ""
    if "station" not in df.columns:
        df["station"] = float("nan")
    for c in ["bf", "hf"]:
        if c not in df.columns:
            df[c] = float("nan")

    # fallback de material sem esconder a origem
    mat = df["material"].astype(str).str.strip()
    empty = mat.str.lower().isin(["", "nan", "none"])
    df.loc[empty, "material"] = DEFAULT_CONCRETE_CLASS
    df["material_source"] = ["fallback" if v else "tabela" for v in empty]
    return df


# ============================================================
# Envelopes por viga/caso
# ============================================================
def _max_abs_with_pos(grp: pd.DataFrame, col: str) -> Tuple[float, str]:
    if col not in grp.columns or grp.empty:
        return 0.0, ""
    vals = grp[col].map(lambda x: finite(x, 0.0)).abs()
    idx = vals.idxmax()
    pos = grp.loc[idx].get("station", "")
    if not math.isfinite(finite(pos, float("nan"))):
        pos = grp.loc[idx].get("node", "")
    return float(vals.loc[idx]), str(pos)


def _max_signed(grp: pd.DataFrame, col: str, positive: bool = True) -> Tuple[float, str]:
    if col not in grp.columns or grp.empty:
        return 0.0, ""
    series = grp[col].map(lambda x: finite(x, 0.0))
    if positive:
        vals = series[series > 0]
        if vals.empty:
            return 0.0, ""
        idx = vals.idxmax()
        value = float(vals.loc[idx])
    else:
        vals = series[series < 0]
        if vals.empty:
            return 0.0, ""
        idx = vals.abs().idxmax()
        value = float(abs(series.loc[idx]))
    pos = grp.loc[idx].get("station", "")
    if not math.isfinite(finite(pos, float("nan"))):
        pos = grp.loc[idx].get("node", "")
    return value, str(pos)


def build_beam_envelopes(df: pd.DataFrame, moment_axis="my", shear_axis="fz", torsion_axis="mx") -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    axis_cols = [moment_axis, shear_axis, torsion_axis]
    for c in axis_cols:
        if c not in df.columns:
            df[c] = 0.0

    rows = []
    group_cols = ["member", "case", "name", "story"]
    for _, grp in df.groupby(group_cols, dropna=False):
        grp = grp.sort_values("__row_order")
        r0 = grp.iloc[0]
        m_pos, m_pos_at = _max_signed(grp, moment_axis, positive=True)
        m_neg, m_neg_at = _max_signed(grp, moment_axis, positive=False)
        m_abs, m_abs_at = _max_abs_with_pos(grp, moment_axis)
        v_ed, v_at = _max_abs_with_pos(grp, shear_axis)
        t_ed, t_at = _max_abs_with_pos(grp, torsion_axis)
        n_ed, n_at = _max_abs_with_pos(grp, "fx")

        # extremidades preservadas para auditoria
        first = grp.iloc[0]
        last = grp.iloc[-1]
        row = {
            "member": r0.get("member", ""),
            "case": r0.get("case", ""),
            "combination_number": extract_combination_number(r0.get("case", "")),
            "limit_state": classify_limit_state(r0.get("case", "")),
            "name": r0.get("name", ""),
            "story": r0.get("story", ""),
            "node_i": first.get("node", ""),
            "node_j": last.get("node", ""),
            "station_i": first.get("station", ""),
            "station_j": last.get("station", ""),
            "n_points_found": len(grp),
            "length": finite(r0.get("length", 0.0), 0.0),
            "material": r0.get("material", DEFAULT_CONCRETE_CLASS),
            "material_source": r0.get("material_source", "tabela"),
            "hy": finite(r0.get("hy", 0.0), 0.0),
            "hz": finite(r0.get("hz", 0.0), 0.0),
            "bf": finite(r0.get("bf", float("nan")), float("nan")),
            "hf": finite(r0.get("hf", float("nan")), float("nan")),
            "moment_axis": moment_axis.upper(),
            "shear_axis": shear_axis.upper(),
            "torsion_axis": torsion_axis.upper(),
            "m_pos_ed_kNm": m_pos,
            "m_pos_at": m_pos_at,
            "m_neg_ed_kNm": m_neg,
            "m_neg_at": m_neg_at,
            "m_abs_ed_kNm": m_abs,
            "m_abs_at": m_abs_at,
            "v_ed_kN": v_ed,
            "v_at": v_at,
            "t_ed_kNm": t_ed,
            "t_at": t_at,
            "n_ed_kN": n_ed,
            "n_at": n_at,
            "__row_order": finite(r0.get("__row_order", 0), 0),
        }
        # Se só houver momento com um sinal, dimensiona o lado correspondente e mantém mínimo no outro.
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["__row_order", "member", "case"]).reset_index(drop=True)


def reduce_to_governing_cases(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    work = df.copy()
    for c in ["m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm", "m_abs_ed_kNm"]:
        if c not in work.columns:
            work[c] = 0.0
    work["_score"] = (
        work["m_abs_ed_kNm"].abs().fillna(0.0)
        + 0.45 * work["v_ed_kN"].abs().fillna(0.0)
        + 0.75 * work["t_ed_kNm"].abs().fillna(0.0)
    )
    selected = set()
    for _, grp in work.groupby(["member", "name", "story"], dropna=False):
        if grp.empty:
            continue
        for col in ["_score", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm", "m_abs_ed_kNm"]:
            selected.add(grp[col].abs().idxmax())
        els = grp[grp["limit_state"].astype(str).eq("ELS")]
        if not els.empty:
            selected.add(els["_score"].idxmax())
    out = work.loc[sorted(selected)].copy().sort_values(["member", "story", "name", "case"]).reset_index(drop=True)
    out.drop(columns=[c for c in out.columns if c.startswith("_")], inplace=True, errors="ignore")
    return out


# ============================================================
# Modelos de cálculo
# ============================================================
@dataclass
class RebarChoice:
    n_bars: int
    phi_mm: float
    area_mm2: float
    layers: int
    bars_per_layer: int
    centroid_from_edge_mm: float
    clear_spacing_mm: float
    status: str

    @property
    def label(self) -> str:
        if self.n_bars <= 0 or self.phi_mm <= 0:
            return "-"
        return f"{self.n_bars}Ø{int(self.phi_mm)}"


@dataclass
class BeamSection:
    bw_mm: float
    h_mm: float
    bf_mm: float = 0.0
    hf_mm: float = 0.0

    @property
    def is_t(self) -> bool:
        return self.bf_mm > self.bw_mm and self.hf_mm > 0.0

    @property
    def b_compression_rect(self) -> float:
        return self.bf_mm if self.is_t else self.bw_mm

    @property
    def section_type(self) -> str:
        return "T" if self.is_t else "Rectangular"


@dataclass
class BeamDesignResult:
    member: str
    case: str
    name: str
    story: str
    status: str
    failure_reason: str = ""


class BeamDesigner:
    def __init__(
        self,
        cover_mm=35.0,
        agg_mm=20.0,
        fyk=500.0,
        gamma_c=1.5,
        gamma_s=1.15,
        alpha_cc=1.0,
        cot_theta=2.0,
        crack_limit_mm=0.30,
        deflection_ld_limit=20.0,
        calc_mode="dimensionamento",
    ):
        self.cover_mm = float(cover_mm)
        self.agg_mm = float(agg_mm)
        self.fyk = float(fyk)
        self.gamma_c = float(gamma_c)
        self.gamma_s = float(gamma_s)
        self.alpha_cc = float(alpha_cc)
        self.cot_theta = max(1.0, min(2.5, float(cot_theta)))
        self.crack_limit_mm = float(crack_limit_mm)
        self.deflection_ld_limit = float(deflection_ld_limit)
        self.calc_mode = calc_mode
        self.long_diams = [10.0, 12.0, 14.0, 16.0, 20.0, 25.0, 32.0]
        self.stirrup_diams = [6.0, 8.0, 10.0, 12.0]
        self.stirrup_legs = [2, 4, 6]
        self.spacing_candidates_mm = [75.0, 100.0, 125.0, 150.0, 175.0, 200.0, 225.0, 250.0, 300.0]

    # ---------------------- geometrias/armaduras ----------------------
    def choose_stirrup_diameter(self, phi_long: float) -> float:
        req = max(6.0, phi_long / 4.0)
        for phi in self.stirrup_diams:
            if phi >= req:
                return phi
        return self.stirrup_diams[-1]

    def min_clear_spacing(self, phi: float) -> float:
        return max(20.0, phi, self.agg_mm + 5.0)

    def max_bars_per_layer(self, b_mm: float, phi: float, phi_st: float) -> int:
        edge = self.cover_mm + phi_st + phi / 2.0
        inner = b_mm - 2.0 * edge
        if inner <= phi:
            return 0
        req = self.min_clear_spacing(phi)
        # n barras: clear = inner/(n-1)-phi >= req; para n=1 aceita mas vigas precisam de 2 barras.
        nmax = int(math.floor(inner / (phi + req) + 1.0 + 1e-9))
        return max(0, nmax)

    def choose_longitudinal_bars(self, As_req: float, b_mm: float, prefer_phi: Optional[float] = None, max_layers: int = 3) -> RebarChoice:
        As_req = max(0.0, As_req)
        best: Optional[RebarChoice] = None
        diams = self.long_diams if prefer_phi is None else [prefer_phi] + [d for d in self.long_diams if d != prefer_phi]
        for phi in diams:
            phi_st = self.choose_stirrup_diameter(phi)
            nmax_layer = self.max_bars_per_layer(b_mm, phi, phi_st)
            if nmax_layer < 2:
                continue
            for n in range(2, nmax_layer * max_layers + 1):
                area = n * bar_area_mm2(phi)
                if area + 1e-9 < As_req:
                    continue
                layers = int(math.ceil(n / nmax_layer))
                if layers > max_layers:
                    continue
                bars_per = min(n, nmax_layer)
                req = self.min_clear_spacing(phi)
                edge = self.cover_mm + phi_st + phi / 2.0
                if bars_per > 1:
                    inner = b_mm - 2.0 * edge
                    clear = inner / (bars_per - 1) - phi
                else:
                    clear = 999.0
                # centroide das camadas assumindo distribuição por camadas cheias excepto última
                remaining = n
                sum_a_y = 0.0
                sum_a = 0.0
                for layer in range(layers):
                    n_layer = min(nmax_layer, remaining)
                    y = edge + layer * (phi + req)
                    a = n_layer * bar_area_mm2(phi)
                    sum_a += a
                    sum_a_y += a * y
                    remaining -= n_layer
                centroid = sum_a_y / max(sum_a, 1e-9)
                choice = RebarChoice(n, phi, area, layers, bars_per, centroid, clear, "OK")
                key = (choice.area_mm2, choice.layers, choice.phi_mm, choice.n_bars)
                if best is None or key < (best.area_mm2, best.layers, best.phi_mm, best.n_bars):
                    best = choice
                break
        if best is None:
            return RebarChoice(0, 0.0, 0.0, 0, 0, 0.0, 0.0, "Não cabe")
        return best

    # ---------------------- EC2 flexão ----------------------
    def as_min_beam(self, bt_mm: float, d_mm: float, fctm: float, fyk: float) -> float:
        return max(0.26 * fctm / fyk * bt_mm * d_mm, 0.0013 * bt_mm * d_mm)

    def compression_block(self, a_mm: float, section: BeamSection, fcd: float) -> Tuple[float, float]:
        """Retorna Cc [N] e ybar [mm] a partir do topo comprimido para bloco retangular de profundidade a=0.8x."""
        a = max(0.0, min(a_mm, section.h_mm))
        if a <= 0:
            return 0.0, 0.0
        if not section.is_t or a <= section.hf_mm:
            C = section.b_compression_rect * a * fcd
            y = a / 2.0
            return C, y
        C1 = section.bf_mm * section.hf_mm * fcd
        a_web = a - section.hf_mm
        C2 = section.bw_mm * a_web * fcd
        C = C1 + C2
        y = (C1 * (section.hf_mm / 2.0) + C2 * (section.hf_mm + a_web / 2.0)) / max(C, 1e-9)
        return C, y

    def a_for_concrete_force(self, C_req: float, section: BeamSection, fcd: float) -> float:
        C_req = max(0.0, C_req)
        if C_req <= 0:
            return 0.0
        # Capacidade até h
        Cmax, _ = self.compression_block(section.h_mm, section, fcd)
        if C_req > Cmax:
            raise ValueError("secção de betão comprimido insuficiente")
        lo, hi = 0.0, section.h_mm
        for _ in range(80):
            mid = 0.5 * (lo + hi)
            Cmid, _ = self.compression_block(mid, section, fcd)
            if Cmid < C_req:
                lo = mid
            else:
                hi = mid
        return 0.5 * (lo + hi)

    def flexural_required(self, MEd_kNm: float, section: BeamSection, d_mm: float, d2_mm: float, fcd: float, fyd: float) -> Dict[str, float]:
        MEd = max(0.0, MEd_kNm) * 1e6
        if MEd <= 1e-9:
            return {"As_req": 0.0, "As_comp_req": 0.0, "x_mm": 0.0, "z_mm": 0.9 * d_mm, "MRd_req_kNm": 0.0, "ductility_status": "OK"}
        x_lim = 0.45 * d_mm
        a_lim = 0.8 * x_lim
        C_lim, y_lim = self.compression_block(a_lim, section, fcd)
        M_lim = C_lim * (d_mm - y_lim)
        if MEd <= M_lim + 1e-6:
            # resolver a por bissecção
            lo, hi = 0.0, a_lim
            for _ in range(100):
                mid = 0.5 * (lo + hi)
                C, y = self.compression_block(mid, section, fcd)
                M = C * (d_mm - y)
                if M < MEd:
                    lo = mid
                else:
                    hi = mid
            a = 0.5 * (lo + hi)
            C, y = self.compression_block(a, section, fcd)
            As_req = C / fyd
            x = a / 0.8
            z = d_mm - y
            return {"As_req": As_req, "As_comp_req": 0.0, "x_mm": x, "z_mm": z, "MRd_req_kNm": MEd / 1e6, "ductility_status": "OK"}
        # Armadura dupla aproximada quando ultrapassa xlim
        lever_cs = max(d_mm - d2_mm, 1e-9)
        delta_M = MEd - M_lim
        As_comp = delta_M / (fyd * lever_cs)
        As_tens = C_lim / fyd + As_comp
        return {"As_req": As_tens, "As_comp_req": As_comp, "x_mm": x_lim, "z_mm": d_mm - y_lim, "MRd_req_kNm": MEd / 1e6, "ductility_status": "Armadura dupla"}

    def flexural_capacity(self, As_tens: float, As_comp: float, section: BeamSection, d_mm: float, d2_mm: float, fcd: float, fyd: float) -> Dict[str, float]:
        if As_tens <= 0:
            return {"MRd_kNm": 0.0, "x_mm": 0.0, "z_mm": 0.0, "eta": None}
        T = As_tens * fyd
        Cs = max(0.0, As_comp) * fyd
        Cc_req = max(T - Cs, 0.0)
        try:
            a = self.a_for_concrete_force(Cc_req, section, fcd)
            Cc, y = self.compression_block(a, section, fcd)
        except Exception:
            return {"MRd_kNm": 0.0, "x_mm": None, "z_mm": None, "eta": None}
        M = Cc * (d_mm - y) + Cs * max(d_mm - d2_mm, 0.0)
        return {"MRd_kNm": M / 1e6, "x_mm": a / 0.8 if a else 0.0, "z_mm": d_mm - y, "eta": None}

    # ---------------------- EC2 corte/torção ----------------------
    def vrd_c(self, VEd_kN: float, bw: float, d: float, Asl: float, fck: float, fcd: float) -> Dict[str, float]:
        C_Rdc = 0.18 / self.gamma_c
        k1 = 0.15
        k = min(2.0, 1.0 + math.sqrt(200.0 / max(d, 1e-9)))
        rho_l = min(max(Asl / max(bw * d, 1e-9), 0.0), 0.02)
        sigma_cp = 0.0
        vmin = 0.035 * k ** 1.5 * math.sqrt(fck)
        vrdc = max(C_Rdc * k * (100.0 * rho_l * fck) ** (1.0 / 3.0) + k1 * sigma_cp, vmin + k1 * sigma_cp)
        VRdc = vrdc * bw * d / 1000.0
        return {"k": k, "rho_l": rho_l, "vmin_MPa": vmin, "VRd_c_kN": VRdc}

    def shear_requirements(self, VEd_kN: float, bw: float, d: float, Asl: float, fck: float, fcd: float, fyd: float) -> Dict[str, float | str]:
        V = abs(VEd_kN) * 1000.0
        z = 0.9 * d
        cot = self.cot_theta
        tan = 1.0 / cot
        nu1 = 0.6 * (1.0 - fck / 250.0)
        VRdmax = bw * z * nu1 * fcd / (cot + tan)
        vrdc = self.vrd_c(VEd_kN, bw, d, Asl, fck, fcd)
        VRdc_N = vrdc["VRd_c_kN"] * 1000.0
        if V <= VRdc_N:
            Asw_s = 0.0
            status = "OK sem armadura adicional por V"
        elif V <= VRdmax:
            Asw_s = V / max(z * fyd * cot, 1e-9)  # mm²/mm
            status = "Requer armadura de esforço transverso"
        else:
            Asw_s = V / max(z * fyd * cot, 1e-9)
            status = "Não conforme: VEd > VRd,max"
        rho_w_min = 0.08 * math.sqrt(fck) / self.fyk
        Asw_s_min = rho_w_min * bw  # mm²/mm
        return {
            **vrdc,
            "VRd_max_kN": VRdmax / 1000.0,
            "Asw_s_shear_req_mm2_per_mm": Asw_s,
            "Asw_s_min_mm2_per_mm": Asw_s_min,
            "shear_status": status,
            "cot_theta": cot,
        }

    def torsion_requirements(self, TEd_kNm: float, section: BeamSection, fck: float, fcd: float, fyd: float) -> Dict[str, float | str]:
        T = abs(TEd_kNm) * 1e6
        if T <= 1e-9:
            return {
                "TRd_max_kNm": None,
                "Asw_s_torsion_req_mm2_per_mm": 0.0,
                "Asl_torsion_req_mm2": 0.0,
                "torsion_status": "Sem torção relevante",
                "tef_mm": None,
                "Ak_mm2": None,
                "uk_mm": None,
            }
        b = section.bw_mm
        h = section.h_mm
        A = b * h
        u = 2.0 * (b + h)
        tef = max(A / max(u, 1e-9), 2.0 * self.cover_mm, 50.0)
        tef = min(tef, min(b, h) / 2.0 - 1.0)
        Ak = max((b - tef) * (h - tef), 1.0)
        uk = 2.0 * max(b + h - 2.0 * tef, 1.0)
        cot = 1.0  # opção conservadora usual para torção
        nu1 = 0.6 * (1.0 - fck / 250.0)
        TRdmax = 2.0 * nu1 * fcd * Ak * tef / (cot + 1.0 / cot)
        Asw_s = T / max(2.0 * Ak * fyd * cot, 1e-9)
        Asl = T * uk * cot / max(2.0 * Ak * fyd, 1e-9)
        if T <= TRdmax:
            status = "Requer armadura de torção"
        else:
            status = "Não conforme: TEd > TRd,max"
        return {
            "TRd_max_kNm": TRdmax / 1e6,
            "Asw_s_torsion_req_mm2_per_mm": Asw_s,
            "Asl_torsion_req_mm2": Asl,
            "torsion_status": status,
            "tef_mm": tef,
            "Ak_mm2": Ak,
            "uk_mm": uk,
        }

    def choose_stirrups(self, Asw_s_total: float, b: float, h: float, d: float, torsion: bool) -> Dict[str, float | str]:
        s_lim_v = min(0.75 * d, 600.0)
        s_lim_t = min((2.0 * (b + h)) / 8.0, min(b, h), 350.0) if torsion else 999.0
        s_lim = min(s_lim_v, s_lim_t)
        best = None
        for phi in self.stirrup_diams:
            for legs in self.stirrup_legs:
                area = legs * bar_area_mm2(phi)
                for s in sorted(self.spacing_candidates_mm, reverse=True):
                    if s > s_lim + 1e-9:
                        continue
                    provided = area / s
                    if provided + 1e-12 >= Asw_s_total:
                        key = (area / s, phi, legs, -s)
                        cand = {"phi_st_mm": phi, "stirrup_legs": legs, "s_st_mm": s, "Asw_s_prov_mm2_per_mm": provided, "s_lim_mm": s_lim, "stirrup_status": "OK"}
                        if best is None or key < (best["Asw_s_prov_mm2_per_mm"], best["phi_st_mm"], best["stirrup_legs"], -best["s_st_mm"]):
                            best = cand
        if best is None:
            # seleccionar a maior solução disponível e assinalar
            phi = self.stirrup_diams[-1]
            legs = self.stirrup_legs[-1]
            s = min([x for x in self.spacing_candidates_mm if x <= s_lim] or [75.0])
            area = legs * bar_area_mm2(phi)
            return {"phi_st_mm": phi, "stirrup_legs": legs, "s_st_mm": s, "Asw_s_prov_mm2_per_mm": area / s, "s_lim_mm": s_lim, "stirrup_status": "Não cabe"}
        return best

    # ---------------------- ELS ----------------------
    def serviceability(self, row: pd.Series, As_bot: float, As_top: float, d_bot: float, d_top: float, section: BeamSection, cp: Dict[str, float]) -> Dict[str, float | str]:
        L = max(finite(row.get("length", 0.0), 0.0) * 1000.0, 1e-9)
        Mserv = finite(row.get("m_abs_ed_kNm", max(finite(row.get("m_pos_ed_kNm"),0), finite(row.get("m_neg_ed_kNm"),0))), 0.0)
        As_use = max(As_bot, As_top, 1e-9)
        d_use = max(d_bot, d_top, 1e-9)
        z = 0.9 * d_use
        sigma_s = abs(Mserv) * 1e6 / max(As_use * z, 1e-9)
        # estimativa expedita de wk; deliberadamente conservadora e auditável
        wk_est = 0.00085 * sigma_s if sigma_s > 0 else 0.0
        wk_status = "OK" if wk_est <= self.crack_limit_mm else "Verificar fendilhação"
        ld_ratio = L / max(d_use, 1e-9)
        defl_status = "OK" if ld_ratio <= self.deflection_ld_limit else "Verificar deformação"
        stress_status = "OK" if sigma_s <= 0.80 * self.fyk else "Verificar tensão no aço"
        service_status = "OK" if wk_status == "OK" and defl_status == "OK" and stress_status == "OK" else "Verificar"
        if classify_limit_state(row.get("case", "")) != "ELS":
            note = "Informativo — caso não identificado como ELS"
        else:
            note = "Combinação ELS identificada"
        return {
            "service_sigma_s_MPa": sigma_s,
            "service_wk_est_mm": wk_est,
            "service_wk_lim_mm": self.crack_limit_mm,
            "service_L_over_d": ld_ratio,
            "service_L_over_d_lim": self.deflection_ld_limit,
            "service_crack_status": wk_status,
            "service_deflection_status": defl_status,
            "service_stress_status": stress_status,
            "service_status": service_status,
            "service_note": note,
        }

    # ---------------------- pormenorização ----------------------
    def detailing_check(self, section: BeamSection, bot: RebarChoice, top: RebarChoice, stir: Dict[str, float | str], As_skin_face: float) -> Dict[str, float | str]:
        issues = []
        if bot.status != "OK":
            issues.append("armadura inferior não cabe")
        if top.status != "OK":
            issues.append("armadura superior não cabe")
        if bot.n_bars and bot.n_bars < 2:
            issues.append("mínimo de 2 varões inferiores não cumprido")
        if top.n_bars and top.n_bars < 2:
            issues.append("mínimo de 2 varões superiores não cumprido")
        if float(stir.get("s_st_mm", 999.0)) > float(stir.get("s_lim_mm", 0.0)) + 1e-9:
            issues.append("espaçamento dos estribos superior ao limite")
        if stir.get("stirrup_status") != "OK":
            issues.append("armadura transversal requerida não cabe nas opções disponíveis")
        if section.h_mm > 1000 and As_skin_face <= 0:
            issues.append("prever armadura de pele nas faces laterais")
        status = "OK" if not issues else "Verificar"
        return {"detailing_status": status, "detailing_issues": "; ".join(issues) if issues else "-"}

    # ---------------------- dimensionamento de uma linha envelope ----------------------
    def design_one(self, row: pd.Series) -> Dict:
        material = str(row.get("material", DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
        fck = parse_concrete_strength(material)
        cp = concrete_props(fck, alpha_cc=self.alpha_cc, gamma_c=self.gamma_c)
        sp = steel_props(self.fyk, gamma_s=self.gamma_s)
        fcd = cp["fcd"]
        fyd = sp["fyd"]

        bw = cm_to_mm(row.get("hy", 0.0))
        h = cm_to_mm(row.get("hz", 0.0))
        bf = cm_to_mm(row.get("bf", float("nan")))
        hf = cm_to_mm(row.get("hf", float("nan")))
        if bw <= 0 or h <= 0:
            return {
                "member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""),
                "status": "Falha", "failure_reason": "Dados incompletos: dimensões HY/HZ inválidas", "failure_type": "dados_incompletos",
            }
        if not math.isfinite(bf) or bf <= bw or not math.isfinite(hf) or hf <= 0:
            bf = 0.0
            hf = 0.0
        section = BeamSection(bw, h, bf, hf)

        # Estimativa inicial de d com Ø16 e estribo Ø8; será actualizada após escolha real.
        phi_initial = 16.0
        phi_st_initial = self.choose_stirrup_diameter(phi_initial)
        d_bot_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
        d_top_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
        d2_initial = self.cover_mm + phi_st_initial + phi_initial / 2.0
        if d_bot_initial <= 0:
            return {
                "member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""),
                "status": "Falha", "failure_reason": "Dados geométricos incompatíveis com recobrimento", "failure_type": "dados_incompletos",
            }

        Mpos = finite(row.get("m_pos_ed_kNm"), 0.0)
        Mneg = finite(row.get("m_neg_ed_kNm"), 0.0)
        VEd = finite(row.get("v_ed_kN"), 0.0)
        TEd = finite(row.get("t_ed_kNm"), 0.0)

        # Armadura mínima por face traccionada. Para negativo, usar bw como largura de tracção conservadora.
        As_min_bot = self.as_min_beam(section.bw_mm, d_bot_initial, cp["fctm"], self.fyk)
        As_min_top = self.as_min_beam(section.bw_mm, d_top_initial, cp["fctm"], self.fyk)

        flex_pos = self.flexural_required(Mpos, section, d_bot_initial, d2_initial, fcd, fyd)
        # Para momento negativo, a compressão está na zona inferior; para vigas T, o banzo superior não ajuda. Usar secção rectangular bw x h.
        section_neg = BeamSection(section.bw_mm, section.h_mm, 0.0, 0.0)
        flex_neg = self.flexural_required(Mneg, section_neg, d_top_initial, d2_initial, fcd, fyd)

        tors = self.torsion_requirements(TEd, section, fck, fcd, fyd)
        Asl_t = float(tors.get("Asl_torsion_req_mm2") or 0.0)
        As_torsion_top = 0.25 * Asl_t
        As_torsion_bot = 0.25 * Asl_t
        As_skin_face = 0.25 * Asl_t
        if h > 1000.0:
            As_skin_face = max(As_skin_face, 0.001 * section.bw_mm * (h - 1000.0) / 2.0)

        As_req_bot = max(As_min_bot, flex_pos["As_req"] + As_torsion_bot)
        As_req_top = max(As_min_top, flex_neg["As_req"] + flex_pos.get("As_comp_req", 0.0) + As_torsion_top)

        bot = self.choose_longitudinal_bars(As_req_bot, bw, max_layers=3)
        top = self.choose_longitudinal_bars(As_req_top, bw, max_layers=3)

        # Recalcular d com centroides reais
        phi_st_for_d = self.choose_stirrup_diameter(max(bot.phi_mm, top.phi_mm, 10.0))
        d_bot = h - bot.centroid_from_edge_mm if bot.n_bars else d_bot_initial
        d_top = h - top.centroid_from_edge_mm if top.n_bars else d_top_initial
        d2_top = top.centroid_from_edge_mm if top.n_bars else d2_initial
        d2_bot = bot.centroid_from_edge_mm if bot.n_bars else d2_initial

        # Capacidades fornecidas com armaduras adoptadas
        Mrd_pos = self.flexural_capacity(bot.area_mm2, top.area_mm2, section, d_bot, d2_top, fcd, fyd)
        Mrd_neg = self.flexural_capacity(top.area_mm2, bot.area_mm2, section_neg, d_top, d2_bot, fcd, fyd)
        eta_m_pos = Mpos / max(Mrd_pos.get("MRd_kNm") or 0.0, 1e-9) if Mpos > 0 else 0.0
        eta_m_neg = Mneg / max(Mrd_neg.get("MRd_kNm") or 0.0, 1e-9) if Mneg > 0 else 0.0

        shear = self.shear_requirements(VEd, bw, min(d_bot, d_top), max(bot.area_mm2, top.area_mm2), fck, fcd, fyd)
        Asw_s_total = max(float(shear["Asw_s_min_mm2_per_mm"]), float(shear["Asw_s_shear_req_mm2_per_mm"])) + float(tors["Asw_s_torsion_req_mm2_per_mm"] or 0.0)
        stir = self.choose_stirrups(Asw_s_total, bw, h, min(d_bot, d_top), torsion=TEd > 1e-9)

        els = self.serviceability(row, bot.area_mm2, top.area_mm2, d_bot, d_top, section, cp)
        det = self.detailing_check(section, bot, top, stir, As_skin_face)

        failure_reasons = []
        if bot.status != "OK" or top.status != "OK":
            failure_reasons.append("armadura longitudinal não cabe")
        if eta_m_pos > 1.0 + 1e-6:
            failure_reasons.append("flexão positiva não verifica")
        if eta_m_neg > 1.0 + 1e-6:
            failure_reasons.append("flexão negativa não verifica")
        if "Não conforme" in str(shear["shear_status"]):
            failure_reasons.append("esforço transverso excede VRd,max")
        if "Não conforme" in str(tors["torsion_status"]):
            failure_reasons.append("torção excede TRd,max")
        if stir.get("stirrup_status") != "OK":
            failure_reasons.append("armadura transversal não cabe")
        status = "OK" if not failure_reasons else "Falha"
        failure_reason = "; ".join(failure_reasons)
        failure_type = ""
        if failure_reasons:
            txt = failure_reason.lower()
            if "flexão" in txt:
                failure_type = "flexao"
            elif "transverso" in txt:
                failure_type = "esforco_transverso"
            elif "torção" in txt:
                failure_type = "torcao"
            elif "cabe" in txt:
                failure_type = "pormenorizacao"
            else:
                failure_type = "outra"

        sol = f"Inf.: {bot.label}; Sup.: {top.label}; Estribos Ø{int(stir['phi_st_mm'])}/{int(stir['stirrup_legs'])}r // {float(stir['s_st_mm'])/10:.1f} cm"
        if As_skin_face > 1e-6:
            sol += f"; pele ≈ {As_skin_face/100:.2f} cm²/face"

        recs = []
        if eta_m_pos > 0.90 or eta_m_neg > 0.90:
            recs.append("avaliar aumento de altura ou armadura longitudinal")
        if "Requer" in str(shear["shear_status"]):
            recs.append("confirmar estribos por zonas junto aos apoios")
        if "torção" in str(tors["torsion_status"]).lower() and TEd > 0:
            recs.append("garantir estribos fechados e armadura longitudinal de torção")
        if det["detailing_status"] != "OK":
            recs.append("rever pormenorização construtiva")

        return {
            "member": row.get("member", ""),
            "case": row.get("case", ""),
            "combination_number": row.get("combination_number", extract_combination_number(row.get("case", ""))),
            "limit_state": row.get("limit_state", classify_limit_state(row.get("case", ""))),
            "name": row.get("name", ""),
            "story": row.get("story", ""),
            "node_i": row.get("node_i", ""),
            "node_j": row.get("node_j", ""),
            "n_points_found": row.get("n_points_found", None),
            "length_m": finite(row.get("length"), 0.0),
            "material": material,
            "material_source": row.get("material_source", "tabela"),
            "section_type": section.section_type,
            "bw_cm": bw / 10.0,
            "h_cm": h / 10.0,
            "bf_cm": bf / 10.0 if bf > 0 else None,
            "hf_cm": hf / 10.0 if hf > 0 else None,
            "cover_mm": self.cover_mm,
            "fck_MPa": fck,
            "fcd_MPa": fcd,
            "fyk_MPa": self.fyk,
            "fyd_MPa": fyd,
            "moment_axis": row.get("moment_axis", "MY"),
            "shear_axis": row.get("shear_axis", "FZ"),
            "torsion_axis": row.get("torsion_axis", "MX"),
            "m_pos_ed_kNm": Mpos,
            "m_pos_at": row.get("m_pos_at", ""),
            "m_neg_ed_kNm": Mneg,
            "m_neg_at": row.get("m_neg_at", ""),
            "v_ed_kN": VEd,
            "v_at": row.get("v_at", ""),
            "t_ed_kNm": TEd,
            "t_at": row.get("t_at", ""),
            "as_min_bot_mm2": As_min_bot,
            "as_min_top_mm2": As_min_top,
            "as_req_bot_mm2": As_req_bot,
            "as_req_top_mm2": As_req_top,
            "as_prov_bot_mm2": bot.area_mm2,
            "as_prov_top_mm2": top.area_mm2,
            "bot_rebar": bot.label,
            "top_rebar": top.label,
            "bot_layers": bot.layers,
            "top_layers": top.layers,
            "d_bot_mm": d_bot,
            "d_top_mm": d_top,
            "mrd_pos_kNm": Mrd_pos.get("MRd_kNm"),
            "mrd_neg_kNm": Mrd_neg.get("MRd_kNm"),
            "eta_m_pos": eta_m_pos,
            "eta_m_neg": eta_m_neg,
            "x_pos_mm": Mrd_pos.get("x_mm"),
            "x_neg_mm": Mrd_neg.get("x_mm"),
            "ductility_pos": flex_pos.get("ductility_status"),
            "ductility_neg": flex_neg.get("ductility_status"),
            "VRd_c_kN": shear.get("VRd_c_kN"),
            "VRd_max_kN": shear.get("VRd_max_kN"),
            "Asw_s_shear_req_mm2_per_m": float(shear.get("Asw_s_shear_req_mm2_per_mm", 0.0)) * 1000.0,
            "Asw_s_min_mm2_per_m": float(shear.get("Asw_s_min_mm2_per_mm", 0.0)) * 1000.0,
            "TRd_max_kNm": tors.get("TRd_max_kNm"),
            "Asw_s_torsion_req_mm2_per_m": float(tors.get("Asw_s_torsion_req_mm2_per_mm", 0.0) or 0.0) * 1000.0,
            "Asl_torsion_req_mm2": tors.get("Asl_torsion_req_mm2"),
            "Asw_s_total_req_mm2_per_m": Asw_s_total * 1000.0,
            "phi_st_mm": stir.get("phi_st_mm"),
            "stirrup_legs": stir.get("stirrup_legs"),
            "s_st_mm": stir.get("s_st_mm"),
            "Asw_s_prov_mm2_per_m": float(stir.get("Asw_s_prov_mm2_per_mm", 0.0)) * 1000.0,
            "shear_status": shear.get("shear_status"),
            "torsion_status": tors.get("torsion_status"),
            "stirrup_status": stir.get("stirrup_status"),
            "skin_reinf_face_mm2": As_skin_face,
            **els,
            **det,
            "status": status,
            "failure_reason": failure_reason,
            "failure_type": failure_type,
            "recommendations": "; ".join(dict.fromkeys(recs)),
            "solution": sol,
            "shortlist_text": f"Bot {bot.label}: As={bot.area_mm2:.0f} mm², layers={bot.layers}; Top {top.label}: As={top.area_mm2:.0f} mm², layers={top.layers}; Stirrups Asw/s={float(stir.get('Asw_s_prov_mm2_per_mm',0))*1000:.0f} mm²/m",
        }

    def design_dataframe(self, df: pd.DataFrame, progress_callback=None) -> pd.DataFrame:
        rows = []
        total = len(df)
        for i, (_, row) in enumerate(df.iterrows(), start=1):
            rows.append(self.design_one(row))
            if progress_callback and (i == total or i % 10 == 0):
                progress_callback(i, total)
        return pd.DataFrame(rows)


# ============================================================
# Validação, resumo e notas
# ============================================================
def build_data_validation(df_clean: pd.DataFrame, df_env: pd.DataFrame, df_results: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    rows = []
    required = ["member", "case", "fx", "fy", "fz", "mx", "my", "mz", "length", "material", "hy", "hz", "name"]
    for c in required:
        ok = df_clean is not None and not df_clean.empty and c in df_clean.columns
        rows.append({"Categoria": "Colunas", "Item": c, "Estado": "OK" if ok else "Não conforme", "Resultado": "presente" if ok else "em falta", "Nota": "coluna reconhecida" if ok else "corrigir cabeçalho"})
    if df_clean is not None and not df_clean.empty:
        rows.append({"Categoria": "Tabela", "Item": "linhas", "Estado": "OK", "Resultado": len(df_clean), "Nota": "linhas importadas"})
        if "member" in df_clean.columns:
            rows.append({"Categoria": "Tabela", "Item": "members", "Estado": "OK", "Resultado": df_clean["member"].astype(str).nunique(), "Nota": "vigas distintas"})
    if df_env is not None and not df_env.empty:
        rows.append({"Categoria": "Envelopes", "Item": "member/case", "Estado": "OK", "Resultado": len(df_env), "Nota": "envelopes criados"})
        single = int((df_env.get("n_points_found", pd.Series(dtype=float)).fillna(0).astype(float) < 2).sum())
        rows.append({"Categoria": "Envelopes", "Item": "casos com <2 pontos", "Estado": "OK" if single == 0 else "Verificar", "Resultado": single, "Nota": "para vigas, recomenda-se exportar estações ao longo da barra"})
        t_missing = int(((df_env.get("bf", pd.Series(dtype=float)).fillna(0) <= 0) | (df_env.get("hf", pd.Series(dtype=float)).fillna(0) <= 0)).sum())
        rows.append({"Categoria": "Secções T", "Item": "BF/HF ausente", "Estado": "Informativo", "Resultado": t_missing, "Nota": "sem BF/HF, a viga é tratada como rectangular"})
    if df_results is not None and not df_results.empty:
        n_fail = int((df_results.get("status", pd.Series(dtype=str)) == "Falha").sum())
        rows.append({"Categoria": "Cálculo", "Item": "falhas", "Estado": "OK" if n_fail == 0 else "Verificar", "Resultado": n_fail, "Nota": "ver separador Falhas"})
    return pd.DataFrame(rows)


def build_normative_notes() -> pd.DataFrame:
    notes = [
        ("Âmbito", "NP EN 1992-1-1", "Vigas de betão armado em ELU e verificações ELS simplificadas."),
        ("Flexão", "EC2 6.1", "Dimensionamento a flexão positiva e negativa; MRd é recalculado com a armadura adoptada."),
        ("Secções T", "EC2 5/6", "Apenas são tratadas como T quando BF/HF são fornecidos; não é gerado banzo colaborante automático."),
        ("Esforço transverso", "EC2 6.2", "Cálculo de VRd,c, VRd,max e Asw/s requerido com modelo de treliça de ângulo variável."),
        ("Torção", "EC2 6.3", "Cálculo de TRd,max, estribos fechados e armadura longitudinal de torção por secção de parede fina equivalente."),
        ("Corte + torção", "EC2 6.2/6.3", "Asw/s transversal resulta da soma da parcela de corte e da parcela de torção, acrescida dos mínimos."),
        ("ELS", "EC2 7", "Controlo expedito de tensão no aço, fendilhação estimada e L/d; casos críticos devem ser confirmados por cálculo detalhado."),
        ("Pormenorização", "EC2 8/9.2", "Controla número mínimo de varões, camadas, espaçamento livre, estribos e armadura de pele quando aplicável."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])


def build_summary_by_member(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame()
    tmp = results.copy()
    tmp["_eta"] = tmp[[c for c in ["eta_m_pos", "eta_m_neg"] if c in tmp.columns]].max(axis=1).fillna(0.0)
    tmp["_vt"] = tmp.get("v_ed_kN", pd.Series(0, index=tmp.index)).abs().fillna(0.0) + tmp.get("t_ed_kNm", pd.Series(0, index=tmp.index)).abs().fillna(0.0)
    tmp = tmp.sort_values(["member", "story", "name", "status", "_eta", "_vt"], ascending=[True, True, True, True, False, False])
    out = tmp.groupby(["member", "story", "name"], dropna=False, as_index=False).first()
    out.drop(columns=["_eta", "_vt"], inplace=True, errors="ignore")
    return out



# ============================================================
# GUI
# ============================================================
class BeamsEC2App(tk.Tk):
    TEMPLATE_COLUMNS = [
        "Member/Node/Case", "Station (m)", "FX (kN)", "FY (kN)", "FZ (kN)", "MX (kNm)", "MY (kNm)", "MZ (kNm)",
        "Length (m)", "Material", "HY (cm)", "HZ (cm)", "BF (cm)", "HF (cm)", "Name", "Story"
    ]

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1280x780")
        self.minsize(980, 620)

        self.df_raw = pd.DataFrame()
        self.df_clean = pd.DataFrame()
        self.df_env = pd.DataFrame()
        self.df_calc_input = pd.DataFrame()
        self.df_results = pd.DataFrame()
        self.df_summary = pd.DataFrame()
        self.df_failures = pd.DataFrame()
        self.df_ok = pd.DataFrame()
        self.df_validation = pd.DataFrame()
        self.df_notes = build_normative_notes()
        self.df_filtered = pd.DataFrame()
        self.input_file_path = ""

        self.var_cover = tk.StringVar(value="35")
        self.var_agg = tk.StringVar(value="20")
        self.var_fyk = tk.StringVar(value="500")
        self.var_gamma_c = tk.StringVar(value="1.50")
        self.var_gamma_s = tk.StringVar(value="1.15")
        self.var_cot_theta = tk.StringVar(value="2.0")
        self.var_crack_limit = tk.StringVar(value="0.30")
        self.var_ld_limit = tk.StringVar(value="20")
        self.var_moment_axis = tk.StringVar(value="MY")
        self.var_shear_axis = tk.StringVar(value="FZ")
        self.var_torsion_axis = tk.StringVar(value="MX")
        self.var_reduce_cases = tk.BooleanVar(value=True)
        self.var_summary = tk.BooleanVar(value=True)
        self.var_calc_mode = tk.StringVar(value="dimensionamento")
        self.var_filter_status = tk.StringVar(value="Todos")
        self.var_filter_fail = tk.StringVar(value="Todos")
        self.var_filter_member = tk.StringVar(value="")

        self.status_var = tk.StringVar(value="Cole ou importe a tabela de esforços.")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_text_var = tk.StringVar(value="0%")
        self.progress_var.trace_add("write", lambda *args: self.progress_text_var.set(f"{self.progress_var.get():.0f}%"))

        self._build_ui()

    # --------------------------- UI ---------------------------
    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        bg = style.lookup("TFrame", "background") or "#f3f5f7"
        self.configure(background=bg)
        style.configure("TLabelframe", padding=8)
        style.configure("TLabelframe.Label", font=("Segoe UI", 9, "bold"))
        style.configure("TButton", padding=(8, 6), font=("Segoe UI", 9))
        style.configure("Primary.TButton", padding=(10, 8), font=("Segoe UI", 9, "bold"))
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TNotebook.Tab", padding=(10, 6), font=("Segoe UI", 9))
        style.configure("Header.TLabel", font=("Segoe UI Semibold", 11))
        style.configure("Subtle.TLabel", font=("Segoe UI", 8), foreground="#5f6b7a")

        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=0)
        self.columnconfigure(0, weight=1)

        root = ttk.Frame(self, padding=8)
        root.grid(row=0, column=0, sticky="nsew")
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        paned = ttk.Panedwindow(root, orient="horizontal")
        paned.grid(row=0, column=0, sticky="nsew")

        sidebar_host = ttk.Frame(paned, width=390)
        sidebar_host.pack_propagate(False)
        sidebar_host.rowconfigure(0, weight=1)
        sidebar_host.columnconfigure(0, weight=1)

        canvas = tk.Canvas(sidebar_host, highlightthickness=0, borderwidth=0, background=bg)
        sb_y = ttk.Scrollbar(sidebar_host, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb_y.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sidebar = ttk.Frame(canvas, padding=(0, 0, 6, 0))
        window = canvas.create_window((0, 0), window=sidebar, anchor="nw")

        def sync(_=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfigure(window, width=canvas.winfo_width())
        sidebar.bind("<Configure>", sync)
        canvas.bind("<Configure>", sync)

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        right = ttk.Frame(paned)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        paned.add(sidebar_host, weight=0)
        paned.add(right, weight=1)

        bottom = ttk.Frame(self, padding=(8, 4, 8, 8))
        bottom.grid(row=1, column=0, sticky="ew")
        bottom.columnconfigure(1, weight=1)
        ttk.Label(bottom, text="Estado:", style="Header.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(bottom, textvariable=self.status_var).grid(row=0, column=1, sticky="ew")
        ttk.Progressbar(bottom, variable=self.progress_var, maximum=100, length=260).grid(row=0, column=2, sticky="e", padx=(12, 6))
        ttk.Label(bottom, textvariable=self.progress_text_var, width=6, anchor="e").grid(row=0, column=3, sticky="e")

        self._build_sidebar(sidebar)
        self._build_tabs(right)
        self.after(120, lambda: paned.sashpos(0, 405))

    def _build_sidebar(self, parent):
        hero = ttk.LabelFrame(parent, text="BeamsEC2")
        hero.pack(fill="x", pady=(0, 8))
        link = ttk.Label(hero, text="BeamsEC2", style="Header.TLabel", cursor="hand2")
        link.pack(anchor="w")
        link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
        ttk.Label(hero, text="Dimensionamento de vigas de betão armado (EC2)", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
        ttk.Label(hero, text="Importa esforços, cria envelopes por viga/caso, dimensiona flexão positiva/negativa, corte, torção, ELS e exporta memória de cálculo/PDF.", style="Subtle.TLabel", wraplength=340, justify="left").pack(anchor="w", pady=(2, 0))

        data = ttk.LabelFrame(parent, text="1. Entrada")
        data.pack(fill="x", pady=(0, 8))
        ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

        params = ttk.LabelFrame(parent, text="2. Parâmetros EC2")
        params.pack(fill="x", pady=(0, 8))
        self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
        self._add_label_entry(params, "Agregado dg [mm]", self.var_agg, 1)
        ttk.Label(params, text="Aço fyk [MPa]").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        self._add_label_entry(params, "γc", self.var_gamma_c, 3)
        self._add_label_entry(params, "γs", self.var_gamma_s, 4)
        self._add_label_entry(params, "cot θ", self.var_cot_theta, 5)
        self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 6)
        self._add_label_entry(params, "L/d limite", self.var_ld_limit, 7)
        ttk.Label(params, text="Modo").grid(row=8, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(params, textvariable=self.var_calc_mode, values=["pre_dimensionamento", "dimensionamento"], state="readonly").grid(row=8, column=1, sticky="ew", padx=6, pady=4)
        ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=9, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=10, column=0, columnspan=2, sticky="w", padx=6, pady=3)
        params.columnconfigure(1, weight=1)

        axes = ttk.LabelFrame(parent, text="3. Eixos locais / esforços")
        axes.pack(fill="x", pady=(0, 8))
        for i, (lab, var, values) in enumerate([
            ("Momento principal", self.var_moment_axis, ["MY", "MZ"]),
            ("Corte vertical", self.var_shear_axis, ["FZ", "FY"]),
            ("Torção", self.var_torsion_axis, ["MX"]),
        ]):
            ttk.Label(axes, text=lab).grid(row=i, column=0, sticky="w", padx=6, pady=4)
            ttk.Combobox(axes, textvariable=var, values=values, state="readonly").grid(row=i, column=1, sticky="ew", padx=6, pady=4)
        axes.columnconfigure(1, weight=1)

        filters = ttk.LabelFrame(parent, text="4. Filtros")
        filters.pack(fill="x", pady=(0, 8))
        ttk.Label(filters, text="Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
        ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "esforco_transverso", "torcao", "pormenorizacao", "dados_incompletos", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
        filters.columnconfigure(1, weight=1)

        actions = ttk.LabelFrame(parent, text="5. Cálculo e exportação")
        actions.pack(fill="x", pady=(0, 8))
        ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=2, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
        actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

        notes = ttk.LabelFrame(parent, text="6. Notas rápidas")
        notes.pack(fill="x", pady=(0, 8))
        ttk.Label(notes, text="• Para vigas, exportar várias estações ao longo da barra melhora o envelope.\n• MY positivo é tratado como flexão positiva/inferior por defeito.\n• Para secções T, preencher BF e HF; caso contrário, a secção é rectangular.\n• A memória de cálculo contém a memória completa; o PDF é sintético.", wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)

    def _add_label_entry(self, parent, label, var, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(parent, textvariable=var, width=14).grid(row=row, column=1, sticky="ew", padx=6, pady=4)

    def _build_tabs(self, parent):
        nb = ttk.Notebook(parent)
        nb.grid(row=0, column=0, sticky="nsew")
        self.tab_instructions = ttk.Frame(nb); self.tab_paste = ttk.Frame(nb); self.tab_input = ttk.Frame(nb)
        self.tab_env = ttk.Frame(nb); self.tab_validation = ttk.Frame(nb); self.tab_results = ttk.Frame(nb)
        self.tab_summary = ttk.Frame(nb); self.tab_failures = ttk.Frame(nb); self.tab_shortlists = ttk.Frame(nb)
        self.tab_report = ttk.Frame(nb); self.tab_notes = ttk.Frame(nb)
        for frame, title in [(self.tab_instructions,"Instruções"),(self.tab_paste,"Colar"),(self.tab_input,"Tabela"),(self.tab_env,"Envelopes"),(self.tab_validation,"Validação"),(self.tab_results,"Resultados"),(self.tab_summary,"Resumo"),(self.tab_failures,"Falhas"),(self.tab_shortlists,"Shortlists"),(self.tab_report,"Relatório"),(self.tab_notes,"Notas EC2")]:
            nb.add(frame, text=title)
        self._build_instructions_tab(self.tab_instructions)
        self._build_paste_tab(self.tab_paste)
        self.tree_input = self._make_tree(self.tab_input)
        self.tree_env = self._make_tree(self.tab_env)
        self.tree_validation = self._make_tree(self.tab_validation)
        self.tree_results = self._make_tree(self.tab_results)
        self.tree_summary = self._make_tree(self.tab_summary)
        self.tree_failures = self._make_tree(self.tab_failures)
        self.tree_shortlists = self._make_tree(self.tab_shortlists)
        self.report_txt = self._make_text_view(self.tab_report)
        self.tree_notes = self._make_tree(self.tab_notes)

    def _build_instructions_tab(self, parent):
        outer = ttk.Frame(parent, padding=10)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="Instruções de utilização do BeamsEC2", style="Header.TLabel").pack(anchor="w", pady=(0, 8))
        txt = self._make_text_view(outer)
        content = (
            "OBJECTIVO DO PROGRAMA\n"
            "BeamsEC2 dimensiona e verifica vigas de betão armado a partir de esforços exportados de software de análise estrutural ou de tabelas. "
            "O cálculo inclui envelopes por viga/caso, flexão positiva e negativa, corte, torção, ELS por cálculo, pormenorização e exportações .xlsx/.pdf.\n\n"
            "COLUNAS RECOMENDADAS\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
            "UNIDADES\nFX, FY, FZ em kN; MX, MY, MZ em kNm; Station e Length em m; HY/HZ/BF/HF em cm.\n\n"
            "CONVENÇÃO\nO eixo X local é longitudinal. Por defeito, MY é o momento principal da viga, FZ é o corte vertical e MX é a torção. "
            "Momentos MY positivos são tratados como flexão positiva, com tracção inferior; momentos negativos como flexão negativa, com tracção superior.\n\n"
            "SECÇÕES T\nPara cálculo como secção T, preencher BF e HF. O programa não calcula automaticamente a largura efectiva do banzo.\n\n"
            "UTILIZAÇÃO PRÁTICA\n1) Exportar esforços por estações ao longo da viga, não apenas nos nós extremos.\n2) Importar ou colar a tabela.\n3) Confirmar eixos locais.\n4) Calcular.\n5) Rever o separador Validação e Falhas.\n6) Exportar a memória de cálculo e o relatório PDF.\n"
        )
        txt.insert("1.0", content); txt.config(state="disabled")

    def _build_paste_tab(self, parent):
        top = ttk.Frame(parent, padding=6); top.pack(fill="x")
        ttk.Label(top, text="Cole aqui a tabela de esforços e clique em 'Ler caixa de texto'.").pack(side="left")
        ttk.Button(top, text="Ler caixa de texto", command=self.load_from_textbox).pack(side="right")
        ttk.Button(top, text="Limpar", command=lambda: self.txt_paste.delete("1.0", "end")).pack(side="right", padx=(0, 6))
        frame = ttk.Frame(parent, padding=(6,0,6,6)); frame.pack(fill="both", expand=True)
        self.txt_paste = tk.Text(frame, wrap="none", undo=True)
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.txt_paste.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.txt_paste.xview)
        self.txt_paste.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.txt_paste.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)

    def _make_text_view(self, parent):
        frame = ttk.Frame(parent); frame.pack(fill="both", expand=True)
        frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)
        txt = tk.Text(frame, wrap="word", undo=False, font=("Segoe UI", 9))
        vsb = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        txt.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns")
        return txt

    def _make_tree(self, parent):
        frame = ttk.Frame(parent); frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)
        return tree

    def show_df(self, tree: ttk.Treeview, df: pd.DataFrame):
        tree.delete(*tree.get_children())
        if df is None or df.empty:
            tree["columns"] = []
            return
        cols = list(df.columns)
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(85, min(240, len(str(c))*9)), anchor="center")
        for _, row in df.head(MAX_PREVIEW_ROWS).iterrows():
            vals = []
            for c in cols:
                v = row[c]
                if isinstance(v, float):
                    vals.append("" if not math.isfinite(v) else f"{v:.3f}")
                else:
                    vals.append("" if pd.isna(v) else str(v))
            tree.insert("", "end", values=vals)

    # --------------------------- importação ---------------------------
    def load_from_textbox(self):
        text = self.txt_paste.get("1.0", "end").strip()
        if not text:
            messagebox.showwarning("Aviso", "Cole primeiro a tabela na caixa de texto.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.load_df(df, source="tabela colada")

    def paste_clipboard(self):
        try:
            text = self.clipboard_get()
        except Exception:
            messagebox.showwarning("Aviso", "Não foi possível ler a área de transferência.")
            return
        df = parse_pasted_table(text)
        if df.empty:
            messagebox.showwarning("Aviso", "A tabela colada não foi reconhecida.")
            return
        self.txt_paste.delete("1.0", "end"); self.txt_paste.insert("1.0", text)
        self.load_df(df, source="área de transferência")

    def import_file(self):
        path = filedialog.askopenfilename(title="Importar tabela", filetypes=[("Ficheiros de tabela", "*.xlsx *.xls"), ("CSV", "*.csv"), ("Todos", "*.*")])
        if not path:
            return
        try:
            if path.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(path, dtype=str)
            else:
                try:
                    df = pd.read_csv(path, dtype=str)
                except Exception:
                    df = pd.read_csv(path, sep=";", dtype=str)
            self.input_file_path = path
            self.load_df(df, source=os.path.basename(path))
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível importar o ficheiro.\n\n{err}")

    def load_df(self, df: pd.DataFrame, source: str = ""):
        self.df_raw = df.copy()
        self.df_clean = clean_dataframe(df)
        self.df_env = build_beam_envelopes(
            self.df_clean,
            moment_axis=self.var_moment_axis.get().lower(),
            shear_axis=self.var_shear_axis.get().lower(),
            torsion_axis=self.var_torsion_axis.get().lower(),
        )
        self.df_calc_input = pd.DataFrame(); self.df_results = pd.DataFrame(); self.df_summary = pd.DataFrame(); self.df_failures = pd.DataFrame(); self.df_ok = pd.DataFrame(); self.df_filtered = pd.DataFrame()
        self.df_validation = build_data_validation(self.df_clean, self.df_env)
        self.df_notes = build_normative_notes()
        self.show_df(self.tree_input, self.df_clean); self.show_df(self.tree_env, self.df_env); self.show_df(self.tree_validation, self.df_validation); self.show_df(self.tree_results, self.df_results); self.show_df(self.tree_summary, self.df_summary); self.show_df(self.tree_failures, self.df_failures); self.show_df(self.tree_shortlists, pd.DataFrame()); self.show_df(self.tree_notes, self.df_notes)
        self.update_report(); self.progress_var.set(0.0)
        self.status_var.set(f"Tabela carregada ({source}): {len(self.df_clean)} linhas; {len(self.df_env)} envelopes member/case.")

    def validate_inputs(self) -> Optional[str]:
        if self.df_env is None or self.df_env.empty:
            return "Cole ou importe uma tabela de esforços primeiro."
        if finite(self.var_cover.get(), 0.0) <= 0:
            return "Recobrimento inválido."
        if finite(self.var_fyk.get(), 0.0) <= 0:
            return "fyk inválido."
        return None

    # --------------------------- cálculo ---------------------------
    def run_design(self):
        err = self.validate_inputs()
        if err:
            messagebox.showwarning("Aviso", err); return
        designer = BeamDesigner(
            cover_mm=finite(self.var_cover.get(), 35.0),
            agg_mm=finite(self.var_agg.get(), 20.0),
            fyk=finite(self.var_fyk.get(), 500.0),
            gamma_c=finite(self.var_gamma_c.get(), 1.5),
            gamma_s=finite(self.var_gamma_s.get(), 1.15),
            cot_theta=finite(self.var_cot_theta.get(), 2.0),
            crack_limit_mm=finite(self.var_crack_limit.get(), 0.30),
            deflection_ld_limit=finite(self.var_ld_limit.get(), 20.0),
            calc_mode=self.var_calc_mode.get(),
        )
        input_df = reduce_to_governing_cases(self.df_env) if self.var_reduce_cases.get() else self.df_env.copy()
        self.df_calc_input = input_df.copy()
        self.progress_var.set(0.0); self.status_var.set("Análise em curso...")

        def progress(done, total):
            pct = 0.0 if total <= 0 else 100.0 * done / total
            self.after(0, lambda p=pct: self.progress_var.set(p))
            self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} envelopes"))

        def worker():
            try:
                results = designer.design_dataframe(input_df, progress_callback=progress)
                summary = build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
                failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
                ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
                validation = build_data_validation(self.df_clean, self.df_env, results)
                def finish():
                    self.df_results = results; self.df_summary = summary; self.df_failures = failures; self.df_ok = ok; self.df_validation = validation; self.df_notes = build_normative_notes(); self.df_filtered = pd.DataFrame()
                    self.show_df(self.tree_results, self.df_results); self.show_df(self.tree_summary, self.df_summary); self.show_df(self.tree_failures, self.df_failures); self.show_df(self.tree_shortlists, self.build_shortlists_df()); self.show_df(self.tree_validation, self.df_validation); self.show_df(self.tree_notes, self.df_notes)
                    self.update_report(); self.progress_var.set(100.0); self.status_var.set(f"Cálculo concluído: {len(results)} envelopes; {len(summary)} vigas resumidas; {len(failures)} falhas.")
                self.after(0, finish)
            except Exception as err:
                msg = str(err)
                self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
                self.after(0, lambda: self.status_var.set("Falha na análise.")); self.after(0, lambda: self.progress_var.set(0.0))
        threading.Thread(target=worker, daemon=True).start()

    # --------------------------- filtros/relatório ---------------------------
    def apply_filters(self):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()
        member = self.var_filter_member.get().strip()
        st = self.var_filter_status.get()
        ft = self.var_filter_fail.get()
        if member:
            df = df[df["member"].astype(str).str.contains(member, case=False, na=False)]
        if st != "Todos":
            df = df[df["status"] == st]
        if ft != "Todos" and "failure_type" in df.columns:
            df = df[df["failure_type"] == ft]
        self.df_filtered = df
        self.show_df(self.tree_results, df)
        self.status_var.set(f"Filtros aplicados: {len(df)} linhas visíveis.")

    def clear_filters(self):
        self.var_filter_member.set(""); self.var_filter_status.set("Todos"); self.var_filter_fail.set("Todos")
        self.df_filtered = pd.DataFrame(); self.show_df(self.tree_results, self.df_results); self.status_var.set("Filtros removidos.")

    def build_shortlists_df(self) -> pd.DataFrame:
        if self.df_results is None or self.df_results.empty:
            return pd.DataFrame()
        cols = ["member", "case", "name", "story", "status", "failure_type", "shortlist_text", "recommendations"]
        return self.df_results[[c for c in cols if c in self.df_results.columns]].copy()

    def update_report(self):
        self.report_txt.delete("1.0", "end")
        if self.df_results is None or self.df_results.empty:
            self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
            return
        source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
        n_total = len(self.df_results)
        n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
        n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
        lines = [f"{APP_NAME} {APP_VERSION}\n", "Relatório resumido de dimensionamento de vigas\n\n", f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n", f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | cotθ: {self.var_cot_theta.get()}\n\n"]
        for _, r in source.head(80).iterrows():
            lines.append(f"Viga {r.get('member','')} | Caso {r.get('case','')} | Piso {r.get('story','')} | {r.get('section_type','')}\n")
            lines.append(f"  M+Ed={finite(r.get('m_pos_ed_kNm')):.2f} kNm | M-Ed={finite(r.get('m_neg_ed_kNm')):.2f} kNm | VEd={finite(r.get('v_ed_kN')):.2f} kN | TEd={finite(r.get('t_ed_kNm')):.2f} kNm\n")
            lines.append(f"  MRd+={finite(r.get('mrd_pos_kNm')):.2f} kNm | MRd-={finite(r.get('mrd_neg_kNm')):.2f} kNm | η+={finite(r.get('eta_m_pos')):.3f} | η-={finite(r.get('eta_m_neg')):.3f}\n")
            lines.append(f"  As inf={finite(r.get('as_prov_bot_mm2')):.0f} mm² ({r.get('bot_rebar','')}) | As sup={finite(r.get('as_prov_top_mm2')):.0f} mm² ({r.get('top_rebar','')})\n")
            lines.append(f"  Solução: {r.get('solution','')} | Estado: {r.get('status','')}\n")
            if str(r.get("failure_reason", "") or "").strip():
                lines.append(f"  Motivo: {r.get('failure_reason','')}\n")
            if str(r.get("recommendations", "") or "").strip():
                lines.append(f"  Recomendações: {r.get('recommendations','')}\n")
            lines.append("\n")
        self.report_txt.insert("1.0", "".join(lines))

    # --------------------------- exportação memória de cálculo/PDF ---------------------------
    def _metadata_df(self) -> pd.DataFrame:
        return pd.DataFrame([
            ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
            ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
            ["Norma de referência", "Eurocódigo 2 / NP EN 1992-1-1"], ["Âmbito", "Dimensionamento/verificação de vigas de betão armado"], ["Descrição", APP_TABLE_DESCRIPTION]
        ], columns=["Campo", "Valor"])

    def _parameters_df(self) -> pd.DataFrame:
        return pd.DataFrame([
            ["Recobrimento [mm]", self.var_cover.get()], ["Agregado dg [mm]", self.var_agg.get()], ["Aço fyk [MPa]", self.var_fyk.get()], ["γc", self.var_gamma_c.get()], ["γs", self.var_gamma_s.get()], ["cotθ", self.var_cot_theta.get()], ["wk,lim [mm]", self.var_crack_limit.get()], ["L/d limite", self.var_ld_limit.get()], ["Momento principal", self.var_moment_axis.get()], ["Corte vertical", self.var_shear_axis.get()], ["Torção", self.var_torsion_axis.get()], ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"]
        ], columns=["Parâmetro", "Valor"])

    def export_excel(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar."); return
        path = filedialog.asksaveasfilename(title="Exportar resultados", defaultextension=".xlsx", filetypes=[("Ficheiro de tabela", "*.xlsx")])
        if not path: return
        if not path.lower().endswith(".xlsx"): path += ".xlsx"
        try:
            self._write_excel(path); self.status_var.set(f"Resultados exportados para: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível exportar.\n\n{err}")

    def _write_excel(self, path: str):
        els_cols = [c for c in ["member","case","combination_number","limit_state","service_status","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_L_over_d","service_crack_status","service_deflection_status"] if c in self.df_results.columns]
        vt_cols = [c for c in ["member","case","v_ed_kN","VRd_c_kN","VRd_max_kN","Asw_s_shear_req_mm2_per_m","t_ed_kNm","TRd_max_kNm","Asw_s_torsion_req_mm2_per_m","Asl_torsion_req_mm2","Asw_s_total_req_mm2_per_m","shear_status","torsion_status"] if c in self.df_results.columns]
        sheets = {
            "00_Info": self._metadata_df(), "01_Parametros": self._parameters_df(), "02_Entrada_Dados": self.df_clean,
            "03_Envelopes": self.df_env, "04_Casos_Calculo": self.df_calc_input, "05_Resultados": self.df_results,
            "06_Resumo_Vigas": self.df_summary, "07_Falhas": self.df_failures, "08_OK": self.df_ok,
            "09_Shortlists": self.build_shortlists_df(), "10_ELS": self.df_results[els_cols].copy() if els_cols else pd.DataFrame(),
            "11_V_Torcao": self.df_results[vt_cols].copy() if vt_cols else pd.DataFrame(), "12_Validacao": self.df_validation, "13_Notas_EC2": self.df_notes,
        }
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, df in sheets.items():
                (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=name[:31], index=False)
            wb = writer.book
            props = wb.properties
            props.title = APP_NAME; props.subject = APP_SUBJECT; props.creator = APP_AUTHOR; props.keywords = APP_KEYWORDS; props.category = APP_CATEGORY; props.description = APP_TABLE_DESCRIPTION; props.lastModifiedBy = APP_AUTHOR
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                header_fill = PatternFill("solid", fgColor="1F4E5F"); header_font = Font(color="FFFFFF", bold=True)
                thin = Side(style="thin", color="D9E2E7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for ws in wb.worksheets:
                    ws.sheet_view.showGridLines = False; ws.freeze_panes = "A2"
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
                    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 5000)):
                        for cell in row:
                            cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                    for col_idx, col in enumerate(ws.columns, start=1):
                        values = [str(c.value) for c in col[:200] if c.value is not None]
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 48)
            except Exception:
                pass

    def export_pdf_report(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning("Aviso", "Não há resultados para exportar."); return
        path = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not path: return
        if not path.lower().endswith(".pdf"): path += ".pdf"
        try:
            self._write_pdf(path); self.status_var.set(f"PDF exportado: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível exportar PDF.\n\n{err}")

    def _write_pdf(self, path: str):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
        doc.title = APP_NAME; doc.author = APP_AUTHOR; doc.subject = APP_SUBJECT
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Title"], alignment=1, fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=colors.HexColor("#1F4E5F")))
        styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=7, leading=9))
        styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=6, leading=8))
        story = [Paragraph(APP_NAME, styles["TitleCenter"]), Paragraph("Relatório resumido de dimensionamento de vigas segundo o Eurocódigo 2", styles["BodyText"]), Spacer(1, 5*mm)]
        n_total=len(self.df_results); n_ok=int((self.df_results["status"]=="OK").sum()) if "status" in self.df_results.columns else 0; n_fail=int((self.df_results["status"]=="Falha").sum()) if "status" in self.df_results.columns else 0
        meta = [["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR], ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Casos", str(n_total)], ["OK", str(n_ok), "Falhas", str(n_fail)], ["fyk", f"{self.var_fyk.get()} MPa", "cotθ", self.var_cot_theta.get()]]
        t = Table(meta, colWidths=[38*mm, 90*mm, 38*mm, 90*mm]); t.setStyle(self._pdf_table_style(header=False)); story.append(t); story.append(Spacer(1, 6*mm))
        summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
        story.append(Paragraph("Resumo por viga", styles["Heading2"]))
        story.append(self._pdf_df_table(summary, ["member","story","case","section_type","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm","bot_rebar","top_rebar","solution","status"], 30))
        if self.df_failures is not None and not self.df_failures.empty:
            story.append(PageBreak()); story.append(Paragraph("Falhas e recomendações", styles["Heading2"])); story.append(self._pdf_df_table(self.df_failures, ["member","story","case","failure_type","failure_reason","recommendations"], 45))
        story.append(Spacer(1, 5*mm)); story.append(Paragraph("Nota: a memória de cálculo exportado contém a memória completa, incluindo envelopes, ELS, corte, torção e pormenorização.", styles["Small"]))
        def footer(canvas, doc_obj):
            canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(APP_NAME); canvas.setSubject(APP_SUBJECT); canvas.setFont("Helvetica",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
        doc.build(story, onFirstPage=footer, onLaterPages=footer)

    def _pdf_table_style(self, header=True):
        from reportlab.lib import colors
        from reportlab.platypus import TableStyle
        cmds = [("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#D9E2E7")), ("VALIGN",(0,0),(-1,-1),"TOP"), ("FONTNAME",(0,0),(-1,-1),"Helvetica"), ("FONTSIZE",(0,0),(-1,-1),7), ("LEFTPADDING",(0,0),(-1,-1),3), ("RIGHTPADDING",(0,0),(-1,-1),3)]
        if header:
            cmds += [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E5F")), ("TEXTCOLOR",(0,0),(-1,0),colors.white), ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]
        return TableStyle(cmds)

    def _pdf_df_table(self, df: pd.DataFrame, cols: List[str], max_rows: int = 30):
        from reportlab.platypus import Table, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        styles = getSampleStyleSheet(); pstyle = ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=6, leading=8)
        present = [c for c in cols if c in df.columns]
        data = [[Paragraph(str(c), pstyle) for c in present]]
        for _, r in df.head(max_rows).iterrows():
            row=[]
            for c in present:
                v = r.get(c, "")
                if isinstance(v, float): txt = "" if not math.isfinite(v) else f"{v:.2f}"
                else: txt = "" if pd.isna(v) else str(v)
                txt = txt.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                row.append(Paragraph(txt, pstyle))
            data.append(row)
        total_width = 270*mm; widths = [total_width/max(1,len(present))]*max(1,len(present))
        t = Table(data, colWidths=widths, repeatRows=1); t.setStyle(self._pdf_table_style(header=True)); return t


    def export_template(self):
        path = filedialog.asksaveasfilename(title="Guardar modelo de importação", defaultextension=".xlsx", filetypes=[("Ficheiro de tabela", "*.xlsx")])
        if not path: return
        if not path.lower().endswith(".xlsx"): path += ".xlsx"
        sample = pd.DataFrame([
            {"Member/Node/Case":"B1/101/ULS_101", "Station (m)":"0.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"120", "MX (kNm)":"5", "MY (kNm)":"-80", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"30", "HZ (cm)":"60", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
            {"Member/Node/Case":"B1/102/ULS_101", "Station (m)":"3.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"20", "MX (kNm)":"2", "MY (kNm)":"95", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"30", "HZ (cm)":"60", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
            {"Member/Node/Case":"B1/103/ULS_101", "Station (m)":"6.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"-120", "MX (kNm)":"5", "MY (kNm)":"-75", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"30", "HZ (cm)":"60", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
        ])
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                sample.to_excel(writer, sheet_name="FOLHA_IMPORTACAO_TIPO", index=False)
                writer.book.properties.title = f"{APP_NAME} - folha de importação tipo"; writer.book.properties.creator = APP_AUTHOR
            self.status_var.set(f"Modelo de tabela guardado: {path}")
        except Exception as err:
            messagebox.showerror("Erro", f"Não foi possível guardar o modelo.\n\n{err}")



