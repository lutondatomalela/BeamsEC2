# -*- coding: utf-8 -*-
"""Geometrias rectangulares, T e I."""

from . import reporting as _previous
globals().update({k: v for k, v in vars(_previous).items() if not k.startswith("__")})
APP_VERSION = "v0.1.1"

# Geometrias rectangulares, T e I
# ============================================================
# ============================================================

# Colunas adicionais reconhecidas. As propriedades de secção são lidas da
# tabela de propriedades da secção; a geometria explícita é opcional e prevalece sobre a
# reconstrução automática.
COLUMN_ALIASES.update({
    "vy_sec": ["vy (cm)", "vy sec (cm)", "vy_geom", "vy section"],
    "vz_sec": ["vz (cm)", "vz sec (cm)", "vz_geom", "vz section"],
    "vpy_sec": ["vpy (cm)", "vpy sec (cm)", "vpy_geom", "vpy section"],
    "vpz_sec": ["vpz (cm)", "vpz sec (cm)", "vpz_geom", "vpz section"],
    "ax": ["ax (cm2)", "ax (cm²)", "ax", "area (cm2)", "area (cm²)"],
    "ay": ["ay (cm2)", "ay (cm²)", "ay"],
    "az": ["az (cm2)", "az (cm²)", "az"],
    "ix": ["ix (cm4)", "ix (cm⁴)", "ix"],
    "iy": ["iy (cm4)", "iy (cm⁴)", "iy"],
    "iz": ["iz (cm4)", "iz (cm⁴)", "iz"],
    "section_type_input": ["section type", "section_type", "tipo secção", "tipo de secção", "geometria"],
    "b_top": ["b top (cm)", "b_top (cm)", "bsup (cm)", "b superior (cm)", "largura banzo superior (cm)"],
    "tf_top": ["tf top (cm)", "tf_top (cm)", "tfsup (cm)", "espessura banzo superior (cm)"],
    "b_bottom": ["b bottom (cm)", "b_bottom (cm)", "binf (cm)", "b inferior (cm)", "largura banzo inferior (cm)"],
    "tf_bottom": ["tf bottom (cm)", "tf_bottom (cm)", "tfinf (cm)", "espessura banzo inferior (cm)"],
    "tw": ["tw (cm)", "bw web (cm)", "espessura alma (cm)", "alma (cm)"],
    "i_top": ["i top", "topo i", "orientação i", "orientacao i"],
})

_EXTRA_NUMERIC_Geometry = [
    "vy_sec", "vz_sec", "vpy_sec", "vpz_sec", "ax", "ay", "az", "ix", "iy", "iz",
    "b_top", "tf_top", "b_bottom", "tf_bottom", "tw",
]
_EXTRA_GEOM_Geometry = _EXTRA_NUMERIC_Geometry + ["section_type_input", "i_top"]

_clean_dataframe_full_reports = clean_dataframe

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = _clean_dataframe_full_reports(df)
    for c in _EXTRA_NUMERIC_Geometry:
        if c in out.columns:
            out[c] = out[c].map(safe_float)
        else:
            out[c] = float("nan")
    for c in ["section_type_input", "i_top"]:
        if c not in out.columns:
            out[c] = ""
    return out


_build_beam_envelopes_full_reports = build_beam_envelopes

def build_beam_envelopes(df: pd.DataFrame, moment_axis="my", shear_axis="fz", torsion_axis="mx") -> pd.DataFrame:
    env = _build_beam_envelopes_full_reports(df, moment_axis=moment_axis, shear_axis=shear_axis, torsion_axis=torsion_axis)
    if env is None or env.empty or df is None or df.empty:
        return env
    group_cols = ["member", "case", "name", "story"]
    first = df.sort_values("__row_order").groupby(group_cols, dropna=False).first().reset_index()
    keep = group_cols + [c for c in _EXTRA_GEOM_Geometry if c in first.columns]
    first = first[keep]
    env = env.merge(first, on=group_cols, how="left", suffixes=("", "_geom"))
    # A classificação é visível já no separador Envelopes, antes do cálculo.
    types=[]; summaries=[]; sources=[]; confidences=[]; errors=[]; notes=[]
    for _, rr in env.iterrows():
        try:
            sec=_section_from_row_geometry(rr)
            types.append(sec.section_type); summaries.append(sec.geometry_summary); sources.append(sec.source)
            confidences.append(sec.confidence); errors.append(sec.fit_error_pct); notes.append(sec.orientation_note)
        except Exception as exc:
            types.append("Não identificada"); summaries.append(""); sources.append("erro")
            confidences.append("Baixa"); errors.append(None); notes.append(str(exc))
    env["section_type"]=types; env["section_geometry_summary"]=summaries; env["section_geometry_source"]=sources
    env["section_geometry_confidence"]=confidences; env["section_fit_error_pct"]=errors; env["section_orientation_note"]=notes
    return env


@dataclass
class SectionProfileGeometry:
    h_mm: float
    web_mm: float
    top_width_mm: float
    top_thickness_mm: float
    bottom_width_mm: float
    bottom_thickness_mm: float
    kind: str = "Rectangular"
    source: str = "HY/HZ"
    confidence: str = "Alta"
    fit_error_pct: float = 0.0
    orientation_note: str = ""

    @property
    def bw_mm(self) -> float:
        return self.web_mm

    @property
    def bf_mm(self) -> float:
        return self.top_width_mm if self.kind != "Rectangular" else 0.0

    @property
    def hf_mm(self) -> float:
        return self.top_thickness_mm if self.kind != "Rectangular" else 0.0

    @property
    def is_t(self) -> bool:
        return self.kind.startswith("T")

    @property
    def is_i(self) -> bool:
        return self.kind.startswith("I")

    @property
    def section_type(self) -> str:
        return self.kind

    @property
    def outer_width_mm(self) -> float:
        return max(self.top_width_mm, self.bottom_width_mm, self.web_mm)

    @property
    def web_height_mm(self) -> float:
        return max(self.h_mm - self.top_thickness_mm - self.bottom_thickness_mm, 0.0)

    @property
    def segments(self):
        if self.kind == "Rectangular":
            return [(0.0, self.h_mm, self.web_mm)]
        segs = []
        z = 0.0
        if self.top_thickness_mm > 0:
            segs.append((z, z + self.top_thickness_mm, self.top_width_mm))
            z += self.top_thickness_mm
        if self.web_height_mm > 0:
            segs.append((z, z + self.web_height_mm, self.web_mm))
            z += self.web_height_mm
        if self.bottom_thickness_mm > 0:
            segs.append((z, self.h_mm, self.bottom_width_mm))
        return segs

    def width_at_depth(self, z_from_top_mm: float) -> float:
        z = min(max(float(z_from_top_mm), 0.0), self.h_mm)
        for z0, z1, b in self.segments:
            if z >= z0 - 1e-9 and z <= z1 + 1e-9:
                return b
        return self.web_mm

    def flipped(self):
        note = self.orientation_note
        if note:
            note += "; secção invertida para momento negativo"
        return SectionProfileGeometry(
            self.h_mm, self.web_mm,
            self.bottom_width_mm, self.bottom_thickness_mm,
            self.top_width_mm, self.top_thickness_mm,
            self.kind, self.source, self.confidence, self.fit_error_pct, note,
        )

    def gross_properties(self):
        A = 0.0
        Sz = 0.0
        for z0, z1, b in self.segments:
            a = b * (z1 - z0)
            A += a
            Sz += a * (z0 + z1) / 2.0
        if A <= 0:
            return 0.0, self.h_mm / 2.0, 0.0, 0.0
        zbar = Sz / A
        Iy = 0.0
        Iz = 0.0
        for z0, z1, b in self.segments:
            t = z1 - z0
            a = b * t
            zc = (z0 + z1) / 2.0
            Iy += b * t ** 3 / 12.0 + a * (zc - zbar) ** 2
            Iz += t * b ** 3 / 12.0
        return A, zbar, Iy, Iz

    @property
    def geometry_summary(self) -> str:
        if self.kind == "Rectangular":
            return f"Rectangular {self.web_mm/10:.1f}x{self.h_mm/10:.1f} cm"
        if self.is_t:
            return (f"T: bf={self.top_width_mm/10:.1f} cm; hf={self.top_thickness_mm/10:.1f} cm; "
                    f"bw={self.web_mm/10:.1f} cm; h={self.h_mm/10:.1f} cm")
        return (f"{self.kind}: Bsup={self.top_width_mm/10:.1f} cm; tfsup={self.top_thickness_mm/10:.1f} cm; "
                f"bw={self.web_mm/10:.1f} cm; Binf={self.bottom_width_mm/10:.1f} cm; "
                f"tfinf={self.bottom_thickness_mm/10:.1f} cm; h={self.h_mm/10:.1f} cm")


def _i_props_geometry(B, H, tt, tb, tw):
    hw = H - tt - tb
    if min(B, H, tt, tb, tw, hw) <= 0:
        return None
    A1, z1 = B * tt, tt / 2.0
    A2, z2 = tw * hw, tt + hw / 2.0
    A3, z3 = B * tb, H - tb / 2.0
    A = A1 + A2 + A3
    zbar = (A1*z1 + A2*z2 + A3*z3) / A
    Iy = (B*tt**3/12.0 + A1*(z1-zbar)**2 +
          tw*hw**3/12.0 + A2*(z2-zbar)**2 +
          B*tb**3/12.0 + A3*(z3-zbar)**2)
    Iz = tt*B**3/12.0 + hw*tw**3/12.0 + tb*B**3/12.0
    return A, zbar, Iy, Iz


def _infer_equal_flange_i_geometry(B, H, A_t, z_t, Iy_t, Iz_t):
    """Reconstrói uma secção I com banzos de igual largura e espessuras distintas.

    A solução é obtida por Gauss-Newton amortecido e validada contra A, posição do
    centro de gravidade, Iy e Iz. Não é aceite quando o ajustamento não é inequívoco.
    """
    try:
        import numpy as np
    except Exception:
        return None
    targets = np.array([A_t, z_t, Iy_t, Iz_t], dtype=float)
    if not np.all(np.isfinite(targets)) or min(B, H, A_t, Iy_t, Iz_t) <= 0:
        return None
    scales = np.array([max(A_t,1.0), max(H,1.0), max(Iy_t,1.0), max(Iz_t,1.0)])

    def residual(x):
        tt, tb, tw = [float(v) for v in x]
        p = _i_props_geometry(B,H,tt,tb,tw)
        if p is None or tw >= B or tt+tb >= 0.90*H:
            return np.ones(4)*1e3
        return (np.array(p)-targets)/scales

    starts = [
        [0.14*H,0.14*H,0.25*B], [0.10*H,0.18*H,0.30*B],
        [0.18*H,0.10*H,0.30*B], [0.20*H,0.20*H,0.15*B],
        [0.08*H,0.08*H,0.40*B],
    ]
    best = None
    for start in starts:
        x = np.array(start, dtype=float)
        lam = 1e-3
        for _ in range(120):
            r = residual(x)
            eps = np.array([max(0.05,abs(v)*1e-4) for v in x])
            Jcols = []
            for j in range(3):
                xp = x.copy(); xp[j] += eps[j]
                Jcols.append((residual(xp)-r)/eps[j])
            J = np.column_stack(Jcols)
            try:
                dx = np.linalg.solve(J.T@J + lam*np.eye(3), -J.T@r)
            except Exception:
                break
            trial = x + dx
            trial[0] = np.clip(trial[0], 5.0, 0.42*H)
            trial[1] = np.clip(trial[1], 5.0, 0.42*H)
            trial[2] = np.clip(trial[2], 5.0, 0.95*B)
            if trial[0]+trial[1] > 0.88*H:
                fac = 0.88*H/(trial[0]+trial[1]); trial[0]*=fac; trial[1]*=fac
            if np.linalg.norm(residual(trial)) < np.linalg.norm(r):
                x = trial; lam = max(lam/2.0,1e-9)
            else:
                lam = min(lam*10.0,1e9)
            if np.linalg.norm(dx) < 1e-6:
                break
        rv = residual(x)
        score = float(np.linalg.norm(rv))
        if best is None or score < best[0]:
            best = (score, x, rv)
    if best is None:
        return None
    _, x, rv = best
    tt,tb,tw = [float(v) for v in x]
    p = _i_props_geometry(B,H,tt,tb,tw)
    if p is None:
        return None
    rel = [abs((p[i]-targets[i])/max(abs(targets[i]),1.0))*100.0 for i in range(4)]
    maxerr = max(rel)
    if maxerr > 2.0 or tw >= 0.80*B or min(tt,tb) < 5.0:
        return None
    return {"tt":tt,"tb":tb,"tw":tw,"fit_error_pct":maxerr,"errors_pct":rel}


def _section_from_row_geometry(row: pd.Series) -> SectionProfileGeometry:
    B = cm_to_mm(row.get("hy",0.0))
    H = cm_to_mm(row.get("hz",0.0))
    if B <= 0 or H <= 0:
        raise ValueError("dimensões HY/HZ inválidas")
    st = normalize_text(row.get("section_type_input", ""))

    # Geometria I explícita: preferida quando todas as dimensões são fornecidas.
    bt = cm_to_mm(row.get("b_top", float("nan")))
    tt = cm_to_mm(row.get("tf_top", float("nan")))
    bb = cm_to_mm(row.get("b_bottom", float("nan")))
    tb = cm_to_mm(row.get("tf_bottom", float("nan")))
    tw = cm_to_mm(row.get("tw", float("nan")))
    explicit_i = all(math.isfinite(v) and v > 0 for v in [bt,tt,bb,tb,tw]) and tt+tb < H and tw <= max(bt,bb)
    if explicit_i or st in {"i", "i assimetrica", "i assimétrica", "duplo t", "double t"}:
        if not explicit_i:
            raise ValueError("secção I indicada, mas faltam Bsup/tfsup/Binf/tfinf/tw")
        kind = "I assimétrica" if abs(bt-bb)>1.0 or abs(tt-tb)>1.0 else "I simétrica"
        return SectionProfileGeometry(H,tw,bt,tt,bb,tb,kind,"geometria explícita", "Alta",0.0,
                                  "topo conforme colunas de geometria")

    # Secção T explícita continua suportada.
    bf = cm_to_mm(row.get("bf", float("nan")))
    hf = cm_to_mm(row.get("hf", float("nan")))
    if math.isfinite(bf) and math.isfinite(hf) and bf > B and 0 < hf < H:
        return SectionProfileGeometry(H,B,bf,hf,B,0.0,"T","BF/HF", "Alta",0.0,"")

    # Verificação rectangular pelas propriedades importadas.
    A = finite(row.get("ax",float("nan")),float("nan"))*100.0
    Iy = finite(row.get("iy",float("nan")),float("nan"))*10000.0
    Iz = finite(row.get("iz",float("nan")),float("nan"))*10000.0
    Arect = B*H; Iyrect=B*H**3/12.0; Izrect=H*B**3/12.0
    if all(math.isfinite(v) and v>0 for v in [A,Iy,Iz]):
        rect_err = max(abs(A-Arect)/Arect, abs(Iy-Iyrect)/Iyrect, abs(Iz-Izrect)/Izrect)
        if rect_err <= 0.03:
            return SectionProfileGeometry(H,B,B,0.0,B,0.0,"Rectangular","propriedades AX/IY/IZ", "Alta",rect_err*100.0,"")

    # Reconstrução automática I assimétrica com banzos de igual largura externa.
    vz = finite(row.get("vz_sec",float("nan")),float("nan"))*10.0
    vpz = finite(row.get("vpz_sec",float("nan")),float("nan"))*10.0
    top_sel = normalize_text(row.get("i_top", ""))
    if top_sel in {"vpz", "-z", "negativo", "inferior"}:
        ztop = vpz
        orient = "topo associado a VPZ por indicação da tabela"
    else:
        ztop = vz
        orient = "topo associado a VZ; confirmar orientação do eixo local Z"
    if all(math.isfinite(v) and v>0 for v in [A,Iy,Iz,ztop]) and A < 0.88*Arect:
        inferred = _infer_equal_flange_i_geometry(B,H,A,ztop,Iy,Iz)
        if inferred:
            tt,tb,tw = inferred["tt"],inferred["tb"],inferred["tw"]
            kind = "I assimétrica" if abs(tt-tb) > max(1.0,0.005*H) else "I simétrica"
            conf = "Alta" if inferred["fit_error_pct"] <= 0.5 else "Média"
            return SectionProfileGeometry(H,tw,B,tt,B,tb,kind,"reconstrução AX/IY/IZ/VZ/VPZ",conf,
                                      inferred["fit_error_pct"],orient)

    # Não assumir uma geometria não rectangular sem ajustamento confiável.
    return SectionProfileGeometry(H,B,B,0.0,B,0.0,"Rectangular","HY/HZ (fallback)","Baixa",0.0,
                              "propriedades insuficientes ou incompatíveis para classificação automática")


def _compression_block_profile_geometry(self, a_mm: float, section, fcd: float):
    if not isinstance(section, SectionProfileGeometry):
        return _compression_block_full_reports(self, a_mm, section, fcd)
    a = max(0.0,min(float(a_mm),section.h_mm))
    C = 0.0; Sy = 0.0
    for z0,z1,b in section.segments:
        t = max(0.0,min(z1,a)-z0)
        if t <= 0:
            continue
        area = b*t
        yc = z0+t/2.0
        force = area*fcd
        C += force; Sy += force*yc
    return C, (Sy/C if C>0 else 0.0)

_compression_block_full_reports = BeamDesigner.compression_block
BeamDesigner.compression_block = _compression_block_profile_geometry

_section_second_moment_y_full_reports = _section_second_moment_y_reporting

def _section_second_moment_y_reporting(section) -> float:
    if isinstance(section, SectionProfileGeometry):
        return max(section.gross_properties()[2],1.0)
    return _section_second_moment_y_full_reports(section)


_torsion_requirements_full_reports = BeamDesigner.torsion_requirements

def _torsion_requirements_profile_geometry(self, TEd_kNm, section, fck, fcd, fyd):
    if isinstance(section, SectionProfileGeometry) and section.is_i:
        # Para secções I abertas/concavas, a envolvente BxH seria não conservadora.
        # Adopta-se a alma como rectângulo equivalente e o relatório explicita a hipótese.
        proxy = BeamSection(section.web_mm, section.h_mm,0.0,0.0)
        out = _torsion_requirements_full_reports(self,TEd_kNm,proxy,fck,fcd,fyd)
        out["torsion_geometry_model"] = "rectângulo equivalente da alma (conservador)"
        return out
    out = _torsion_requirements_full_reports(self,TEd_kNm,section,fck,fcd,fyd)
    out["torsion_geometry_model"] = "secção efectiva corrente"
    return out

BeamDesigner.torsion_requirements = _torsion_requirements_profile_geometry


def _choose_face_rebar_geometry(self, As_req, section: SectionProfileGeometry, face: str) -> RebarChoice:
    req = max(float(As_req or 0.0),0.0)
    phi_st = 8.0
    candidates = []
    diameter_sets = [[12.0,16.0,20.0],[25.0]]
    for diameters in diameter_sets:
        for phi in diameters:
            aphi = bar_area_mm2(phi)
            n0 = max(2,int(math.ceil(req/max(aphi,1e-9))))
            for n in range(n0,min(n0+10,24)):
                area=n*aphi
                for layers in [1,2]:
                    if layers==2 and n<4:
                        continue
                    counts=[n] if layers==1 else [(n+1)//2,n//2]
                    if min(counts)<2:
                        continue
                    req_clear=self.min_clear_spacing(phi)
                    layer_depths=[]; clears=[]; widths=[]; ok=True
                    for i,count in enumerate(counts):
                        depth=self.cover_mm+phi_st+phi/2.0+i*(phi+req_clear)
                        z=depth if face=="top" else section.h_mm-depth
                        if depth >= section.h_mm-self.cover_mm or z<0 or z>section.h_mm:
                            ok=False; break
                        width=section.width_at_depth(z)
                        avail=width-2.0*(self.cover_mm+phi_st)
                        if count==1:
                            clear=avail-phi
                        else:
                            clear=(avail-count*phi)/(count-1)
                        if clear+1e-9<req_clear:
                            ok=False; break
                        layer_depths.append(depth); clears.append(clear); widths.append(width)
                    if not ok:
                        continue
                    centroid=sum(c*d for c,d in zip(counts,layer_depths))/n
                    # Preferir uma camada, menos varões e menor excesso. Ø25 só entra
                    # quando as gamas correntes não produzem solução.
                    score=(layers,n,area-req,phi)
                    reb=RebarChoice(n,phi,area,layers,max(counts),centroid,min(clears),"OK")
                    reb.layer_counts=counts
                    reb.layer_widths_mm=widths
                    reb.layout="+".join(str(c) for c in counts)
                    candidates.append((score,reb))
        if candidates:
            break
    if candidates:
        candidates.sort(key=lambda x:x[0])
        return candidates[0][1]
    fail=RebarChoice(0,0.0,0.0,0,0,0.0,0.0,"Não cabe")
    fail.layer_counts=[]; fail.layer_widths_mm=[]; fail.layout="-"
    return fail


def _detailing_profile_geometry(self, section, bot, top, stir):
    issues=[]
    for label,reb in [("inferior",bot),("superior",top)]:
        if reb.status!="OK": issues.append(f"armadura {label} não cabe na geometria real")
        if reb.n_bars and reb.n_bars<2: issues.append(f"armadura {label}: mínimo de 2 varões")
        if reb.layers>2: issues.append(f"armadura {label}: mais de 2 camadas")
        if reb.clear_spacing_mm and reb.clear_spacing_mm<self.min_clear_spacing(reb.phi_mm)-1e-9:
            issues.append(f"armadura {label}: espaçamento livre insuficiente")
    if finite(stir.get("stirrup_legs"),0)<2: issues.append("estribos com menos de 2 ramos")
    if stir.get("stirrup_status")!="OK": issues.append("armadura transversal não cabe")
    if finite(stir.get("s_st_mm"),0)>finite(stir.get("s_lim_mm"),9999)+1e-9:
        issues.append("espaçamento dos estribos superior ao limite")
    return {"detailing_status":"OK" if not issues else "Não conforme",
            "detailing_issues":"; ".join(issues) if issues else "-"}


def _design_one_geometry(self, row: pd.Series) -> Dict:
    material=str(row.get("material",DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
    fck=parse_concrete_strength(material)
    cp=concrete_props(fck,alpha_cc=self.alpha_cc,gamma_c=self.gamma_c)
    fyd=steel_props(self.fyk,gamma_s=self.gamma_s)["fyd"]
    try:
        section=_section_from_row_geometry(row)
    except Exception as exc:
        return {"member":row.get("member",""),"case":row.get("case",""),"name":row.get("name",""),"story":row.get("story",""),
                "status":"Falha","failure_reason":f"Geometria inválida: {exc}","failure_type":"dados_incompletos"}
    h=section.h_mm; bw=section.web_mm
    phi0=16.0; phi_st0=8.0
    d0=h-self.cover_mm-phi_st0-phi0/2.0; d20=self.cover_mm+phi_st0+phi0/2.0
    if d0<=0:
        return {"member":row.get("member",""),"case":row.get("case",""),"name":row.get("name",""),"story":row.get("story",""),
                "status":"Falha","failure_reason":"Geometria incompatível com o recobrimento","failure_type":"dados_incompletos"}
    Mpos=finite(row.get("m_pos_ed_kNm"),0.0); Mneg=finite(row.get("m_neg_ed_kNm"),0.0)
    VEd=finite(row.get("v_ed_kN"),0.0); TEd=finite(row.get("t_ed_kNm"),0.0)
    As_min_bot=self.as_min_beam(bw,d0,cp["fctm"],self.fyk)
    As_min_top=self.as_min_beam(bw,d0,cp["fctm"],self.fyk)
    flex_pos=self.flexural_required(Mpos,section,d0,d20,cp["fcd"],fyd)
    section_neg=section.flipped()
    flex_neg=self.flexural_required(Mneg,section_neg,d0,d20,cp["fcd"],fyd)
    tors=self.torsion_requirements(TEd,section,fck,cp["fcd"],fyd)
    torsion_active=str(tors.get("torsion_considered","Não"))=="Sim"
    Asl_t=float(tors.get("Asl_torsion_req_mm2") or 0.0)
    if torsion_active and h>=SKIN_REINF_MIN_HEIGHT_MM:
        Atop=Abot=0.25*Asl_t; As_skin_req=0.25*Asl_t
        As_skin_req=max(As_skin_req,0.001*bw*max(h-SKIN_REINF_MIN_HEIGHT_MM,0.0)/2.0)
        skin_note="Aplicável: h >= 40 cm"
    elif torsion_active:
        Atop=Abot=0.50*Asl_t; As_skin_req=0.0; skin_note="Não aplicável: h < 40 cm"
    elif h>SKIN_REINF_MIN_HEIGHT_MM:
        Atop=Abot=0.0; As_skin_req=0.001*bw*(h-SKIN_REINF_MIN_HEIGHT_MM)/2.0
        skin_note="Armadura de pele mínima para h > 40 cm"
    else:
        Atop=Abot=As_skin_req=0.0; skin_note="Não aplicável"
    skin_choice=_choose_skin_rebar_detailing(As_skin_req)
    As_req_bot=max(As_min_bot,flex_pos["As_req"]+Abot)
    As_req_top=max(As_min_top,flex_neg["As_req"]+flex_pos.get("As_comp_req",0.0)+Atop)
    bot=_choose_face_rebar_geometry(self,As_req_bot,section,"bottom")
    top=_choose_face_rebar_geometry(self,As_req_top,section,"top")
    d_bot=h-bot.centroid_from_edge_mm if bot.n_bars else d0
    d_top=h-top.centroid_from_edge_mm if top.n_bars else d0
    d2_top=top.centroid_from_edge_mm if top.n_bars else d20
    d2_bot=bot.centroid_from_edge_mm if bot.n_bars else d20
    Mrd_pos=self.flexural_capacity(bot.area_mm2,top.area_mm2,section,d_bot,d2_top,cp["fcd"],fyd)
    Mrd_neg=self.flexural_capacity(top.area_mm2,bot.area_mm2,section_neg,d_top,d2_bot,cp["fcd"],fyd)
    eta_pos=Mpos/max(finite(Mrd_pos.get("MRd_kNm")),1e-9) if Mpos>0 else 0.0
    eta_neg=Mneg/max(finite(Mrd_neg.get("MRd_kNm")),1e-9) if Mneg>0 else 0.0
    shear=self.shear_requirements(VEd,bw,min(d_bot,d_top),max(bot.area_mm2,top.area_mm2),fck,cp["fcd"],fyd)
    Asw_total=max(float(shear["Asw_s_min_mm2_per_mm"]),float(shear["Asw_s_shear_req_mm2_per_mm"]))+float(tors.get("Asw_s_torsion_req_mm2_per_mm") or 0.0)
    stir=self.choose_stirrups(Asw_total,bw,h,min(d_bot,d_top),torsion=torsion_active)
    els=self.serviceability(row,bot.area_mm2,top.area_mm2,d_bot,d_top,section,cp)
    det=_detailing_profile_geometry(self,section,bot,top,stir)
    failures=[]
    if bot.status!="OK" or top.status!="OK": failures.append("armadura longitudinal não cabe na geometria real")
    if eta_pos>1.0+1e-6: failures.append("flexão positiva não verifica")
    if eta_neg>1.0+1e-6: failures.append("flexão negativa não verifica")
    if "Não conforme" in str(shear.get("shear_status","")): failures.append("esforço transverso excede VRd,max")
    if "Não conforme" in str(tors.get("torsion_status","")): failures.append("torção excede TRd,max")
    if stir.get("stirrup_status")!="OK": failures.append("armadura transversal não cabe")
    if det["detailing_status"]!="OK": failures.append("pormenorização não conforme")
    if section.confidence=="Baixa" and section.kind!="Rectangular": failures.append("geometria não confirmada")
    status="OK" if not failures else "Falha"
    reason="; ".join(dict.fromkeys(failures))
    ftype=""
    if failures:
        low=reason.lower()
        ftype="flexao" if "flexão" in low else ("esforco_transverso" if "transverso" in low else ("torcao" if "torção" in low else ("pormenorizacao" if "pormenoriz" in low or "cabe" in low else "dados")))
    stir_label="Estribos fechados" if torsion_active else "Estribos"
    sol=f"Inf.: {bot.label}; Sup.: {top.label}; {stir_label} Ø{int(stir['phi_st_mm'])}/{int(stir['stirrup_legs'])}r // {float(stir['s_st_mm'])/10:.1f} cm"
    if As_skin_req>1e-6: sol+=f"; pele/alma: {skin_choice['skin_rebar']}"
    Acalc,zbar,Iycalc,Izcalc=section.gross_properties()
    eta_s=VEd/max(finite(shear.get("VRd_max_kN")),1e-9) if VEd>0 else 0.0
    eta_t=TEd/max(finite(tors.get("TRd_max_kNm")),1e-9) if TEd>0 else 0.0
    out={
        "member":row.get("member",""),"case":row.get("case",""),"combination_number":row.get("combination_number",extract_combination_number(row.get("case",""))),
        "limit_state":row.get("limit_state",classify_limit_state(row.get("case",""))),"name":row.get("name",""),"story":row.get("story",""),
        "node_i":row.get("node_i",""),"node_j":row.get("node_j",""),"n_points_found":row.get("n_points_found"),"length_m":finite(row.get("length"),0.0),
        "material":material,"material_source":row.get("material_source","tabela"),"section_type":section.section_type,"section_geometry_source":section.source,
        "section_geometry_confidence":section.confidence,"section_fit_error_pct":section.fit_error_pct,"section_orientation_note":section.orientation_note,
        "section_geometry_summary":section.geometry_summary,"bw_cm":bw/10.0,"h_cm":h/10.0,"b_top_cm":section.top_width_mm/10.0,
        "tf_top_cm":section.top_thickness_mm/10.0,"b_bottom_cm":section.bottom_width_mm/10.0,"tf_bottom_cm":section.bottom_thickness_mm/10.0,
        "web_height_cm":section.web_height_mm/10.0,"ax_calc_cm2":Acalc/100.0,"iy_calc_cm4":Iycalc/10000.0,"iz_calc_cm4":Izcalc/10000.0,
        "centroid_from_top_cm":zbar/10.0,"ax_import_cm2":finite(row.get("ax",float("nan")),float("nan")),"iy_import_cm4":finite(row.get("iy",float("nan")),float("nan")),
        "iz_import_cm4":finite(row.get("iz",float("nan")),float("nan")),"cover_mm":self.cover_mm,"fck_MPa":fck,"fcd_MPa":cp["fcd"],"fyk_MPa":self.fyk,"fyd_MPa":fyd,
        "moment_axis":row.get("moment_axis","MY"),"shear_axis":row.get("shear_axis","FZ"),"torsion_axis":row.get("torsion_axis","MX"),
        "m_pos_ed_kNm":Mpos,"m_pos_at":row.get("m_pos_at",""),"m_neg_ed_kNm":Mneg,"m_neg_at":row.get("m_neg_at",""),
        "v_ed_kN":VEd,"v_at":row.get("v_at",""),"t_ed_kNm":TEd,"t_at":row.get("t_at",""),
        "as_min_bot_mm2":As_min_bot,"as_min_top_mm2":As_min_top,"as_req_bot_mm2":As_req_bot,"as_req_top_mm2":As_req_top,
        "as_prov_bot_mm2":bot.area_mm2,"as_prov_top_mm2":top.area_mm2,"bot_rebar":bot.label,"top_rebar":top.label,
        "bot_layout":getattr(bot,"layout",""),"top_layout":getattr(top,"layout",""),"bot_layers":bot.layers,"top_layers":top.layers,
        "bot_bars_per_layer":bot.bars_per_layer,"top_bars_per_layer":top.bars_per_layer,"bot_clear_spacing_mm":bot.clear_spacing_mm,"top_clear_spacing_mm":top.clear_spacing_mm,
        "d_bot_mm":d_bot,"d_top_mm":d_top,"mrd_pos_kNm":Mrd_pos.get("MRd_kNm"),"mrd_neg_kNm":Mrd_neg.get("MRd_kNm"),
        "eta_m_pos":eta_pos,"eta_m_neg":eta_neg,"x_pos_mm":Mrd_pos.get("x_mm"),"x_neg_mm":Mrd_neg.get("x_mm"),
        "ductility_pos":flex_pos.get("ductility_status"),"ductility_neg":flex_neg.get("ductility_status"),
        "VRd_c_kN":shear.get("VRd_c_kN"),"VRd_max_kN":shear.get("VRd_max_kN"),"cot_theta_shear":shear.get("cot_theta_shear"),
        "Asw_s_shear_req_mm2_per_m":float(shear.get("Asw_s_shear_req_mm2_per_mm",0.0))*1000.0,"Asw_s_min_mm2_per_m":float(shear.get("Asw_s_min_mm2_per_mm",0.0))*1000.0,
        "TRd_max_kNm":tors.get("TRd_max_kNm"),"cot_theta_torsion":tors.get("cot_theta_torsion"),"torsion_geometry_model":tors.get("torsion_geometry_model"),
        "Asw_s_torsion_req_mm2_per_m":float(tors.get("Asw_s_torsion_req_mm2_per_mm",0.0) or 0.0)*1000.0,"Asl_torsion_req_mm2":tors.get("Asl_torsion_req_mm2"),
        "torsion_considered":tors.get("torsion_considered"),"eta_torsion_design":tors.get("eta_torsion_design"),"torsion_relevance_limit":TORSION_RELEVANCE_ETA_Detailing,
        "Asw_s_total_req_mm2_per_m":Asw_total*1000.0,"phi_st_mm":stir.get("phi_st_mm"),"stirrup_legs":stir.get("stirrup_legs"),"s_st_mm":stir.get("s_st_mm"),
        "Asw_s_prov_mm2_per_m":float(stir.get("Asw_s_prov_mm2_per_mm",0.0))*1000.0,"shear_status":shear.get("shear_status"),"torsion_status":tors.get("torsion_status"),
        "stirrup_status":stir.get("stirrup_status"),"skin_reinf_face_mm2":As_skin_req,"skin_reinf_face_prov_mm2":float(skin_choice.get("skin_area_prov_face_mm2",0.0) or 0.0),
        "skin_rebar":skin_choice.get("skin_rebar"),"skin_n_per_face":skin_choice.get("skin_n_per_face"),"skin_phi_mm":skin_choice.get("skin_phi_mm"),
        "skin_reinf_threshold_cm":SKIN_REINF_MIN_HEIGHT_MM/10.0,"skin_reinf_note":skin_note,"norma":NORMATIVE_SUPPORT,
        "design_basis":"Geometria real identificada; suporte normativo fixo para Portugal","calc_version":APP_VERSION,
        "flexure_method":"NP EN 1992-1-1:2010, 6.1; integração por segmentos da secção e MRd com armadura adoptada",
        "shear_method":"NP EN 1992-1-1:2010, 6.2; largura resistente da alma","torsion_method":"NP EN 1992-1-1:2010, 6.3; modelo conservador pela alma para secções I",
        "sls_method":"NP EN 1992-1-1:2010, Secção 7; ELS expedito com Iy da geometria identificada","detailing_method":"NP EN 1992-1-1:2010, Secções 8 e 9.2; espaçamentos verificados na largura local de cada camada",
        "eta_flexure_max":max(eta_pos,eta_neg),"eta_shear_max":eta_s,"eta_torsion_max":eta_t,"eta_global":max(eta_pos,eta_neg,eta_s,eta_t),
        **els,**det,"status":status,"failure_reason":reason,"failure_type":ftype,"recommendations":"","solution":sol,
        "shortlist_text":f"Inf {bot.label} ({getattr(bot,'layout','')}); Sup {top.label} ({getattr(top,'layout','')}); geometria {section.geometry_summary}",
    }
    out["beam_ref"]=_beam_label_labels(out)
    return out

BeamDesigner.design_one = _design_one_geometry


# Validação adicional da geometria identificada.
_build_data_validation_full_reports = build_data_validation

def _build_data_validation_geometry(df_clean: pd.DataFrame, df_env: pd.DataFrame, df_results: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    base=_build_data_validation_full_reports(df_clean,df_env,df_results)
    rows=[]
    env=df_env if df_env is not None else pd.DataFrame()
    if not env.empty:
        kinds={}; low=0; failed=0
        for _,r in env.iterrows():
            try:
                sec=_section_from_row_geometry(r)
                kinds[sec.section_type]=kinds.get(sec.section_type,0)+1
                if sec.confidence=="Baixa": low+=1
            except Exception:
                failed+=1
        for k,v in kinds.items():
            rows.append({"Categoria":"Geometria","Item":k,"Estado":"OK","Resultado":v,"Nota":"secções identificadas"})
        rows.append({"Categoria":"Geometria","Item":"confiança baixa","Estado":"OK" if low==0 else "Verificar","Resultado":low,"Nota":"confirmar geometria antes do dimensionamento final"})
        rows.append({"Categoria":"Geometria","Item":"não identificadas","Estado":"OK" if failed==0 else "Não conforme","Resultado":failed,"Nota":"preencher geometria explícita"})
    extra=pd.DataFrame(rows)
    return pd.concat([base,extra],ignore_index=True) if base is not None and not base.empty else extra

build_data_validation = _build_data_validation_geometry


# Memória de cálculo com geometria real.
def _calc_memory_df_geometry(results: pd.DataFrame) -> pd.DataFrame:
    cols=["Viga","Caso","Piso","Secção","Etapa","Item","Valor","Unidade","Critério/Referência","Estado/Nota"]
    rows=[]
    if results is None or results.empty: return pd.DataFrame(columns=cols)
    for _,r in results.iterrows():
        base={"Viga":_beam_label_labels(r),"Caso":r.get("case",""),"Piso":r.get("story",""),
              "Secção":r.get("section_geometry_summary",f"{_fmt_report_base(r.get('bw_cm'),0)}x{_fmt_report_base(r.get('h_cm'),0)} cm")}
        def add(e,i,v,u,c,n=""): rows.append({**base,"Etapa":e,"Item":i,"Valor":v,"Unidade":u,"Critério/Referência":c,"Estado/Nota":n})
        add("Geometria","Tipo",r.get("section_type",""),"-","identificação automática/entrada",f"{r.get('section_geometry_source','')}; confiança {r.get('section_geometry_confidence','')}")
        if str(r.get("section_type","")).startswith("I"):
            add("Geometria","Bsup / tfsup",f"{_fmt_report_base(r.get('b_top_cm'))} / {_fmt_report_base(r.get('tf_top_cm'))}","cm","propriedades da secção",r.get("section_orientation_note",""))
            add("Geometria","bw / Binf / tfinf",f"{_fmt_report_base(r.get('bw_cm'))} / {_fmt_report_base(r.get('b_bottom_cm'))} / {_fmt_report_base(r.get('tf_bottom_cm'))}","cm","propriedades da secção",f"erro máx.={_fmt_report_base(r.get('section_fit_error_pct'),3)}%")
        add("Dados","Material",r.get("material",""),"-","NP EN 1992-1-1, Secção 3",r.get("material_source",""))
        add("Dados","Comprimento",finite(r.get("length_m")),"m","entrada")
        add("Esforços","M+Ed",finite(r.get("m_pos_ed_kNm")),"kNm","envelope",f"x={_fmt_report_base(r.get('m_pos_at'))} m")
        add("Esforços","M-Ed",finite(r.get("m_neg_ed_kNm")),"kNm","envelope",f"x={_fmt_report_base(r.get('m_neg_at'))} m")
        add("Esforços","VEd",finite(r.get("v_ed_kN")),"kN","envelope",f"x={_fmt_report_base(r.get('v_at'))} m")
        add("Esforços","TEd",finite(r.get("t_ed_kNm")),"kNm","envelope",f"x={_fmt_report_base(r.get('t_at'))} m")
        add("Flexão +","As,req / As,prov",f"{_fmt_report_base(r.get('as_req_bot_mm2'),0)} / {_fmt_report_base(r.get('as_prov_bot_mm2'),0)}","mm²","NP EN 1992-1-1, 6.1 / 9.2",f"{r.get('bot_rebar','')}; MRd={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_pos'),3)}")
        add("Flexão -","As,req / As,prov",f"{_fmt_report_base(r.get('as_req_top_mm2'),0)} / {_fmt_report_base(r.get('as_prov_top_mm2'),0)}","mm²","NP EN 1992-1-1, 6.1 / 9.2",f"{r.get('top_rebar','')}; MRd={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_neg'),3)}")
        add("Corte","VRd,c / VRd,max",f"{_fmt_report_base(r.get('VRd_c_kN'))} / {_fmt_report_base(r.get('VRd_max_kN'))}","kN","NP EN 1992-1-1, 6.2",r.get("shear_status",""))
        add("Torção","TEd / TRd,max",f"{_fmt_report_base(r.get('t_ed_kNm'))} / {_fmt_report_base(r.get('TRd_max_kNm'))}","kNm","NP EN 1992-1-1, 6.3",f"{r.get('torsion_status','')}; {r.get('torsion_geometry_model','')}")
        add("Corte/Torção","Asw/s req / prov",f"{_fmt_report_base(r.get('Asw_s_total_req_mm2_per_m'),0)} / {_fmt_report_base(r.get('Asw_s_prov_mm2_per_m'),0)}","mm²/m","NP EN 1992-1-1, 6.2 + 6.3")
        add("ELS","wk",f"{_fmt_report_base(r.get('service_wk_est_mm'),3)} / {_fmt_report_base(r.get('service_wk_lim_mm'),3)}","mm","NP EN 1992-1-1, 7.3",r.get("service_crack_status",""))
        add("ELS","Flecha estimada / limite",f"{_fmt_report_base(r.get('service_deflection_est_mm'),2)} / {_fmt_report_base(r.get('service_deflection_lim_mm'),2)}","mm",r.get("service_deflection_limit","L/250"),r.get("service_deflection_status",""))
        add("Pormenorização","Solução",r.get("solution",""),"-","NP EN 1992-1-1, Secções 8 e 9.2",r.get("detailing_status",""))
        add("Resultado","Estado final",r.get("status",""),"-","síntese",r.get("failure_reason", ""))
    return pd.DataFrame(rows,columns=cols)

globals()["_calc_memory_df_labels"]=_calc_memory_df_geometry
globals()["_calc_memory_df_reporting"]=_calc_memory_df_geometry
globals()["_calc_memory_df_report_base"]=_calc_memory_df_geometry


# Instruções actualizadas, sem redundância.
def _build_instructions_tab_geometry(self,parent):
    outer=ttk.Frame(parent,padding=10); outer.pack(fill="both",expand=True); outer.rowconfigure(1,weight=1); outer.columnconfigure(0,weight=1)
    ttk.Label(outer,text="Instruções de utilização",style="Header.TLabel").grid(row=0,column=0,sticky="w",pady=(0,8))
    host=ttk.Frame(outer); host.grid(row=1,column=0,sticky="nsew"); txt=self._make_text_view(host)
    content=(
        "OBJECTIVO\n"
        f"Dimensionamento e verificação de vigas de betão armado segundo {NORMATIVE_SUPPORT_EXTENDED}.\n\n"
        "ENTRADA\n"
        "Importe uma tabela por estações com Member/Node/Case, esforços, comprimento, material, HY/HZ e identificação da viga. "
        "FX/FY/FZ são introduzidos em kN; MX/MY/MZ em kNm; comprimentos em m; propriedades geométricas em cm, cm² e cm⁴.\n\n"
        "GEOMETRIA DA SECÇÃO\n"
        "Rectangular: HY/HZ e propriedades compatíveis. T: preencher BF/HF. I: o programa reconstrói automaticamente uma secção I com banzos de igual largura a partir de HY, HZ, AX, IY, IZ, VZ e VPZ. "
        "Para geometrias I gerais ou para dispensar a reconstrução, preencher B Top, TF Top, B Bottom, TF Bottom e TW. A orientação adoptada é indicada nos resultados e deve ser confirmada com os eixos locais do modelo.\n\n"
        "CÁLCULO\n"
        "Confirme o eixo do momento principal, do corte vertical e da torção. O programa cria envelopes por viga/caso, dimensiona flexão positiva e negativa, corte, torção, ELS e pormenorização. "
        "Em secções I, a flexão usa a geometria segmentada, o corte usa a largura da alma e a torção usa um rectângulo equivalente da alma, de forma conservadora.\n\n"
        "ELS E EXPORTAÇÃO\n"
        "A combinação ELS é opcional. O limite de flecha é introduzido como L/n. O memória de cálculo mantém a auditoria completa; os PDFs são exportados separadamente como Resumo executivo, Relatório técnico ou Memória de cálculo."
    )
    txt.insert("1.0",content); txt.config(state="disabled")

BeamsEC2App._build_instructions_tab=_build_instructions_tab_geometry


_build_normative_notes_global_full_reports = build_normative_notes

def _geometry_notes_geometry():
    return pd.DataFrame([
        ("Geometria","Secções I","Reconstrução automática por AX/IY/IZ/VZ/VPZ ou utilização de dimensões explícitas dos banzos e da alma."),
        ("Flexão","Secção segmentada","O bloco comprimido é integrado na largura real dos banzos e da alma, em ambos os sentidos de flexão."),
        ("Corte","Alma","Em secções I, bw corresponde à espessura da alma."),
        ("Torção","Hipótese conservadora","Em secções I, a verificação usa um rectângulo equivalente da alma; a hipótese fica identificada no relatório."),
    ],columns=["Tema","Referência","Nota"])

def build_normative_notes():
    base=_build_normative_notes_global_full_reports()
    extra=_geometry_notes_geometry()
    return pd.concat([base,extra],ignore_index=True) if base is not None and not base.empty else extra

def _build_normative_notes_geometry(self):
    return build_normative_notes()

BeamsEC2App.build_normative_notes=_build_normative_notes_geometry


# Relatório interno com geometria completa.
_update_report_labels_base=BeamsEC2App.update_report

def _update_report_geometry(self):
    _update_report_labels_base(self)
    if self.df_results is None or self.df_results.empty:
        return
    text=self.report_txt.get("1.0","end")
    # A informação completa já fica na solução e nas tabelas; acrescentar uma linha
    # curta por viga para tornar a identificação da secção inequívoca.
    src=self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    appendix="\nGEOMETRIA IDENTIFICADA\n"
    for _,r in src.head(120).iterrows():
        appendix+=f"{_beam_label_labels(r)}: {r.get('section_geometry_summary',r.get('section_type',''))} | origem: {r.get('section_geometry_source','')} | confiança: {r.get('section_geometry_confidence','')}\n"
    self.report_txt.insert("end",appendix)

BeamsEC2App.update_report=_update_report_geometry


# Sidebar: actualizar apenas a nota rápida, preservando a interface validada.
_build_sidebar_geometry_base = BeamsEC2App._build_sidebar

def _build_sidebar_geometry(self, parent):
    _build_sidebar_geometry_base(self, parent)
    def walk(widget):
        for child in widget.winfo_children():
            try:
                txt=str(child.cget("text"))
                if "Para secções T, preencher BF e HF" in txt or "Para secções T, preencher BF/HF" in txt:
                    child.configure(text=(
                        f"• Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}.\n"
                        "• Secções rectangulares e T são reconhecidas pelas dimensões; secções I podem ser reconstruídas por AX/IY/IZ/VZ/VPZ ou definidas explicitamente.\n"
                        "• Confirmar a orientação do eixo local Z nas secções I assimétricas.\n"
                        "• Em secções I, o corte usa a alma e a torção adopta um modelo conservador pela alma.\n"
                        "• A armadura de alma/pele é reportada a partir de h >= 40 cm."
                    ))
            except Exception:
                pass
            walk(child)
    walk(parent)

BeamsEC2App._build_sidebar = _build_sidebar_geometry


# O botão "Modelo de tabela" exporta a versão  incluída no pacote.
def _export_template_geometry(self):
    path=filedialog.asksaveasfilename(title="Guardar modelo de importação",defaultextension=".xlsx",filetypes=[("Ficheiro de tabela","*.xlsx")])
    if not path: return
    if not path.lower().endswith(".xlsx"): path += ".xlsx"
    try:
        import shutil
        bundled=Path(__file__).with_name("BeamsEC2_template.xlsx")
        if bundled.exists():
            shutil.copyfile(bundled,path)
        else:
            raise FileNotFoundError("modelo incluído no pacote não encontrado")
        self.status_var.set(f"Modelo de tabela guardado: {path}")
    except Exception as err:
        messagebox.showerror("Erro",f"Não foi possível guardar o modelo.\n\n{err}")

BeamsEC2App.export_template = _export_template_geometry


# memória de cálculo: acrescentar uma folha dedicada à geometria identificada.
_write_excel_geometry_base = BeamsEC2App._write_excel

def _write_excel_geometry(self, path: str):
    _write_excel_geometry_base(self,path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb=load_workbook(path)
        if "05A_Geometria" in wb.sheetnames:
            del wb["05A_Geometria"]
        ws=wb.create_sheet("05A_Geometria",6)
        src=self.df_results if self.df_results is not None and not self.df_results.empty else self.df_env
        cols=["member","name","story","case","section_type","section_geometry_summary","section_geometry_source","section_geometry_confidence","section_fit_error_pct","section_orientation_note","bw_cm","h_cm","b_top_cm","tf_top_cm","b_bottom_cm","tf_bottom_cm","web_height_cm","ax_import_cm2","ax_calc_cm2","iy_import_cm4","iy_calc_cm4","iz_import_cm4","iz_calc_cm4"]
        present=[c for c in cols if c in src.columns]
        ws.append(present)
        for _,rr in src[present].iterrows():
            ws.append([None if pd.isna(rr[c]) else rr[c] for c in present])
        fill=PatternFill("solid",fgColor="1F4E5F"); thin=Side(style="thin",color="D9E2E7"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
        for cell in ws[1]:
            cell.fill=fill; cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border=border; cell.alignment=Alignment(vertical="top",wrap_text=True)
        ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
        for j,col in enumerate(present,1):
            vals=[str(ws.cell(i,j).value) for i in range(1,min(ws.max_row,200)+1) if ws.cell(i,j).value is not None]
            ws.column_dimensions[get_column_letter(j)].width=min(max([len(v) for v in vals]+[10])+2,45)
        wb.save(path)
    except Exception:
        pass

BeamsEC2App._write_excel = _write_excel_geometry


# PDF técnico: acrescentar uma tabela própria de geometria antes dos envelopes.
_write_pdf_geometry_base = BeamsEC2App._write_pdf

def _write_pdf_geometry(self, path: str):
    # Para resumo e memória, a versão base já contém section_type/geometry_summary.
    # No relatório técnico, gerar temporariamente uma página de geometria no início
    # através da rotina completa abaixo.
    if not hasattr(self,"var_pdf_scope") or self.var_pdf_scope.get() != "Relatório técnico":
        return _write_pdf_geometry_base(self,path)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles=_pdf_styles_report_base()
    results=self.df_results if self.df_results is not None else pd.DataFrame()
    summary=self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures=self.df_failures if self.df_failures is not None else pd.DataFrame()
    doc=SimpleDocTemplate(path,pagesize=landscape(A4),rightMargin=12*mm,leftMargin=12*mm,topMargin=12*mm,bottomMargin=12*mm)
    doc.title=APP_NAME; doc.author=""; doc.subject=APP_SUBJECT
    story=[_program_link_pdf_reporting(styles)]
    n_total=len(results); n_ok=int((results.get("status",pd.Series(dtype=str))=="OK").sum()); n_fail=int((results.get("status",pd.Series(dtype=str))=="Falha").sum())
    meta=[["Data",datetime.now().strftime("%Y-%m-%d %H:%M"),"Suporte normativo","NP EN 1992-1-1:2010"],["Envelopes",str(n_total),"OK/Falhas",f"{n_ok}/{n_fail}"]]
    t=Table(meta,colWidths=[38*mm,90*mm,38*mm,105*mm]); t.setStyle(self._pdf_table_style(header=False)); story += [t,Spacer(1,5*mm),Paragraph("Relatório técnico",styles["Section"])]
    geom=summary.copy()
    if geom is not None and not geom.empty:
        geom=geom.assign(viga=geom.apply(_beam_label_labels,axis=1))
    story += [Paragraph("Geometria identificada",styles["BodyCourier"]),self._pdf_df_table(geom,["viga","story","section_type","section_geometry_summary","section_geometry_source","section_geometry_confidence","section_fit_error_pct"],max_rows=None),Spacer(1,5*mm)]
    env=self.df_env.copy() if self.df_env is not None else pd.DataFrame()
    if not env.empty: env=env.assign(viga=env.apply(_beam_label_labels,axis=1))
    story += [Paragraph("Critérios de cálculo",styles["BodyCourier"]),self._pdf_df_table(self._parameters_df(),["Parâmetro","Valor"],max_rows=None,widths=[90,180]),Spacer(1,5*mm),
              Paragraph("Envelopes de esforços",styles["BodyCourier"]),self._pdf_df_table(env,["viga","story","case","section_type","n_points_found","length","material","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Flexão",styles["BodyCourier"]),self._pdf_df_table(_flexure_audit_df_calculation(summary),["viga","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Esforço transverso e torção",styles["BodyCourier"]),self._pdf_df_table(_vt_audit_df_calculation(summary),["viga","case","v_ed_kN","VRd_c_kN","VRd_max_kN","cot_theta_shear","t_ed_kNm","TRd_max_kNm","cot_theta_torsion","torsion_considered","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"],max_rows=None),Spacer(1,5*mm),
              Paragraph("ELS",styles["BodyCourier"]),self._pdf_df_table(_sls_audit_df_calculation(summary),["viga","case","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_deflection_est_mm","service_deflection_lim_mm","service_deflection_limit","service_status","service_note"],max_rows=None),Spacer(1,5*mm),
              Paragraph("Pormenorização",styles["BodyCourier"]),self._pdf_df_table(_detailing_audit_df_calculation(summary),["viga","case","bot_rebar","top_rebar","bot_clear_spacing_mm","top_clear_spacing_mm","phi_st_mm","stirrup_legs","s_st_mm","skin_rebar","detailing_status","detailing_issues"],max_rows=None)]
    if failures is not None and not failures.empty:
        story += [PageBreak(),Paragraph("Falhas",styles["Section"]),self._pdf_df_table(failures.assign(viga=failures.apply(_beam_label_labels,axis=1)),["viga","story","case","failure_type","failure_reason"],max_rows=None)]
    footer_date=datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas,doc_obj):
        canvas.saveState(); canvas.setTitle(APP_NAME); canvas.setAuthor(""); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm,7*mm,f"{APP_NAME} | {footer_date}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_geometry


