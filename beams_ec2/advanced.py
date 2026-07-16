# -*- coding: utf-8 -*-
"""Condições de apoio, histórico e validação."""

from . import serviceability as _previous
globals().update({k: v for k, v in vars(_previous).items() if not k.startswith("__")})
APP_VERSION = "v0.1.1"

# Condições de apoio, histórico e validação
# ============================================================
# histórico de carregamento, armadura longitudinal e validação
# ============================================================
APP_TABLE_DESCRIPTION = (
    "Workbook técnico de cálculo de vigas de betão armado com ELU, envelope integral de combinações ELS, "
    "curvaturas, fluência, retracção, condições de apoio, histórico de carregamento, "
    "armadura longitudinal por estações, pormenorização e validação automática."
)

# Colunas opcionais para a utilização avançada. Mantém-se compatibilidade total
# com os modelos anteriores.
COLUMN_ALIASES.update({
    "support_condition": [
        "support condition", "boundary condition", "bc", "support", "supports",
        "condição de apoio", "condicao de apoio", "apoio", "apoios",
    ],
    "as_bot_local_mm2": [
        "as bot local (mm2)", "as_bot_local_mm2", "as inf local (mm2)",
        "as inferior local (mm2)", "as bot (mm2)",
    ],
    "as_top_local_mm2": [
        "as top local (mm2)", "as_top_local_mm2", "as sup local (mm2)",
        "as superior local (mm2)", "as top (mm2)",
    ],
    "bot_rebar_local": [
        "bot rebar local", "armadura inferior local", "arm inf local", "rebar bot local",
    ],
    "top_rebar_local": [
        "top rebar local", "armadura superior local", "arm sup local", "rebar top local",
    ],
})

_clean_dataframe_serviceability = clean_dataframe

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = _clean_dataframe_serviceability(df)
    for c in ["as_bot_local_mm2", "as_top_local_mm2"]:
        if c in out.columns:
            out[c] = out[c].map(safe_float)
        else:
            out[c] = float("nan")
    for c in ["support_condition", "bot_rebar_local", "top_rebar_local"]:
        if c not in out.columns:
            out[c] = ""
    return out

# Actualizar as colunas apresentadas nas instruções/modelo.
BeamsEC2App.TEMPLATE_COLUMNS = [
    "Member/Node/Case", "Station (m)", "FX (kN)", "FY (kN)", "FZ (kN)",
    "MX (kNm)", "MY (kNm)", "MZ (kNm)", "Length (m)", "Material",
    "HY (cm)", "HZ (cm)", "BF (cm)", "HF (cm)", "Name", "Story",
    "Support Condition", "As Bot Local (mm2)", "As Top Local (mm2)",
    "Bot Rebar Local", "Top Rebar Local",
    "VY (cm)", "VZ (cm)", "VPY (cm)", "VPZ (cm)", "AX (cm2)",
    "AY (cm2)", "AZ (cm2)", "IX (cm4)", "IY (cm4)", "IZ (cm4)",
    "Section Type", "B Top (cm)", "TF Top (cm)", "B Bottom (cm)",
    "TF Bottom (cm)", "TW (cm)", "I Top",
]


@dataclass
class SLSParametersAdvanced(SLSParametersServiceability):
    auto_regime: bool = True
    support_condition: str = "Automática / por tabela"
    history_mode: str = "Uma idade efectiva"
    permanent_fraction_pct: float = 70.0
    t0_permanent_days: float = 14.0
    t0_variable_days: float = 365.0


def _case_regime_advanced(case_value, fallback: str) -> str:
    s = normalize_text(str(case_value or ""))
    if any(k in s for k in ["quase", "cqc", "q.p", " qp", "qp "]):
        return "Quase-permanente / longo prazo"
    if any(k in s for k in ["frequente", "freq", "repetida", "repet"]):
        return "Frequente / repetida"
    if any(k in s for k in ["rara", "caracteristica", "characteristic", "char"]):
        return "Característica / curta duração"
    return fallback


def _params_for_case_advanced(base: SLSParametersAdvanced, case_value) -> SLSParametersAdvanced:
    return SLSParametersAdvanced(
        regime=_case_regime_advanced(case_value, base.regime) if base.auto_regime else base.regime,
        rh_pct=base.rh_pct,
        t0_days=base.t0_days,
        t_days=base.t_days,
        ts_days=base.ts_days,
        cement_class=base.cement_class,
        h0_override_mm=base.h0_override_mm,
        include_shrinkage=base.include_shrinkage,
        deflection_limit_denominator=base.deflection_limit_denominator,
        crack_limit_mm=base.crack_limit_mm,
        auto_regime=base.auto_regime,
        support_condition=base.support_condition,
        history_mode=base.history_mode,
        permanent_fraction_pct=base.permanent_fraction_pct,
        t0_permanent_days=base.t0_permanent_days,
        t0_variable_days=base.t0_variable_days,
    )


def _creep_history_advanced(cp: Dict[str, float], params: SLSParametersAdvanced, h0: float) -> Dict[str, float]:
    if not params.long_term:
        return {"phi": 0.0, "phi_g": 0.0, "phi_q": 0.0, "fraction_g": 0.0, "mode": "curta duração"}
    if "duas" not in normalize_text(params.history_mode):
        phi = _ec2_creep_coefficient_serviceability(
            cp["fcm"], params.rh_pct, h0, params.t0_days, params.t_days, params.cement_class
        )
        return {"phi": phi, "phi_g": phi, "phi_q": phi, "fraction_g": 1.0, "mode": "uma idade efectiva"}
    fg = min(max(params.permanent_fraction_pct / 100.0, 0.0), 1.0)
    phi_g = _ec2_creep_coefficient_serviceability(
        cp["fcm"], params.rh_pct, h0, params.t0_permanent_days, params.t_days, params.cement_class
    ) if params.t_days > params.t0_permanent_days else 0.0
    phi_q = _ec2_creep_coefficient_serviceability(
        cp["fcm"], params.rh_pct, h0, params.t0_variable_days, params.t_days, params.cement_class
    ) if params.t_days > params.t0_variable_days else 0.0
    phi = fg * phi_g + (1.0 - fg) * phi_q
    return {"phi": phi, "phi_g": phi_g, "phi_q": phi_q, "fraction_g": fg, "mode": "duas fases G+Q"}


def _normalize_support_advanced(value: str) -> str:
    s = normalize_text(str(value or ""))
    if "direita" in s and ("consola" in s or "cantilever" in s or "encastr" in s):
        return "Consola - encastrada à direita"
    if "esquerda" in s and ("consola" in s or "cantilever" in s or "encastr" in s):
        return "Consola - encastrada à esquerda"
    if "consola" in s or "cantilever" in s:
        return "Consola - encastrada à esquerda"
    if ("encastre" in s or "fixed" in s) and ("duplo" in s or "ambos" in s or "fixed-fixed" in s or "encastre-encastre" in s):
        return "Encastre-encastre"
    if "encastre-encastre" in s:
        return "Encastre-encastre"
    if "entre apoios" in s or "biapoi" in s or "simples" in s or "pinned" in s or "support-support" in s:
        return "Entre apoios (y0=yL=0)"
    if "autom" in s or "tabela" in s or not s:
        return "Automática / por tabela"
    return "Entre apoios (y0=yL=0)"


def _resolve_support_advanced(app, group: pd.DataFrame, params: SLSParametersAdvanced) -> Tuple[str, str]:
    if group is not None and not group.empty and "support_condition" in group.columns:
        vals = [str(v).strip() for v in group["support_condition"].tolist() if str(v).strip() and str(v).strip().lower() not in {"nan", "none"}]
        if vals:
            support = _normalize_support_advanced(vals[0])
            if support != "Automática / por tabela":
                return support, "tabela importada"
    support = _normalize_support_advanced(params.support_condition)
    if support == "Automática / por tabela":
        support = "Entre apoios (y0=yL=0)"
        return support, "fallback: vão entre apoios"
    return support, "parâmetro global"


def _integrate_curvature_advanced(x_mm, kappa, support_condition: str):
    import numpy as np
    x = np.asarray(x_mm, dtype=float)
    k = np.asarray(kappa, dtype=float)
    if len(x) == 0:
        return np.array([]), np.array([]), {"disp_residual_mm": 0.0, "rotation_residual_rad": 0.0}
    theta0 = np.zeros_like(x)
    y0 = np.zeros_like(x)
    for i in range(1, len(x)):
        dx = x[i] - x[i - 1]
        theta0[i] = theta0[i - 1] + 0.5 * (k[i] + k[i - 1]) * dx
        y0[i] = y0[i - 1] + 0.5 * (theta0[i] + theta0[i - 1]) * dx
    xr = x - x[0]
    L = max(float(xr[-1]), 1e-9)
    support = _normalize_support_advanced(support_condition)
    if support == "Consola - encastrada à esquerda":
        c0 = -y0[0]
        c1 = -theta0[0]
        relevant_y = [0]
        relevant_t = [0]
    elif support == "Consola - encastrada à direita":
        c1 = -theta0[-1]
        c0 = -y0[-1] - c1 * L
        relevant_y = [-1]
        relevant_t = [-1]
    elif support == "Encastre-encastre":
        # Ajuste de mínimos quadrados, ponderando rotações pela extensão do vão.
        A = np.array([[1.0, 0.0], [0.0, L], [1.0, L], [0.0, L]], dtype=float)
        b = np.array([-y0[0], -theta0[0] * L, -y0[-1], -theta0[-1] * L], dtype=float)
        c0, c1 = np.linalg.lstsq(A, b, rcond=None)[0]
        relevant_y = [0, -1]
        relevant_t = [0, -1]
    else:
        # Vão entre dois apoios, incluindo barras de uma viga contínua.
        c0 = -y0[0]
        c1 = -(y0[-1] + c0) / L
        relevant_y = [0, -1]
        relevant_t = []
    theta = theta0 + c1
    y = y0 + c0 + c1 * xr
    disp_res = max([abs(float(y[i])) for i in relevant_y] + [0.0])
    rot_res = max([abs(float(theta[i])) for i in relevant_t] + [0.0])
    return theta, y, {"disp_residual_mm": disp_res, "rotation_residual_rad": rot_res}


def _service_groups_advanced(app, r: pd.Series, target: str):
    df = app.df_clean if getattr(app, "df_clean", None) is not None else pd.DataFrame()
    if df.empty:
        return [], "sem dados importados"
    cand = df[df["member"].astype(str) == str(r.get("member", ""))].copy()
    story = str(r.get("story", "") or "")
    if story and "story" in cand.columns:
        same = cand[cand["story"].astype(str) == story]
        if not same.empty:
            cand = same
    if cand.empty:
        return [], "viga não encontrada"
    target = str(target or "").strip()
    if target and normalize_text(target) not in {"todas", "todos", "all"}:
        tokens = [t.strip() for t in re.split(r"[;,|]+", target) if t.strip()]
        selected = []
        for case, grp in cand.groupby("case", dropna=False, sort=False):
            cstr = str(case)
            cnum = str(extract_combination_number(case))
            if any(cstr == t or cnum == t or normalize_text(t) in normalize_text(cstr) for t in tokens):
                selected.append((cstr, grp.copy()))
        return selected, "combinação(ões) indicada(s)"
    els = cand[cand["case"].map(_case_is_els_serviceability)]
    if els.empty:
        return [], "sem combinação ELS identificada"
    return [(str(case), grp.copy()) for case, grp in els.groupby("case", dropna=False, sort=False)], "todas as combinações ELS reconhecidas"


def _bar_label_area_advanced(label: str) -> float:
    s = str(label or "")
    m = re.search(r"(\d+)\s*[Øφ]\s*(\d+(?:[.,]\d+)?)", s)
    if not m:
        return float("nan")
    n = int(m.group(1))
    phi = float(m.group(2).replace(",", "."))
    return n * bar_area_mm2(phi)


def _group_positions_advanced(group: pd.DataFrame, length_m: float):
    import numpy as np
    g = group.sort_values("__row_order").copy()
    L = max(float(length_m), 1e-6)
    st = np.array([safe_float(v, float("nan")) for v in g.get("station", pd.Series([float("nan")] * len(g)))], dtype=float)
    if np.isfinite(st).sum() >= 2:
        stf = st.copy()
        finite_mask = np.isfinite(stf)
        fallback = np.linspace(0.0, L, len(g))
        stf[~finite_mask] = fallback[~finite_mask]
        if np.nanmax(stf) <= 1.001 and L > 1.5:
            stf = stf * L
        else:
            stf = stf - np.nanmin(stf)
        if np.nanmax(stf) > L * 1.10 or np.nanmax(stf) <= 1e-9:
            stf = fallback
        elif np.nanmax(stf) < L * 0.95:
            stf = stf * (L / max(np.nanmax(stf), 1e-9))
        x = stf
    else:
        x = np.linspace(0.0, L, len(g))
    order = np.argsort(x)
    return g.iloc[order].reset_index(drop=True), np.asarray(x)[order]


def _local_rebar_profiles_advanced(group: pd.DataFrame, length_m: float, r: pd.Series):
    import numpy as np
    g, x = _group_positions_advanced(group, length_m)
    defaults = {
        "as_bot": finite(r.get("as_prov_bot_mm2"), 0.0),
        "as_top": finite(r.get("as_prov_top_mm2"), 0.0),
        "phi_bot": _rebar_phi_serviceability(r.get("bot_rebar", ""), 12.0),
        "phi_top": _rebar_phi_serviceability(r.get("top_rebar", ""), 12.0),
    }
    result = {}
    for face in ["bot", "top"]:
        area_col = f"as_{face}_local_mm2"
        label_col = f"{face}_rebar_local"
        vals = []
        phis = []
        sources = []
        for _, rr in g.iterrows():
            area = safe_float(rr.get(area_col), float("nan"))
            label = str(rr.get(label_col, "") or "").strip()
            if not math.isfinite(area) or area <= 0:
                area = _bar_label_area_advanced(label)
            phi = _rebar_phi_serviceability(label, defaults[f"phi_{face}"]) if label else defaults[f"phi_{face}"]
            if math.isfinite(area) and area > 0:
                sources.append("armadura local importada")
            else:
                area = defaults[f"as_{face}"]
                sources.append("armadura uniforme dimensionada")
            vals.append(area)
            phis.append(phi)
        result[f"as_{face}"] = np.asarray(vals, dtype=float)
        result[f"phi_{face}"] = np.asarray(phis, dtype=float)
        result[f"source_{face}"] = sources
    return x, result


def _calculate_member_sls_advanced(app, r: pd.Series, group: pd.DataFrame, case_label: str, source_note: str, params: SLSParametersAdvanced):
    import numpy as np
    p = _profile_from_result_serviceability(r)
    A, _, _, _ = p.gross_properties()
    u = _profile_exposed_perimeter_serviceability(p)
    h0_auto = 2.0 * A / max(u, 1e-9)
    h0 = params.h0_override_mm if params.h0_override_mm > 0 else h0_auto
    fck = finite(r.get("fck_MPa"), parse_concrete_strength(r.get("material", DEFAULT_CONCRETE_CLASS)))
    cp = concrete_props(fck)
    Ecm = cp["Ecm"]
    history = _creep_history_advanced(cp, params, h0)
    phi = history["phi"]
    Ec_eff = Ecm / (1.0 + phi)
    shrink = _ec2_shrinkage_strain_serviceability(fck, cp["fcm"], params.rh_pct, h0, params.t_days, params.ts_days, params.cement_class)
    eps_cs = shrink["eps_cs"] if params.long_term and params.include_shrinkage else 0.0
    d_bot = finite(r.get("d_bot_mm"), p.h_mm - 50.0)
    d_top = finite(r.get("d_top_mm"), p.h_mm - 50.0)
    axis_var = getattr(app, "var_moment_axis", None)
    moment_col = str(axis_var.get() if axis_var is not None else "MY").lower()
    x_m, M_pts, quality, diagram_source = _moment_diagram_serviceability(group, finite(r.get("length_m"), 0.0), moment_col)
    if len(x_m) < 2:
        return {"service_status": "Dados insuficientes", "service_note": "Flecha não verificada: diagrama de momentos ELS indisponível.", "service_diagram_quality": "Insuficiente"}, []
    L_m = max(finite(r.get("length_m"), 0.0), float(x_m[-1]))
    x_grid_m = np.linspace(0.0, L_m, max(201, 40 * (len(x_m) - 1) + 1))
    M_grid = np.interp(x_grid_m, x_m, M_pts)
    x_rebar, local = _local_rebar_profiles_advanced(group, L_m, r)
    asb_grid = np.interp(x_grid_m, x_rebar, local["as_bot"])
    ast_grid = np.interp(x_grid_m, x_rebar, local["as_top"])
    phib_grid = np.interp(x_grid_m, x_rebar, local["phi_bot"])
    phit_grid = np.interp(x_grid_m, x_rebar, local["phi_top"])
    local_rebar_used = bool(
        ("as_bot_local_mm2" in group.columns and group["as_bot_local_mm2"].map(lambda v: math.isfinite(safe_float(v, float("nan"))) and safe_float(v, 0) > 0).any())
        or ("as_top_local_mm2" in group.columns and group["as_top_local_mm2"].map(lambda v: math.isfinite(safe_float(v, float("nan"))) and safe_float(v, 0) > 0).any())
        or ("bot_rebar_local" in group.columns and group["bot_rebar_local"].astype(str).str.contains("Ø|φ", regex=True).any())
        or ("top_rebar_local" in group.columns and group["top_rebar_local"].astype(str).str.contains("Ø|φ", regex=True).any())
    )
    clear_bot = finite(r.get("bot_clear_spacing_mm"), 999.0)
    clear_top = finite(r.get("top_clear_spacing_mm"), 999.0)
    phi_st = finite(r.get("phi_st_mm"), 8.0)
    cover = finite(r.get("cover_mm"), 35.0)
    point_rows = []
    k_short = []
    k_final = []
    wk_vals = []
    sigs_vals = []
    sigc_vals = []
    state_cache = {}
    for xx, mm, Asb, Ast, phib, phit in zip(x_grid_m, M_grid, asb_grid, ast_grid, phib_grid, phit_grid):
        key = (round(float(Asb), 3), round(float(Ast), 3))
        if key not in state_cache:
            state_cache[key] = _section_sls_states_serviceability(p, float(Asb), float(Ast), d_bot, d_top, Ecm, Ec_eff, cp["fctm"], eps_cs)
        states = state_cache[key]
        is_pos = mm >= 0
        phi_t = float(phib if is_pos else phit)
        clear = clear_bot if is_pos else clear_top
        rs = _local_sls_response_serviceability(mm, states, "short", 1.0, cp["fctm"], Ecm, 200000.0, p, phi_t, clear, cover, phi_st, 0.6, 0.0)
        rl = _local_sls_response_serviceability(mm, states, "long", params.beta_tension_stiffening, cp["fctm"], Ecm, 200000.0, p, phi_t, clear, cover, phi_st, params.kt_crack, eps_cs)
        k_short.append(rs["curvature_total_1_per_mm"])
        k_final.append(rl["curvature_total_1_per_mm"] if params.long_term else rs["curvature_total_1_per_mm"])
        active = rl if params.long_term else rs
        wk_vals.append(active["wk_mm"])
        sigs_vals.append(active["sigma_s_MPa"])
        sigc_vals.append(active["sigma_c_MPa"])
        point_rows.append({
            "viga": _beam_label_labels(r), "member": r.get("member", ""), "name": r.get("name", ""), "story": r.get("story", ""),
            "service_case": case_label, "service_regime": params.regime, "x_m": float(xx), "M_service_kNm": float(mm),
            "As_bot_local_mm2": float(Asb), "As_top_local_mm2": float(Ast),
            "reinforcement_source": "local importada" if local_rebar_used else "uniforme dimensionada",
            "section_state": active["section_state"], "zeta": active["zeta"], "Mcr_kNm": active["Mcr_kNm"],
            "x_cr_mm": active["x_cr_mm"], "I_un_mm4": active["I_un_mm4"], "I_cr_mm4": active["I_cr_mm4"],
            "sigma_s_MPa": active["sigma_s_MPa"], "sigma_c_MPa": active["sigma_c_MPa"], "wk_mm": active["wk_mm"],
            "sr_max_mm": active["sr_max_mm"], "rho_p_eff": active["rho_p_eff"],
            "curvature_load_1_per_m": active["curvature_load_1_per_mm"] * 1000.0,
            "curvature_shrinkage_1_per_m": active["curvature_shrinkage_1_per_mm"] * 1000.0,
            "curvature_total_1_per_m": active["curvature_total_1_per_mm"] * 1000.0,
        })
    support, support_source = _resolve_support_advanced(app, group, params)
    x_mm = x_grid_m * 1000.0
    _, y_short, residual_short = _integrate_curvature_advanced(x_mm, k_short, support)
    _, y_final, residual_final = _integrate_curvature_advanced(x_mm, k_final, support)
    for pr, yi, yf in zip(point_rows, y_short, y_final):
        pr["support_condition"] = support
        pr["deflection_instant_mm"] = float(yi)
        pr["deflection_final_mm"] = float(yf)
    delta_inst = float(np.max(np.abs(y_short)))
    delta_final = float(np.max(np.abs(y_final)))
    delta_used = delta_final if params.long_term else delta_inst
    limit = L_m * 1000.0 / max(params.deflection_limit_denominator, 1.0)
    wk_max = float(max(wk_vals) if wk_vals else 0.0)
    sigs_max = float(max(sigs_vals) if sigs_vals else 0.0)
    sigc_max = float(max(sigc_vals) if sigc_vals else 0.0)
    crack_status = "OK" if wk_max <= params.crack_limit_mm + 1e-9 else "Não verifica"
    defl_status = "OK" if delta_used <= limit + 1e-9 else "Não verifica"
    steel_lim = 0.8 * finite(r.get("fyk_MPa"), 500.0)
    steel_status = "OK" if sigs_max <= steel_lim + 1e-9 else "Não verifica"
    concrete_lim = 0.45 * fck if params.long_term else 0.60 * fck
    concrete_status = "OK" if sigc_max <= concrete_lim + 1e-9 else "Não verifica"
    axial_service = max(abs(finite(v, 0.0)) for v in group.get("fx", pd.Series([0.0]))) if "fx" in group.columns else 0.0
    axial_ratio = axial_service * 1000.0 / max(A * cp["fctm"], 1e-9)
    checks = [crack_status, defl_status, steel_status, concrete_status]
    status = "OK" if all(v == "OK" for v in checks) else "Não verifica"
    note = "Cálculo por curvaturas com envelope ELS, condição de apoio e efeitos diferidos."
    if quality == "Insuficiente":
        status = "Dados insuficientes"
        note = "Flecha não validada: importar pelo menos 3 estações nesta combinação ELS."
    elif axial_ratio > 0.10:
        status = "Dados insuficientes"
        note = "Esforço axial de serviço relevante; é necessária verificação N-M dedicada."
    # No encastre-encastre, resíduos elevados indicam que o diagrama importado e as
    # condições cinemáticas seleccionadas não são compatíveis.
    residual = residual_final if params.long_term else residual_short
    residual_limit = max(0.25, L_m * 1000.0 / 20000.0)
    if support == "Encastre-encastre" and residual["disp_residual_mm"] > residual_limit and status == "OK":
        status = "Dados insuficientes"
        note = "Diagrama de curvaturas incompatível com encastre-encastre; rever apoios/estações."
    return {
        "service_combination": case_label,
        "service_case_source": source_note,
        "service_regime": params.regime,
        "service_support_condition": support,
        "service_support_source": support_source,
        "service_boundary_disp_residual_mm": residual["disp_residual_mm"],
        "service_boundary_rotation_residual_rad": residual["rotation_residual_rad"],
        "service_points_imported": len(x_m),
        "service_points_integrated": len(x_grid_m),
        "service_diagram_quality": quality,
        "service_diagram_source": diagram_source,
        "service_RH_pct": params.rh_pct,
        "service_t0_days": params.t0_days,
        "service_t_days": params.t_days,
        "service_ts_days": params.ts_days,
        "service_cement_class": params.cement_class,
        "service_h0_mm": h0,
        "service_h0_source": "utilizador" if params.h0_override_mm > 0 else "2Ac/u automático",
        "service_history_mode": history["mode"],
        "service_permanent_fraction_pct": history["fraction_g"] * 100.0,
        "service_phi_creep": phi,
        "service_phi_creep_G": history["phi_g"],
        "service_phi_creep_Q": history["phi_q"],
        "service_Ecm_MPa": Ecm,
        "service_Ec_eff_MPa": Ec_eff,
        "service_eps_cd_permille": shrink["eps_cd"] * 1000.0,
        "service_eps_ca_permille": shrink["eps_ca"] * 1000.0,
        "service_eps_cs_permille": eps_cs * 1000.0,
        "service_reinforcement_model": "armadura local por estação" if local_rebar_used else "armadura uniforme dimensionada",
        "service_sigma_s_MPa": sigs_max,
        "service_sigma_s_lim_MPa": steel_lim,
        "service_sigma_c_MPa": sigc_max,
        "service_sigma_c_lim_MPa": concrete_lim,
        "service_wk_est_mm": wk_max,
        "service_wk_lim_mm": params.crack_limit_mm,
        "service_deflection_instant_mm": delta_inst,
        "service_deflection_final_mm": delta_final,
        "service_deflection_est_mm": delta_used,
        "service_deflection_lim_mm": limit,
        "service_deflection_limit": f"L/{params.deflection_limit_denominator:.0f}",
        "service_crack_status": crack_status,
        "service_deflection_status": defl_status,
        "service_stress_status": steel_status,
        "service_concrete_status": concrete_status,
        "service_axial_ratio": axial_ratio,
        "service_status": status,
        "service_note": note,
        "sls_method": "NP EN 1992-1-1:2010, 7.2, 7.3.4 e 7.4.3; Anexo B; envelope por combinação",
    }, point_rows


def _util_advanced(d: Dict, value_key: str, limit_key: str) -> float:
    return finite(d.get(value_key), 0.0) / max(finite(d.get(limit_key), 0.0), 1e-12)


def _governing_sls_advanced(items: List[Dict]) -> Dict:
    if not items:
        return {"service_status": "Dados insuficientes", "service_note": "Sem combinações ELS calculadas."}
    valid = [d for d in items if d.get("service_status") != "Dados insuficientes"]
    pool = valid or items
    crack = max(pool, key=lambda d: _util_advanced(d, "service_wk_est_mm", "service_wk_lim_mm"))
    defl = max(pool, key=lambda d: _util_advanced(d, "service_deflection_est_mm", "service_deflection_lim_mm"))
    steel = max(pool, key=lambda d: _util_advanced(d, "service_sigma_s_MPa", "service_sigma_s_lim_MPa"))
    concrete = max(pool, key=lambda d: _util_advanced(d, "service_sigma_c_MPa", "service_sigma_c_lim_MPa"))
    scored = []
    for d in pool:
        scored.append((max(
            _util_advanced(d, "service_wk_est_mm", "service_wk_lim_mm"),
            _util_advanced(d, "service_deflection_est_mm", "service_deflection_lim_mm"),
            _util_advanced(d, "service_sigma_s_MPa", "service_sigma_s_lim_MPa"),
            _util_advanced(d, "service_sigma_c_MPa", "service_sigma_c_lim_MPa"),
        ), d))
    overall = max(scored, key=lambda x: x[0])[1].copy()
    for prefix, src in [("crack", crack), ("deflection", defl), ("steel", steel), ("concrete", concrete)]:
        overall[f"service_combination_{prefix}"] = src.get("service_combination", "")
    # Cada verificação conserva o valor da respectiva combinação governante.
    for key in ["service_wk_est_mm", "service_wk_lim_mm", "service_crack_status"]:
        overall[key] = crack.get(key)
    for key in ["service_deflection_instant_mm", "service_deflection_final_mm", "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit", "service_deflection_status"]:
        overall[key] = defl.get(key)
    for key in ["service_sigma_s_MPa", "service_sigma_s_lim_MPa", "service_stress_status"]:
        overall[key] = steel.get(key)
    for key in ["service_sigma_c_MPa", "service_sigma_c_lim_MPa", "service_concrete_status"]:
        overall[key] = concrete.get(key)
    combos = [str(d.get("service_combination", "")) for d in items]
    overall["service_combinations_checked"] = len(items)
    overall["service_combinations_list"] = "; ".join(combos)
    overall["service_governing_combination"] = overall.get("service_combination", "")
    if any(d.get("service_status") == "Não verifica" for d in items):
        overall["service_status"] = "Não verifica"
    elif any(d.get("service_status") == "Dados insuficientes" for d in items):
        overall["service_status"] = "Dados insuficientes"
    else:
        overall["service_status"] = "OK"
    overall["service_note"] = (
        f"Envelope de {len(items)} combinação(ões) ELS. Governantes: "
        f"wk={overall.get('service_combination_crack','')}; "
        f"flecha={overall.get('service_combination_deflection','')}; "
        f"aço={overall.get('service_combination_steel','')}; "
        f"betão={overall.get('service_combination_concrete','')}."
    )
    return overall


def _sls_params_from_app_advanced(app) -> SLSParametersAdvanced:
    return SLSParametersAdvanced(
        regime=app.var_service_regime.get() if hasattr(app, "var_service_regime") else "Quase-permanente / longo prazo",
        rh_pct=finite(app.var_service_rh.get(), 70.0) if hasattr(app, "var_service_rh") else 70.0,
        t0_days=finite(app.var_service_t0.get(), 28.0) if hasattr(app, "var_service_t0") else 28.0,
        t_days=finite(app.var_service_t.get(), 18250.0) if hasattr(app, "var_service_t") else 18250.0,
        ts_days=finite(app.var_service_ts.get(), 7.0) if hasattr(app, "var_service_ts") else 7.0,
        cement_class=app.var_service_cement.get() if hasattr(app, "var_service_cement") else "N",
        h0_override_mm=finite(app.var_service_h0.get(), 0.0) if hasattr(app, "var_service_h0") else 0.0,
        include_shrinkage=bool(app.var_service_shrinkage.get()) if hasattr(app, "var_service_shrinkage") else True,
        deflection_limit_denominator=finite(app.var_ld_limit.get(), 250.0),
        crack_limit_mm=finite(app.var_crack_limit.get(), 0.30),
        auto_regime=bool(app.var_service_auto_regime.get()) if hasattr(app, "var_service_auto_regime") else True,
        support_condition=app.var_support_condition.get() if hasattr(app, "var_support_condition") else "Automática / por tabela",
        history_mode=app.var_history_mode.get() if hasattr(app, "var_history_mode") else "Uma idade efectiva",
        permanent_fraction_pct=finite(app.var_permanent_fraction.get(), 70.0) if hasattr(app, "var_permanent_fraction") else 70.0,
        t0_permanent_days=finite(app.var_t0_permanent.get(), 14.0) if hasattr(app, "var_t0_permanent") else 14.0,
        t0_variable_days=finite(app.var_t0_variable.get(), 365.0) if hasattr(app, "var_t0_variable") else 365.0,
    )


def _apply_serviceability_advanced(app, results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        app.df_sls_points = pd.DataFrame()
        app.df_sls_parameters = pd.DataFrame()
        app.df_sls_combinations = pd.DataFrame()
        return results
    out = results.copy()
    target = app.var_service_case.get().strip() if hasattr(app, "var_service_case") else ""
    base_params = _sls_params_from_app_advanced(app)
    all_points = []
    all_combos = []
    for idx, r in out.iterrows():
        groups, source_note = _service_groups_advanced(app, r, target)
        if not groups:
            out.at[idx, "service_status"] = "Dados insuficientes"
            out.at[idx, "service_note"] = "Sem combinação ELS correspondente; resultado de serviço não validado."
            out.at[idx, "service_case_source"] = source_note
            out.at[idx, "service_diagram_quality"] = "Insuficiente"
            if str(out.at[idx, "status"] if "status" in out.columns else "OK") == "OK":
                out.at[idx, "status"] = "Verificar"
            continue
        member_items = []
        for case_label, group in groups:
            params = _params_for_case_advanced(base_params, case_label)
            sls, points = _calculate_member_sls_advanced(app, r, group, case_label, source_note, params)
            member_items.append(sls)
            all_points.extend(points)
            all_combos.append({
                "viga": _beam_label_labels(r), "member": r.get("member", ""), "name": r.get("name", ""), "story": r.get("story", ""),
                **sls,
            })
        governing = _governing_sls_advanced(member_items)
        for k, v in governing.items():
            out.at[idx, k] = v
        current = str(out.at[idx, "status"] if "status" in out.columns else "OK")
        if governing.get("service_status") == "Não verifica":
            out.at[idx, "status"] = "Falha"
            existing = str(out.at[idx, "failure_reason"] if "failure_reason" in out.columns else "").strip()
            out.at[idx, "failure_reason"] = "; ".join(x for x in [existing, "ELS não verifica"] if x)
            out.at[idx, "failure_type"] = "els"
        elif governing.get("service_status") == "Dados insuficientes" and current == "OK":
            out.at[idx, "status"] = "Verificar"
    app.df_sls_points = pd.DataFrame(all_points)
    app.df_sls_combinations = pd.DataFrame(all_combos)
    app.df_sls_parameters = pd.DataFrame([
        ["Regime base", base_params.regime], ["Regime automático", "Sim" if base_params.auto_regime else "Não"],
        ["Combinações ELS", target or "todas as reconhecidas"], ["Condição de apoio", base_params.support_condition],
        ["Histórico de carregamento", base_params.history_mode], ["Fração permanente [%]", base_params.permanent_fraction_pct],
        ["t0,G [dias]", base_params.t0_permanent_days], ["t0,Q [dias]", base_params.t0_variable_days],
        ["RH [%]", base_params.rh_pct], ["t0 efectivo [dias]", base_params.t0_days], ["t [dias]", base_params.t_days],
        ["ts [dias]", base_params.ts_days], ["Classe de cimento", base_params.cement_class],
        ["h0 [mm]", base_params.h0_override_mm or "automático 2Ac/u"], ["Retracção", "Sim" if base_params.include_shrinkage else "Não"],
        ["Limite de flecha", f"L/{base_params.deflection_limit_denominator:.0f}"], ["wk,lim [mm]", base_params.crack_limit_mm],
    ], columns=["Parâmetro", "Valor"])
    return out

# Substituir a rotina ELS principal.
globals()["_apply_serviceability_serviceability"] = _apply_serviceability_advanced


def _sls_audit_df_advanced(results: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "viga", "member", "name", "story", "case", "service_governing_combination",
        "service_combinations_checked", "service_combination_crack", "service_combination_deflection",
        "service_combination_steel", "service_combination_concrete", "service_regime",
        "service_support_condition", "service_history_mode", "service_reinforcement_model",
        "service_points_imported", "service_diagram_quality", "service_phi_creep",
        "service_phi_creep_G", "service_phi_creep_Q", "service_eps_cs_permille", "service_Ec_eff_MPa",
        "service_sigma_c_MPa", "service_sigma_c_lim_MPa", "service_sigma_s_MPa", "service_sigma_s_lim_MPa",
        "service_wk_est_mm", "service_wk_lim_mm", "service_deflection_instant_mm",
        "service_deflection_final_mm", "service_deflection_est_mm", "service_deflection_lim_mm",
        "service_deflection_limit", "service_crack_status", "service_deflection_status",
        "service_concrete_status", "service_stress_status", "service_status", "service_note",
    ]
    if results is None or results.empty:
        return pd.DataFrame()
    df = results.copy()
    if "viga" not in df.columns:
        df["viga"] = df.apply(_beam_label_labels, axis=1)
    return df[[c for c in cols if c in df.columns]].copy()

globals()["_sls_audit_df_serviceability"] = _sls_audit_df_advanced
globals()["_sls_audit_df_calculation"] = _sls_audit_df_advanced
globals()["_sls_audit_df_report_base"] = _sls_audit_df_advanced


# --------------------------- pormenorização longitudinal ---------------------------
def _rebar_count_phi_advanced(label: str) -> Tuple[int, float]:
    s = str(label or "")
    m = re.search(r"(\d+)\s*[Øφ]\s*(\d+(?:[.,]\d+)?)", s)
    return (int(m.group(1)), float(m.group(2).replace(",", "."))) if m else (0, 0.0)


def _anchorage_advanced(phi: float, fck: float, fyd: float, gamma_c: float, bond_eta1: float) -> Dict[str, float]:
    if phi <= 0:
        return {"fbd_MPa": 0.0, "lb_rqd_mm": 0.0, "lbd_mm": 0.0}
    fctm = concrete_props(fck)["fctm"]
    fctd = 0.7 * fctm / max(gamma_c, 1e-9)
    eta2 = 1.0 if phi <= 32.0 else max((132.0 - phi) / 100.0, 0.7)
    fbd = 2.25 * bond_eta1 * eta2 * fctd
    lb_rqd = phi * fyd / max(4.0 * fbd, 1e-9)
    lb_min = max(0.30 * lb_rqd, 10.0 * phi, 100.0)
    return {"fbd_MPa": fbd, "lb_rqd_mm": lb_rqd, "lbd_mm": max(lb_rqd, lb_min)}


def _longitudinal_detailing_advanced(app, results: pd.DataFrame):
    import numpy as np
    if results is None or results.empty:
        return pd.DataFrame(), pd.DataFrame()
    rows = []
    profiles = []
    designer = BeamDesigner(
        cover_mm=finite(app.var_cover.get(), 35.0), agg_mm=DEFAULT_AGGREGATE_MM_Reporting,
        fyk=finite(app.var_fyk.get(), 500.0), gamma_c=DEFAULT_GAMMA_C_Reporting, gamma_s=DEFAULT_GAMMA_S_Reporting,
        cot_theta=2.0, crack_limit_mm=finite(app.var_crack_limit.get(), 0.30),
        deflection_ld_limit=finite(app.var_ld_limit.get(), 250.0), calc_mode=app.var_calc_mode.get(),
    )
    raw = app.df_clean if getattr(app, "df_clean", None) is not None else pd.DataFrame()
    moment_col = str(app.var_moment_axis.get()).lower() if hasattr(app, "var_moment_axis") else "my"
    for _, r in results.iterrows():
        p = _profile_from_result_serviceability(r)
        fck = finite(r.get("fck_MPa"), parse_concrete_strength(r.get("material", DEFAULT_CONCRETE_CLASS)))
        fyd = finite(r.get("fyd_MPa"), finite(app.var_fyk.get(), 500.0) / DEFAULT_GAMMA_S_Reporting)
        nb, phib = _rebar_count_phi_advanced(r.get("bot_rebar", ""))
        nt, phit = _rebar_count_phi_advanced(r.get("top_rebar", ""))
        bot_anchor = _anchorage_advanced(phib, fck, fyd, DEFAULT_GAMMA_C_Reporting, 1.0)
        # Condição de aderência desfavorável para varões superiores em vigas altas.
        eta_top = 0.7 if p.h_mm > 250.0 else 1.0
        top_anchor = _anchorage_advanced(phit, fck, fyd, DEFAULT_GAMMA_C_Reporting, eta_top)
        z = 0.9 * min(finite(r.get("d_bot_mm"), p.h_mm - 50.0), finite(r.get("d_top_mm"), p.h_mm - 50.0))
        cot = finite(r.get("cot_theta_shear"), 2.0)
        al = 0.5 * z * cot
        rows.append({
            "viga": _beam_label_labels(r), "member": r.get("member", ""), "name": r.get("name", ""), "story": r.get("story", ""),
            "bot_rebar": r.get("bot_rebar", ""), "top_rebar": r.get("top_rebar", ""),
            "fbd_bot_MPa": bot_anchor["fbd_MPa"], "lbd_bot_mm": bot_anchor["lbd_mm"],
            "fbd_top_MPa": top_anchor["fbd_MPa"], "lbd_top_mm": top_anchor["lbd_mm"],
            "shift_al_mm": al, "extension_bot_mm": bot_anchor["lbd_mm"] + al,
            "extension_top_mm": top_anchor["lbd_mm"] + al,
            "min_bottom_bars_into_support": max(2, int(math.ceil(0.25 * max(nb, 0)))) if nb else 0,
            "detailing_basis": "EC2 8.4 e 9.2.1.3; comprimentos sem reduções por confinamento/ganchos",
            "detailing_note": "Confirmar comprimentos disponíveis, nós, emendas e desenho final.",
        })
        if raw.empty:
            continue
        cand = raw[(raw["member"].astype(str) == str(r.get("member", ""))) & (~raw["case"].map(_case_is_els_serviceability))].copy()
        story = str(r.get("story", "") or "")
        if story and "story" in cand.columns:
            same = cand[cand["story"].astype(str) == story]
            if not same.empty:
                cand = same
        if cand.empty:
            continue
        L = max(finite(r.get("length_m"), 0.0), 1e-6)
        xgrid = np.linspace(0.0, L, 101)
        pos_env = np.zeros_like(xgrid)
        neg_env = np.zeros_like(xgrid)
        for case, grp in cand.groupby("case", dropna=False):
            x, M, _, _ = _moment_diagram_serviceability(grp, L, moment_col)
            if len(x) < 2:
                continue
            mg = np.interp(xgrid, x, M)
            pos_env = np.maximum(pos_env, np.maximum(mg, 0.0))
            neg_env = np.maximum(neg_env, np.maximum(-mg, 0.0))
        cp = concrete_props(fck, gamma_c=DEFAULT_GAMMA_C_Reporting)
        bw = p.web_mm
        d_bot = finite(r.get("d_bot_mm"), p.h_mm - 50.0)
        d_top = finite(r.get("d_top_mm"), p.h_mm - 50.0)
        d2 = finite(r.get("d2_top_mm"), 50.0)
        Asmin_b = designer.as_min_beam(bw, d_bot, cp["fctm"], finite(app.var_fyk.get(), 500.0))
        Asmin_t = designer.as_min_beam(bw, d_top, cp["fctm"], finite(app.var_fyk.get(), 500.0))
        for xx, mp, mn in zip(xgrid, pos_env, neg_env):
            fb = designer.flexural_required(float(mp), p, d_bot, d2, cp["fcd"], fyd)
            ft = designer.flexural_required(float(mn), p.flipped(), d_top, d2, cp["fcd"], fyd)
            Asb = max(Asmin_b, finite(fb.get("As_req"), 0.0))
            Ast = max(Asmin_t, finite(ft.get("As_req"), 0.0))
            nreq_b = max(2, int(math.ceil(Asb / max(bar_area_mm2(phib), 1e-9)))) if phib else 0
            nreq_t = max(2, int(math.ceil(Ast / max(bar_area_mm2(phit), 1e-9)))) if phit else 0
            profiles.append({
                "viga": _beam_label_labels(r), "member": r.get("member", ""), "story": r.get("story", ""), "x_m": float(xx),
                "Mpos_env_kNm": float(mp), "Mneg_env_kNm": float(mn),
                "Asreq_bot_mm2": Asb, "Asreq_top_mm2": Ast,
                "n_bot_req": nreq_b, "phi_bot_mm": phib, "n_top_req": nreq_t, "phi_top_mm": phit,
                "n_bot_adopted": nb, "n_top_adopted": nt,
                "bot_reduction_possible": "Sim" if nb and nreq_b < nb else "Não",
                "top_reduction_possible": "Sim" if nt and nreq_t < nt else "Não",
                "cutoff_extension_bot_mm": bot_anchor["lbd_mm"] + al,
                "cutoff_extension_top_mm": top_anchor["lbd_mm"] + al,
            })
    return pd.DataFrame(rows), pd.DataFrame(profiles)


# --------------------------- torção em secções I ---------------------------
def _rect_torsion_j_advanced(b: float, h: float) -> float:
    a = max(float(b), float(h))
    t = min(float(b), float(h))
    if t <= 0 or a <= 0:
        return 0.0
    r = t / a
    return (a * t ** 3 / 3.0) * max(1.0 - 0.63 * r + 0.052 * r ** 5, 0.05)


_torsion_requirements_geometry = BeamDesigner.torsion_requirements

def _torsion_requirements_advanced(self, TEd_kNm, section, fck, fcd, fyd):
    if not isinstance(section, SectionProfileGeometry) or not section.is_i:
        return _torsion_requirements_geometry(self, TEd_kNm, section, fck, fcd, fyd)
    T = abs(float(TEd_kNm or 0.0))
    # Preservar o limiar de torção não relevante.
    proxy_outer = BeamSection(section.outer_width_mm, section.h_mm, 0.0, 0.0)
    check = _torsion_requirements_geometry(self, T, proxy_outer, fck, fcd, fyd)
    if str(check.get("torsion_considered", "Não")) != "Sim":
        check["torsion_geometry_model"] = "secção I subdividida; torção não relevante"
        return check
    parts = []
    if section.top_thickness_mm > 0:
        parts.append(("banzo superior", section.top_width_mm, section.top_thickness_mm))
    if section.web_height_mm > 0:
        parts.append(("alma", section.web_mm, section.web_height_mm))
    if section.bottom_thickness_mm > 0:
        parts.append(("banzo inferior", section.bottom_width_mm, section.bottom_thickness_mm))
    weights = [_rect_torsion_j_advanced(b, h) for _, b, h in parts]
    sw = sum(weights)
    if sw <= 0:
        return _torsion_requirements_geometry(self, T, BeamSection(section.web_mm, section.h_mm, 0.0, 0.0), fck, fcd, fyd)
    Asw = 0.0
    Asl = 0.0
    TRdmax = 0.0
    details = []
    for (name, b, h), w in zip(parts, weights):
        Ti = T * w / sw
        rr = _torsion_requirements_geometry(self, Ti, BeamSection(b, h, 0.0, 0.0), fck, fcd, fyd)
        Asw += finite(rr.get("Asw_s_torsion_req_mm2_per_mm"), 0.0)
        Asl += finite(rr.get("Asl_torsion_req_mm2"), 0.0)
        TRdmax += finite(rr.get("TRd_max_kNm"), 0.0)
        details.append(f"{name}:{Ti:.2f} kNm")
    status = "Requer armadura de torção por subsecções" if T <= TRdmax + 1e-9 else "Não conforme: TEd > soma TRd,max"
    return {
        "torsion_considered": "Sim", "torsion_ratio": T / max(TRdmax, 1e-9),
        "TRd_max_kNm": TRdmax, "Asw_s_torsion_req_mm2_per_mm": Asw,
        "Asl_torsion_req_mm2": Asl, "torsion_status": status,
        "tef_mm": None, "Ak_mm2": None, "uk_mm": None,
        "torsion_geometry_model": "secção I subdividida em banzos e alma; T distribuído por rigidez de Saint-Venant",
        "torsion_distribution": "; ".join(details),
    }

BeamDesigner.torsion_requirements = _torsion_requirements_advanced


# --------------------------- testes internos ---------------------------
def _internal_tests_advanced() -> pd.DataFrame:
    import numpy as np
    tests = []
    def add(name, calc, ref, tol, unit="-"):
        err = abs(calc - ref)
        rel = err / max(abs(ref), 1e-12)
        tests.append({"Teste": name, "Calculado": calc, "Referência": ref, "Unidade": unit, "Erro relativo": rel, "Tolerância": tol, "Estado": "OK" if rel <= tol else "Falha"})
    # Integração para curvatura constante.
    L = 6000.0
    k = 2.0e-8
    x = np.linspace(0.0, L, 401)
    kap = np.full_like(x, k)
    _, yss, _ = _integrate_curvature_advanced(x, kap, "Entre apoios (y0=yL=0)")
    add("Integração - vão entre apoios", float(np.max(np.abs(yss))), k * L ** 2 / 8.0, 2e-3, "mm")
    _, ycant, _ = _integrate_curvature_advanced(x, kap, "Consola - encastrada à esquerda")
    add("Integração - consola", abs(float(ycant[-1])), k * L ** 2 / 2.0, 2e-3, "mm")
    # Propriedades de secção rectangular.
    p = SectionProfileGeometry(600.0, 300.0, 300.0, 0.0, 300.0, 0.0, "Rectangular")
    A, z, I, _ = p.gross_properties()
    add("Área rectangular", A, 300.0 * 600.0, 1e-12, "mm²")
    add("Centroide rectangular", z, 300.0, 1e-12, "mm")
    add("Inércia rectangular", I, 300.0 * 600.0 ** 3 / 12.0, 1e-12, "mm⁴")
    # Fluência/retracção: verificações de sanidade e monotonia.
    phi1 = _ec2_creep_coefficient_serviceability(38.0, 70.0, 200.0, 28.0, 365.0, "N")
    phi2 = _ec2_creep_coefficient_serviceability(38.0, 70.0, 200.0, 28.0, 18250.0, "N")
    tests.append({"Teste": "Fluência positiva e crescente", "Calculado": phi2, "Referência": phi1, "Unidade": "-", "Erro relativo": 0.0, "Tolerância": 0.0, "Estado": "OK" if phi2 >= phi1 > 0 else "Falha"})
    sh = _ec2_shrinkage_strain_serviceability(30.0, 38.0, 70.0, 200.0, 18250.0, 7.0, "N")["eps_cs"]
    tests.append({"Teste": "Retracção positiva", "Calculado": sh, "Referência": 0.0, "Unidade": "-", "Erro relativo": 0.0, "Tolerância": 0.0, "Estado": "OK" if sh > 0 else "Falha"})
    return pd.DataFrame(tests)


# --------------------------- GUI ---------------------------
def _ensure_sls_vars_advanced(self):
    _ensure_sls_vars_serviceability(self)
    defaults = {
        "var_service_auto_regime": (tk.BooleanVar, True),
        "var_support_condition": (tk.StringVar, "Automática / por tabela"),
        "var_history_mode": (tk.StringVar, "Duas fases G+Q"),
        "var_permanent_fraction": (tk.StringVar, "70"),
        "var_t0_permanent": (tk.StringVar, "14"),
        "var_t0_variable": (tk.StringVar, "365"),
    }
    for name, (cls, value) in defaults.items():
        if not hasattr(self, name):
            setattr(self, name, cls(value=value))


def _build_sidebar_advanced(self, parent):
    _ensure_sls_vars_advanced(self)
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    if finite(self.var_ld_limit.get(), 20.0) < 100:
        self.var_ld_limit.set("250")
    hero = ttk.LabelFrame(parent, text="BeamsEC2")
    hero.pack(fill="x", pady=(0, 8))
    link = ttk.Label(hero, text=APP_NAME, style="Header.TLabel", cursor="hand2")
    link.pack(anchor="w"); link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
    ttk.Label(hero, text="Dimensionamento de vigas de betão armado", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(hero, text="ELU, envelope ELS, condições de apoio, efeitos diferidos e auditoria segundo a NP EN 1992-1-1.", style="Subtle.TLabel", wraplength=340, justify="left").pack(anchor="w", pady=(2, 0))

    data = ttk.LabelFrame(parent, text="1. Entrada")
    data.pack(fill="x", pady=(0, 8))
    ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

    params = ttk.LabelFrame(parent, text="2. Parâmetros de cálculo")
    params.pack(fill="x", pady=(0, 8))
    self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
    ttk.Label(params, text="Aço fyk [MPa]").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 2)
    self._add_label_entry(params, "Limite de flecha L/", self.var_ld_limit, 3)
    ttk.Label(params, text="Momento principal").grid(row=4, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_moment_axis, values=["MY", "MZ"], state="readonly").grid(row=4, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Corte vertical").grid(row=5, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_shear_axis, values=["FZ", "FY"], state="readonly").grid(row=5, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Torção").grid(row=6, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_torsion_axis, values=["MX", "MY", "MZ", "Nenhuma"], state="readonly").grid(row=6, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    params.columnconfigure(1, weight=1)

    sls = ttk.LabelFrame(parent, text="3. ELS e deformação")
    sls.pack(fill="x", pady=(0, 8))
    ttk.Label(sls, text="Combinações ELS").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(sls, textvariable=self.var_service_case).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(sls, text="Regime automático pelo nome da combinação", variable=self.var_service_auto_regime).grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Label(sls, text="Regime base").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_service_regime, values=["Quase-permanente / longo prazo", "Frequente / repetida", "Característica / curta duração"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(sls, text="Condição de apoio").grid(row=3, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_support_condition, values=["Automática / por tabela", "Entre apoios (y0=yL=0)", "Consola - encastrada à esquerda", "Consola - encastrada à direita", "Encastre-encastre"], state="readonly").grid(row=3, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(sls, text="Histórico").grid(row=4, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_history_mode, values=["Duas fases G+Q", "Uma idade efectiva"], state="readonly").grid(row=4, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(sls, "Fração permanente [%]", self.var_permanent_fraction, 5)
    self._add_label_entry(sls, "t0,G [d]", self.var_t0_permanent, 6)
    self._add_label_entry(sls, "t0,Q [d]", self.var_t0_variable, 7)
    self._add_label_entry(sls, "Humidade RH [%]", self.var_service_rh, 8)
    self._add_label_entry(sls, "t0 efectivo [d]", self.var_service_t0, 9)
    self._add_label_entry(sls, "Idade final t [d]", self.var_service_t, 10)
    self._add_label_entry(sls, "Início secagem ts [d]", self.var_service_ts, 11)
    ttk.Label(sls, text="Classe de cimento").grid(row=12, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(sls, textvariable=self.var_service_cement, values=["S", "N", "R"], state="readonly", width=8).grid(row=12, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(sls, "h0 [mm] (0=auto)", self.var_service_h0, 13)
    ttk.Checkbutton(sls, text="Considerar retracção", variable=self.var_service_shrinkage).grid(row=14, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Label(sls, text="Em branco: verifica todas as combinações ELS reconhecidas. Podem indicar-se várias, separadas por ;", style="Subtle.TLabel", wraplength=320, justify="left").grid(row=15, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 4))
    sls.columnconfigure(1, weight=1)

    filters = ttk.LabelFrame(parent, text="4. Filtros")
    filters.pack(fill="x", pady=(0, 8))
    ttk.Label(filters, text="Viga/Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha", "Verificar"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "corte", "torcao", "pormenorizacao", "els", "dados", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    filters.columnconfigure(1, weight=1)

    actions = ttk.LabelFrame(parent, text="5. Cálculo e exportação")
    actions.pack(fill="x", pady=(0, 8))
    ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ttk.Label(actions, text="PDF").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(actions, textvariable=self.var_pdf_scope, values=PDF_SCOPES_Reporting, state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(actions, text="Exportar .csv", command=self.export_csv).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

    status_box = ttk.LabelFrame(parent, text="6. Estado")
    status_box.pack(fill="x", pady=(0, 8))
    ttk.Label(status_box, textvariable=self.status_var, wraplength=340, justify="left").pack(fill="x", padx=6, pady=(4, 2))
    ttk.Progressbar(status_box, variable=self.progress_var, maximum=100).pack(fill="x", padx=6, pady=(2, 2))
    ttk.Label(status_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

    notes = ttk.LabelFrame(parent, text="7. Notas rápidas")
    notes.pack(fill="x", pady=(0, 8))
    ttk.Label(notes, text=(
        f"• Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}.\n"
        "• O ELS é calculado para todas as combinações reconhecidas, com governantes separados.\n"
        "• A condição de apoio pode ser definida globalmente ou por viga na tabela.\n"
        "• As colunas opcionais de armadura local permitem modelar variações ao longo do vão."
    ), wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)

BeamsEC2App._build_sidebar = _build_sidebar_advanced


def _build_instructions_tab_advanced(self, parent):
    outer = ttk.Frame(parent, padding=10); outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1); outer.columnconfigure(0, weight=1)
    ttk.Label(outer, text="Instruções de utilização", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    host = ttk.Frame(outer); host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(host)
    content = (
        "OBJECTIVO\n"
        "BeamsEC2 dimensiona e verifica vigas de betão armado segundo a NP EN 1992-1-1. O ELS cria um envelope de todas as combinações de serviço reconhecidas e identifica separadamente as combinações governantes de fendilhação, deformação e tensões.\n\n"
        "DADOS DE IMPORTAÇÃO\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
        "COMBINAÇÕES ELS\n"
        "Deixe o campo vazio para calcular todas as combinações identificadas como ELS/SLS/CQC/QP/FREQ/RARA. Para combinações apenas numéricas, indique os números separados por ponto e vírgula. Cada combinação deve conter pelo menos três estações por viga; recomendam-se cinco ou mais.\n\n"
        "CONDIÇÕES DE APOIO\n"
        "Pode seleccionar vão entre apoios, consola à esquerda/direita ou encastre-encastre. A coluna Support Condition permite uma definição específica por viga e prevalece sobre o parâmetro global.\n\n"
        "HISTÓRICO DE CARREGAMENTO\n"
        "O modo Duas fases G+Q calcula coeficientes de fluência distintos para acções permanentes e variáveis e adopta um coeficiente efectivo ponderado pela fracção permanente. Introduza t0,G, t0,Q e a idade final t.\n\n"
        "ARMADURA AO LONGO DO VÃO\n"
        "As Bot Local/As Top Local ou Bot Rebar Local/Top Rebar Local são opcionais e podem variar por Station. Quando não são fornecidas, o ELS utiliza a armadura uniforme dimensionada. A exportação inclui o perfil de armadura necessária, comprimentos de amarração e decalagem al.\n\n"
        "RESULTADOS E VALIDAÇÃO\n"
        "O memória de cálculo inclui o envelope ELS, todas as combinações, resultados ponto a ponto, pormenorização longitudinal e testes internos. Casos com dados insuficientes ou condições cinemáticas incompatíveis são marcados como Verificar."
    )
    txt.insert("1.0", content); txt.config(state="disabled")

BeamsEC2App._build_instructions_tab = _build_instructions_tab_advanced


_validate_inputs_advanced_base = BeamsEC2App.validate_inputs

def _validate_inputs_advanced(self):
    _ensure_sls_vars_advanced(self)
    err = _validate_inputs_advanced_base(self)
    if err:
        return err
    fg = finite(self.var_permanent_fraction.get(), -1.0)
    tg = finite(self.var_t0_permanent.get(), -1.0)
    tq = finite(self.var_t0_variable.get(), -1.0)
    t = finite(self.var_service_t.get(), -1.0)
    if not 0.0 <= fg <= 100.0:
        return "Fração permanente inválida; utilizar um valor entre 0 % e 100 %."
    if "duas" in normalize_text(self.var_history_mode.get()):
        if tg <= 0 or tq <= 0 or tg >= t or tq >= t:
            return "Histórico G+Q inválido: deve verificar-se 0 < t0,G < t e 0 < t0,Q < t."
    return None

BeamsEC2App.validate_inputs = _validate_inputs_advanced


_run_design_serviceability_base = BeamsEC2App.run_design

def _run_design_advanced(self):
    _ensure_sls_vars_advanced(self)
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err); return
    designer = BeamDesigner(
        cover_mm=finite(self.var_cover.get(), 35.0), agg_mm=DEFAULT_AGGREGATE_MM_Reporting,
        fyk=finite(self.var_fyk.get(), 500.0), gamma_c=DEFAULT_GAMMA_C_Reporting, gamma_s=DEFAULT_GAMMA_S_Reporting,
        cot_theta=2.0, crack_limit_mm=finite(self.var_crack_limit.get(), 0.30),
        deflection_ld_limit=finite(self.var_ld_limit.get(), 250.0), calc_mode=self.var_calc_mode.get(),
    )
    elu_env = self.df_env[~self.df_env["case"].map(_case_is_els_serviceability)].copy() if "case" in self.df_env.columns else self.df_env.copy()
    if elu_env.empty:
        messagebox.showwarning("Aviso", "Não foram identificadas combinações ELU para o dimensionamento.")
        return
    input_df = reduce_to_governing_cases(elu_env) if self.var_reduce_cases.get() else elu_env
    self.df_calc_input = input_df.copy(); self.progress_var.set(0.0); self.status_var.set("Análise ELU/ELS em curso...")
    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular ELU... {d}/{t} envelopes"))
    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            self.after(0, lambda: self.status_var.set("A calcular envelope ELS..."))
            results = _apply_serviceability_advanced(self, results)
            self.df_longitudinal_detailing, self.df_longitudinal_profile = _longitudinal_detailing_advanced(self, results)
            self.df_internal_tests = _internal_tests_advanced()
            summary = build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
            failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
            ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
            validation = build_data_validation(self.df_clean, self.df_env, results)
            if self.df_internal_tests is not None and not self.df_internal_tests.empty:
                internal = self.df_internal_tests.copy()
                internal.insert(0, "Origem", "Teste interno")
                # Harmonização simples com a aba de validação.
                internal = internal.rename(columns={"Teste": "Verificação", "Estado": "Resultado"})
                validation = pd.concat([validation, internal], ignore_index=True, sort=False)
            def finish():
                self.df_results=results; self.df_summary=summary; self.df_failures=failures; self.df_ok=ok
                self.df_validation=validation; self.df_notes=build_normative_notes(); self.df_filtered=pd.DataFrame()
                self.show_df(self.tree_results,self.df_results); self.show_df(self.tree_summary,self.df_summary)
                self.show_df(self.tree_failures,self.df_failures); self.show_df(self.tree_shortlists,self.build_shortlists_df())
                self.show_df(self.tree_validation,self.df_validation); self.show_df(self.tree_notes,self.df_notes)
                self.update_report(); self.progress_var.set(100.0)
                n_ver=int((results["status"]=="Verificar").sum()) if "status" in results.columns else 0
                n_tests_fail=int((self.df_internal_tests["Estado"]=="Falha").sum()) if not self.df_internal_tests.empty else 0
                self.status_var.set(f"Cálculo concluído: {len(results)} envelopes; {len(failures)} falhas; {n_ver} a verificar; testes internos com {n_tests_fail} falhas.")
            self.after(0, finish)
        except Exception as exc:
            msg=str(exc); self.after(0,lambda m=msg:messagebox.showerror("Erro",m)); self.after(0,lambda:self.status_var.set("Falha na análise.")); self.after(0,lambda:self.progress_var.set(0.0))
    threading.Thread(target=worker,daemon=True).start()

BeamsEC2App.run_design = _run_design_advanced


# --------------------------- metadados e exportações ---------------------------
_parameters_df_advanced_base = BeamsEC2App._parameters_df

def _parameters_df_advanced(self):
    base = _parameters_df_advanced_base(self)
    p = _sls_params_from_app_advanced(self)
    extra = pd.DataFrame([
        ["Envelope ELS", "todas as combinações reconhecidas" if not self.var_service_case.get().strip() else self.var_service_case.get().strip()],
        ["Regime automático", "Sim" if p.auto_regime else "Não"], ["Condição de apoio", p.support_condition],
        ["Histórico", p.history_mode], ["Fração permanente [%]", p.permanent_fraction_pct],
        ["t0,G [dias]", p.t0_permanent_days], ["t0,Q [dias]", p.t0_variable_days],
        ["Armadura ELS", "local por Station quando fornecida; caso contrário uniforme"],
    ], columns=["Parâmetro", "Valor"])
    return pd.concat([base, extra], ignore_index=True)

BeamsEC2App._parameters_df = _parameters_df_advanced


_write_excel_advanced_base = BeamsEC2App._write_excel

def _write_excel_advanced(self, path: str):
    _write_excel_advanced_base(self, path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path)
        datasets = [
            ("11C_ELS_Combinacoes", getattr(self, "df_sls_combinations", pd.DataFrame())),
            ("12A_Porm_Longitudinal", getattr(self, "df_longitudinal_detailing", pd.DataFrame())),
            ("12B_Perfil_Armadura", getattr(self, "df_longitudinal_profile", pd.DataFrame())),
            ("16A_Testes_Internos", getattr(self, "df_internal_tests", pd.DataFrame())),
        ]
        for name, _ in datasets:
            if name in wb.sheetnames:
                del wb[name]
        fill=PatternFill("solid",fgColor="1F4E5F"); thin=Side(style="thin",color="D9E2E7"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
        for name, df in datasets:
            ws=wb.create_sheet(name)
            if df is None: df=pd.DataFrame()
            ws.append(list(df.columns))
            for _, rr in df.iterrows(): ws.append([None if pd.isna(v) else v for v in rr.tolist()])
            for cell in ws[1]: cell.fill=fill; cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
            for row in ws.iter_rows(min_row=2):
                for cell in row: cell.border=border; cell.alignment=Alignment(vertical="top",wrap_text=True)
            ws.freeze_panes="A2"; ws.sheet_view.showGridLines=False
            for j in range(1,ws.max_column+1):
                vals=[str(ws.cell(i,j).value) for i in range(1,min(ws.max_row,250)+1) if ws.cell(i,j).value is not None]
                ws.column_dimensions[get_column_letter(j)].width=min(max([len(v) for v in vals]+[10])+2,42)
        wb.properties.title = APP_NAME; wb.properties.creator = ""; wb.properties.lastModifiedBy = ""
        wb.save(path)
    except Exception:
        pass

BeamsEC2App._write_excel = _write_excel_advanced


# Memória de cálculo: acrescentar os governantes ELS e a condição de apoio.
_calc_memory_advanced_base = _calc_memory_df_serviceability

def _calc_memory_df_advanced(results: pd.DataFrame) -> pd.DataFrame:
    base = _calc_memory_advanced_base(results)
    rows = []
    if results is not None and not results.empty:
        for _, r in results.iterrows():
            common = {"Viga": _beam_label_labels(r), "Caso": r.get("case", ""), "Piso": r.get("story", ""), "Secção": r.get("section_geometry_summary", r.get("section_type", ""))}
            rows.extend([
                {**common, "Etapa": "ELS envelope", "Item": "Combinações verificadas", "Valor": r.get("service_combinations_list", ""), "Unidade": "-", "Critério/Referência": "NP EN 1992-1-1, Secção 7", "Estado/Nota": r.get("service_status", "")},
                {**common, "Etapa": "ELS envelope", "Item": "Governantes wk / flecha", "Valor": f"{r.get('service_combination_crack','')} / {r.get('service_combination_deflection','')}", "Unidade": "-", "Critério/Referência": "7.3.4 / 7.4.3", "Estado/Nota": r.get("service_support_condition", "")},
                {**common, "Etapa": "Pormenorização", "Item": "Armadura ELS ao longo do vão", "Valor": r.get("service_reinforcement_model", ""), "Unidade": "-", "Critério/Referência": "Modelo de rigidez", "Estado/Nota": ""},
            ])
    return pd.concat([base, pd.DataFrame(rows)], ignore_index=True) if rows else base

globals()["_calc_memory_df_serviceability"] = _calc_memory_df_advanced
globals()["_calc_memory_df_geometry"] = _calc_memory_df_advanced
globals()["_calc_memory_df_labels"] = _calc_memory_df_advanced
globals()["_calc_memory_df_reporting"] = _calc_memory_df_advanced
globals()["_calc_memory_df_report_base"] = _calc_memory_df_advanced


# Notas normativas da versão.
_build_notes_advanced_base = build_normative_notes

def build_normative_notes_advanced() -> pd.DataFrame:
    notes = _build_notes_advanced_base()
    extra = pd.DataFrame([
        ("Envelope ELS", "NP EN 1992-1-1, Secção 7", "Todas as combinações de serviço reconhecidas são calculadas; wk, flecha e tensões podem ter combinações governantes distintas."),
        ("Condições de apoio", "Compatibilidade cinemática", "A integração da curvatura considera vão entre apoios, consola esquerda/direita ou encastre-encastre; a condição pode ser definida por viga."),
        ("Histórico de carregamento", "NP EN 1992-1-1, Anexo B", "O modo G+Q usa coeficientes de fluência distintos por idade de aplicação e um coeficiente efectivo ponderado pela fracção permanente."),
        ("Pormenorização longitudinal", "NP EN 1992-1-1, 8.4 e 9.2.1.3", "São reportados comprimentos de amarração, decalagem al e perfil de armadura necessária; confirmar nós, emendas e disponibilidade geométrica."),
        ("Torção em I", "NP EN 1992-1-1, 6.3", "A secção I é subdividida em banzos e alma, distribuindo T pela rigidez de Saint-Venant; rever a pormenorização específica quando a torção for relevante."),
        ("Validação automática", "Controlo interno", "O workbook inclui testes de integração, propriedades geométricas e sanidade dos modelos de fluência/retracção."),
    ], columns=["Tema", "Referência", "Nota"])
    return pd.concat([notes, extra], ignore_index=True)

globals()["build_normative_notes"] = build_normative_notes_advanced



# Ajustes finais : estado de torção em I, relatório interno e PDF técnico.
_design_one_advanced_base = BeamDesigner.design_one

def _design_one_advanced(self, row: pd.Series) -> Dict:
    out = _design_one_advanced_base(self, row)
    if str(out.get("section_type", "")).startswith("I"):
        out["torsion_method"] = (
            "NP EN 1992-1-1:2010, 6.3; secção subdividida em banzos e alma, "
            "com distribuição de T pela rigidez de Saint-Venant"
        )
        if str(out.get("torsion_considered", "Não")) == "Sim" and out.get("status") == "OK":
            out["status"] = "Verificar"
            out["detailing_issues"] = "; ".join(x for x in [
                str(out.get("detailing_issues", "")).strip(" -;"),
                "confirmar disposição dos estribos fechados e armadura longitudinal de torção nas subsecções da secção I",
            ] if x)
    return out

BeamDesigner.design_one = _design_one_advanced


def _update_report_advanced(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
        return
    src = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results)
    n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
    n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [
        f"{APP_NAME} {APP_VERSION}\n",
        "Relatório resumido de dimensionamento de vigas\n\n",
        f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n",
        f"Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}\n",
        f"ELS: {self.var_service_case.get().strip() if hasattr(self,'var_service_case') and self.var_service_case.get().strip() else 'todas as combinações reconhecidas'}\n\n",
    ]
    for _, r in src.head(120).iterrows():
        lines.append(f"Viga {_beam_label_labels(r)} | Caso ELU {r.get('case','')} | Piso {r.get('story','')}\n")
        lines.append(f"  Secção: {r.get('section_geometry_summary', r.get('section_type',''))} | Material: {r.get('material','')}\n")
        lines.append(f"  Armaduras: {r.get('solution','')}\n")
        lines.append(
            f"  ELS: {r.get('service_status','')} | combinações={r.get('service_combinations_checked','')} | "
            f"wk: {r.get('service_combination_crack','')} | flecha: {r.get('service_combination_deflection','')} | "
            f"apoios: {r.get('service_support_condition','')}\n"
        )
        lines.append(
            f"  wk={_fmt_report_base(r.get('service_wk_est_mm'),3)}/{_fmt_report_base(r.get('service_wk_lim_mm'),3)} mm; "
            f"flecha={_fmt_report_base(r.get('service_deflection_est_mm'),2)}/{_fmt_report_base(r.get('service_deflection_lim_mm'),2)} mm\n"
        )
        if str(r.get("failure_reason", "")).strip():
            lines.append(f"  Falha: {r.get('failure_reason','')}\n")
        if str(r.get("detailing_issues", "")).strip(" -"):
            lines.append(f"  Pormenorização: {r.get('detailing_issues','')}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))

BeamsEC2App.update_report = _update_report_advanced


_write_pdf_advanced_base = BeamsEC2App._write_pdf

def _write_pdf_advanced(self, path: str):
    if not hasattr(self, "var_pdf_scope") or self.var_pdf_scope.get() != "Relatório técnico":
        return _write_pdf_advanced_base(self, path)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_report_base()
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME; doc.author = ""; doc.subject = APP_SUBJECT
    story = [_program_link_pdf_reporting(styles)]
    n_total=len(results); n_ok=int((results.get("status",pd.Series(dtype=str))=="OK").sum()); n_fail=int((results.get("status",pd.Series(dtype=str))=="Falha").sum())
    meta=[["Data",datetime.now().strftime("%Y-%m-%d %H:%M"),"Suporte normativo","NP EN 1992-1-1:2010"],["Envelopes",str(n_total),"OK/Falhas",f"{n_ok}/{n_fail}"]]
    t=Table(meta,colWidths=[38*mm,90*mm,38*mm,105*mm]); t.setStyle(self._pdf_table_style(header=False))
    story += [t,Spacer(1,5*mm),Paragraph("Relatório técnico",styles["Section"])]
    geom=summary.copy(); geom=geom.assign(viga=geom.apply(_beam_label_labels,axis=1)) if not geom.empty else geom
    env=self.df_env.copy() if self.df_env is not None else pd.DataFrame(); env=env.assign(viga=env.apply(_beam_label_labels,axis=1)) if not env.empty else env
    sls=_sls_audit_df_advanced(summary)
    detail=getattr(self,"df_longitudinal_detailing",pd.DataFrame())
    tests=getattr(self,"df_internal_tests",pd.DataFrame())
    story += [
        Paragraph("Geometria identificada",styles["BodyCourier"]),self._pdf_df_table(geom,["viga","story","section_type","section_geometry_summary","section_geometry_source","section_geometry_confidence","section_fit_error_pct"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Critérios de cálculo",styles["BodyCourier"]),self._pdf_df_table(self._parameters_df(),["Parâmetro","Valor"],max_rows=None,widths=[90,180]),Spacer(1,5*mm),
        Paragraph("Envelopes de esforços",styles["BodyCourier"]),self._pdf_df_table(env,["viga","story","case","section_type","n_points_found","length","material","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Flexão",styles["BodyCourier"]),self._pdf_df_table(_flexure_audit_df_calculation(summary),["viga","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Esforço transverso e torção",styles["BodyCourier"]),self._pdf_df_table(_vt_audit_df_calculation(summary),["viga","case","v_ed_kN","VRd_c_kN","VRd_max_kN","cot_theta_shear","t_ed_kNm","TRd_max_kNm","torsion_considered","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Envelope ELS",styles["BodyCourier"]),self._pdf_df_table(sls,["viga","service_combinations_checked","service_combination_crack","service_combination_deflection","service_combination_steel","service_combination_concrete","service_support_condition","service_history_mode","service_reinforcement_model","service_wk_est_mm","service_wk_lim_mm","service_deflection_est_mm","service_deflection_lim_mm","service_sigma_s_MPa","service_sigma_c_MPa","service_status"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Pormenorização longitudinal",styles["BodyCourier"]),self._pdf_df_table(detail,["viga","bot_rebar","top_rebar","lbd_bot_mm","lbd_top_mm","shift_al_mm","extension_bot_mm","extension_top_mm","min_bottom_bars_into_support","detailing_note"],max_rows=None),Spacer(1,5*mm),
        Paragraph("Testes internos",styles["BodyCourier"]),self._pdf_df_table(tests,["Teste","Calculado","Referência","Erro relativo","Tolerância","Estado"],max_rows=None),
    ]
    if failures is not None and not failures.empty:
        story += [PageBreak(),Paragraph("Falhas",styles["Section"]),self._pdf_df_table(failures.assign(viga=failures.apply(_beam_label_labels,axis=1)),["viga","story","case","failure_type","failure_reason"],max_rows=None)]
    footer_date=datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas,doc_obj):
        canvas.saveState(); canvas.setTitle(APP_NAME); canvas.setAuthor(""); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm,7*mm,f"{APP_NAME} | {footer_date}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_advanced

