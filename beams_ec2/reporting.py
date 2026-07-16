# -*- coding: utf-8 -*-
"""Relatórios, exportações e critérios de apresentação."""

from . import design as _previous
globals().update({k: v for k, v in vars(_previous).items() if not k.startswith("__")})
APP_VERSION = "v0.1"

# Relatórios, exportações e critérios de apresentação
# ============================================================
# ============================================================
NORMATIVE_SUPPORT = (
    "NP EN 1992-1-1:2010 - Eurocódigo 2 - Projecto de estruturas de betão - "
    "Parte 1-1: Regras gerais e regras para edifícios"
)
NORMATIVE_SUPPORT_EXTENDED = (
    NORMATIVE_SUPPORT + " (inclui Anexo Nacional, AC:2012 e A1:2019 quando aplicável)"
)
DEFAULT_AGGREGATE_MM_Reporting = 20.0
DEFAULT_GAMMA_C_Reporting = 1.50
DEFAULT_GAMMA_S_Reporting = 1.15
PDF_SCOPES_Reporting = ["Resumo executivo", "Relatório técnico", "Memória de cálculo"]


def _program_link_pdf_reporting(styles):
    from reportlab.platypus import Paragraph
    return Paragraph(f'<a href="{GITHUB_URL}">{APP_NAME}</a>', styles["ReportTitle"])


def _section_second_moment_y_reporting(section: BeamSection) -> float:
    """Momento de inércia aproximado em torno do eixo forte da viga, em mm4."""
    b = max(float(section.bw_mm), 1.0)
    h = max(float(section.h_mm), 1.0)
    if not section.is_t:
        return b * h ** 3 / 12.0
    bf = max(float(section.bf_mm), b)
    hf = max(min(float(section.hf_mm), h), 0.0)
    hw = max(h - hf, 1.0)
    Af = bf * hf
    Aw = b * hw
    yf = hf / 2.0 if hf > 0 else 0.0
    yw = hf + hw / 2.0
    A = max(Af + Aw, 1e-9)
    ybar = (Af * yf + Aw * yw) / A
    If = bf * hf ** 3 / 12.0 + Af * (yf - ybar) ** 2 if hf > 0 else 0.0
    Iw = b * hw ** 3 / 12.0 + Aw * (yw - ybar) ** 2
    return max(If + Iw, b * h ** 3 / 40.0)


# ELS : substituir o antigo controlo L/d por limite de flecha L/n.
def _serviceability_reporting(self, row: pd.Series, As_bot: float, As_top: float, d_bot: float, d_top: float, section: BeamSection, cp: Dict[str, float]) -> Dict[str, float | str]:
    L = max(finite(row.get("length", 0.0), 0.0) * 1000.0, 1e-9)
    Mserv = finite(row.get("m_abs_ed_kNm", max(finite(row.get("m_pos_ed_kNm"), 0), finite(row.get("m_neg_ed_kNm"), 0))), 0.0)
    As_use = max(As_bot, As_top, 1e-9)
    d_use = max(d_bot, d_top, 1e-9)
    z = 0.9 * d_use
    sigma_s = abs(Mserv) * 1e6 / max(As_use * z, 1e-9)
    wk_est = 0.00085 * sigma_s if sigma_s > 0 else 0.0
    wk_status = "OK" if wk_est <= self.crack_limit_mm else "Verificar fendilhação"

    denom = finite(getattr(self, "deflection_ld_limit", 250.0), 250.0)
    # Compatibilidade com versões anteriores: se vier 20, tratar como não configurado e usar L/250.
    if denom < 100.0:
        denom = 250.0
    denom = max(denom, 1.0)
    Ecm = max(float(cp.get("Ecm", 30000.0)), 1e-9)
    Ig = _section_second_moment_y_reporting(section)
    # Estimativa expedita: rigidez efectiva reduzida para ter em conta fendilhação provável.
    Ieff = 0.50 * Ig if wk_est > 0 else Ig
    delta_est = abs(Mserv) * 1e6 * L ** 2 / max(10.0 * Ecm * Ieff, 1e-9)
    delta_lim = L / denom
    defl_status = "OK" if delta_est <= delta_lim + 1e-9 else "Verificar deformação"

    stress_status = "OK" if sigma_s <= 0.80 * self.fyk else "Verificar tensão no aço"
    service_status = "OK" if wk_status == "OK" and defl_status == "OK" and stress_status == "OK" else "Verificar"
    note = "Combinação ELS identificada" if classify_limit_state(row.get("case", "")) == "ELS" else "Informativo - caso não identificado como ELS"
    return {
        "service_sigma_s_MPa": sigma_s,
        "service_wk_est_mm": wk_est,
        "service_wk_lim_mm": self.crack_limit_mm,
        "service_L_over_d": L / max(d_use, 1e-9),  # mantido para auditoria histórica
        "service_L_over_d_lim": None,
        "service_deflection_est_mm": delta_est,
        "service_deflection_lim_mm": delta_lim,
        "service_deflection_limit": f"L/{denom:.0f}",
        "service_crack_status": wk_status,
        "service_deflection_status": defl_status,
        "service_stress_status": stress_status,
        "service_status": service_status,
        "service_note": note,
        "sls_method": "NP EN 1992-1-1:2010, Secção 7; tensão, fendilhação estimada e limite de flecha L/n",
    }


BeamDesigner.serviceability = _serviceability_reporting


# Auditorias .
def _sls_audit_df_reporting(results: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "member", "case", "combination_number", "limit_state", "story",
        "service_sigma_s_MPa", "service_wk_est_mm", "service_wk_lim_mm",
        "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit",
        "service_crack_status", "service_deflection_status", "service_stress_status", "service_status", "service_note",
    ]
    return results[[c for c in cols if c in results.columns]].copy() if results is not None and not results.empty else pd.DataFrame()


def _calc_memory_df_reporting(results: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if results is None or results.empty:
        return pd.DataFrame(columns=["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"])
    for _, r in results.iterrows():
        base = {
            "Viga": r.get("member", ""),
            "Caso": r.get("case", ""),
            "Piso": r.get("story", ""),
            "Secção": f"{_fmt_report_base(r.get('bw_cm'),0)}x{_fmt_report_base(r.get('h_cm'),0)} cm {r.get('section_type','')}",
        }
        def add(etapa, item, valor, unidade, criterio, nota=""):
            rows.append({**base, "Etapa": etapa, "Item": item, "Valor": valor, "Unidade": unidade, "Critério/Referência": criterio, "Estado/Nota": nota})
        add("Dados", "Material", r.get("material", ""), "-", "NP EN 1992-1-1, Secção 3", r.get("material_source", ""))
        add("Dados", "Comprimento", finite(r.get("length_m")), "m", "entrada")
        add("Esforços", "M+Ed", finite(r.get("m_pos_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_pos_at'))} m")
        add("Esforços", "M-Ed", finite(r.get("m_neg_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_neg_at'))} m")
        add("Esforços", "VEd", finite(r.get("v_ed_kN")), "kN", "envelope", f"x={_fmt_report_base(r.get('v_at'))} m")
        add("Esforços", "TEd", finite(r.get("t_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('t_at'))} m")
        add("Flexão +", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_bot_mm2'),0)} / {_fmt_report_base(r.get('as_prov_bot_mm2'),0)}", "mm²", "NP EN 1992-1-1, 6.1 / 9.2", f"{r.get('bot_rebar','')}; MRd={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_pos'),3)}")
        add("Flexão -", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_top_mm2'),0)} / {_fmt_report_base(r.get('as_prov_top_mm2'),0)}", "mm²", "NP EN 1992-1-1, 6.1 / 9.2", f"{r.get('top_rebar','')}; MRd={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_neg'),3)}")
        add("Corte", "VRd,c / VRd,max", f"{_fmt_report_base(r.get('VRd_c_kN'))} / {_fmt_report_base(r.get('VRd_max_kN'))}", "kN", "NP EN 1992-1-1, 6.2", r.get("shear_status", ""))
        add("Torção", "TEd / TRd,max", f"{_fmt_report_base(r.get('t_ed_kNm'))} / {_fmt_report_base(r.get('TRd_max_kNm'))}", "kNm", "NP EN 1992-1-1, 6.3", r.get("torsion_status", ""))
        add("Corte/Torção", "Asw/s req / prov", f"{_fmt_report_base(r.get('Asw_s_total_req_mm2_per_m'),0)} / {_fmt_report_base(r.get('Asw_s_prov_mm2_per_m'),0)}", "mm²/m", "NP EN 1992-1-1, 6.2 + 6.3")
        add("ELS", "wk", f"{_fmt_report_base(r.get('service_wk_est_mm'),3)} / {_fmt_report_base(r.get('service_wk_lim_mm'),3)}", "mm", "NP EN 1992-1-1, 7.3", r.get("service_crack_status", ""))
        add("ELS", "Flecha estimada / limite", f"{_fmt_report_base(r.get('service_deflection_est_mm'),2)} / {_fmt_report_base(r.get('service_deflection_lim_mm'),2)}", "mm", f"{r.get('service_deflection_limit','L/250')}", r.get("service_deflection_status", ""))
        add("Pormenorização", "Solução", r.get("solution", ""), "-", "NP EN 1992-1-1, Secções 8 e 9.2", r.get("detailing_status", ""))
        add("Resultado", "Estado final", r.get("status", ""), "-", "síntese", r.get("failure_reason", "") or r.get("recommendations", ""))
    return pd.DataFrame(rows)


globals()["_sls_audit_df_report_base"] = _sls_audit_df_reporting
globals()["_calc_memory_df_report_base"] = _calc_memory_df_reporting


# Metadados: sem autor e sem URL visível. O link fica no nome do programa dentro do memória de cálculo/PDF.
def _metadata_df_reporting(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Versão", APP_VERSION],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", self.input_file_path or "-"],
        ["Suporte normativo", NORMATIVE_SUPPORT_EXTENDED],
        ["Âmbito", "Dimensionamento/verificação ELU e ELS expedito de vigas de betão armado"],
        ["Critério de torção", f"Armadura específica apenas se TEd/TRd,max >= {TORSION_RELEVANCE_ETA_Detailing:.2f}"],
        ["Diâmetros principais", "Ø12, Ø16, Ø20, Ø25"],
        ["Diâmetros de estribos", "Ø6, Ø8, Ø10; mínimo 2 ramos"],
        ["Descrição", APP_TABLE_DESCRIPTION],
    ], columns=["Campo", "Valor"])


def _parameters_df_reporting(self) -> pd.DataFrame:
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Relatório técnico"
    if scope == "Completo":
        scope = "Relatório técnico"
    return pd.DataFrame([
        ["Suporte normativo", NORMATIVE_SUPPORT_EXTENDED],
        ["Recobrimento [mm]", self.var_cover.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()],
        ["cotθ", self.var_cot_theta.get()],
        ["wk,lim [mm]", self.var_crack_limit.get()],
        ["Limite de flecha", f"L/{self.var_ld_limit.get()}"],
        ["Momento principal", self.var_moment_axis.get()],
        ["Corte vertical", self.var_shear_axis.get()],
        ["Torção", self.var_torsion_axis.get()],
        ["Diâmetros principais adoptados", "Ø12, Ø16, Ø20, Ø25"],
        ["Diâmetros de estribos adoptados", "Ø6, Ø8, Ø10; mínimo 2 ramos"],
        ["Máximo de camadas principais", MAX_MAIN_LAYERS_Detailing],
        ["Limiar de torção", f"TEd/TRd,max >= {TORSION_RELEVANCE_ETA_Detailing:.2f}"],
        ["Armadura de pele/alma", "reportada em formato nØ/face quando aplicável"],
        ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
        ["Tipo de relatório PDF", scope],
    ], columns=["Parâmetro", "Valor"])


BeamsEC2App._metadata_df = _metadata_df_reporting
BeamsEC2App._parameters_df = _parameters_df_reporting


# Notas normativas sem referência a autor, agregados ou coeficientes parciais no UI.
def _build_normative_notes_reporting() -> pd.DataFrame:
    notes = [
        ("Suporte normativo", NORMATIVE_SUPPORT_EXTENDED, "O programa usa por defeito o Eurocódigo 2 adoptado em Portugal para vigas de betão armado."),
        ("Materiais", "Classe de betão", "A classe de betão é lida da coluna Material; se estiver vazia, é usado C30/37 como fallback interno."),
        ("Flexão", "NP EN 1992-1-1, 6.1", "Dimensionamento à flexão positiva e negativa; MRd é recalculado com a armadura adoptada."),
        ("Esforço transverso", "NP EN 1992-1-1, 6.2", "Cálculo de VRd,c, VRd,max e Asw/s; estribos com mínimo de 2 ramos."),
        ("Torção", "NP EN 1992-1-1, 6.3", "A armadura específica de torção só é considerada quando TEd/TRd,max atinge o limiar de relevância definido."),
        ("ELS", "NP EN 1992-1-1, Secção 7", "Inclui tensão no aço, fendilhação estimada e controlo expedito de flecha por limite L/n definido pelo utilizador."),
        ("Pormenorização", "NP EN 1992-1-1, Secções 8 e 9.2", "Verifica espaçamentos livres, camadas, diâmetros adoptados, estribos e armadura de pele/alma quando aplicável."),
        ("Relatórios", "Critério do utilizador", "O PDF é exportado como Resumo executivo, Relatório técnico ou Memória de cálculo; o memória de cálculo mantém a auditoria completa."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])


globals()["build_normative_notes"] = _build_normative_notes_reporting


# Sidebar : ocultar gamma_c, gamma_s e agregado; remover PDF completo.
def _build_sidebar_reporting(self, parent):
    if not hasattr(self, "var_pdf_scope"):
        self.var_pdf_scope = tk.StringVar(value="Relatório técnico")
    if self.var_pdf_scope.get() == "Completo":
        self.var_pdf_scope.set("Relatório técnico")
    try:
        if finite(self.var_ld_limit.get(), 20.0) < 100:
            self.var_ld_limit.set("250")
    except Exception:
        self.var_ld_limit.set("250")
    # valores internos fixos por defeito do EC2
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")

    hero = ttk.LabelFrame(parent, text="BeamsEC2")
    hero.pack(fill="x", pady=(0, 8))
    link = ttk.Label(hero, text=APP_NAME, style="Header.TLabel", cursor="hand2")
    link.pack(anchor="w")
    link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
    ttk.Label(hero, text="Dimensionamento de vigas de betão armado", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(
        hero,
        text=f"Importa esforços, cria envelopes por viga/caso e dimensiona segundo {NORMATIVE_SUPPORT_EXTENDED}. Exporta workbook memória de cálculo detalhado e relatórios PDF separados.",
        style="Subtle.TLabel", wraplength=340, justify="left",
    ).pack(anchor="w", pady=(2, 0))

    data = ttk.LabelFrame(parent, text="1. Entrada")
    data.pack(fill="x", pady=(0, 8))
    ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

    params = ttk.LabelFrame(parent, text="2. Parâmetros de cálculo")
    params.pack(fill="x", pady=(0, 8))
    self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
    ttk.Label(params, text="Aço fyk [MPa]").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(params, "cotθ", self.var_cot_theta, 2)
    self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 3)
    self._add_label_entry(params, "Limite de flecha L/", self.var_ld_limit, 4)
    ttk.Label(params, text="Momento principal").grid(row=5, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_moment_axis, values=["MY", "MZ"], state="readonly").grid(row=5, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Corte vertical").grid(row=6, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_shear_axis, values=["FZ", "FY"], state="readonly").grid(row=6, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Torção").grid(row=7, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_torsion_axis, values=["MX", "MY", "MZ", "Nenhuma"], state="readonly").grid(row=7, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=9, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    params.columnconfigure(1, weight=1)

    filters = ttk.LabelFrame(parent, text="3. Filtros")
    filters.pack(fill="x", pady=(0, 8))
    ttk.Label(filters, text="Viga/Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha", "Verificar"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "corte", "torcao", "pormenorizacao", "els", "dados", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    filters.columnconfigure(1, weight=1)

    actions = ttk.LabelFrame(parent, text="4. Cálculo e exportação")
    actions.pack(fill="x", pady=(0, 8))
    ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ttk.Label(actions, text="Tipo de PDF").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(actions, textvariable=self.var_pdf_scope, values=PDF_SCOPES_Reporting, state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(actions, text="Exportar .csv", command=self.export_csv).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

    status_box = ttk.LabelFrame(parent, text="5. Estado")
    status_box.pack(fill="x", pady=(0, 8))
    ttk.Label(status_box, textvariable=self.status_var, wraplength=340, justify="left").pack(fill="x", padx=6, pady=(4, 2))
    ttk.Progressbar(status_box, variable=self.progress_var, maximum=100).pack(fill="x", padx=6, pady=(2, 2))
    ttk.Label(status_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

    notes = ttk.LabelFrame(parent, text="6. Notas rápidas")
    notes.pack(fill="x", pady=(0, 8))
    ttk.Label(
        notes,
        text=(
            f"• Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}.\n"
            "• γc, γs e dimensão do agregado são internos e usam valores por defeito do Eurocódigo.\n"
            "• Para vigas, exportar várias estações ao longo da barra melhora o envelope.\n"
            "• Para secções T, preencher BF e HF; caso contrário, a secção é rectangular.\n"
            "• A armadura de alma/pele só é reportada para h >= 40 cm.\n"
            "• O PDF é exportado separadamente como resumo, relatório técnico ou memória de cálculo."
        ),
        wraplength=340, justify="left",
    ).pack(fill="x", padx=6, pady=6)


BeamsEC2App._build_sidebar = _build_sidebar_reporting


# Instruções .
def _build_instructions_tab_reporting(self, parent):
    outer = ttk.Frame(parent, padding=10)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1); outer.columnconfigure(0, weight=1)
    ttk.Label(outer, text="Instruções de utilização", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    host = ttk.Frame(outer); host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(host)
    instructions = (
        "OBJECTIVO DO PROGRAMA\n"
        f"BeamsEC2 dimensiona e verifica vigas de betão armado com suporte normativo fixo: {NORMATIVE_SUPPORT_EXTENDED}. "
        "O cálculo inclui envelopes por viga/caso, flexão positiva e negativa, corte, torção, ELS expedito, pormenorização e exportações .xlsx/.pdf.\n\n"
        "COLUNAS DA FOLHA DE IMPORTAÇÃO TIPO\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
        "UNIDADES ESPERADAS\n"
        "FX, FY, FZ em kN; MX, MY, MZ em kNm; Station e Length em m; dimensões HY/HZ/BF/HF em cm.\n\n"
        "CONVENÇÃO ADOPTADA\n"
        "O eixo longitudinal da viga é o eixo local X. Por defeito, MY é o momento principal, FZ é o esforço transverso vertical e MX é a torção. "
        "MY positivo é tratado como flexão positiva, com armadura inferior; MY negativo é tratado como flexão negativa, com armadura superior.\n\n"
        "PARÂMETROS OCULTOS\n"
        "γc, γs e dimensão do agregado são usados internamente com os valores por defeito do Eurocódigo. O utilizador define apenas os parâmetros práticos necessários para o projecto corrente.\n\n"
        "ELS\n"
        "O limite de deformação é introduzido como denominador L/n. Exemplos usuais: 250 para L/250 ou 500 para L/500.\n\n"
        "RELATÓRIOS\n"
        "O PDF é exportado separadamente como Resumo executivo, Relatório técnico ou Memória de cálculo. A memória de cálculo mantém a auditoria completa, com folhas de entrada, envelopes, resultados, memória, flexão, corte/torção, ELS, pormenorização, falhas, validação e notas EC2.\n"
    )
    txt.insert("1.0", instructions); txt.config(state="disabled")


BeamsEC2App._build_instructions_tab = _build_instructions_tab_reporting


# Validação e run_design .
_old_validate_inputs_reporting = BeamsEC2App.validate_inputs

def _validate_inputs_reporting(self):
    err = _old_validate_inputs_reporting(self)
    if err:
        return err
    if finite(self.var_ld_limit.get(), 0.0) <= 0:
        return "Limite de flecha inválido. Use, por exemplo, 250 para L/250 ou 500 para L/500."
    return None

BeamsEC2App.validate_inputs = _validate_inputs_reporting

_old_run_design_reporting = BeamsEC2App.run_design

def _run_design_reporting(self):
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    if hasattr(self, "var_pdf_scope") and self.var_pdf_scope.get() == "Completo":
        self.var_pdf_scope.set("Relatório técnico")
    if finite(self.var_ld_limit.get(), 20.0) < 100:
        self.var_ld_limit.set("250")
    return _old_run_design_reporting(self)

BeamsEC2App.run_design = _run_design_reporting


# memória de cálculo : manter workbook detalhado, mas sem autor/URL visível e com hiperligação no nome do programa.
_old_write_excel_reporting = BeamsEC2App._write_excel

def _write_excel_reporting(self, path: str):
    _old_write_excel_reporting(self, path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        wb = load_workbook(path)
        props = wb.properties
        props.title = APP_NAME
        props.subject = APP_SUBJECT
        props.creator = ""
        props.lastModifiedBy = ""
        props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY
        props.description = APP_TABLE_DESCRIPTION
        for ws_name in ["00_Info", "01_Parametros"]:
            if ws_name in wb.sheetnames:
                ws = wb[ws_name]
                for row in range(2, ws.max_row + 1):
                    key = str(ws.cell(row, 1).value or "").strip().lower()
                    if key == "programa":
                        cell = ws.cell(row, 2)
                        cell.value = APP_NAME
                        cell.hyperlink = GITHUB_URL
                        cell.font = Font(color="1F4E5F", underline="single", bold=True, name="Segoe UI")
        wb.save(path)
    except Exception:
        pass

BeamsEC2App._write_excel = _write_excel_reporting


# PDF : três relatórios separados, sem autor, sem nota final e com rodapé pedido.
def _write_pdf_reporting(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_report_base()
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Relatório técnico"
    if scope == "Completo" or scope not in PDF_SCOPES_Reporting:
        scope = "Relatório técnico"
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME
    doc.author = ""
    doc.subject = APP_SUBJECT
    story = []
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    n_total = len(results)
    n_ok = int((results["status"] == "OK").sum()) if "status" in results.columns else 0
    n_fail = int((results["status"] == "Falha").sum()) if "status" in results.columns else 0
    story.append(_program_link_pdf_reporting(styles))
    story.append(Paragraph(f"{scope} - {NORMATIVE_SUPPORT}", styles["ReportSubtitle"]))
    meta = [
        ["Programa", APP_NAME, "Tipo de relatório", scope],
        ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Suporte normativo", "NP EN 1992-1-1:2010"],
        ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"],
        ["Diâmetros principais", "Ø12, Ø16, Ø20, Ø25", "Torção", f"ηT >= {TORSION_RELEVANCE_ETA_Detailing:.2f}"],
    ]
    t = Table(meta, colWidths=[38*mm, 90*mm, 38*mm, 105*mm]); t.setStyle(self._pdf_table_style(header=False)); story += [t, Spacer(1, 5*mm)]

    if scope == "Resumo executivo":
        story.extend([
            Paragraph("Resumo executivo", styles["Section"]),
            self._pdf_df_table(_executive_summary_df_report_base(results, summary), ["Indicador", "Valor"], max_rows=25, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Resumo por viga", styles["BodyCourier"]),
            self._pdf_df_table(summary, ["member", "story", "case", "section_type", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm", "bot_rebar", "top_rebar", "skin_rebar", "solution", "status"], max_rows=34),
        ])
    elif scope == "Relatório técnico":
        story.extend([
            Paragraph("Relatório técnico", styles["Section"]),
            Paragraph("Critérios de cálculo", styles["BodyCourier"]),
            self._pdf_df_table(self._parameters_df(), ["Parâmetro", "Valor"], max_rows=45, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Envelopes de esforços", styles["BodyCourier"]),
            self._pdf_df_table(self.df_env, ["member", "story", "case", "n_points_found", "length", "material", "hy", "hz", "bf", "hf", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm"], max_rows=38),
            Spacer(1, 5*mm),
            Paragraph("Flexão", styles["BodyCourier"]),
            self._pdf_df_table(_flexure_audit_df_report_base(summary), ["member", "case", "m_pos_ed_kNm", "mrd_pos_kNm", "eta_m_pos", "bot_rebar", "m_neg_ed_kNm", "mrd_neg_kNm", "eta_m_neg", "top_rebar", "ductility_pos", "ductility_neg"], max_rows=34),
            Spacer(1, 5*mm),
            Paragraph("Esforço transverso e torção", styles["BodyCourier"]),
            self._pdf_df_table(_vt_audit_df_detailing(summary), ["member", "case", "v_ed_kN", "VRd_c_kN", "VRd_max_kN", "t_ed_kNm", "TRd_max_kNm", "eta_torsion_design", "torsion_considered", "Asw_s_total_req_mm2_per_m", "Asw_s_prov_mm2_per_m", "shear_status", "torsion_status"], max_rows=34),
            Spacer(1, 5*mm),
            Paragraph("ELS", styles["BodyCourier"]),
            self._pdf_df_table(_sls_audit_df_reporting(summary), ["member", "case", "service_sigma_s_MPa", "service_wk_est_mm", "service_wk_lim_mm", "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit", "service_status", "service_note"], max_rows=30),
            Spacer(1, 5*mm),
            Paragraph("Pormenorização", styles["BodyCourier"]),
            self._pdf_df_table(_detailing_audit_df_detailing(summary), ["member", "case", "bot_rebar", "top_rebar", "bot_clear_spacing_mm", "top_clear_spacing_mm", "phi_st_mm", "stirrup_legs", "s_st_mm", "skin_rebar", "detailing_status", "detailing_issues"], max_rows=30),
        ])
        if failures is not None and not failures.empty:
            story.extend([PageBreak(), Paragraph("Falhas", styles["Section"]), self._pdf_df_table(failures, ["member", "story", "case", "failure_type", "failure_reason"], max_rows=45)])
    elif scope == "Memória de cálculo":
        mem = _calc_memory_df_reporting(summary)
        story.extend([
            Paragraph("Memória de cálculo", styles["Section"]),
            self._pdf_df_table(mem, ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"], max_rows=105),
        ])

    footer_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setTitle(APP_NAME)
        canvas.setAuthor("")
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"{APP_NAME} | {footer_date}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_reporting


# Modelo de importação : sem autor nos metadados.
def _export_template_reporting(self):
    path = filedialog.asksaveasfilename(title="Guardar modelo de importação", defaultextension=".xlsx", filetypes=[("Ficheiro de tabela", "*.xlsx")])
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    sample = pd.DataFrame([
        {"Member/Node/Case":"B1/101/103 (C)", "Station (m)":"0.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"120", "MX (kNm)":"0.8", "MY (kNm)":"-80", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"25", "HZ (cm)":"40", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
        {"Member/Node/Case":"B1/102/103 (C)", "Station (m)":"3.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"20", "MX (kNm)":"0.3", "MY (kNm)":"95", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"25", "HZ (cm)":"40", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
        {"Member/Node/Case":"B1/103/103 (C)", "Station (m)":"6.00", "FX (kN)":"0", "FY (kN)":"0", "FZ (kN)":"-120", "MX (kNm)":"0.8", "MY (kNm)":"-75", "MZ (kNm)":"0", "Length (m)":"6.00", "Material":"C30/37", "HY (cm)":"25", "HZ (cm)":"40", "BF (cm)":"", "HF (cm)":"", "Name":"V1", "Story":"Piso 1"},
    ])
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sample.to_excel(writer, sheet_name="FOLHA_IMPORTACAO_TIPO", index=False)
            wb = writer.book
            wb.properties.title = f"{APP_NAME} - folha de importação tipo"
            wb.properties.creator = ""
            wb.properties.lastModifiedBy = ""
            ws = wb["FOLHA_IMPORTACAO_TIPO"]
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                fill = PatternFill("solid", fgColor="1F4E5F")
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = Font(color="FFFFFF", bold=True, name="Segoe UI")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = max(12, min(24, len(str(col[0].value)) + 2))
            except Exception:
                pass
        self.status_var.set(f"Modelo de tabela guardado: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível guardar o modelo.\n\n{err}")

BeamsEC2App.export_template = _export_template_reporting



# ============================================================
# ============================================================


def _export_csv_csv_export(self):
    """Exporta a tabela de resultados actualmente disponível para CSV.
    Método recolocado porque a sidebar  mantém o botão Exportar .csv.
    """
    if self.df_results is None or self.df_results.empty:
        messagebox.showwarning("Aviso", "Não há resultados para exportar.")
        return
    path = filedialog.asksaveasfilename(
        title="Exportar CSV",
        defaultextension=".csv",
        filetypes=[("CSV", "*.csv")],
    )
    if not path:
        return
    if not path.lower().endswith(".csv"):
        path += ".csv"
    try:
        source = self.df_filtered if getattr(self, "df_filtered", pd.DataFrame()).shape[0] else self.df_results
        source.to_csv(path, index=False, encoding="utf-8-sig")
        self.status_var.set(f"CSV exportado para: {path}")
    except Exception as err:
        messagebox.showerror("Erro", f"Não foi possível exportar CSV.\n\n{err}")


BeamsEC2App.export_csv = _export_csv_csv_export



# ============================================================
# ============================================================
AUTO_COT_MIN_Calculation = 1.0
AUTO_COT_MAX_Calculation = 2.5
AUTO_COT_STEP_Calculation = 0.05


def _beam_label_calculation(row) -> str:
    member = str(row.get("member", "") or "").strip()
    name = str(row.get("name", "") or "").strip()
    if name and name.lower() not in ["nan", "none", member.lower()]:
        return f"{member} - {name}" if member else name
    return member


def _cot_candidates_calculation():
    n = int(round((AUTO_COT_MAX_Calculation - AUTO_COT_MIN_Calculation) / AUTO_COT_STEP_Calculation))
    vals = [AUTO_COT_MAX_Calculation - i * AUTO_COT_STEP_Calculation for i in range(n + 1)]
    vals.append(AUTO_COT_MIN_Calculation)
    return sorted({round(max(AUTO_COT_MIN_Calculation, min(AUTO_COT_MAX_Calculation, v)), 3) for v in vals}, reverse=True)


def _auto_cot_for_shear_calculation(V_N: float, bw: float, z: float, nu1: float, fcd: float) -> tuple[float, float]:
    """Escolhe o maior cotθ que verifica VEd <= VRd,max, reduzindo cotθ só quando necessário."""
    best_cot = AUTO_COT_MIN_Calculation
    best_vrd = 0.0
    for cot in _cot_candidates_calculation():
        vrd = bw * z * nu1 * fcd / (cot + 1.0 / cot)
        if vrd >= V_N - 1e-9:
            return cot, vrd
        if vrd > best_vrd:
            best_cot, best_vrd = cot, vrd
    return best_cot, best_vrd


def _auto_cot_for_torsion_calculation(T_Nmm: float, Ak: float, tef: float, uk: float, nu1: float, fcd: float, fyd: float) -> tuple[float, float, float, float]:
    """Escolhe cotθ por varrimento entre 1.0 e 2.5, minimizando uma medida prática de armadura total."""
    feasible = []
    best_cot, best_trd, best_asw, best_asl = AUTO_COT_MIN_Calculation, 0.0, 0.0, 0.0
    for cot in _cot_candidates_calculation():
        trd = 2.0 * nu1 * fcd * Ak * tef / (cot + 1.0 / cot)
        asw_s = T_Nmm / max(2.0 * Ak * fyd * cot, 1e-9)
        asl = T_Nmm * uk * cot / max(2.0 * Ak * fyd, 1e-9)
        # score com unidades mistas, apenas para escolher uma solução equilibrada.
        score = asw_s * 1000.0 + 0.002 * asl
        if trd >= T_Nmm - 1e-9:
            feasible.append((score, cot, trd, asw_s, asl))
        if trd > best_trd:
            best_cot, best_trd, best_asw, best_asl = cot, trd, asw_s, asl
    if feasible:
        feasible.sort(key=lambda x: (x[0], -x[1]))
        _, cot, trd, asw_s, asl = feasible[0]
        return cot, trd, asw_s, asl
    return best_cot, best_trd, best_asw, best_asl


def _shear_requirements_calculation(self, VEd_kN: float, bw: float, d: float, Asl: float, fck: float, fcd: float, fyd: float) -> Dict[str, float | str]:
    V = abs(float(VEd_kN or 0.0)) * 1000.0
    z = 0.9 * d
    nu1 = 0.6 * (1.0 - fck / 250.0)
    cot, VRdmax = _auto_cot_for_shear_calculation(V, bw, z, nu1, fcd)
    vrdc = self.vrd_c(VEd_kN, bw, d, Asl, fck, fcd)
    VRdc_N = vrdc["VRd_c_kN"] * 1000.0
    if V <= VRdc_N + 1e-9:
        Asw_s = 0.0
        status = "OK sem armadura adicional por V"
    elif V <= VRdmax + 1e-9:
        Asw_s = V / max(z * fyd * cot, 1e-9)
        status = "Requer armadura de esforço transverso"
    else:
        Asw_s = V / max(z * fyd * cot, 1e-9)
        status = "Não conforme: VEd > VRd,max"
    rho_w_min = 0.08 * math.sqrt(fck) / self.fyk
    Asw_s_min = rho_w_min * bw
    return {
        **vrdc,
        "VRd_max_kN": VRdmax / 1000.0,
        "Asw_s_shear_req_mm2_per_mm": Asw_s,
        "Asw_s_min_mm2_per_mm": Asw_s_min,
        "shear_status": status,
        "cot_theta": cot,
        "cot_theta_shear": cot,
        "cot_theta_source": "calculado automaticamente no intervalo 1.0 <= cotθ <= 2.5",
    }


BeamDesigner.shear_requirements = _shear_requirements_calculation


def _torsion_requirements_calculation(self, TEd_kNm: float, section: BeamSection, fck: float, fcd: float, fyd: float) -> Dict[str, float | str]:
    T = abs(float(TEd_kNm or 0.0)) * 1e6
    b = section.bw_mm
    h = section.h_mm
    if b <= 0 or h <= 0:
        return {
            "TRd_max_kNm": None,
            "Asw_s_torsion_req_mm2_per_mm": 0.0,
            "Asl_torsion_req_mm2": 0.0,
            "torsion_status": "Dados insuficientes",
            "torsion_considered": "Não",
            "eta_torsion_design": 0.0,
            "cot_theta_torsion": None,
            "tef_mm": None,
            "Ak_mm2": None,
            "uk_mm": None,
        }
    A = b * h
    u = 2.0 * (b + h)
    tef = max(A / max(u, 1e-9), 2.0 * self.cover_mm, 50.0)
    tef = min(tef, min(b, h) / 2.0 - 1.0)
    Ak = max((b - tef) * (h - tef), 1.0)
    uk = 2.0 * max(b + h - 2.0 * tef, 1.0)
    nu1 = 0.6 * (1.0 - fck / 250.0)
    cot, TRdmax, Asw_s_auto, Asl_auto = _auto_cot_for_torsion_calculation(T, Ak, tef, uk, nu1, fcd, fyd)
    TRdmax_kNm = TRdmax / 1e6
    eta = (T / max(TRdmax, 1e-9)) if T > 0 else 0.0
    if T <= 1e-9:
        status = "Sem torção relevante"
        considered = "Não"
        Asw_s = 0.0
        Asl = 0.0
    elif eta < TORSION_RELEVANCE_ETA_Detailing:
        status = f"Sem torção relevante (ηT={eta:.3f} < {TORSION_RELEVANCE_ETA_Detailing:.2f})"
        considered = "Não"
        Asw_s = 0.0
        Asl = 0.0
    elif T <= TRdmax + 1e-9:
        status = "Requer verificação/armadura de torção"
        considered = "Sim"
        Asw_s = Asw_s_auto
        Asl = Asl_auto
    else:
        status = "Não conforme: TEd > TRd,max"
        considered = "Sim"
        Asw_s = Asw_s_auto
        Asl = Asl_auto
    return {
        "TRd_max_kNm": TRdmax_kNm,
        "Asw_s_torsion_req_mm2_per_mm": Asw_s,
        "Asl_torsion_req_mm2": Asl,
        "torsion_status": status,
        "torsion_considered": considered,
        "eta_torsion_design": eta,
        "cot_theta_torsion": cot,
        "tef_mm": tef,
        "Ak_mm2": Ak,
        "uk_mm": uk,
    }


BeamDesigner.torsion_requirements = _torsion_requirements_calculation


# ELS : notas técnicas e compatibilidade com combinação indicada.
def _serviceability_calculation(self, row: pd.Series, As_bot: float, As_top: float, d_bot: float, d_top: float, section: BeamSection, cp: Dict[str, float]) -> Dict[str, float | str]:
    L = max(finite(row.get("length", 0.0), 0.0) * 1000.0, 1e-9)
    Mserv = finite(row.get("m_abs_ed_kNm", max(finite(row.get("m_pos_ed_kNm"), 0), finite(row.get("m_neg_ed_kNm"), 0))), 0.0)
    As_use = max(As_bot, As_top, 1e-9)
    d_use = max(d_bot, d_top, 1e-9)
    z = 0.9 * d_use
    sigma_s = abs(Mserv) * 1e6 / max(As_use * z, 1e-9)
    wk_est = 0.00085 * sigma_s if sigma_s > 0 else 0.0
    wk_status = "OK" if wk_est <= self.crack_limit_mm else "Verificar fendilhação"
    denom = finite(getattr(self, "deflection_ld_limit", 250.0), 250.0)
    if denom < 100.0:
        denom = 250.0
    denom = max(denom, 1.0)
    Ecm = max(float(cp.get("Ecm", 30000.0)), 1e-9)
    Ig = _section_second_moment_y_reporting(section)
    Ieff = 0.50 * Ig if wk_est > 0 else Ig
    delta_est = abs(Mserv) * 1e6 * L ** 2 / max(10.0 * Ecm * Ieff, 1e-9)
    delta_lim = L / denom
    defl_status = "OK" if delta_est <= delta_lim + 1e-9 else "Verificar deformação"
    stress_status = "OK" if sigma_s <= 0.80 * self.fyk else "Verificar tensão no aço"
    service_status = "OK" if wk_status == "OK" and defl_status == "OK" and stress_status == "OK" else "Verificar"
    if classify_limit_state(row.get("case", "")) == "ELS":
        note = "Combinação ELS identificada; controlo expedito conforme NP EN 1992-1-1, Secção 7."
    else:
        note = "Sem combinação ELS indicada; verificação informativa com o envelope disponível."
    return {
        "service_sigma_s_MPa": sigma_s,
        "service_wk_est_mm": wk_est,
        "service_wk_lim_mm": self.crack_limit_mm,
        "service_L_over_d": L / max(d_use, 1e-9),
        "service_L_over_d_lim": None,
        "service_deflection_est_mm": delta_est,
        "service_deflection_lim_mm": delta_lim,
        "service_deflection_limit": f"L/{denom:.0f}",
        "service_crack_status": wk_status,
        "service_deflection_status": defl_status,
        "service_stress_status": stress_status,
        "service_status": service_status,
        "service_note": note,
        "sls_method": "NP EN 1992-1-1:2010, Secção 7; tensão, fendilhação estimada e limite de flecha L/n",
    }


BeamDesigner.serviceability = _serviceability_calculation


# Acrescentar campos derivados sem interferir na rotina principal.
_design_one_base_calculation = BeamDesigner.design_one

def _design_one_calculation(self, row: pd.Series) -> Dict:
    out = _design_one_base_calculation(self, row)
    try:
        out["beam_ref"] = _beam_label_calculation(out)
        fck = parse_concrete_strength(out.get("material", DEFAULT_CONCRETE_CLASS))
        cp = concrete_props(fck, alpha_cc=self.alpha_cc, gamma_c=self.gamma_c)
        fyd = steel_props(self.fyk, gamma_s=self.gamma_s)["fyd"]
        section = BeamSection(float(out.get("bw_cm", 0.0) or 0.0) * 10.0, float(out.get("h_cm", 0.0) or 0.0) * 10.0,
                              float(out.get("bf_cm", 0.0) or 0.0) * 10.0, float(out.get("hf_cm", 0.0) or 0.0) * 10.0)
        d_eff = min(float(out.get("d_bot_mm", 0.0) or 0.0), float(out.get("d_top_mm", 0.0) or 0.0))
        sh = self.shear_requirements(float(out.get("v_ed_kN", 0.0) or 0.0), section.bw_mm, d_eff,
                                     max(float(out.get("as_prov_bot_mm2", 0.0) or 0.0), float(out.get("as_prov_top_mm2", 0.0) or 0.0)),
                                     fck, cp["fcd"], fyd)
        tor = self.torsion_requirements(float(out.get("t_ed_kNm", 0.0) or 0.0), section, fck, cp["fcd"], fyd)
        out["cot_theta_shear"] = sh.get("cot_theta_shear", sh.get("cot_theta"))
        out["cot_theta_torsion"] = tor.get("cot_theta_torsion")
        out["cot_theta_source"] = "calculado automaticamente no intervalo 1.0 <= cotθ <= 2.5"
    except Exception:
        pass
    return out


BeamDesigner.design_one = _design_one_calculation


def _apply_service_combination_calculation(app, results: pd.DataFrame) -> pd.DataFrame:
    target = app.var_service_case.get().strip() if hasattr(app, "var_service_case") else ""
    if results is None or results.empty:
        return results
    out = results.copy()
    if not target:
        if "service_note" in out.columns:
            out["service_note"] = out["service_note"].fillna("Sem combinação ELS indicada; verificação informativa com o envelope disponível.")
        out["service_case_source"] = "automático/envelope disponível"
        return out
    env = app.df_env.copy() if getattr(app, "df_env", None) is not None else pd.DataFrame()
    if env.empty:
        out["service_status"] = "Verificar"
        out["service_note"] = f"Combinação de serviço {target} indicada, mas não existem envelopes carregados para ELS."
        out["service_case_source"] = "combinação indicada não encontrada"
        return out
    env["_case_str"] = env.get("case", pd.Series(dtype=str)).astype(str)
    env["_comb_str"] = env.get("combination_number", env["_case_str"].map(extract_combination_number)).astype(str)
    sel = env[(env["_case_str"] == target) | (env["_comb_str"] == target)]
    if sel.empty:
        out["service_status"] = "Verificar"
        out["service_note"] = f"Combinação de serviço {target} indicada, mas não foi encontrada na tabela importada."
        out["service_case_source"] = "combinação indicada não encontrada"
        return out
    # indexar por member/name/story, com fallback por member.
    service_by_key = {}
    service_by_member = {}
    for _, er in sel.iterrows():
        key = (str(er.get("member", "")), str(er.get("name", "")), str(er.get("story", "")))
        service_by_key.setdefault(key, er)
        service_by_member.setdefault(str(er.get("member", "")), er)
    designer = BeamDesigner(
        cover_mm=finite(app.var_cover.get(), 35.0),
        agg_mm=DEFAULT_AGGREGATE_MM_Reporting,
        fyk=finite(app.var_fyk.get(), 500.0),
        gamma_c=DEFAULT_GAMMA_C_Reporting,
        gamma_s=DEFAULT_GAMMA_S_Reporting,
        cot_theta=2.0,
        crack_limit_mm=finite(app.var_crack_limit.get(), 0.30),
        deflection_ld_limit=finite(app.var_ld_limit.get(), 250.0),
        calc_mode="dimensionamento",
    )
    for idx, r in out.iterrows():
        key = (str(r.get("member", "")), str(r.get("name", "")), str(r.get("story", "")))
        # Não usar `A or B` aqui: quando A é uma pandas.Series, o Python tenta avaliar
        # a verdade da Series e lança "The truth value of a Series is ambiguous".
        er = service_by_key.get(key, None)
        if er is None:
            er = service_by_member.get(str(r.get("member", "")), None)
        if er is None:
            out.at[idx, "service_status"] = "Verificar"
            out.at[idx, "service_note"] = f"Combinação de serviço {target} sem envelope correspondente para esta viga."
            out.at[idx, "service_case_source"] = "sem envelope correspondente"
            continue
        try:
            fck = parse_concrete_strength(r.get("material", DEFAULT_CONCRETE_CLASS))
            cp = concrete_props(fck, gamma_c=DEFAULT_GAMMA_C_Reporting)
            section = BeamSection(float(r.get("bw_cm", 0.0) or 0.0) * 10.0, float(r.get("h_cm", 0.0) or 0.0) * 10.0,
                                  float(r.get("bf_cm", 0.0) or 0.0) * 10.0, float(r.get("hf_cm", 0.0) or 0.0) * 10.0)
            els = designer.serviceability(er, float(r.get("as_prov_bot_mm2", 0.0) or 0.0), float(r.get("as_prov_top_mm2", 0.0) or 0.0),
                                          float(r.get("d_bot_mm", 0.0) or 0.0), float(r.get("d_top_mm", 0.0) or 0.0), section, cp)
            for k, v in els.items():
                out.at[idx, k] = v
            out.at[idx, "service_combination"] = target
            out.at[idx, "service_m_abs_kNm"] = finite(er.get("m_abs_ed_kNm"), 0.0)
            out.at[idx, "service_case_source"] = f"combinação indicada pelo utilizador: {target}"
            out.at[idx, "service_note"] = "Combinação de serviço indicada pelo utilizador; controlo expedito conforme NP EN 1992-1-1, Secção 7."
        except Exception as exc:
            out.at[idx, "service_status"] = "Verificar"
            out.at[idx, "service_note"] = f"Não foi possível recalcular ELS para a combinação {target}: {exc}"
            out.at[idx, "service_case_source"] = "erro no recálculo ELS"
    return out


# Auditorias com nome da viga e campos cotθ automáticos.
def _sls_audit_df_calculation(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame()
    df = results.copy()
    df["viga"] = df.apply(_beam_label_calculation, axis=1)
    cols = [
        "viga", "case", "combination_number", "limit_state", "story",
        "service_sigma_s_MPa", "service_wk_est_mm", "service_wk_lim_mm",
        "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit",
        "service_crack_status", "service_deflection_status", "service_stress_status", "service_status", "service_note",
    ]
    return df[[c for c in cols if c in df.columns]].copy()


def _vt_audit_df_calculation(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame()
    df = results.copy()
    df["viga"] = df.apply(_beam_label_calculation, axis=1)
    cols = ["viga","case","story","v_ed_kN","v_at","VRd_c_kN","VRd_max_kN","cot_theta_shear","eta_shear_max","Asw_s_shear_req_mm2_per_m","Asw_s_min_mm2_per_m","t_ed_kNm","t_at","TRd_max_kNm","cot_theta_torsion","eta_torsion_max","eta_torsion_design","torsion_considered","Asw_s_torsion_req_mm2_per_m","Asl_torsion_req_mm2","Asw_s_total_req_mm2_per_m","phi_st_mm","stirrup_legs","s_st_mm","Asw_s_prov_mm2_per_m","shear_status","torsion_status","stirrup_status"]
    return df[[c for c in cols if c in df.columns]].copy()


def _detailing_audit_df_calculation(results: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame()
    df = results.copy()
    df["viga"] = df.apply(_beam_label_calculation, axis=1)
    cols = ["viga","case","story","bot_rebar","top_rebar","bot_layers","top_layers","bot_bars_per_layer","top_bars_per_layer","bot_clear_spacing_mm","top_clear_spacing_mm","phi_st_mm","stirrup_legs","s_st_mm","skin_rebar","skin_reinf_face_mm2","skin_reinf_face_prov_mm2","detailing_status","detailing_issues","solution"]
    return df[[c for c in cols if c in df.columns]].copy()


def _flexure_audit_df_calculation(results: pd.DataFrame) -> pd.DataFrame:
    base = _flexure_audit_df_report_base(results)
    if base is None or base.empty:
        return pd.DataFrame()
    source = results.copy()
    source["viga"] = source.apply(_beam_label_calculation, axis=1)
    base = base.copy()
    if "member" in base.columns:
        labels = source[["member", "case", "viga"]].drop_duplicates()
        base = base.merge(labels, on=["member", "case"], how="left")
    return base


globals()["_sls_audit_df_report_base"] = _sls_audit_df_calculation
globals()["_vt_audit_df_report_base"] = _vt_audit_df_calculation
globals()["_detailing_audit_df_report_base"] = _detailing_audit_df_calculation


# Metadados e parâmetros sem autor, sem tipo de relatório, sem diâmetros principais e sem resumo de torção.
def _metadata_df_calculation(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME],
        ["Versão", APP_VERSION],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Ficheiro de origem", self.input_file_path or "-"],
        ["Suporte normativo", NORMATIVE_SUPPORT_EXTENDED],
        ["Âmbito", "Dimensionamento/verificação ELU e ELS expedito de vigas de betão armado"],
        ["Descrição", APP_TABLE_DESCRIPTION],
    ], columns=["Campo", "Valor"])


def _parameters_df_calculation(self) -> pd.DataFrame:
    service_case = self.var_service_case.get().strip() if hasattr(self, "var_service_case") else ""
    return pd.DataFrame([
        ["Suporte normativo", NORMATIVE_SUPPORT_EXTENDED],
        ["Recobrimento [mm]", self.var_cover.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()],
        ["cotθ", "calculado automaticamente por caso no intervalo 1.0 <= cotθ <= 2.5"],
        ["wk,lim [mm]", self.var_crack_limit.get()],
        ["Limite de flecha", f"L/{self.var_ld_limit.get()}"],
        ["Combinação ELS", service_case or "automática/informativa"],
        ["Momento principal", self.var_moment_axis.get()],
        ["Corte vertical", self.var_shear_axis.get()],
        ["Torção", self.var_torsion_axis.get()],
        ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
    ], columns=["Parâmetro", "Valor"])


BeamsEC2App._metadata_df = _metadata_df_calculation
BeamsEC2App._parameters_df = _parameters_df_calculation


# Sidebar sem campo cotθ e com combinação ELS opcional.
def _build_sidebar_calculation(self, parent):
    if not hasattr(self, "var_pdf_scope"):
        self.var_pdf_scope = tk.StringVar(value="Relatório técnico")
    if not hasattr(self, "var_service_case"):
        self.var_service_case = tk.StringVar(value="")
    if self.var_pdf_scope.get() == "Completo":
        self.var_pdf_scope.set("Relatório técnico")
    try:
        if finite(self.var_ld_limit.get(), 20.0) < 100:
            self.var_ld_limit.set("250")
    except Exception:
        self.var_ld_limit.set("250")
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    self.var_cot_theta.set("2.0")

    hero = ttk.LabelFrame(parent, text="BeamsEC2")
    hero.pack(fill="x", pady=(0, 8))
    link = ttk.Label(hero, text=APP_NAME, style="Header.TLabel", cursor="hand2")
    link.pack(anchor="w")
    link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
    ttk.Label(hero, text="Dimensionamento de vigas de betão armado", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(hero, text=f"Importa esforços, cria envelopes por viga/caso e dimensiona segundo {NORMATIVE_SUPPORT_EXTENDED}.", style="Subtle.TLabel", wraplength=340, justify="left").pack(anchor="w", pady=(2, 0))

    data = ttk.LabelFrame(parent, text="1. Entrada")
    data.pack(fill="x", pady=(0, 8))
    ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

    params = ttk.LabelFrame(parent, text="2. Parâmetros de cálculo")
    params.pack(fill="x", pady=(0, 8))
    self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
    ttk.Label(params, text="Aço fyk [MPa]").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 2)
    self._add_label_entry(params, "Limite de flecha L/", self.var_ld_limit, 3)
    ttk.Label(params, text="Combinação ELS").grid(row=4, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(params, textvariable=self.var_service_case).grid(row=4, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="opcional; exemplo: 103", style="Subtle.TLabel").grid(row=5, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 4))
    ttk.Label(params, text="Momento principal").grid(row=6, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_moment_axis, values=["MY", "MZ"], state="readonly").grid(row=6, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Corte vertical").grid(row=7, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_shear_axis, values=["FZ", "FY"], state="readonly").grid(row=7, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(params, text="Torção").grid(row=8, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_torsion_axis, values=["MX", "MY", "MZ", "Nenhuma"], state="readonly").grid(row=8, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=9, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=10, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    params.columnconfigure(1, weight=1)

    filters = ttk.LabelFrame(parent, text="3. Filtros")
    filters.pack(fill="x", pady=(0, 8))
    ttk.Label(filters, text="Viga/Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha", "Verificar"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "corte", "torcao", "pormenorizacao", "els", "dados", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    filters.columnconfigure(1, weight=1)

    actions = ttk.LabelFrame(parent, text="4. Cálculo e exportação")
    actions.pack(fill="x", pady=(0, 8))
    ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ttk.Label(actions, text="PDF").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(actions, textvariable=self.var_pdf_scope, values=PDF_SCOPES_Reporting, state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(actions, text="Exportar .csv", command=self.export_csv).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

    status_box = ttk.LabelFrame(parent, text="5. Estado")
    status_box.pack(fill="x", pady=(0, 8))
    ttk.Label(status_box, textvariable=self.status_var, wraplength=340, justify="left").pack(fill="x", padx=6, pady=(4, 2))
    ttk.Progressbar(status_box, variable=self.progress_var, maximum=100).pack(fill="x", padx=6, pady=(2, 2))
    ttk.Label(status_box, textvariable=self.progress_text_var, anchor="e").pack(fill="x", padx=6, pady=(0, 4))

    notes = ttk.LabelFrame(parent, text="6. Notas rápidas")
    notes.pack(fill="x", pady=(0, 8))
    ttk.Label(notes, text=(
        f"• Suporte normativo: {NORMATIVE_SUPPORT_EXTENDED}.\n"
        "• γc, γs, cotθ e dimensão do agregado são tratados internamente.\n"
        "• cotθ é calculado por caso no intervalo 1.0 a 2.5.\n"
        "• Para ELS, pode indicar a combinação de serviço; em branco usa o cálculo informativo por envelope.\n"
        "• Para secções T, preencher BF e HF; caso contrário, a secção é rectangular."
    ), wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)


BeamsEC2App._build_sidebar = _build_sidebar_calculation


# Instruções actualizadas.
def _build_instructions_tab_calculation(self, parent):
    outer = ttk.Frame(parent, padding=10)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1); outer.columnconfigure(0, weight=1)
    ttk.Label(outer, text="Instruções de utilização", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    host = ttk.Frame(outer); host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(host)
    instructions = (
        "OBJECTIVO DO PROGRAMA\n"
        f"BeamsEC2 dimensiona e verifica vigas de betão armado com suporte normativo fixo: {NORMATIVE_SUPPORT_EXTENDED}. "
        "O cálculo inclui envelopes por viga/caso, flexão positiva e negativa, esforço transverso, torção, ELS expedito, pormenorização e exportações .xlsx/.pdf.\n\n"
        "COLUNAS DA FOLHA DE IMPORTAÇÃO TIPO\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
        "UNIDADES ESPERADAS\n"
        "FX, FY, FZ em kN; MX, MY, MZ em kNm; Station e Length em m; dimensões HY/HZ/BF/HF em cm.\n\n"
        "ELS\n"
        "O utilizador pode introduzir o número da combinação de serviço relevante. Se o campo ficar vazio, o programa mantém a verificação informativa por envelope. "
        "O limite de deformação é introduzido como denominador L/n; por exemplo, 250 para L/250 ou 500 para L/500.\n\n"
        "COTθ\n"
        "O valor de cotθ deixa de ser introduzido pelo utilizador e é calculado automaticamente por caso no intervalo 1.0 <= cotθ <= 2.5.\n\n"
        "RELATÓRIOS\n"
        "O PDF é exportado como Resumo executivo, Relatório técnico ou Memória de cálculo. A memória de cálculo mantém a auditoria completa por envelope.\n"
    )
    txt.insert("1.0", instructions); txt.config(state="disabled")


BeamsEC2App._build_instructions_tab = _build_instructions_tab_calculation


# Relatório textual interno: member + nome da viga, sem informação construtiva desnecessária de torção.
def _update_report_calculation(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
        return
    source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results)
    n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
    n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [
        f"{APP_NAME} {APP_VERSION}\n",
        f"Dimensionamento de vigas - {NORMATIVE_SUPPORT}\n\n",
        f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n",
        f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | cotθ: automático\n",
        f"ELS: combinação {self.var_service_case.get().strip() if hasattr(self, 'var_service_case') and self.var_service_case.get().strip() else 'automática/informativa'}\n\n",
    ]
    for _, r in source.head(120).iterrows():
        sec = f"{finite(r.get('bw_cm')):.0f} x {finite(r.get('h_cm')):.0f} cm"
        story = str(r.get('story','') or '')
        lines.append(f"Viga {_beam_label_calculation(r)} | Caso {r.get('case','')} | Piso {story}\n")
        lines.append(f"  Secção: {sec} | {r.get('section_type','')} | Material: {r.get('material','')}\n")
        lines.append(f"  Esforços: M+Ed={_fmt_report_base(r.get('m_pos_ed_kNm'),2)} kNm; M-Ed={_fmt_report_base(r.get('m_neg_ed_kNm'),2)} kNm; VEd={_fmt_report_base(r.get('v_ed_kN'),2)} kN; TEd={_fmt_report_base(r.get('t_ed_kNm'),2)} kNm\n")
        lines.append(f"  Flexão: MRd+={_fmt_report_base(r.get('mrd_pos_kNm'),2)} kNm (η={_fmt_report_base(r.get('eta_m_pos'),3)}); MRd-={_fmt_report_base(r.get('mrd_neg_kNm'),2)} kNm (η={_fmt_report_base(r.get('eta_m_neg'),3)})\n")
        lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'),2)} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'),2)} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'),2)} kNm; cotθ,V={_fmt_report_base(r.get('cot_theta_shear'),2)}\n")
        skin = str(r.get('skin_rebar','') or 'n/a')
        skin_txt = f" | Pele/alma: {skin}" if skin and skin.lower() != 'n/a' else ""
        stirrup_prefix = "Estribos fechados" if str(r.get('torsion_considered','Não')) == 'Sim' else "Estribos"
        lines.append(f"  Armaduras: Inf. {r.get('bot_rebar','')} | Sup. {r.get('top_rebar','')} | {stirrup_prefix} Ø{_fmt_report_base(r.get('phi_st_mm'),0)}/{_fmt_report_base(r.get('stirrup_legs'),0)}r // {_fmt_report_base(finite(r.get('s_st_mm'))/10,1)} cm{skin_txt}\n")
        lines.append(f"  ELS: {r.get('service_status','')} | Nota: {r.get('service_note','')} | Estado: {r.get('status','')}\n")
        motivo = str(r.get("failure_reason", "") or "").strip()
        if motivo:
            lines.append(f"  Motivo: {motivo}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))


BeamsEC2App.update_report = _update_report_calculation


# Run design com ELS por combinação do utilizador antes do resumo.
def _run_design_calculation(self):
    self.var_agg.set(f"{DEFAULT_AGGREGATE_MM_Reporting:.0f}")
    self.var_gamma_c.set(f"{DEFAULT_GAMMA_C_Reporting:.2f}")
    self.var_gamma_s.set(f"{DEFAULT_GAMMA_S_Reporting:.2f}")
    self.var_cot_theta.set("2.0")
    if hasattr(self, "var_pdf_scope") and self.var_pdf_scope.get() == "Completo":
        self.var_pdf_scope.set("Relatório técnico")
    if finite(self.var_ld_limit.get(), 20.0) < 100:
        self.var_ld_limit.set("250")
    err = self.validate_inputs()
    if err:
        messagebox.showwarning("Aviso", err)
        return
    designer = BeamDesigner(
        cover_mm=finite(self.var_cover.get(), 35.0),
        agg_mm=DEFAULT_AGGREGATE_MM_Reporting,
        fyk=finite(self.var_fyk.get(), 500.0),
        gamma_c=DEFAULT_GAMMA_C_Reporting,
        gamma_s=DEFAULT_GAMMA_S_Reporting,
        cot_theta=2.0,
        crack_limit_mm=finite(self.var_crack_limit.get(), 0.30),
        deflection_ld_limit=finite(self.var_ld_limit.get(), 250.0),
        calc_mode=self.var_calc_mode.get(),
    )
    input_df = reduce_to_governing_cases(self.df_env) if self.var_reduce_cases.get() else self.df_env.copy()
    self.df_calc_input = input_df.copy()
    self.progress_var.set(0.0)
    self.status_var.set("Análise em curso...")

    def progress(done, total):
        pct = 0.0 if total <= 0 else 100.0 * done / total
        self.after(0, lambda p=pct: self.progress_var.set(p))
        self.after(0, lambda d=done, t=total: self.status_var.set(f"A calcular... {d}/{t} envelopes"))

    def worker():
        try:
            results = designer.design_dataframe(input_df, progress_callback=progress)
            results = _apply_service_combination_calculation(self, results)
            summary = build_summary_by_member(results) if self.var_summary.get() else pd.DataFrame()
            failures = results[results["status"] == "Falha"].copy() if "status" in results.columns else pd.DataFrame()
            ok = results[results["status"] == "OK"].copy() if "status" in results.columns else pd.DataFrame()
            validation = build_data_validation(self.df_clean, self.df_env, results)
            def finish():
                self.df_results = results
                self.df_summary = summary
                self.df_failures = failures
                self.df_ok = ok
                self.df_validation = validation
                self.df_notes = build_normative_notes()
                self.df_filtered = pd.DataFrame()
                self.show_df(self.tree_results, self.df_results)
                self.show_df(self.tree_summary, self.df_summary)
                self.show_df(self.tree_failures, self.df_failures)
                self.show_df(self.tree_shortlists, self.build_shortlists_df())
                self.show_df(self.tree_validation, self.df_validation)
                self.show_df(self.tree_notes, self.df_notes)
                self.update_report()
                self.progress_var.set(100.0)
                self.status_var.set(f"Cálculo concluído: {len(results)} envelopes; {len(summary)} vigas resumidas; {len(failures)} falhas.")
            self.after(0, finish)
        except Exception as err:
            msg = str(err)
            self.after(0, lambda m=msg: messagebox.showerror("Erro", m))
            self.after(0, lambda: self.status_var.set("Falha na análise."))
            self.after(0, lambda: self.progress_var.set(0.0))
    threading.Thread(target=worker, daemon=True).start()


BeamsEC2App.run_design = _run_design_calculation


# PDF : só o nome do programa no topo; sem tipo de relatório, sem diâmetros principais e sem linha geral de torção.
def _write_pdf_calculation(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_report_base()
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Relatório técnico"
    if scope == "Completo" or scope not in PDF_SCOPES_Reporting:
        scope = "Relatório técnico"
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME
    doc.author = ""
    doc.subject = APP_SUBJECT
    story = []
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    n_total = len(results)
    n_ok = int((results["status"] == "OK").sum()) if "status" in results.columns else 0
    n_fail = int((results["status"] == "Falha").sum()) if "status" in results.columns else 0
    story.append(_program_link_pdf_reporting(styles))
    meta = [
        ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Suporte normativo", "NP EN 1992-1-1:2010"],
        ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"],
    ]
    t = Table(meta, colWidths=[38*mm, 90*mm, 38*mm, 105*mm])
    t.setStyle(self._pdf_table_style(header=False))
    story += [t, Spacer(1, 5*mm)]

    if scope == "Resumo executivo":
        exec_df = _executive_summary_df_report_base(results, summary)
        story.extend([
            Paragraph("Resumo executivo", styles["Section"]),
            self._pdf_df_table(exec_df, ["Indicador", "Valor"], max_rows=25, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Resumo por viga", styles["BodyCourier"]),
            self._pdf_df_table(summary.assign(viga=summary.apply(_beam_label_calculation, axis=1)), ["viga", "story", "case", "section_type", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm", "bot_rebar", "top_rebar", "skin_rebar", "solution", "status"], max_rows=34),
        ])
    elif scope == "Relatório técnico":
        flex_df = _flexure_audit_df_calculation(summary)
        story.extend([
            Paragraph("Relatório técnico", styles["Section"]),
            Paragraph("Critérios de cálculo", styles["BodyCourier"]),
            self._pdf_df_table(self._parameters_df(), ["Parâmetro", "Valor"], max_rows=35, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Envelopes de esforços", styles["BodyCourier"]),
            self._pdf_df_table(self.df_env.assign(viga=self.df_env.apply(_beam_label_calculation, axis=1)) if self.df_env is not None and not self.df_env.empty else pd.DataFrame(), ["viga", "story", "case", "n_points_found", "length", "material", "hy", "hz", "bf", "hf", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm"], max_rows=38),
            Spacer(1, 5*mm),
            Paragraph("Flexão", styles["BodyCourier"]),
            self._pdf_df_table(flex_df, ["viga", "case", "m_pos_ed_kNm", "mrd_pos_kNm", "eta_m_pos", "bot_rebar", "m_neg_ed_kNm", "mrd_neg_kNm", "eta_m_neg", "top_rebar", "ductility_pos", "ductility_neg"], max_rows=34),
            Spacer(1, 5*mm),
            Paragraph("Esforço transverso e torção", styles["BodyCourier"]),
            self._pdf_df_table(_vt_audit_df_calculation(summary), ["viga", "case", "v_ed_kN", "VRd_c_kN", "VRd_max_kN", "cot_theta_shear", "t_ed_kNm", "TRd_max_kNm", "cot_theta_torsion", "torsion_considered", "Asw_s_total_req_mm2_per_m", "Asw_s_prov_mm2_per_m", "shear_status", "torsion_status"], max_rows=34),
            Spacer(1, 5*mm),
            Paragraph("ELS", styles["BodyCourier"]),
            self._pdf_df_table(_sls_audit_df_calculation(summary), ["viga", "case", "service_sigma_s_MPa", "service_wk_est_mm", "service_wk_lim_mm", "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit", "service_status", "service_note"], max_rows=30),
            Spacer(1, 5*mm),
            Paragraph("Pormenorização", styles["BodyCourier"]),
            self._pdf_df_table(_detailing_audit_df_calculation(summary), ["viga", "case", "bot_rebar", "top_rebar", "bot_clear_spacing_mm", "top_clear_spacing_mm", "phi_st_mm", "stirrup_legs", "s_st_mm", "skin_rebar", "detailing_status", "detailing_issues"], max_rows=30),
        ])
        if failures is not None and not failures.empty:
            story.extend([PageBreak(), Paragraph("Falhas", styles["Section"]), self._pdf_df_table(failures.assign(viga=failures.apply(_beam_label_calculation, axis=1)), ["viga", "story", "case", "failure_type", "failure_reason"], max_rows=45)])
    elif scope == "Memória de cálculo":
        mem = _calc_memory_df_reporting(summary)
        story.extend([
            Paragraph("Memória de cálculo", styles["Section"]),
            self._pdf_df_table(mem, ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"], max_rows=105),
        ])

    footer_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setTitle(APP_NAME)
        canvas.setAuthor("")
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"{APP_NAME} | {footer_date}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


BeamsEC2App._write_pdf = _write_pdf_calculation


# memória de cálculo: aplicar hyperlink invisível e retirar autor; deixar parâmetros limpos.
def _write_excel_calculation(self, path: str):
    _old_write_excel_reporting(self, path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        wb = load_workbook(path)
        props = wb.properties
        props.title = APP_NAME
        props.subject = APP_SUBJECT
        props.creator = ""
        props.lastModifiedBy = ""
        props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY
        props.description = APP_TABLE_DESCRIPTION
        for ws_name in ["00_Info", "01_Parametros"]:
            if ws_name in wb.sheetnames:
                ws = wb[ws_name]
                for row in range(2, ws.max_row + 1):
                    key = str(ws.cell(row, 1).value or "").strip().lower()
                    if key == "programa":
                        cell = ws.cell(row, 2)
                        cell.value = APP_NAME
                        cell.hyperlink = GITHUB_URL
                        cell.font = Font(color="1F4E5F", underline="single", bold=True, name="Segoe UI")
                    if key in ["autor", "tipo de relatório pdf", "diâmetros principais", "critério de torção", "limiar de torção"]:
                        ws.delete_rows(row, 1)
                        break
        wb.save(path)
    except Exception:
        pass


BeamsEC2App._write_excel = _write_excel_calculation


# Validação final .
_old_validate_inputs_calculation_base = BeamsEC2App.validate_inputs

def _validate_inputs_calculation(self):
    err = _old_validate_inputs_calculation_base(self)
    if err:
        return err
    if finite(self.var_ld_limit.get(), 0.0) <= 0:
        return "Limite de flecha inválido. Use, por exemplo, 250 para L/250 ou 500 para L/500."
    return None


BeamsEC2App.validate_inputs = _validate_inputs_calculation



# ============================================================
# ============================================================
SPACING_CANDIDATES_Stirrups = [75.0, 100.0, 125.0, 150.0, 175.0, 200.0, 250.0]

# Reforçar a lista de espaçamentos admissíveis em toda a aplicação.
_old_beamdesigner_init_stirrups_base = BeamDesigner.__init__

def _beamdesigner_init_stirrups(self, *args, **kwargs):
    _old_beamdesigner_init_stirrups_base(self, *args, **kwargs)
    self.long_diams = list(MAIN_LONG_DIAMS_Detailing)
    self.stirrup_diams = list(STIRRUP_DIAMS_Detailing)
    self.stirrup_legs = list(STIRRUP_LEGS_Detailing)
    self.spacing_candidates_mm = list(SPACING_CANDIDATES_Stirrups)

BeamDesigner.__init__ = _beamdesigner_init_stirrups


def _choose_stirrups_stirrups(self, Asw_s_total: float, b: float, h: float, d: float, torsion: bool) -> Dict[str, float | str]:
    """Escolha construtiva dos estribos.

    Regras adoptadas:
    - espaçamentos normalizados: 7.5, 10, 12.5, 15, 17.5, 20 e 25 cm;
    - mínimo de 2 ramos;
    - preferir 2 ramos sempre que verificam a área requerida;
    - para vigas correntes, preferir Ø8 a soluções menos intuitivas como 4Ø6;
    - aumentar para Ø10 ou para 4/6 ramos apenas quando necessário.
    """
    Asw_s_total = max(0.0, float(Asw_s_total or 0.0))
    s_lim_v = min(0.75 * d, 600.0)
    s_lim_t = min((2.0 * (b + h)) / 8.0, min(b, h), 350.0) if torsion else 999.0
    s_lim = min(s_lim_v, s_lim_t)
    allowed_s = [s for s in SPACING_CANDIDATES_Stirrups if s <= s_lim + 1e-9]
    if not allowed_s:
        allowed_s = [75.0]

    candidates = []
    for legs in STIRRUP_LEGS_Detailing:
        if legs < 2:
            continue
        for phi in STIRRUP_DIAMS_Detailing:
            area = legs * bar_area_mm2(phi)
            for s in sorted(allowed_s, reverse=True):
                provided = area / s
                if provided + 1e-12 < Asw_s_total:
                    continue
                # Score construtivo:
                # 1) máximo espaçamento admissível; 2) menor número de ramos; 3) Ø8 como solução corrente preferencial;
                # 4) depois Ø10 e, por fim, Ø6 quando de facto for a solução mais coerente.
                phi_preference = {8.0: 0, 10.0: 1, 6.0: 2}.get(float(phi), 3)
                excess = max(0.0, provided - Asw_s_total)
                score = (-s, legs, phi_preference, excess, phi)
                cand = {
                    "phi_st_mm": phi,
                    "stirrup_legs": legs,
                    "s_st_mm": s,
                    "Asw_s_prov_mm2_per_mm": provided,
                    "s_lim_mm": s_lim,
                    "stirrup_status": "OK",
                    "stirrup_selection_note": "Escolha construtiva: 2 ramos preferidos; Ø8 adoptado como diâmetro corrente quando suficiente.",
                }
                candidates.append((score, cand))

    if candidates:
        candidates.sort(key=lambda x: x[0])
        return candidates[0][1]

    # Se não couber com a lista normalizada, adoptar a solução máxima no menor espaçamento da lista e assinalar.
    phi = STIRRUP_DIAMS_Detailing[-1]
    legs = STIRRUP_LEGS_Detailing[-1]
    s = min(allowed_s)
    area = legs * bar_area_mm2(phi)
    return {
        "phi_st_mm": phi,
        "stirrup_legs": legs,
        "s_st_mm": s,
        "Asw_s_prov_mm2_per_mm": area / s,
        "s_lim_mm": s_lim,
        "stirrup_status": "Não cabe",
        "stirrup_selection_note": "Área requerida superior à solução máxima normalizada; rever secção ou esforços.",
    }

BeamDesigner.choose_stirrups = _choose_stirrups_stirrups


def _build_instructions_tab_stirrups(self, parent):
    outer = ttk.Frame(parent, padding=10)
    outer.pack(fill="both", expand=True)
    outer.rowconfigure(1, weight=1)
    outer.columnconfigure(0, weight=1)
    ttk.Label(outer, text="Instruções de utilização", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    host = ttk.Frame(outer)
    host.grid(row=1, column=0, sticky="nsew")
    txt = self._make_text_view(host)
    instructions = (
        "BEAMSEC2\n"
        f"Ferramenta para dimensionamento e verificação de vigas de betão armado segundo {NORMATIVE_SUPPORT_EXTENDED}.\n\n"
        "1. Preparação da tabela\n"
        "Importe ou cole a tabela de esforços exportada do modelo de cálculo. A tabela deve identificar a viga, o nó ou estação, o caso de carga/combinação, os esforços internos, o comprimento e a geometria da secção.\n\n"
        "2. Unidades esperadas\n"
        "Esforços axiais e transversos em kN; momentos e torção em kNm; Station e Length em m; dimensões HY, HZ, BF e HF em cm. A classe de betão deve estar na coluna Material, por exemplo C30/37.\n\n"
        "3. Convenção de cálculo\n"
        "O utilizador escolhe o momento principal, o corte vertical e o momento torsor nas opções da esquerda. Para vigas correntes, use MY como momento principal quando esse for o eixo de flexão do modelo, FY ou FZ como corte vertical conforme a convenção local, e MX como torção.\n\n"
        "4. Envelopes\n"
        "O programa agrupa os resultados por viga e combinação, procurando os valores governantes de flexão positiva, flexão negativa, esforço transverso e torção. Exportar várias estações ao longo da viga melhora a qualidade do envelope.\n\n"
        "5. Estados limites de serviço\n"
        "O campo Combinação ELS é opcional. Quando preenchido, o programa recalcula o ELS com essa combinação. Quando vazio, o ELS é tratado como verificação informativa a partir do envelope disponível. O limite de flecha é definido como L/n, por exemplo 250 para L/250 ou 500 para L/500.\n\n"
        "6. Pormenorização\n"
        "A armadura longitudinal principal usa Ø12, Ø16, Ø20 e Ø25. Os estribos usam Ø6, Ø8 ou Ø10, com mínimo de 2 ramos e espaçamentos normalizados de 7.5, 10, 12.5, 15, 17.5, 20 e 25 cm. A escolha evita soluções congestionadas ou pouco construtivas.\n\n"
        "7. Exportação\n"
        "O ficheiro memória de cálculo contém a auditoria detalhada. O PDF pode ser exportado como Resumo executivo, Relatório técnico ou Memória de cálculo, conforme o objectivo da entrega.\n"
    )
    txt.insert("1.0", instructions)
    txt.config(state="disabled")

BeamsEC2App._build_instructions_tab = _build_instructions_tab_stirrups


def _build_normative_notes_stirrups(self) -> pd.DataFrame:
    base = _build_normative_notes_calculation(self) if '_build_normative_notes_calculation' in globals() else pd.DataFrame()
    extra = pd.DataFrame([
        ("Estribos", "Pormenorização", "A selecção privilegia 2 ramos e Ø8 em vigas correntes, aumentando diâmetro/ramos apenas quando a área requerida o exige."),
        ("Espaçamentos", "Normalização construtiva", "Os espaçamentos adoptados para estribos são 7.5, 10, 12.5, 15, 17.5, 20 e 25 cm."),
    ], columns=["Tema", "Referência", "Nota"])
    if base is None or base.empty:
        return extra
    return pd.concat([base, extra], ignore_index=True)

BeamsEC2App.build_normative_notes = _build_normative_notes_stirrups



# ============================================================
# e melhoria da largura das tabelas de memória de cálculo.
# ============================================================


def _beam_label_labels(row) -> str:
    """Identificação amigável para relatórios: member + nome da viga quando disponível.
    Se o nome não existir, mantém a referência como 'Viga <member>' para evitar
    que o relatório pareça conter apenas um identificador interno.
    """
    member = str(row.get("member", "") or "").strip()
    name = str(row.get("name", "") or "").strip()
    invalid = {"", "nan", "none", "-"}
    if name.lower() not in invalid and name.lower() != member.lower():
        return f"{member} - {name}" if member else name
    if member:
        return member if member.lower().startswith("viga") else f"Viga {member}"
    return "-"

# Reaponta o label global usado pelos relatórios /.
globals()["_beam_label_calculation"] = _beam_label_labels


def _calc_memory_df_labels(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"]
    rows = []
    if results is None or results.empty:
        return pd.DataFrame(columns=cols)
    for _, r in results.iterrows():
        base = {
            "Viga": _beam_label_labels(r),
            "Caso": r.get("case", ""),
            "Piso": r.get("story", ""),
            "Secção": f"{_fmt_report_base(r.get('bw_cm'),0)}x{_fmt_report_base(r.get('h_cm'),0)} cm {r.get('section_type','')}",
        }
        def add(etapa, item, valor, unidade, criterio, nota=""):
            rows.append({**base, "Etapa": etapa, "Item": item, "Valor": valor, "Unidade": unidade, "Critério/Referência": criterio, "Estado/Nota": nota})
        add("Dados", "Material", r.get("material", ""), "-", "NP EN 1992-1-1, Secção 3", r.get("material_source", ""))
        add("Dados", "Comprimento", finite(r.get("length_m")), "m", "entrada")
        add("Esforços", "M+Ed", finite(r.get("m_pos_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_pos_at'))} m")
        add("Esforços", "M-Ed", finite(r.get("m_neg_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_neg_at'))} m")
        add("Esforços", "VEd", finite(r.get("v_ed_kN")), "kN", "envelope", f"x={_fmt_report_base(r.get('v_at'))} m")
        add("Esforços", "TEd", finite(r.get("t_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('t_at'))} m")
        add("Flexão +", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_bot_mm2'),0)} / {_fmt_report_base(r.get('as_prov_bot_mm2'),0)}", "mm²", "NP EN 1992-1-1, 6.1 / 9.2", f"{r.get('bot_rebar','')}; MRd={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_pos'),3)}")
        add("Flexão -", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_top_mm2'),0)} / {_fmt_report_base(r.get('as_prov_top_mm2'),0)}", "mm²", "NP EN 1992-1-1, 6.1 / 9.2", f"{r.get('top_rebar','')}; MRd={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_neg'),3)}")
        add("Corte", "VRd,c / VRd,max", f"{_fmt_report_base(r.get('VRd_c_kN'))} / {_fmt_report_base(r.get('VRd_max_kN'))}", "kN", "NP EN 1992-1-1, 6.2", r.get("shear_status", ""))
        add("Torção", "TEd / TRd,max", f"{_fmt_report_base(r.get('t_ed_kNm'))} / {_fmt_report_base(r.get('TRd_max_kNm'))}", "kNm", "NP EN 1992-1-1, 6.3", r.get("torsion_status", ""))
        add("Corte/Torção", "Asw/s req / prov", f"{_fmt_report_base(r.get('Asw_s_total_req_mm2_per_m'),0)} / {_fmt_report_base(r.get('Asw_s_prov_mm2_per_m'),0)}", "mm²/m", "NP EN 1992-1-1, 6.2 + 6.3")
        add("ELS", "wk", f"{_fmt_report_base(r.get('service_wk_est_mm'),3)} / {_fmt_report_base(r.get('service_wk_lim_mm'),3)}", "mm", "NP EN 1992-1-1, 7.3", r.get("service_crack_status", ""))
        add("ELS", "Flecha estimada / limite", f"{_fmt_report_base(r.get('service_deflection_est_mm'),2)} / {_fmt_report_base(r.get('service_deflection_lim_mm'),2)}", "mm", f"{r.get('service_deflection_limit','L/250')}", r.get("service_deflection_status", ""))
        add("Pormenorização", "Solução", r.get("solution", ""), "-", "NP EN 1992-1-1, Secções 8 e 9.2", r.get("detailing_status", ""))
        add("Resultado", "Estado final", r.get("status", ""), "-", "síntese", r.get("failure_reason", "") or r.get("recommendations", ""))
    return pd.DataFrame(rows, columns=cols)

# Reaponta as funções usadas pelos PDFs e memória de cálculo.
globals()["_calc_memory_df_reporting"] = _calc_memory_df_labels
globals()["_calc_memory_df_report_base"] = _calc_memory_df_labels


_old_pdf_df_table_labels_base = BeamsEC2App._pdf_df_table

def _pdf_df_table_labels(self, df: pd.DataFrame, cols: List[str], max_rows: int = 30, widths=None):
    """Define larguras mais adequadas para a memória de cálculo.
    Evita cabeçalhos partidos como 'Critério/Referênci\na' e torna a coluna Viga
    legível quando inclui member + nome.
    """
    memory_cols = ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"]
    if widths is None and list(cols) == memory_cols:
        widths = [28, 20, 18, 28, 26, 32, 42, 16, 38, 22]
    return _old_pdf_df_table_labels_base(self, df, cols, max_rows=max_rows, widths=widths)

BeamsEC2App._pdf_df_table = _pdf_df_table_labels


# Pequena melhoria no relatório textual interno: usar sempre a referência completa da viga.
def _update_report_labels(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
        return
    src = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results)
    n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
    n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [
        f"{APP_NAME} {APP_VERSION}\n",
        "Relatório resumido de dimensionamento de vigas\n\n",
        f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n",
        f"Suporte normativo: NP EN 1992-1-1:2010\n",
        f"ELS: combinação {self.var_service_case.get().strip() if hasattr(self, 'var_service_case') and self.var_service_case.get().strip() else 'automática/informativa'}\n\n",
    ]
    for _, r in src.head(80).iterrows():
        story = r.get("story", "")
        lines.append(f"Viga {_beam_label_labels(r)} | Caso {r.get('case','')} | Piso {story}\n")
        lines.append(f"  Secção: {_fmt_report_base(r.get('bw_cm'),0)} x {_fmt_report_base(r.get('h_cm'),0)} cm | {r.get('section_type','')} | Material: {r.get('material','')}\n")
        lines.append(f"  Esforços: M+Ed={_fmt_report_base(r.get('m_pos_ed_kNm'))} kNm; M-Ed={_fmt_report_base(r.get('m_neg_ed_kNm'))} kNm; VEd={_fmt_report_base(r.get('v_ed_kN'))} kN; TEd={_fmt_report_base(r.get('t_ed_kNm'))} kNm\n")
        lines.append(f"  Flexão: MRd+={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_pos'),3)}); MRd-={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_neg'),3)})\n")
        lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'))} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'))} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'))} kNm\n")
        lines.append(f"  Armaduras: {r.get('solution','')}\n")
        lines.append(f"  ELS: {r.get('service_status','')} | Nota: {r.get('service_note','')} | Estado: {r.get('status','')}\n")
        if str(r.get('failure_reason','')).strip():
            lines.append(f"  Falha: {r.get('failure_reason','')}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))

BeamsEC2App.update_report = _update_report_labels



# ============================================================
# ============================================================


def _all_rows_pdf_limit_full_reports(max_rows=None):
    """Normaliza o limite de linhas para tabelas PDF.
    None significa exportar todas as linhas disponíveis.
    """
    return None if max_rows in (None, "all", "ALL", 0, -1) else max_rows


_old_pdf_df_table_full_reports_base = BeamsEC2App._pdf_df_table

def _pdf_df_table_full_reports(self, df: pd.DataFrame, cols: List[str], max_rows=None, widths=None):
    """Tabela PDF com suporte explícito para exportação integral.

    As versões anteriores usavam limites fixos por secção do relatório técnico
    e da memória de cálculo. Em obras com muitas vigas, as últimas vigas do
    resumo, por exemplo V4, podiam não aparecer nos PDFs, apesar de existirem
    no memória de cálculo. Esta versão permite max_rows=None para incluir todas as linhas.
    """
    memory_cols = ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"]
    if widths is None and list(cols) == memory_cols:
        widths = [30, 20, 18, 28, 26, 32, 42, 16, 38, 20]
    return _old_pdf_df_table_full_reports_base(self, df, cols, max_rows=_all_rows_pdf_limit_full_reports(max_rows), widths=widths)

BeamsEC2App._pdf_df_table = _pdf_df_table_full_reports


def _write_pdf_full_reports(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak

    styles = _pdf_styles_report_base()
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Relatório técnico"
    if scope == "Completo" or scope not in PDF_SCOPES_Reporting:
        scope = "Relatório técnico"

    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME
    doc.author = ""
    doc.subject = APP_SUBJECT

    story = []
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    n_total = len(results)
    n_ok = int((results["status"] == "OK").sum()) if "status" in results.columns else 0
    n_fail = int((results["status"] == "Falha").sum()) if "status" in results.columns else 0

    story.append(_program_link_pdf_reporting(styles))
    meta = [
        ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Suporte normativo", "NP EN 1992-1-1:2010"],
        ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"],
    ]
    t = Table(meta, colWidths=[38*mm, 90*mm, 38*mm, 105*mm])
    t.setStyle(self._pdf_table_style(header=False))
    story += [t, Spacer(1, 5*mm)]

    if scope == "Resumo executivo":
        exec_df = _executive_summary_df_report_base(results, summary)
        summary_pdf = summary.assign(viga=summary.apply(_beam_label_labels, axis=1)) if summary is not None and not summary.empty else pd.DataFrame()
        story.extend([
            Paragraph("Resumo executivo", styles["Section"]),
            self._pdf_df_table(exec_df, ["Indicador", "Valor"], max_rows=None, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Resumo por viga", styles["BodyCourier"]),
            self._pdf_df_table(summary_pdf, ["viga", "story", "case", "section_type", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm", "bot_rebar", "top_rebar", "skin_rebar", "solution", "status"], max_rows=None),
        ])

    elif scope == "Relatório técnico":
        flex_df = _flexure_audit_df_calculation(summary)
        env_pdf = self.df_env.assign(viga=self.df_env.apply(_beam_label_labels, axis=1)) if self.df_env is not None and not self.df_env.empty else pd.DataFrame()
        story.extend([
            Paragraph("Relatório técnico", styles["Section"]),
            Paragraph("Critérios de cálculo", styles["BodyCourier"]),
            self._pdf_df_table(self._parameters_df(), ["Parâmetro", "Valor"], max_rows=None, widths=[90, 180]),
            Spacer(1, 5*mm),
            Paragraph("Envelopes de esforços", styles["BodyCourier"]),
            self._pdf_df_table(env_pdf, ["viga", "story", "case", "n_points_found", "length", "material", "hy", "hz", "bf", "hf", "m_pos_ed_kNm", "m_neg_ed_kNm", "v_ed_kN", "t_ed_kNm"], max_rows=None),
            Spacer(1, 5*mm),
            Paragraph("Flexão", styles["BodyCourier"]),
            self._pdf_df_table(flex_df, ["viga", "case", "m_pos_ed_kNm", "mrd_pos_kNm", "eta_m_pos", "bot_rebar", "m_neg_ed_kNm", "mrd_neg_kNm", "eta_m_neg", "top_rebar", "ductility_pos", "ductility_neg"], max_rows=None),
            Spacer(1, 5*mm),
            Paragraph("Esforço transverso e torção", styles["BodyCourier"]),
            self._pdf_df_table(_vt_audit_df_calculation(summary), ["viga", "case", "v_ed_kN", "VRd_c_kN", "VRd_max_kN", "cot_theta_shear", "t_ed_kNm", "TRd_max_kNm", "cot_theta_torsion", "torsion_considered", "Asw_s_total_req_mm2_per_m", "Asw_s_prov_mm2_per_m", "shear_status", "torsion_status"], max_rows=None),
            Spacer(1, 5*mm),
            Paragraph("ELS", styles["BodyCourier"]),
            self._pdf_df_table(_sls_audit_df_calculation(summary), ["viga", "case", "service_sigma_s_MPa", "service_wk_est_mm", "service_wk_lim_mm", "service_deflection_est_mm", "service_deflection_lim_mm", "service_deflection_limit", "service_status", "service_note"], max_rows=None),
            Spacer(1, 5*mm),
            Paragraph("Pormenorização", styles["BodyCourier"]),
            self._pdf_df_table(_detailing_audit_df_calculation(summary), ["viga", "case", "bot_rebar", "top_rebar", "bot_clear_spacing_mm", "top_clear_spacing_mm", "phi_st_mm", "stirrup_legs", "s_st_mm", "skin_rebar", "detailing_status", "detailing_issues"], max_rows=None),
        ])
        if failures is not None and not failures.empty:
            story.extend([
                PageBreak(),
                Paragraph("Falhas", styles["Section"]),
                self._pdf_df_table(failures.assign(viga=failures.apply(_beam_label_labels, axis=1)), ["viga", "story", "case", "failure_type", "failure_reason"], max_rows=None),
            ])

    elif scope == "Memória de cálculo":
        mem = _calc_memory_df_labels(summary)
        story.extend([
            Paragraph("Memória de cálculo", styles["Section"]),
            self._pdf_df_table(mem, ["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"], max_rows=None),
        ])

    footer_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setTitle(APP_NAME)
        canvas.setAuthor("")
        canvas.setSubject(APP_SUBJECT)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(12*mm, 7*mm, f"{APP_NAME} | {footer_date}")
        canvas.drawRightString(285*mm, 7*mm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_full_reports


def _build_normative_notes_full_reports(self) -> pd.DataFrame:
    base = _build_normative_notes_stirrups(self) if '_build_normative_notes_stirrups' in globals() else pd.DataFrame()
    extra = pd.DataFrame([
        ("PDF", "Exportação integral", "Os PDFs deixam de truncar tabelas por limite fixo de linhas; todas as vigas resumidas e todos os envelopes aplicáveis são exportados."),
    ], columns=["Tema", "Referência", "Nota"])
    if base is None or base.empty:
        return extra
    return pd.concat([base, extra], ignore_index=True)

BeamsEC2App.build_normative_notes = _build_normative_notes_full_reports



