# -*- coding: utf-8 -*-
"""Dimensionamento, pormenorização e auditoria."""

from . import base as _previous
globals().update({k: v for k, v in vars(_previous).items() if not k.startswith("__")})
APP_VERSION = "v0.1.1"

# Dimensionamento, pormenorização e auditoria
# ============================================================
# ============================================================
APP_SUBJECT = "Dimensionamento e verificação de vigas de betão armado segundo a NP EN 1992-1-1"
APP_TABLE_DESCRIPTION = (
    "Workbook técnico detalhado de cálculo de vigas de betão armado segundo a NP EN 1992-1-1, "
    "com entrada, validação, envelopes, memória de cálculo, flexão, esforço transverso, torção, ELS, "
    "pormenorização, resumo executivo, falhas e notas normativas."
)


def _fmt_report_base(value, digits=2, empty="-"):
    try:
        if value is None or pd.isna(value):
            return empty
        if isinstance(value, (int, float)):
            v = float(value)
            if not math.isfinite(v):
                return empty
            if abs(v) >= 1000:
                return f"{v:.0f}"
            return f"{v:.{digits}f}"
        return str(value) if str(value).strip() else empty
    except Exception:
        return empty


# Capacidade de flexão corrigida: a armadura comprimida é limitada pelo equilíbrio de forças.
def _flexural_capacity_report_base(self, As_tens: float, As_comp: float, section: BeamSection, d_mm: float, d2_mm: float, fcd: float, fyd: float) -> Dict[str, float]:
    if As_tens <= 0 or d_mm <= 0:
        return {"MRd_kNm": 0.0, "x_mm": 0.0, "z_mm": 0.0, "capacity_note": "sem armadura traccionada"}
    T = max(0.0, As_tens) * fyd
    Cs_yield = max(0.0, As_comp) * fyd
    Cs_eff = min(Cs_yield, 0.85 * T)
    Cc_req = max(T - Cs_eff, 0.0)
    try:
        a = self.a_for_concrete_force(Cc_req, section, fcd)
        Cc, y = self.compression_block(a, section, fcd)
    except Exception:
        return {"MRd_kNm": 0.0, "x_mm": None, "z_mm": None, "capacity_note": "compressão de betão insuficiente"}
    M = Cc * max(d_mm - y, 0.0) + Cs_eff * max(d_mm - d2_mm, 0.0)
    note = "OK" if Cs_yield <= Cs_eff + 1e-6 else "aço comprimido limitado por equilíbrio"
    return {"MRd_kNm": M / 1e6, "x_mm": a / 0.8 if a else 0.0, "z_mm": d_mm - y, "capacity_note": note}

BeamDesigner.flexural_capacity = _flexural_capacity_report_base

_original_design_one_v10 = BeamDesigner.design_one

def _design_one_report_base(self, row: pd.Series) -> Dict:
    out = _original_design_one_v10(self, row)
    if not isinstance(out, dict):
        return out
    out["norma"] = "NP EN 1992-1-1"
    out["design_basis"] = "EC2 Portugal — norma fixa por defeito"
    out["calc_version"] = APP_VERSION
    out["flexure_method"] = "EC2 6.1; bloco rectangular equivalente; MRd com armadura adoptada"
    out["shear_method"] = "EC2 6.2; VRd,c, VRd,max e Asw/s por treliça de ângulo variável"
    out["torsion_method"] = "EC2 6.3; secção de parede fina equivalente; estribos fechados + Asl"
    out["sls_method"] = "EC2 7; tensão, fendilhação estimada e L/d"
    out["detailing_method"] = "EC2 8 e 9.2; pormenorização construtiva"
    out["eta_flexure_max"] = max(finite(out.get("eta_m_pos")), finite(out.get("eta_m_neg")))
    out["eta_shear_max"] = finite(out.get("v_ed_kN")) / max(finite(out.get("VRd_max_kN")), 1e-9)
    out["eta_torsion_max"] = finite(out.get("t_ed_kNm")) / max(finite(out.get("TRd_max_kNm")), 1e-9)
    out["eta_global"] = max(finite(out.get("eta_flexure_max")), finite(out.get("eta_shear_max")), finite(out.get("eta_torsion_max")))
    recs = [r.strip() for r in str(out.get("recommendations", "") or "").split(";") if r.strip()]
    if str(out.get("service_status", "")).lower().startswith("verificar"):
        recs.append("confirmar ELS com combinação de serviço e cálculo detalhado")
    if str(out.get("detailing_status", "")).lower() != "ok":
        recs.append("rever pormenorização construtiva antes da emissão")
    out["recommendations"] = "; ".join(dict.fromkeys(recs))
    return out

BeamDesigner.design_one = _design_one_report_base


def build_normative_notes_report_base() -> pd.DataFrame:
    notes = [
        ("Norma", "NP EN 1992-1-1", "O programa calcula por defeito apenas com o Eurocódigo 2 adoptado em Portugal; não existe selector de normas."),
        ("Materiais", "EC2 3", "A classe de betão é lida da coluna Material; C30/37 é fallback interno se o campo estiver vazio."),
        ("Flexão", "EC2 6.1", "Dimensionamento para momento positivo e negativo; MRd é recalculado com a armadura adoptada."),
        ("Secções T", "EC2 5/6", "A secção só é tratada como T quando BF e HF são fornecidos; a largura efectiva não é calculada automaticamente."),
        ("Esforço transverso", "EC2 6.2", "Inclui VRd,c, VRd,max, armadura mínima e Asw/s requerido."),
        ("Torção", "EC2 6.3", "Inclui TRd,max, Asw/s de torção e armadura longitudinal adicional."),
        ("Corte + torção", "EC2 6.2/6.3", "A armadura transversal adoptada cobre corte + torção + mínimos."),
        ("ELS", "EC2 7", "Controlo expedito de tensão no aço, fendilhação estimada e L/d."),
        ("Pormenorização", "EC2 8 e 9.2", "Controla camadas, espaçamento livre, estribos e armadura de pele."),
        ("Auditoria", "Memória de cálculo", "A memória de cálculo e o PDF incluem dados, esforços, resistências, utilizações e recomendações."),
    ]
    return pd.DataFrame(notes, columns=["Tema", "Referência", "Nota"])

build_normative_notes = build_normative_notes_report_base


def build_data_validation_report_base(df_clean: pd.DataFrame, df_env: pd.DataFrame, df_results: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    rows = []
    df = df_clean if df_clean is not None else pd.DataFrame()
    env = df_env if df_env is not None else pd.DataFrame()
    required = ["member", "case", "fx", "fy", "fz", "mx", "my", "mz", "length", "material", "hy", "hz", "name"]
    for c in required:
        ok = not df.empty and c in df.columns
        rows.append({"Categoria": "Colunas obrigatórias", "Item": c, "Estado": "OK" if ok else "Não conforme", "Resultado": "presente" if ok else "em falta", "Nota": "necessária para cálculo EC2" if ok else "corrigir cabeçalho"})
    if not df.empty:
        rows.append({"Categoria": "Tabela", "Item": "linhas importadas", "Estado": "OK", "Resultado": len(df), "Nota": "linhas lidas"})
        if "member" in df.columns:
            rows.append({"Categoria": "Tabela", "Item": "vigas distintas", "Estado": "OK", "Resultado": df["member"].astype(str).nunique(), "Nota": "members detectados"})
        for c, unit, low, high in [("length", "m", 0.0, 40.0), ("hy", "cm", 10.0, 200.0), ("hz", "cm", 10.0, 250.0)]:
            if c in df.columns:
                vals = pd.to_numeric(df[c], errors="coerce")
                bad = int(((vals <= low) | (vals > high) | vals.isna()).sum())
                rows.append({"Categoria": "Unidades", "Item": f"{c} [{unit}]", "Estado": "OK" if bad == 0 else "Verificar", "Resultado": bad, "Nota": f"intervalo prático esperado ({low}; {high}]"})
        if "material" in df.columns:
            bad_mat = int((~df["material"].astype(str).str.contains(r"C\s*\d+\s*/\s*\d+", case=False, regex=True, na=False)).sum())
            rows.append({"Categoria": "Materiais", "Item": "classe de betão", "Estado": "OK" if bad_mat == 0 else "Verificar", "Resultado": bad_mat, "Nota": "formato esperado C30/37"})
    if not env.empty:
        rows.append({"Categoria": "Envelopes", "Item": "member/case", "Estado": "OK", "Resultado": len(env), "Nota": "envelopes criados"})
        n_single = int((env.get("n_points_found", pd.Series(dtype=float)).fillna(0).astype(float) < 2).sum()) if "n_points_found" in env.columns else 0
        rows.append({"Categoria": "Envelopes", "Item": "casos com <2 estações", "Estado": "OK" if n_single == 0 else "Verificar", "Resultado": n_single, "Nota": "para vigas, recomenda-se várias estações ao longo do vão"})
        if "bf" in env.columns and "hf" in env.columns:
            inc = int(((env.get("bf", pd.Series(dtype=float)).fillna(0) > 0) ^ (env.get("hf", pd.Series(dtype=float)).fillna(0) > 0)).sum())
            rows.append({"Categoria": "Secções T", "Item": "BF/HF incompletos", "Estado": "OK" if inc == 0 else "Verificar", "Resultado": inc, "Nota": "para secção T, preencher BF e HF"})
    if df_results is not None and not df_results.empty:
        n_fail = int((df_results.get("status", pd.Series(dtype=str)) == "Falha").sum())
        n_sls = int(df_results.get("service_status", pd.Series(dtype=str)).astype(str).str.contains("Verificar", case=False, na=False).sum()) if "service_status" in df_results.columns else 0
        rows.append({"Categoria": "Cálculo", "Item": "casos calculados", "Estado": "OK", "Resultado": len(df_results), "Nota": "linhas processadas"})
        rows.append({"Categoria": "Cálculo", "Item": "falhas ELU", "Estado": "OK" if n_fail == 0 else "Verificar", "Resultado": n_fail, "Nota": "consultar Falhas"})
        rows.append({"Categoria": "Cálculo", "Item": "avisos ELS", "Estado": "OK" if n_sls == 0 else "Verificar", "Resultado": n_sls, "Nota": "verificar casos de serviço"})
    return pd.DataFrame(rows)

build_data_validation = build_data_validation_report_base


def _executive_summary_df_report_base(results: pd.DataFrame, summary: pd.DataFrame) -> pd.DataFrame:
    if results is None or results.empty:
        return pd.DataFrame(columns=["Indicador", "Valor", "Nota"])
    n_total = len(results)
    n_ok = int((results.get("status", pd.Series(dtype=str)) == "OK").sum())
    n_fail = int((results.get("status", pd.Series(dtype=str)) == "Falha").sum())
    n_sls = int(results.get("service_status", pd.Series(dtype=str)).astype(str).str.contains("Verificar", case=False, na=False).sum()) if "service_status" in results.columns else 0
    n_vigas = results.get("member", pd.Series(dtype=str)).astype(str).nunique() if "member" in results.columns else 0
    eta_max = results.get("eta_global", pd.Series(dtype=float)).apply(lambda x: finite(x, 0.0)).max() if "eta_global" in results.columns else 0.0
    return pd.DataFrame([
        ["Norma de cálculo", "NP EN 1992-1-1", "configuração fixa"],
        ["Vigas distintas", n_vigas, "por member"],
        ["Envelopes analisados", n_total, "member/case/name/story"],
        ["Casos OK", n_ok, "ELU verificado"],
        ["Falhas ELU", n_fail, "consultar Falhas"],
        ["Avisos ELS", n_sls, "confirmar em casos críticos"],
        ["Utilização global máxima", float(eta_max), "máximo entre flexão, VEd/VRd,max e TEd/TRd,max"],
        ["Resumo por viga", len(summary) if summary is not None else 0, "caso governante por viga"],
    ], columns=["Indicador", "Valor", "Nota"])


def _calc_memory_df_report_base(results: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if results is None or results.empty:
        return pd.DataFrame(columns=["Viga", "Caso", "Piso", "Secção", "Etapa", "Item", "Valor", "Unidade", "Critério/Referência", "Estado/Nota"])
    for _, r in results.iterrows():
        base = {"Viga": r.get("member", ""), "Caso": r.get("case", ""), "Piso": r.get("story", ""), "Secção": f"{_fmt_report_base(r.get('bw_cm'),0)}x{_fmt_report_base(r.get('h_cm'),0)} cm {r.get('section_type','')}"}
        def add(etapa, item, valor, unidade, criterio, nota=""):
            rows.append({**base, "Etapa": etapa, "Item": item, "Valor": valor, "Unidade": unidade, "Critério/Referência": criterio, "Estado/Nota": nota})
        add("Dados", "Material", r.get("material", ""), "-", "EC2 3", r.get("material_source", ""))
        add("Dados", "Comprimento", finite(r.get("length_m")), "m", "entrada")
        add("Esforços", "M+Ed", finite(r.get("m_pos_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_pos_at'))} m")
        add("Esforços", "M-Ed", finite(r.get("m_neg_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('m_neg_at'))} m")
        add("Esforços", "VEd", finite(r.get("v_ed_kN")), "kN", "envelope", f"x={_fmt_report_base(r.get('v_at'))} m")
        add("Esforços", "TEd", finite(r.get("t_ed_kNm")), "kNm", "envelope", f"x={_fmt_report_base(r.get('t_at'))} m")
        add("Flexão +", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_bot_mm2'),0)} / {_fmt_report_base(r.get('as_prov_bot_mm2'),0)}", "mm²", "EC2 6.1 / 9.2", f"{r.get('bot_rebar','')}; MRd={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_pos'),3)}")
        add("Flexão -", "As,req / As,prov", f"{_fmt_report_base(r.get('as_req_top_mm2'),0)} / {_fmt_report_base(r.get('as_prov_top_mm2'),0)}", "mm²", "EC2 6.1 / 9.2", f"{r.get('top_rebar','')}; MRd={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm; η={_fmt_report_base(r.get('eta_m_neg'),3)}")
        add("Corte", "VRd,c / VRd,max", f"{_fmt_report_base(r.get('VRd_c_kN'))} / {_fmt_report_base(r.get('VRd_max_kN'))}", "kN", "EC2 6.2", r.get("shear_status", ""))
        add("Torção", "TEd / TRd,max", f"{_fmt_report_base(r.get('t_ed_kNm'))} / {_fmt_report_base(r.get('TRd_max_kNm'))}", "kNm", "EC2 6.3", r.get("torsion_status", ""))
        add("Corte/Torção", "Asw/s req / prov", f"{_fmt_report_base(r.get('Asw_s_total_req_mm2_per_m'),0)} / {_fmt_report_base(r.get('Asw_s_prov_mm2_per_m'),0)}", "mm²/m", "EC2 6.2 + 6.3")
        add("ELS", "wk / L/d", f"{_fmt_report_base(r.get('service_wk_est_mm'),3)} / {_fmt_report_base(r.get('service_L_over_d'),2)}", "mm / -", "EC2 7", r.get("service_status", ""))
        add("Pormenorização", "Solução", r.get("solution", ""), "-", "EC2 8/9.2", r.get("detailing_status", ""))
        add("Resultado", "Estado final", r.get("status", ""), "-", "síntese", r.get("failure_reason", "") or r.get("recommendations", ""))
    return pd.DataFrame(rows)


def _cols_existing(df, cols):
    return [c for c in cols if df is not None and not df.empty and c in df.columns]


def _flexure_audit_df_report_base(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","story","section_type","bw_cm","h_cm","bf_cm","hf_cm","m_pos_ed_kNm","m_pos_at","m_neg_ed_kNm","m_neg_at","as_req_bot_mm2","as_prov_bot_mm2","bot_rebar","mrd_pos_kNm","eta_m_pos","x_pos_mm","ductility_pos","as_req_top_mm2","as_prov_top_mm2","top_rebar","mrd_neg_kNm","eta_m_neg","x_neg_mm","ductility_neg","flexure_method"]
    return results[_cols_existing(results, cols)].copy() if results is not None and not results.empty else pd.DataFrame()


def _vt_audit_df_report_base(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","story","v_ed_kN","v_at","VRd_c_kN","VRd_max_kN","eta_shear_max","Asw_s_shear_req_mm2_per_m","Asw_s_min_mm2_per_m","t_ed_kNm","t_at","TRd_max_kNm","eta_torsion_max","Asw_s_torsion_req_mm2_per_m","Asl_torsion_req_mm2","Asw_s_total_req_mm2_per_m","phi_st_mm","stirrup_legs","s_st_mm","Asw_s_prov_mm2_per_m","shear_status","torsion_status","stirrup_status"]
    return results[_cols_existing(results, cols)].copy() if results is not None and not results.empty else pd.DataFrame()


def _sls_audit_df_report_base(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","combination_number","limit_state","story","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_L_over_d","service_L_over_d_lim","service_crack_status","service_deflection_status","service_stress_status","service_status","service_note"]
    return results[_cols_existing(results, cols)].copy() if results is not None and not results.empty else pd.DataFrame()


def _detailing_audit_df_report_base(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","story","bot_rebar","top_rebar","bot_layers","top_layers","phi_st_mm","stirrup_legs","s_st_mm","skin_reinf_face_mm2","detailing_status","detailing_issues","solution"]
    return results[_cols_existing(results, cols)].copy() if results is not None and not results.empty else pd.DataFrame()


def _metadata_df_report_base(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Norma de referência", "NP EN 1992-1-1 / Eurocódigo 2"],
        ["Âmbito", "Vigas de betão armado: ELU, ELS expedito, corte, torção e pormenorização"],
        ["Descrição", APP_TABLE_DESCRIPTION], ["Suporte normativo", "NP EN 1992-1-1:2010 + Anexo Nacional Português, se aplicável"],
    ], columns=["Campo", "Valor"])

BeamsEC2App._metadata_df = _metadata_df_report_base


def _parameters_df_report_base(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Norma", "NP EN 1992-1-1"], ["Recobrimento [mm]", self.var_cover.get()], ["Agregado dg [mm]", self.var_agg.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()], ["γc", self.var_gamma_c.get()], ["γs", self.var_gamma_s.get()], ["αcc", "1.00"],
        ["cotθ", self.var_cot_theta.get()], ["wk,lim [mm]", self.var_crack_limit.get()], ["L/d limite", self.var_ld_limit.get()],
        ["Momento principal", self.var_moment_axis.get()], ["Corte vertical", self.var_shear_axis.get()], ["Torção", self.var_torsion_axis.get()],
        ["Modo de cálculo", self.var_calc_mode.get()], ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
    ], columns=["Parâmetro", "Valor"])

BeamsEC2App._parameters_df = _parameters_df_report_base


def _write_excel_report_base(self, path: str):
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None else pd.DataFrame()
    sheets = {
        "00_Info": self._metadata_df(),
        "01_Parametros": self._parameters_df(),
        "02_Resumo_Executivo": _executive_summary_df_report_base(results, summary),
        "03_Entrada_Dados": self.df_clean,
        "04_Envelopes": self.df_env,
        "05_Casos_Calculo": self.df_calc_input,
        "06_Resultados_Completos": results,
        "07_Resumo_Vigas": summary,
        "08_Memoria_Calculo": _calc_memory_df_report_base(results),
        "09_Flexao": _flexure_audit_df_report_base(results),
        "10_Corte_Torcao": _vt_audit_df_report_base(results),
        "11_ELS": _sls_audit_df_report_base(results),
        "12_Pormenorizacao": _detailing_audit_df_report_base(results),
        "13_Falhas": self.df_failures,
        "14_OK": self.df_ok,
        "15_Shortlists": self.build_shortlists_df(),
        "16_Validacao": self.df_validation,
        "17_Notas_EC2": self.df_notes,
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=name[:31], index=False)
        wb = writer.book
        props = wb.properties
        props.title = APP_NAME; props.subject = APP_SUBJECT; props.creator = APP_AUTHOR; props.keywords = APP_KEYWORDS
        props.category = APP_CATEGORY; props.description = APP_TABLE_DESCRIPTION; props.lastModifiedBy = APP_AUTHOR
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            header_fill = PatternFill("solid", fgColor="1F4E5F"); header_font = Font(color="FFFFFF", bold=True, name="Segoe UI")
            light_fill = PatternFill("solid", fgColor="EAF2F5"); fail_fill = PatternFill("solid", fgColor="FDE9D9")
            thin = Side(style="thin", color="D9E2E7"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.sheet_view.showGridLines = False; ws.freeze_panes = "A2"
                if ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
                for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 6000)):
                    for cell in row:
                        cell.border = border; cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if isinstance(cell.value, float): cell.number_format = "0.00"
                for col_idx, col in enumerate(ws.columns, start=1):
                    values = [str(c.value) for c in col[:250] if c.value is not None]
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max([len(v) for v in values] + [10]) + 2, 48)
                if ws.title in ["00_Info", "01_Parametros", "02_Resumo_Executivo"]:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        if row: row[0].fill = light_fill; row[0].font = Font(bold=True, name="Segoe UI")
                if ws.title == "13_Falhas" and ws.max_row > 1:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row: cell.fill = fail_fill
                try: ws.auto_filter.ref = ws.dimensions
                except Exception: pass
        except Exception:
            pass

BeamsEC2App._write_excel = _write_excel_report_base


def _pdf_styles_report_base():
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontName="Courier-Bold", fontSize=14, leading=21, spaceAfter=10))
    styles.add(ParagraphStyle(name="ReportSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontName="Courier", fontSize=10, leading=15, textColor=colors.darkgrey, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyCourier", parent=styles["Normal"], fontName="Courier", fontSize=9, leading=13.5, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontName="Courier", fontSize=7.2, leading=10.8))
    styles.add(ParagraphStyle(name="Cell", parent=styles["Normal"], fontName="Courier", fontSize=6.5, leading=9.75, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Courier-Bold", fontSize=11, leading=16.5, spaceBefore=10, spaceAfter=16, textColor=colors.HexColor("#1F4E5F")))
    return styles


def _pdf_table_style_report_base(self, header=True):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    cmds = [("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey), ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTNAME", (0,0), (-1,-1), "Courier"), ("FONTSIZE", (0,0), (-1,-1), 6.5), ("LEFTPADDING", (0,0), (-1,-1), 3), ("RIGHTPADDING", (0,0), (-1,-1), 3)]
    if header:
        cmds += [("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EFEFEF")), ("FONTNAME", (0,0), (-1,0), "Courier-Bold")]
    return TableStyle(cmds)

BeamsEC2App._pdf_table_style = _pdf_table_style_report_base


def _pdf_df_table_report_base(self, df: pd.DataFrame, cols: List[str], max_rows: int = 30, widths=None):
    from reportlab.platypus import Table, Paragraph
    from reportlab.lib.units import mm
    styles = _pdf_styles_report_base(); pstyle = styles["Cell"]
    if df is None or df.empty: return Paragraph("Sem dados.", styles["Small"])
    present = [c for c in cols if c in df.columns]
    if not present: return Paragraph("Sem colunas aplicáveis.", styles["Small"])
    data = [[Paragraph(str(c), pstyle) for c in present]]
    for _, r in df.head(max_rows).iterrows():
        row=[]
        for c in present:
            txt = _fmt_report_base(r.get(c, ""), 2).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            row.append(Paragraph(txt, pstyle))
        data.append(row)
    total_width = 270*mm
    col_widths = [w*mm for w in widths] if widths and len(widths) == len(present) else [total_width/max(1,len(present))]*max(1,len(present))
    t = Table(data, colWidths=col_widths, repeatRows=1); t.setStyle(self._pdf_table_style(header=True)); return t

BeamsEC2App._pdf_df_table = _pdf_df_table_report_base


def _write_pdf_report_base(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_report_base()
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME; doc.author = APP_AUTHOR; doc.subject = APP_SUBJECT
    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    exec_df = _executive_summary_df_report_base(results, summary)
    mem_df = _calc_memory_df_report_base(summary.head(20) if summary is not None and not summary.empty else results.head(20))
    story = [Paragraph(APP_NAME, styles["ReportTitle"]), Paragraph("Dimensionamento e verificação de vigas de betão armado segundo a NP EN 1992-1-1", styles["ReportSubtitle"]), Spacer(1,4*mm)]
    n_total = len(results); n_ok = int((results.get("status", pd.Series(dtype=str)) == "OK").sum()) if not results.empty else 0; n_fail = int((results.get("status", pd.Series(dtype=str)) == "Falha").sum()) if not results.empty else 0
    meta = [["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR], ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Norma", "NP EN 1992-1-1"], ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"], ["fyk", f"{self.var_fyk.get()} MPa", "cotθ", self.var_cot_theta.get()], ["Ficheiro", self.input_file_path or "-", "Modo", self.var_calc_mode.get()]]
    t = Table(meta, colWidths=[35*mm,102*mm,35*mm,98*mm]); t.setStyle(self._pdf_table_style(header=False)); story += [t, Spacer(1,7*mm)]
    story += [Paragraph("1. Resumo executivo", styles["Section"]), self._pdf_df_table(exec_df, ["Indicador","Valor","Nota"], max_rows=20, widths=[70,55,145]), Spacer(1,5*mm)]
    story += [Paragraph("Vigas governantes", styles["Section"]), self._pdf_df_table(summary, ["member","story","case","section_type","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm","eta_global","solution","status"], max_rows=32)]
    story += [PageBreak(), Paragraph("2. Relatório técnico", styles["Section"]), Paragraph("Critérios de cálculo", styles["BodyCourier"]), self._pdf_df_table(self._parameters_df(), ["Parâmetro","Valor"], max_rows=40, widths=[90,180]), Spacer(1,5*mm), Paragraph("Envelopes de esforços", styles["BodyCourier"]), self._pdf_df_table(self.df_env, ["member","story","case","n_points_found","length","material","hy","hz","bf","hf","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"], max_rows=38)]
    story += [PageBreak(), Paragraph("3. Verificações principais", styles["Section"]), Paragraph("Flexão", styles["BodyCourier"]), self._pdf_df_table(_flexure_audit_df_report_base(summary), ["member","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"], max_rows=34), Spacer(1,5*mm), Paragraph("Esforço transverso e torção", styles["BodyCourier"]), self._pdf_df_table(_vt_audit_df_report_base(summary), ["member","case","v_ed_kN","VRd_c_kN","VRd_max_kN","t_ed_kNm","TRd_max_kNm","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"], max_rows=34)]
    story += [PageBreak(), Paragraph("4. Memória de cálculo", styles["Section"]), Paragraph("Resumo por viga governante dos dados, esforços, armaduras, resistências, ELS e estado final.", styles["BodyCourier"]), self._pdf_df_table(mem_df, ["Viga","Caso","Piso","Etapa","Item","Valor","Unidade","Critério/Referência","Estado/Nota"], max_rows=95)]
    story += [PageBreak(), Paragraph("5. ELS, pormenorização e falhas", styles["Section"]), Paragraph("Verificações em serviço", styles["BodyCourier"]), self._pdf_df_table(_sls_audit_df_report_base(summary), ["member","case","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_L_over_d","service_status","service_note"], max_rows=30), Spacer(1,5*mm), Paragraph("Pormenorização", styles["BodyCourier"]), self._pdf_df_table(_detailing_audit_df_report_base(summary), ["member","case","bot_rebar","top_rebar","phi_st_mm","stirrup_legs","s_st_mm","skin_reinf_face_mm2","detailing_status","detailing_issues"], max_rows=30)]
    if failures is not None and not failures.empty:
        story += [PageBreak(), Paragraph("Falhas e recomendações", styles["Section"]), self._pdf_df_table(failures, ["member","story","case","failure_type","failure_reason","recommendations"], max_rows=45)]
    story += [PageBreak(), Paragraph("6. Notas EC2 e limitações", styles["Section"]), self._pdf_df_table(self.df_notes if self.df_notes is not None else build_normative_notes(), ["Tema","Referência","Nota"], max_rows=40, widths=[45,45,180]), Spacer(1,4*mm), Paragraph("O relatório contém resumo executivo, relatório técnico e memória de cálculo. A revisão final deve confirmar convenções de eixos, combinações, modelo estrutural, largura efectiva de banzos, disposições construtivas e casos críticos.", styles["Small"])]
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(APP_NAME); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)

BeamsEC2App._write_pdf = _write_pdf_report_base


def _update_report_report_base(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo."); return
    source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results); n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0; n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [f"{APP_NAME} {APP_VERSION}\n", "Relatório técnico de dimensionamento de vigas — NP EN 1992-1-1\n\n", f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n", f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | cotθ: {self.var_cot_theta.get()}\n\n"]
    for _, r in source.head(80).iterrows():
        lines.append(f"Viga {r.get('member','')} | Caso {r.get('case','')} | Piso {r.get('story','')} | NP EN 1992-1-1\n")
        lines.append(f"  Secção: {_fmt_report_base(r.get('bw_cm'),0)} x {_fmt_report_base(r.get('h_cm'),0)} cm | {r.get('section_type','')} | Material: {r.get('material','')}\n")
        lines.append(f"  Esforços: M+Ed={_fmt_report_base(r.get('m_pos_ed_kNm'))} kNm; M-Ed={_fmt_report_base(r.get('m_neg_ed_kNm'))} kNm; VEd={_fmt_report_base(r.get('v_ed_kN'))} kN; TEd={_fmt_report_base(r.get('t_ed_kNm'))} kNm\n")
        lines.append(f"  Flexão: MRd+={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_pos'),3)}); MRd-={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_neg'),3)})\n")
        lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'))} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'))} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'))} kNm\n")
        lines.append(f"  Armaduras: Inf. {r.get('bot_rebar','')} | Sup. {r.get('top_rebar','')} | Estribos Ø{_fmt_report_base(r.get('phi_st_mm'),0)}/{_fmt_report_base(r.get('stirrup_legs'),0)}r // {_fmt_report_base(finite(r.get('s_st_mm'))/10,1)} cm\n")
        lines.append(f"  ELS: {r.get('service_status','')} | Pormenorização: {r.get('detailing_status','')} | Estado: {r.get('status','')}\n")
        if str(r.get("failure_reason", "") or "").strip(): lines.append(f"  Motivo: {r.get('failure_reason','')}\n")
        if str(r.get("recommendations", "") or "").strip(): lines.append(f"  Recomendações: {r.get('recommendations','')}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))

BeamsEC2App.update_report = _update_report_report_base

_original_build_instructions_tab_v10 = BeamsEC2App._build_instructions_tab

def _build_instructions_tab_report_base(self, parent):
    outer = ttk.Frame(parent, padding=10); outer.pack(fill="both", expand=True)
    ttk.Label(outer, text="Instruções de utilização do BeamsEC2", style="Header.TLabel").pack(anchor="w", pady=(0, 8))
    txt = self._make_text_view(outer)
    content = ("OBJECTIVO DO PROGRAMA\nBeamsEC2 dimensiona e verifica vigas de betão armado segundo a NP EN 1992-1-1, usando uma configuração fixa do Eurocódigo 2 adoptado em Portugal. O cálculo inclui envelopes por viga/caso, flexão positiva e negativa, corte, torção, ELS expedito, pormenorização e exportações .xlsx/.pdf.\n\n" +
               "COLUNAS RECOMENDADAS\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n" +
               "UNIDADES\nFX, FY, FZ em kN; MX, MY, MZ em kNm; Station e Length em m; HY/HZ/BF/HF em cm.\n\n" +
               "CONVENÇÃO\nO eixo X local é longitudinal. Por defeito, MY é o momento principal da viga, FZ é o corte vertical e MX é a torção. Momentos positivos são tratados como tracção inferior; momentos negativos como tracção superior.\n\n" +
               "RELATÓRIOS\nO PDF exportado inclui resumo executivo, relatório técnico e memória de cálculo. A memória de cálculo exportado inclui folhas detalhadas para entrada, envelopes, resultados, memória, flexão, corte/torção, ELS, pormenorização, falhas, validação e notas EC2.\n")
    txt.insert("1.0", content); txt.config(state="disabled")

BeamsEC2App._build_instructions_tab = _build_instructions_tab_report_base

_original_build_sidebar_v10 = BeamsEC2App._build_sidebar

def _build_sidebar_report_base(self, parent):
    _original_build_sidebar_v10(self, parent)
    info = ttk.LabelFrame(parent, text="7. Relatórios técnicos")
    info.pack(fill="x", pady=(0, 8))
    ttk.Label(info, text="Norma fixa: NP EN 1992-1-1.\nPDF: resumo executivo + relatório técnico + memória de cálculo.\nMemória de cálculo: workbook detalhado e auditável.", wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)

BeamsEC2App._build_sidebar = _build_sidebar_report_base



# ============================================================
# ============================================================
NORMATIVE_SUPPORT = "NP EN 1992-1-1:2010 + Anexo Nacional Português, se aplicável"
SKIN_REINF_MIN_HEIGHT_MM = 400.0

# Guardar implementação  para reutilizar o cálculo base.
_design_one_report_base_base = BeamDesigner.design_one


def _design_one_design(self, row: pd.Series) -> Dict:
    """Ajuste : armadura de alma/pele só é atribuída a vigas com h >= 40 cm.

    Para vigas com h < 40 cm, a parcela longitudinal de torção é absorvida nas faces
    superior/inferior e a solução deixa de reportar armadura de pele lateral.
    """
    out = _design_one_report_base_base(self, row)
    try:
        h_cm = finite(out.get("h_cm"), 0.0)
        h_mm = h_cm * 10.0
        skin = finite(out.get("skin_reinf_face_mm2"), 0.0)
        if h_mm < SKIN_REINF_MIN_HEIGHT_MM:
            # Remover armadura lateral/pele em vigas baixas. Mantém os resultados ELU;
            # evita apenas reportar pormenorização lateral que não é pretendida nesta ferramenta.
            out["skin_reinf_face_mm2"] = 0.0
            sol = str(out.get("solution", "") or "")
            sol = re.sub(r";\s*pele\s*≈\s*[^;]+", "", sol).strip()
            out["solution"] = sol
            out["skin_reinf_note"] = "Não aplicável: h < 40 cm"
        else:
            out["skin_reinf_note"] = "Aplicável: h >= 40 cm" if skin > 0 else "Verificar necessidade: h >= 40 cm"
        # Normalizar problemas de pormenorização relacionados com pele em vigas baixas.
        issues = str(out.get("detailing_issues", "") or "")
        if h_mm < SKIN_REINF_MIN_HEIGHT_MM and "pele" in issues.lower():
            parts = [p.strip() for p in issues.split(";") if p.strip() and "pele" not in p.lower()]
            out["detailing_issues"] = "; ".join(parts) if parts else "-"
            if out["detailing_issues"] == "-":
                out["detailing_status"] = "OK"
    except Exception:
        pass
    return out


BeamDesigner.design_one = _design_one_design


# Pormenorização : o aviso de armadura de pele só existe a partir de 40 cm.
def _detailing_check_design(self, section: BeamSection, bot: RebarChoice, top: RebarChoice, stir: Dict[str, float | str], As_skin_face: float) -> Dict[str, float | str]:
    issues = []
    if bot.status != "OK":
        issues.append("armadura inferior não cabe")
    if top.status != "OK":
        issues.append("armadura superior não cabe")
    if bot.n_bars and bot.n_bars < 2:
        issues.append("mínimo de 2 varões inferiores não cumprido")
    if top.n_bars and top.n_bars < 2:
        issues.append("mínimo de 2 varões superiores não cumprido")
    if float(stir.get("s_st_mm", 999.0)) > float(stir.get("s_lim_mm", 0.0)) + 1e-9:
        issues.append("espaçamento dos estribos superior ao limite")
    if stir.get("stirrup_status") != "OK":
        issues.append("armadura transversal requerida não cabe nas opções disponíveis")
    if section.h_mm >= SKIN_REINF_MIN_HEIGHT_MM and As_skin_face <= 0:
        # Não força falha, apenas deixa nota técnica para vigas com altura relevante.
        issues.append("avaliar armadura de alma/pele nas faces laterais")
    status = "OK" if not issues else "Verificar"
    return {"detailing_status": status, "detailing_issues": "; ".join(issues) if issues else "-"}


BeamDesigner.detailing_check = _detailing_check_design


# Metadata e parâmetros : suporte normativo explícito.
def _metadata_df_design(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Suporte normativo", NORMATIVE_SUPPORT], ["Âmbito", "Dimensionamento/verificação ELU e ELS expedito de vigas de betão armado"],
        ["Descrição", APP_TABLE_DESCRIPTION],
        ["Limitações", "Confirmar eixos locais, combinações, largura efectiva de banzos, aberturas, redistribuições, ancoragens e casos críticos."],
    ], columns=["Campo", "Valor"])


def _parameters_df_design(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Suporte normativo", NORMATIVE_SUPPORT],
        ["Recobrimento [mm]", self.var_cover.get()], ["Agregado dg [mm]", self.var_agg.get()],
        ["Aço fyk [MPa]", self.var_fyk.get()], ["Betão", "lido da coluna Material; fallback interno C30/37 quando ausente"],
        ["γc", self.var_gamma_c.get()], ["γs", self.var_gamma_s.get()], ["cot θ", self.var_cot_theta.get()],
        ["wk,lim [mm]", self.var_crack_limit.get()], ["L/d limite", self.var_ld_limit.get()],
        ["Modo de cálculo", self.var_calc_mode.get()], ["Relatório PDF", (self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Completo")],
        ["Armadura de alma/pele", "considerada apenas para vigas com h >= 40 cm"],
        ["Momento principal", self.var_moment_axis.get()], ["Corte vertical", self.var_shear_axis.get()], ["Torção", self.var_torsion_axis.get()],
    ], columns=["Parâmetro", "Valor"])


BeamsEC2App._metadata_df = _metadata_df_design
BeamsEC2App._parameters_df = _parameters_df_design


# Notas normativas .
def _build_normative_notes_design() -> pd.DataFrame:
    return pd.DataFrame([
        ("Suporte normativo", NORMATIVE_SUPPORT, "O programa usa por defeito o Eurocódigo 2 adoptado em Portugal para vigas de betão armado."),
        ("Âmbito", "Vigas de betão armado", "Flexão positiva/negativa, esforço transverso, torção, ELS expedito e pormenorização."),
        ("Entrada", "Envelopes por viga/caso", "Recomenda-se exportar várias estações ao longo da viga para obter envelopes governantes."),
        ("Flexão", "NP EN 1992-1-1:2010, Secção 6.1", "MRd é recalculado com a armadura adoptada, separando zonas de tracção inferior e superior."),
        ("Esforço transverso", "NP EN 1992-1-1:2010, Secção 6.2", "Calcula VRd,c, VRd,max, Asw/s requerido e espaçamento adoptado."),
        ("Torção", "NP EN 1992-1-1:2010, Secção 6.3", "A armadura transversal de torção é combinada com a de corte; a armadura longitudinal de torção é auditada."),
        ("ELS", "NP EN 1992-1-1:2010, Secção 7", "Inclui controlo expedito de tensão, fendilhação e L/d; não substitui análise detalhada de deformações."),
        ("Pormenorização", "NP EN 1992-1-1:2010, Secções 8 e 9.2", "Controla camadas, espaçamento livre, estribos e armadura de alma/pele apenas para h >= 40 cm."),
        ("Relatórios", "Critério do utilizador", "O PDF pode ser exportado como resumo executivo, relatório técnico, memória de cálculo ou relatório completo."),
    ], columns=["Tema", "Referência", "Nota"])


build_normative_notes = _build_normative_notes_design


# Sidebar : sem DXF e com selecção do tipo de relatório PDF.
def _build_sidebar_design(self, parent):
    if not hasattr(self, "var_pdf_scope"):
        self.var_pdf_scope = tk.StringVar(value="Completo")

    hero = ttk.LabelFrame(parent, text="BeamsEC2")
    hero.pack(fill="x", pady=(0, 8))
    link = ttk.Label(hero, text="BeamsEC2", style="Header.TLabel", cursor="hand2")
    link.pack(anchor="w")
    link.bind("<Button-1>", lambda _e: webbrowser.open_new(GITHUB_URL))
    ttk.Label(hero, text="Dimensionamento de vigas de betão armado (EC2)", style="Header.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(hero, text=f"Importa esforços, cria envelopes por viga/caso e dimensiona segundo {NORMATIVE_SUPPORT}. Exporta workbook memória de cálculo detalhado e relatórios PDF seleccionáveis.", style="Subtle.TLabel", wraplength=340, justify="left").pack(anchor="w", pady=(2, 0))

    data = ttk.LabelFrame(parent, text="1. Entrada")
    data.pack(fill="x", pady=(0, 8))
    ttk.Button(data, text="Colar área de transferência", command=self.paste_clipboard).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Importar .xlsx/.csv", command=self.import_file).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Ler caixa de texto", command=self.load_from_textbox).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(data, text="Modelo de tabela", command=self.export_template).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    data.columnconfigure(0, weight=1); data.columnconfigure(1, weight=1)

    params = ttk.LabelFrame(parent, text="2. Parâmetros EC2")
    params.pack(fill="x", pady=(0, 8))
    self._add_label_entry(params, "Recobrimento [mm]", self.var_cover, 0)
    self._add_label_entry(params, "Agregado dg [mm]", self.var_agg, 1)
    ttk.Label(params, text="Aço fyk [MPa]").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_fyk, values=["400", "500"], state="readonly", width=14).grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    self._add_label_entry(params, "γc", self.var_gamma_c, 3)
    self._add_label_entry(params, "γs", self.var_gamma_s, 4)
    self._add_label_entry(params, "cot θ", self.var_cot_theta, 5)
    self._add_label_entry(params, "wk,lim [mm]", self.var_crack_limit, 6)
    self._add_label_entry(params, "L/d limite", self.var_ld_limit, 7)
    ttk.Label(params, text="Modo").grid(row=8, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(params, textvariable=self.var_calc_mode, values=["pre_dimensionamento", "dimensionamento"], state="readonly").grid(row=8, column=1, sticky="ew", padx=6, pady=4)
    ttk.Checkbutton(params, text="Reduzir para casos governantes", variable=self.var_reduce_cases).grid(row=9, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    ttk.Checkbutton(params, text="Resumo por viga", variable=self.var_summary).grid(row=10, column=0, columnspan=2, sticky="w", padx=6, pady=3)
    params.columnconfigure(1, weight=1)

    axes = ttk.LabelFrame(parent, text="3. Eixos locais / esforços")
    axes.pack(fill="x", pady=(0, 8))
    for i, (lab, var, values) in enumerate([
        ("Momento principal", self.var_moment_axis, ["MY", "MZ"]),
        ("Corte vertical", self.var_shear_axis, ["FZ", "FY"]),
        ("Torção", self.var_torsion_axis, ["MX"]),
    ]):
        ttk.Label(axes, text=lab).grid(row=i, column=0, sticky="w", padx=6, pady=4)
        ttk.Combobox(axes, textvariable=var, values=values, state="readonly").grid(row=i, column=1, sticky="ew", padx=6, pady=4)
    axes.columnconfigure(1, weight=1)

    filters = ttk.LabelFrame(parent, text="4. Filtros")
    filters.pack(fill="x", pady=(0, 8))
    ttk.Label(filters, text="Member").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    ttk.Entry(filters, textvariable=self.var_filter_member).grid(row=0, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Estado").grid(row=1, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_status, values=["Todos", "OK", "Falha"], state="readonly").grid(row=1, column=1, sticky="ew", padx=6, pady=4)
    ttk.Label(filters, text="Falha").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(filters, textvariable=self.var_filter_fail, values=["Todos", "flexao", "esforco_transverso", "torcao", "pormenorizacao", "dados_incompletos", "outra"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(filters, text="Aplicar", command=self.apply_filters).grid(row=3, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(filters, text="Limpar", command=self.clear_filters).grid(row=3, column=1, sticky="ew", padx=4, pady=4)
    filters.columnconfigure(1, weight=1)

    actions = ttk.LabelFrame(parent, text="5. Cálculo e exportação")
    actions.pack(fill="x", pady=(0, 8))
    ttk.Button(actions, text="Calcular", command=self.run_design, style="Primary.TButton").grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Exportar .xlsx", command=self.export_excel).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
    ttk.Button(actions, text="Relatório .pdf", command=self.export_pdf_report).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ttk.Label(actions, text="Tipo de PDF").grid(row=2, column=0, sticky="w", padx=6, pady=4)
    ttk.Combobox(actions, textvariable=self.var_pdf_scope, values=["Completo", "Resumo executivo", "Relatório técnico", "Memória de cálculo"], state="readonly").grid(row=2, column=1, sticky="ew", padx=6, pady=4)
    ttk.Button(actions, text="Abrir repositório", command=lambda: webbrowser.open_new(GITHUB_URL)).grid(row=3, column=0, columnspan=2, sticky="ew", padx=4, pady=4)
    actions.columnconfigure(0, weight=1); actions.columnconfigure(1, weight=1)

    notes = ttk.LabelFrame(parent, text="6. Notas rápidas")
    notes.pack(fill="x", pady=(0, 8))
    ttk.Label(notes, text=f"• Suporte normativo: {NORMATIVE_SUPPORT}.\n• Para vigas, exportar várias estações ao longo da barra melhora o envelope.\n• Para secções T, preencher BF e HF; caso contrário, a secção é rectangular.\n• A armadura de alma/pele só é reportada para h >= 40 cm.\n• A memória de cálculo contém a memória completa; o PDF pode ser resumido, técnico, memória ou completo.", wraplength=340, justify="left").pack(fill="x", padx=6, pady=6)


BeamsEC2App._build_sidebar = _build_sidebar_design


# Instruções .
def _build_instructions_tab_design(self, parent):
    outer = ttk.Frame(parent, padding=10); outer.pack(fill="both", expand=True)
    ttk.Label(outer, text="Instruções de utilização do BeamsEC2", style="Header.TLabel").pack(anchor="w", pady=(0, 8))
    txt = self._make_text_view(outer)
    content = ("OBJECTIVO DO PROGRAMA\n"
               f"BeamsEC2 dimensiona e verifica vigas de betão armado com suporte normativo fixo: {NORMATIVE_SUPPORT}. "
               "O cálculo inclui envelopes por viga/caso, flexão positiva e negativa, corte, torção, ELS expedito, pormenorização e exportações .xlsx/.pdf.\n\n"
               "COLUNAS RECOMENDADAS\n" + " | ".join(self.TEMPLATE_COLUMNS) + "\n\n"
               "UNIDADES\nFX, FY, FZ em kN; MX, MY, MZ em kNm; Station e Length em m; HY/HZ/BF/HF em cm.\n\n"
               "CONVENÇÃO\nO eixo X local é longitudinal. Por defeito, MY é o momento principal da viga, FZ é o corte vertical e MX é a torção. Momentos positivos são tratados como tracção inferior; momentos negativos como tracção superior.\n\n"
               "PORMENORIZAÇÃO\nA armadura de alma/pele só é reportada para vigas com altura h >= 40 cm. Em vigas mais baixas, a solução não apresenta armadura lateral de pele por defeito.\n\n"
               "RELATÓRIOS\nO utilizador escolhe o tipo de PDF: Resumo executivo, Relatório técnico, Memória de cálculo ou Completo. A memória de cálculo exportado mantém a auditoria completa, com folhas separadas para entrada, envelopes, resultados, memória, flexão, corte/torção, ELS, pormenorização, falhas, validação e notas EC2.\n")
    txt.insert("1.0", content); txt.config(state="disabled")


BeamsEC2App._build_instructions_tab = _build_instructions_tab_design


# PDF : relatório dividido por critério do utilizador.
def _write_pdf_design(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak

    styles = _pdf_styles_report_base()
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME; doc.author = APP_AUTHOR; doc.subject = APP_SUBJECT

    results = self.df_results if self.df_results is not None else pd.DataFrame()
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    exec_df = _executive_summary_df_report_base(results, summary)
    mem_df = _calc_memory_df_report_base(summary.head(20) if summary is not None and not summary.empty else results.head(20))
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Completo"

    n_total = len(results)
    n_ok = int((results.get("status", pd.Series(dtype=str)) == "OK").sum()) if not results.empty else 0
    n_fail = int((results.get("status", pd.Series(dtype=str)) == "Falha").sum()) if not results.empty else 0

    def meta_block(report_name):
        meta = [
            ["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR],
            ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Suporte normativo", NORMATIVE_SUPPORT],
            ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"],
            ["fyk", f"{self.var_fyk.get()} MPa", "cotθ", self.var_cot_theta.get()],
            ["Ficheiro", self.input_file_path or "-", "Tipo de relatório", report_name],
        ]
        t = Table(meta, colWidths=[35*mm,102*mm,35*mm,98*mm]); t.setStyle(self._pdf_table_style(header=False))
        return t

    story = [Paragraph(APP_NAME, styles["ReportTitle"]), Paragraph(f"{scope} - dimensionamento de vigas de betão armado", styles["ReportSubtitle"]), Spacer(1,4*mm), meta_block(scope), Spacer(1,7*mm)]

    def add_exec():
        story.extend([Paragraph("Resumo executivo", styles["Section"]), self._pdf_df_table(exec_df, ["Indicador","Valor","Nota"], max_rows=20, widths=[70,55,145]), Spacer(1,5*mm), Paragraph("Vigas governantes", styles["Section"]), self._pdf_df_table(summary, ["member","story","case","section_type","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm","eta_global","solution","status"], max_rows=32)])

    def add_technical():
        story.extend([Paragraph("Relatório técnico", styles["Section"]), Paragraph("Critérios de cálculo", styles["BodyCourier"]), self._pdf_df_table(self._parameters_df(), ["Parâmetro","Valor"], max_rows=40, widths=[90,180]), Spacer(1,5*mm), Paragraph("Envelopes de esforços", styles["BodyCourier"]), self._pdf_df_table(self.df_env, ["member","story","case","n_points_found","length","material","hy","hz","bf","hf","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"], max_rows=38), Spacer(1,5*mm), Paragraph("Flexão", styles["BodyCourier"]), self._pdf_df_table(_flexure_audit_df_report_base(summary), ["member","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"], max_rows=34), Spacer(1,5*mm), Paragraph("Esforço transverso e torção", styles["BodyCourier"]), self._pdf_df_table(_vt_audit_df_report_base(summary), ["member","case","v_ed_kN","VRd_c_kN","VRd_max_kN","t_ed_kNm","TRd_max_kNm","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"], max_rows=34)])

    def add_memory():
        story.extend([Paragraph("Memória de cálculo", styles["Section"]), Paragraph("Resumo por viga governante dos dados, esforços, armaduras, resistências, ELS e estado final.", styles["BodyCourier"]), self._pdf_df_table(mem_df, ["Viga","Caso","Piso","Etapa","Item","Valor","Unidade","Critério/Referência","Estado/Nota"], max_rows=95)])

    def add_service_detailing_failures():
        story.extend([Paragraph("ELS e pormenorização", styles["Section"]), Paragraph("Verificações em serviço", styles["BodyCourier"]), self._pdf_df_table(_sls_audit_df_report_base(summary), ["member","case","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_L_over_d","service_status","service_note"], max_rows=30), Spacer(1,5*mm), Paragraph("Pormenorização", styles["BodyCourier"]), self._pdf_df_table(_detailing_audit_df_report_base(summary), ["member","case","bot_rebar","top_rebar","phi_st_mm","stirrup_legs","s_st_mm","skin_reinf_face_mm2","detailing_status","detailing_issues"], max_rows=30)])
        if failures is not None and not failures.empty:
            story.extend([PageBreak(), Paragraph("Falhas e recomendações", styles["Section"]), self._pdf_df_table(failures, ["member","story","case","failure_type","failure_reason","recommendations"], max_rows=45)])

    def add_notes():
        story.extend([Paragraph("Notas EC2 e limitações", styles["Section"]), self._pdf_df_table(self.df_notes if self.df_notes is not None else build_normative_notes(), ["Tema","Referência","Nota"], max_rows=40, widths=[45,65,160]), Spacer(1,4*mm), Paragraph("A revisão final deve confirmar convenções de eixos, combinações, modelo estrutural, largura efectiva de banzos, disposições construtivas e casos críticos.", styles["Small"])])

    if scope == "Resumo executivo":
        add_exec(); story.extend([Spacer(1,5*mm)]); add_notes()
    elif scope == "Relatório técnico":
        add_technical(); story.extend([PageBreak()]); add_service_detailing_failures(); story.extend([PageBreak()]); add_notes()
    elif scope == "Memória de cálculo":
        add_memory(); story.extend([PageBreak()]); add_notes()
    else:
        add_exec(); story.extend([PageBreak()]); add_technical(); story.extend([PageBreak()]); add_memory(); story.extend([PageBreak()]); add_service_detailing_failures(); story.extend([PageBreak()]); add_notes()

    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(APP_NAME); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


BeamsEC2App._write_pdf = _write_pdf_design


# Relatório em texto .
def _update_report_design(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo."); return
    source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results); n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0; n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [f"{APP_NAME} {APP_VERSION}\n", f"Relatório técnico de dimensionamento de vigas - {NORMATIVE_SUPPORT}\n\n", f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n", f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | cotθ: {self.var_cot_theta.get()}\n", f"PDF seleccionado: {(self.var_pdf_scope.get() if hasattr(self, 'var_pdf_scope') else 'Completo')}\n\n"]
    for _, r in source.head(80).iterrows():
        lines.append(f"Viga {r.get('member','')} | Caso {r.get('case','')} | Piso {r.get('story','')}\n")
        lines.append(f"  Secção: {_fmt_report_base(r.get('bw_cm'),0)} x {_fmt_report_base(r.get('h_cm'),0)} cm | {r.get('section_type','')} | Material: {r.get('material','')}\n")
        lines.append(f"  Esforços: M+Ed={_fmt_report_base(r.get('m_pos_ed_kNm'))} kNm; M-Ed={_fmt_report_base(r.get('m_neg_ed_kNm'))} kNm; VEd={_fmt_report_base(r.get('v_ed_kN'))} kN; TEd={_fmt_report_base(r.get('t_ed_kNm'))} kNm\n")
        lines.append(f"  Flexão: MRd+={_fmt_report_base(r.get('mrd_pos_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_pos'),3)}); MRd-={_fmt_report_base(r.get('mrd_neg_kNm'))} kNm (η={_fmt_report_base(r.get('eta_m_neg'),3)})\n")
        lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'))} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'))} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'))} kNm\n")
        skin = finite(r.get('skin_reinf_face_mm2'), 0.0)
        skin_txt = f" | Pele/alma={skin/100:.2f} cm²/face" if skin > 0 else " | Pele/alma=n/a"
        lines.append(f"  Armaduras: Inf. {r.get('bot_rebar','')} | Sup. {r.get('top_rebar','')} | Estribos Ø{_fmt_report_base(r.get('phi_st_mm'),0)}/{_fmt_report_base(r.get('stirrup_legs'),0)}r // {_fmt_report_base(finite(r.get('s_st_mm'))/10,1)} cm{skin_txt}\n")
        lines.append(f"  ELS: {r.get('service_status','')} | Pormenorização: {r.get('detailing_status','')} | Estado: {r.get('status','')}\n")
        if str(r.get("failure_reason", "") or "").strip(): lines.append(f"  Motivo: {r.get('failure_reason','')}\n")
        if str(r.get("recommendations", "") or "").strip(): lines.append(f"  Recomendações: {r.get('recommendations','')}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))


BeamsEC2App.update_report = _update_report_design


# Override final : recalcula a distribuição da armadura longitudinal de torção
# para não remover armadura de pele sem redistribuir a parcela resistente.
def _design_one_design_final(self, row: pd.Series) -> Dict:
    material = str(row.get("material", DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, alpha_cc=self.alpha_cc, gamma_c=self.gamma_c)
    sp = steel_props(self.fyk, gamma_s=self.gamma_s)
    fcd = cp["fcd"]
    fyd = sp["fyd"]

    bw = cm_to_mm(row.get("hy", 0.0))
    h = cm_to_mm(row.get("hz", 0.0))
    bf = cm_to_mm(row.get("bf", float("nan")))
    hf = cm_to_mm(row.get("hf", float("nan")))
    if bw <= 0 or h <= 0:
        return {
            "member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""),
            "status": "Falha", "failure_reason": "Dados incompletos: dimensões HY/HZ inválidas", "failure_type": "dados_incompletos",
        }
    if not math.isfinite(bf) or bf <= bw or not math.isfinite(hf) or hf <= 0:
        bf = 0.0
        hf = 0.0
    section = BeamSection(bw, h, bf, hf)

    phi_initial = 16.0
    phi_st_initial = self.choose_stirrup_diameter(phi_initial)
    d_bot_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
    d_top_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
    d2_initial = self.cover_mm + phi_st_initial + phi_initial / 2.0
    if d_bot_initial <= 0:
        return {
            "member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""),
            "status": "Falha", "failure_reason": "Dados geométricos incompatíveis com recobrimento", "failure_type": "dados_incompletos",
        }

    Mpos = finite(row.get("m_pos_ed_kNm"), 0.0)
    Mneg = finite(row.get("m_neg_ed_kNm"), 0.0)
    VEd = finite(row.get("v_ed_kN"), 0.0)
    TEd = finite(row.get("t_ed_kNm"), 0.0)

    As_min_bot = self.as_min_beam(section.bw_mm, d_bot_initial, cp["fctm"], self.fyk)
    As_min_top = self.as_min_beam(section.bw_mm, d_top_initial, cp["fctm"], self.fyk)

    flex_pos = self.flexural_required(Mpos, section, d_bot_initial, d2_initial, fcd, fyd)
    section_neg = BeamSection(section.bw_mm, section.h_mm, 0.0, 0.0)
    flex_neg = self.flexural_required(Mneg, section_neg, d_top_initial, d2_initial, fcd, fyd)

    tors = self.torsion_requirements(TEd, section, fck, fcd, fyd)
    Asl_t = float(tors.get("Asl_torsion_req_mm2") or 0.0)
    if h >= SKIN_REINF_MIN_HEIGHT_MM:
        As_torsion_top = 0.25 * Asl_t
        As_torsion_bot = 0.25 * Asl_t
        As_skin_face = 0.25 * Asl_t
        if h > SKIN_REINF_MIN_HEIGHT_MM:
            As_skin_face = max(As_skin_face, 0.001 * section.bw_mm * (h - SKIN_REINF_MIN_HEIGHT_MM) / 2.0)
        skin_note = "Aplicável: h >= 40 cm"
    else:
        # Em vigas baixas, não se reporta armadura lateral de pele. A parcela longitudinal
        # de torção é redistribuída para as faces inferior/superior.
        As_torsion_top = 0.50 * Asl_t
        As_torsion_bot = 0.50 * Asl_t
        As_skin_face = 0.0
        skin_note = "Não aplicável: h < 40 cm"

    As_req_bot = max(As_min_bot, flex_pos["As_req"] + As_torsion_bot)
    As_req_top = max(As_min_top, flex_neg["As_req"] + flex_pos.get("As_comp_req", 0.0) + As_torsion_top)

    bot = self.choose_longitudinal_bars(As_req_bot, bw, max_layers=3)
    top = self.choose_longitudinal_bars(As_req_top, bw, max_layers=3)

    d_bot = h - bot.centroid_from_edge_mm if bot.n_bars else d_bot_initial
    d_top = h - top.centroid_from_edge_mm if top.n_bars else d_top_initial
    d2_top = top.centroid_from_edge_mm if top.n_bars else d2_initial
    d2_bot = bot.centroid_from_edge_mm if bot.n_bars else d2_initial

    Mrd_pos = self.flexural_capacity(bot.area_mm2, top.area_mm2, section, d_bot, d2_top, fcd, fyd)
    Mrd_neg = self.flexural_capacity(top.area_mm2, bot.area_mm2, section_neg, d_top, d2_bot, fcd, fyd)
    eta_m_pos = Mpos / max(Mrd_pos.get("MRd_kNm") or 0.0, 1e-9) if Mpos > 0 else 0.0
    eta_m_neg = Mneg / max(Mrd_neg.get("MRd_kNm") or 0.0, 1e-9) if Mneg > 0 else 0.0

    shear = self.shear_requirements(VEd, bw, min(d_bot, d_top), max(bot.area_mm2, top.area_mm2), fck, fcd, fyd)
    Asw_s_total = max(float(shear["Asw_s_min_mm2_per_mm"]), float(shear["Asw_s_shear_req_mm2_per_mm"])) + float(tors["Asw_s_torsion_req_mm2_per_mm"] or 0.0)
    stir = self.choose_stirrups(Asw_s_total, bw, h, min(d_bot, d_top), torsion=TEd > 1e-9)

    els = self.serviceability(row, bot.area_mm2, top.area_mm2, d_bot, d_top, section, cp)
    det = self.detailing_check(section, bot, top, stir, As_skin_face)

    failure_reasons = []
    if bot.status != "OK" or top.status != "OK":
        failure_reasons.append("armadura longitudinal não cabe")
    if eta_m_pos > 1.0 + 1e-6:
        failure_reasons.append("flexão positiva não verifica")
    if eta_m_neg > 1.0 + 1e-6:
        failure_reasons.append("flexão negativa não verifica")
    if "Não conforme" in str(shear["shear_status"]):
        failure_reasons.append("esforço transverso excede VRd,max")
    if "Não conforme" in str(tors["torsion_status"]):
        failure_reasons.append("torção excede TRd,max")
    if stir.get("stirrup_status") != "OK":
        failure_reasons.append("armadura transversal não cabe")
    status = "OK" if not failure_reasons else "Falha"
    failure_reason = "; ".join(failure_reasons)
    failure_type = ""
    if failure_reasons:
        txt = failure_reason.lower()
        if "flexão" in txt:
            failure_type = "flexao"
        elif "transverso" in txt:
            failure_type = "esforco_transverso"
        elif "torção" in txt:
            failure_type = "torcao"
        elif "cabe" in txt:
            failure_type = "pormenorizacao"
        else:
            failure_type = "outra"

    sol = f"Inf.: {bot.label}; Sup.: {top.label}; Estribos Ø{int(stir['phi_st_mm'])}/{int(stir['stirrup_legs'])}r // {float(stir['s_st_mm'])/10:.1f} cm"
    if As_skin_face > 1e-6:
        sol += f"; pele/alma ≈ {As_skin_face/100:.2f} cm²/face"

    recs = []
    if eta_m_pos > 0.90 or eta_m_neg > 0.90:
        recs.append("avaliar aumento de altura ou armadura longitudinal")
    if "Requer" in str(shear["shear_status"]):
        recs.append("confirmar estribos por zonas junto aos apoios")
    if "torção" in str(tors["torsion_status"]).lower() and TEd > 0:
        recs.append("garantir estribos fechados e armadura longitudinal de torção")
    if det["detailing_status"] != "OK":
        recs.append("rever pormenorização construtiva")
    if str(els.get("service_status", "")).lower().startswith("verificar"):
        recs.append("confirmar ELS com combinação de serviço e cálculo detalhado")

    eta_shear_max = VEd / max(finite(shear.get("VRd_max_kN")), 1e-9) if VEd > 0 else 0.0
    eta_torsion_max = TEd / max(finite(tors.get("TRd_max_kNm")), 1e-9) if TEd > 0 else 0.0
    eta_flexure_max = max(eta_m_pos, eta_m_neg)
    eta_global = max(eta_flexure_max, eta_shear_max, eta_torsion_max)

    return {
        "member": row.get("member", ""),
        "case": row.get("case", ""),
        "combination_number": row.get("combination_number", extract_combination_number(row.get("case", ""))),
        "limit_state": row.get("limit_state", classify_limit_state(row.get("case", ""))),
        "name": row.get("name", ""),
        "story": row.get("story", ""),
        "node_i": row.get("node_i", ""),
        "node_j": row.get("node_j", ""),
        "n_points_found": row.get("n_points_found", None),
        "length_m": finite(row.get("length"), 0.0),
        "material": material,
        "material_source": row.get("material_source", "tabela"),
        "section_type": section.section_type,
        "bw_cm": bw / 10.0,
        "h_cm": h / 10.0,
        "bf_cm": bf / 10.0 if bf > 0 else None,
        "hf_cm": hf / 10.0 if hf > 0 else None,
        "cover_mm": self.cover_mm,
        "fck_MPa": fck,
        "fcd_MPa": fcd,
        "fyk_MPa": self.fyk,
        "fyd_MPa": fyd,
        "moment_axis": row.get("moment_axis", "MY"),
        "shear_axis": row.get("shear_axis", "FZ"),
        "torsion_axis": row.get("torsion_axis", "MX"),
        "m_pos_ed_kNm": Mpos,
        "m_pos_at": row.get("m_pos_at", ""),
        "m_neg_ed_kNm": Mneg,
        "m_neg_at": row.get("m_neg_at", ""),
        "v_ed_kN": VEd,
        "v_at": row.get("v_at", ""),
        "t_ed_kNm": TEd,
        "t_at": row.get("t_at", ""),
        "as_min_bot_mm2": As_min_bot,
        "as_min_top_mm2": As_min_top,
        "as_req_bot_mm2": As_req_bot,
        "as_req_top_mm2": As_req_top,
        "as_prov_bot_mm2": bot.area_mm2,
        "as_prov_top_mm2": top.area_mm2,
        "bot_rebar": bot.label,
        "top_rebar": top.label,
        "bot_layers": bot.layers,
        "top_layers": top.layers,
        "d_bot_mm": d_bot,
        "d_top_mm": d_top,
        "mrd_pos_kNm": Mrd_pos.get("MRd_kNm"),
        "mrd_neg_kNm": Mrd_neg.get("MRd_kNm"),
        "eta_m_pos": eta_m_pos,
        "eta_m_neg": eta_m_neg,
        "x_pos_mm": Mrd_pos.get("x_mm"),
        "x_neg_mm": Mrd_neg.get("x_mm"),
        "ductility_pos": flex_pos.get("ductility_status"),
        "ductility_neg": flex_neg.get("ductility_status"),
        "VRd_c_kN": shear.get("VRd_c_kN"),
        "VRd_max_kN": shear.get("VRd_max_kN"),
        "Asw_s_shear_req_mm2_per_m": float(shear.get("Asw_s_shear_req_mm2_per_mm", 0.0)) * 1000.0,
        "Asw_s_min_mm2_per_m": float(shear.get("Asw_s_min_mm2_per_mm", 0.0)) * 1000.0,
        "TRd_max_kNm": tors.get("TRd_max_kNm"),
        "Asw_s_torsion_req_mm2_per_m": float(tors.get("Asw_s_torsion_req_mm2_per_mm", 0.0) or 0.0) * 1000.0,
        "Asl_torsion_req_mm2": tors.get("Asl_torsion_req_mm2"),
        "Asw_s_total_req_mm2_per_m": Asw_s_total * 1000.0,
        "phi_st_mm": stir.get("phi_st_mm"),
        "stirrup_legs": stir.get("stirrup_legs"),
        "s_st_mm": stir.get("s_st_mm"),
        "Asw_s_prov_mm2_per_m": float(stir.get("Asw_s_prov_mm2_per_mm", 0.0)) * 1000.0,
        "shear_status": shear.get("shear_status"),
        "torsion_status": tors.get("torsion_status"),
        "stirrup_status": stir.get("stirrup_status"),
        "skin_reinf_face_mm2": As_skin_face,
        "skin_reinf_threshold_cm": SKIN_REINF_MIN_HEIGHT_MM / 10.0,
        "skin_reinf_note": skin_note,
        "norma": NORMATIVE_SUPPORT,
        "design_basis": "Suporte normativo fixo para Portugal",
        "calc_version": APP_VERSION,
        "flexure_method": "NP EN 1992-1-1:2010, Secção 6.1; MRd com armadura adoptada",
        "shear_method": "NP EN 1992-1-1:2010, Secção 6.2; VRd,c, VRd,max e Asw/s",
        "torsion_method": "NP EN 1992-1-1:2010, Secção 6.3; secção equivalente de parede fina",
        "sls_method": "NP EN 1992-1-1:2010, Secção 7; ELS expedito",
        "detailing_method": "NP EN 1992-1-1:2010, Secções 8 e 9.2; pormenorização construtiva",
        "eta_flexure_max": eta_flexure_max,
        "eta_shear_max": eta_shear_max,
        "eta_torsion_max": eta_torsion_max,
        "eta_global": eta_global,
        **els,
        **det,
        "status": status,
        "failure_reason": failure_reason,
        "failure_type": failure_type,
        "recommendations": "; ".join(dict.fromkeys(recs)),
        "solution": sol,
        "shortlist_text": f"Bot {bot.label}: As={bot.area_mm2:.0f} mm², layers={bot.layers}; Top {top.label}: As={top.area_mm2:.0f} mm², layers={top.layers}; Stirrups Asw/s={float(stir.get('Asw_s_prov_mm2_per_mm',0))*1000:.0f} mm²/m; Pele/alma={As_skin_face/100:.2f} cm²/face",
    }


BeamDesigner.design_one = _design_one_design_final


# ============================================================
# ============================================================
MAIN_LONG_DIAMS_Detailing = [12.0, 16.0, 20.0, 25.0]
SKIN_DIAMS_Detailing = [10.0, 12.0, 16.0]
STIRRUP_DIAMS_Detailing = [6.0, 8.0, 10.0]
STIRRUP_LEGS_Detailing = [2, 4, 6]
TORSION_RELEVANCE_ETA_Detailing = 0.05  # abaixo deste valor TEd/TRd,max não se dimensiona armadura específica de torção
MAX_MAIN_LAYERS_Detailing = 2


def _practical_max_bars_per_layer_detailing(b_mm: float) -> int:
    """Limite construtivo prático para evitar soluções como 7Ø10 em vigas estreitas."""
    if b_mm <= 300.0:
        return 4
    if b_mm <= 450.0:
        return 5
    return 6


_original_beamdesigner_init_detailing = BeamDesigner.__init__


def _beamdesigner_init_detailing(self, *args, **kwargs):
    _original_beamdesigner_init_detailing(self, *args, **kwargs)
    self.long_diams = list(MAIN_LONG_DIAMS_Detailing)
    self.stirrup_diams = list(STIRRUP_DIAMS_Detailing)
    self.stirrup_legs = list(STIRRUP_LEGS_Detailing)


BeamDesigner.__init__ = _beamdesigner_init_detailing


def _choose_stirrup_diameter_detailing(self, phi_long: float) -> float:
    req = max(6.0, phi_long / 4.0)
    for phi in STIRRUP_DIAMS_Detailing:
        if phi >= req:
            return phi
    return STIRRUP_DIAMS_Detailing[-1]


BeamDesigner.choose_stirrup_diameter = _choose_stirrup_diameter_detailing


def _choose_longitudinal_bars_detailing(self, As_req: float, b_mm: float, prefer_phi: Optional[float] = None, max_layers: int = MAX_MAIN_LAYERS_Detailing) -> RebarChoice:
    """Escolha robusta da armadura principal.

    Regras de implementação:
    - armadura principal limitada a Ø12, Ø16, Ø20 e Ø25;
    - Ø25 só aparece quando soluções com Ø12/Ø16/Ø20 não satisfazem a área/arranjo;
    - máximo de 2 camadas para vigas correntes;
    - número de varões por camada limitado de forma prática para evitar congestionamento;
    - espaçamento livre mínimo: max(20 mm, Ø, dg+5 mm).
    """
    As_req = max(0.0, float(As_req or 0.0))
    max_layers_eff = min(int(max_layers or MAX_MAIN_LAYERS_Detailing), MAX_MAIN_LAYERS_Detailing)
    base_diams = list(MAIN_LONG_DIAMS_Detailing)
    if prefer_phi in base_diams:
        diams = [prefer_phi] + [d for d in base_diams if d != prefer_phi]
    else:
        diams = base_diams

    def build_candidates(use_practical_limit: bool = True):
        candidates = []
        for phi in diams:
            phi_st = self.choose_stirrup_diameter(phi)
            nmax_geom = self.max_bars_per_layer(b_mm, phi, phi_st)
            if nmax_geom < 2:
                continue
            nmax_layer = nmax_geom
            if use_practical_limit:
                nmax_layer = min(nmax_layer, _practical_max_bars_per_layer_detailing(b_mm))
            if nmax_layer < 2:
                continue
            for n in range(2, nmax_layer * max_layers_eff + 1):
                area = n * bar_area_mm2(phi)
                if area + 1e-9 < As_req:
                    continue
                layers = int(math.ceil(n / nmax_layer))
                if layers > max_layers_eff:
                    continue
                bars_per = min(n, nmax_layer)
                req = self.min_clear_spacing(phi)
                edge = self.cover_mm + phi_st + phi / 2.0
                inner = b_mm - 2.0 * edge
                clear = inner / (bars_per - 1) - phi if bars_per > 1 else 999.0
                if clear + 1e-9 < req:
                    continue
                remaining = n
                sum_a_y = 0.0
                sum_a = 0.0
                for layer in range(layers):
                    n_layer = min(nmax_layer, remaining)
                    y = edge + layer * (phi + req)
                    a = n_layer * bar_area_mm2(phi)
                    sum_a += a
                    sum_a_y += a * y
                    remaining -= n_layer
                centroid = sum_a_y / max(sum_a, 1e-9)
                choice = RebarChoice(n, phi, area, layers, bars_per, centroid, clear, "OK")
                # Score: uma camada e menor congestionamento são preferidos; depois área. Ø25 recebe penalização leve.
                phi25_penalty = 400.0 if phi >= 25.0 else 0.0
                congestion_penalty = 80.0 * max(0, bars_per - 3)
                layer_penalty = 10000.0 * (layers - 1)
                score = (layer_penalty + area + congestion_penalty + phi25_penalty, layers, bars_per, phi, n)
                candidates.append((score, choice))
                break
        return candidates

    candidates = build_candidates(use_practical_limit=True)
    if not candidates:
        # fallback normativo por espaçamento geométrico, ainda limitado a 2 camadas e diâmetros aceites
        candidates = build_candidates(use_practical_limit=False)
    if not candidates:
        return RebarChoice(0, 0.0, 0.0, 0, 0, 0.0, 0.0, "Não cabe")
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


BeamDesigner.choose_longitudinal_bars = _choose_longitudinal_bars_detailing


def _torsion_requirements_detailing(self, TEd_kNm: float, section: BeamSection, fck: float, fcd: float, fyd: float) -> Dict[str, float | str]:
    """Torção EC2 com limiar de relevância para evitar armadura específica para torções residuais.

    O programa calcula sempre TRd,max para auditoria, mas só dimensiona armadura específica
    quando TEd/TRd,max excede o limiar definido em TORSION_RELEVANCE_ETA_Detailing.
    """
    T = abs(float(TEd_kNm or 0.0)) * 1e6
    b = section.bw_mm
    h = section.h_mm
    if b <= 0 or h <= 0:
        return {
            "TRd_max_kNm": None,
            "Asw_s_torsion_req_mm2_per_mm": 0.0,
            "Asl_torsion_req_mm2": 0.0,
            "torsion_status": "Dados insuficientes",
            "torsion_considered": "Não",
            "eta_torsion_design": 0.0,
            "tef_mm": None,
            "Ak_mm2": None,
            "uk_mm": None,
        }
    A = b * h
    u = 2.0 * (b + h)
    tef = max(A / max(u, 1e-9), 2.0 * self.cover_mm, 50.0)
    tef = min(tef, min(b, h) / 2.0 - 1.0)
    Ak = max((b - tef) * (h - tef), 1.0)
    uk = 2.0 * max(b + h - 2.0 * tef, 1.0)
    cot = 1.0
    nu1 = 0.6 * (1.0 - fck / 250.0)
    TRdmax = 2.0 * nu1 * fcd * Ak * tef / (cot + 1.0 / cot)
    TRdmax_kNm = TRdmax / 1e6
    eta = (T / max(TRdmax, 1e-9)) if T > 0 else 0.0
    if T <= 1e-9:
        status = "Sem torção relevante"
        considered = "Não"
        Asw_s = 0.0
        Asl = 0.0
    elif eta < TORSION_RELEVANCE_ETA_Detailing:
        status = f"Sem torção relevante (ηT={eta:.3f} < {TORSION_RELEVANCE_ETA_Detailing:.2f})"
        considered = "Não"
        Asw_s = 0.0
        Asl = 0.0
    elif T <= TRdmax:
        status = "Requer verificação/armadura de torção"
        considered = "Sim"
        Asw_s = T / max(2.0 * Ak * fyd * cot, 1e-9)
        Asl = T * uk * cot / max(2.0 * Ak * fyd, 1e-9)
    else:
        status = "Não conforme: TEd > TRd,max"
        considered = "Sim"
        Asw_s = T / max(2.0 * Ak * fyd * cot, 1e-9)
        Asl = T * uk * cot / max(2.0 * Ak * fyd, 1e-9)
    return {
        "TRd_max_kNm": TRdmax_kNm,
        "Asw_s_torsion_req_mm2_per_mm": Asw_s,
        "Asl_torsion_req_mm2": Asl,
        "torsion_status": status,
        "torsion_considered": considered,
        "eta_torsion_design": eta,
        "tef_mm": tef,
        "Ak_mm2": Ak,
        "uk_mm": uk,
    }


BeamDesigner.torsion_requirements = _torsion_requirements_detailing


def _choose_stirrups_detailing(self, Asw_s_total: float, b: float, h: float, d: float, torsion: bool) -> Dict[str, float | str]:
    Asw_s_total = max(0.0, float(Asw_s_total or 0.0))
    s_lim_v = min(0.75 * d, 600.0)
    s_lim_t = min((2.0 * (b + h)) / 8.0, min(b, h), 350.0) if torsion else 999.0
    s_lim = min(s_lim_v, s_lim_t)
    best = None
    for phi in STIRRUP_DIAMS_Detailing:
        for legs in STIRRUP_LEGS_Detailing:
            if legs < 2:
                continue
            area = legs * bar_area_mm2(phi)
            for s in sorted(self.spacing_candidates_mm, reverse=True):
                if s > s_lim + 1e-9:
                    continue
                provided = area / s
                if provided + 1e-12 >= Asw_s_total:
                    # preferir maior espaçamento dentro do limite, depois menor diâmetro e menos ramos
                    score = (-s, phi, legs, provided)
                    cand = {"phi_st_mm": phi, "stirrup_legs": legs, "s_st_mm": s, "Asw_s_prov_mm2_per_mm": provided, "s_lim_mm": s_lim, "stirrup_status": "OK"}
                    if best is None or score < best[0]:
                        best = (score, cand)
    if best is None:
        phi = STIRRUP_DIAMS_Detailing[-1]
        legs = STIRRUP_LEGS_Detailing[-1]
        s = min([x for x in self.spacing_candidates_mm if x <= s_lim] or [75.0])
        area = legs * bar_area_mm2(phi)
        return {"phi_st_mm": phi, "stirrup_legs": legs, "s_st_mm": s, "Asw_s_prov_mm2_per_mm": area / s, "s_lim_mm": s_lim, "stirrup_status": "Não cabe"}
    return best[1]


BeamDesigner.choose_stirrups = _choose_stirrups_detailing


def _choose_skin_rebar_detailing(As_req_face_mm2: float) -> Dict[str, float | str]:
    req = max(0.0, float(As_req_face_mm2 or 0.0))
    if req <= 1e-9:
        return {"skin_rebar": "n/a", "skin_n_per_face": 0, "skin_phi_mm": 0.0, "skin_area_prov_face_mm2": 0.0}
    best = None
    for phi in SKIN_DIAMS_Detailing:
        for n in range(1, 5):
            area = n * bar_area_mm2(phi)
            if area + 1e-9 >= req:
                score = (n, area, phi)
                cand = {"skin_rebar": f"{n}Ø{int(phi)}/face", "skin_n_per_face": n, "skin_phi_mm": phi, "skin_area_prov_face_mm2": area}
                if best is None or score < best[0]:
                    best = (score, cand)
                break
    if best is None:
        phi = SKIN_DIAMS_Detailing[-1]
        n = int(math.ceil(req / bar_area_mm2(phi)))
        area = n * bar_area_mm2(phi)
        return {"skin_rebar": f"{n}Ø{int(phi)}/face", "skin_n_per_face": n, "skin_phi_mm": phi, "skin_area_prov_face_mm2": area}
    return best[1]


def _detailing_check_detailing(self, section: BeamSection, bot: RebarChoice, top: RebarChoice, stir: Dict[str, float | str], As_skin_face: float) -> Dict[str, float | str]:
    issues = []
    for name, reb in [("inferior", bot), ("superior", top)]:
        if reb.n_bars and reb.n_bars < 2:
            issues.append(f"armadura {name}: mínimo construtivo de 2 varões")
        if reb.phi_mm and reb.phi_mm not in MAIN_LONG_DIAMS_Detailing:
            issues.append(f"armadura {name}: diâmetro não aceite nesta versão")
        if reb.layers and reb.layers > MAX_MAIN_LAYERS_Detailing:
            issues.append(f"armadura {name}: mais de {MAX_MAIN_LAYERS_Detailing} camadas")
        if reb.bars_per_layer and reb.bars_per_layer > _practical_max_bars_per_layer_detailing(section.bw_mm):
            issues.append(f"armadura {name}: excesso de varões por camada")
        if reb.clear_spacing_mm and reb.clear_spacing_mm < self.min_clear_spacing(reb.phi_mm) - 1e-9:
            issues.append(f"armadura {name}: espaçamento livre insuficiente")
    if stir.get("phi_st_mm") not in STIRRUP_DIAMS_Detailing:
        issues.append("diâmetro dos estribos fora da gama adoptada")
    if finite(stir.get("stirrup_legs"), 0) < 2:
        issues.append("estribos com menos de 2 ramos")
    if finite(stir.get("s_st_mm"), 0) > finite(stir.get("s_lim_mm"), 9999) + 1e-9:
        issues.append("espaçamento dos estribos superior ao limite")
    status = "OK" if not issues else "Não conforme"
    return {
        "detailing_status": status,
        "detailing_issues": "; ".join(issues) if issues else "-",
        "detailing_min_clear_bot_mm": bot.clear_spacing_mm if bot.n_bars else None,
        "detailing_min_clear_top_mm": top.clear_spacing_mm if top.n_bars else None,
        "detailing_max_bars_per_layer": _practical_max_bars_per_layer_detailing(section.bw_mm),
    }


BeamDesigner.detailing_check = _detailing_check_detailing


def _design_one_detailing(self, row: pd.Series) -> Dict:
    material = str(row.get("material", DEFAULT_CONCRETE_CLASS) or DEFAULT_CONCRETE_CLASS)
    fck = parse_concrete_strength(material)
    cp = concrete_props(fck, alpha_cc=self.alpha_cc, gamma_c=self.gamma_c)
    sp = steel_props(self.fyk, gamma_s=self.gamma_s)
    fcd = cp["fcd"]
    fyd = sp["fyd"]

    bw = cm_to_mm(row.get("hy", 0.0))
    h = cm_to_mm(row.get("hz", 0.0))
    bf = cm_to_mm(row.get("bf", float("nan")))
    hf = cm_to_mm(row.get("hf", float("nan")))
    if bw <= 0 or h <= 0:
        return {"member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""), "status": "Falha", "failure_reason": "Dados incompletos: dimensões HY/HZ inválidas", "failure_type": "dados_incompletos"}
    if not math.isfinite(bf) or bf <= bw or not math.isfinite(hf) or hf <= 0:
        bf = 0.0
        hf = 0.0
    section = BeamSection(bw, h, bf, hf)

    phi_initial = 16.0
    phi_st_initial = self.choose_stirrup_diameter(phi_initial)
    d_bot_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
    d_top_initial = h - self.cover_mm - phi_st_initial - phi_initial / 2.0
    d2_initial = self.cover_mm + phi_st_initial + phi_initial / 2.0
    if d_bot_initial <= 0:
        return {"member": row.get("member", ""), "case": row.get("case", ""), "name": row.get("name", ""), "story": row.get("story", ""), "status": "Falha", "failure_reason": "Dados geométricos incompatíveis com recobrimento", "failure_type": "dados_incompletos"}

    Mpos = finite(row.get("m_pos_ed_kNm"), 0.0)
    Mneg = finite(row.get("m_neg_ed_kNm"), 0.0)
    VEd = finite(row.get("v_ed_kN"), 0.0)
    TEd = finite(row.get("t_ed_kNm"), 0.0)

    As_min_bot = self.as_min_beam(section.bw_mm, d_bot_initial, cp["fctm"], self.fyk)
    As_min_top = self.as_min_beam(section.bw_mm, d_top_initial, cp["fctm"], self.fyk)

    flex_pos = self.flexural_required(Mpos, section, d_bot_initial, d2_initial, fcd, fyd)
    section_neg = BeamSection(section.bw_mm, section.h_mm, 0.0, 0.0)
    flex_neg = self.flexural_required(Mneg, section_neg, d_top_initial, d2_initial, fcd, fyd)

    tors = self.torsion_requirements(TEd, section, fck, fcd, fyd)
    torsion_active = str(tors.get("torsion_considered", "Não")) == "Sim"
    Asl_t = float(tors.get("Asl_torsion_req_mm2") or 0.0)
    if torsion_active and h >= SKIN_REINF_MIN_HEIGHT_MM:
        As_torsion_top = 0.25 * Asl_t
        As_torsion_bot = 0.25 * Asl_t
        As_skin_face_req = 0.25 * Asl_t
        if h > SKIN_REINF_MIN_HEIGHT_MM:
            As_skin_face_req = max(As_skin_face_req, 0.001 * section.bw_mm * (h - SKIN_REINF_MIN_HEIGHT_MM) / 2.0)
        skin_note = "Aplicável: h >= 40 cm"
    elif torsion_active:
        As_torsion_top = 0.50 * Asl_t
        As_torsion_bot = 0.50 * Asl_t
        As_skin_face_req = 0.0
        skin_note = "Não aplicável: h < 40 cm"
    elif h > SKIN_REINF_MIN_HEIGHT_MM:
        # Sem torção relevante, só mantém eventual armadura de pele mínima para vigas mais altas.
        As_torsion_top = 0.0
        As_torsion_bot = 0.0
        As_skin_face_req = 0.001 * section.bw_mm * (h - SKIN_REINF_MIN_HEIGHT_MM) / 2.0
        skin_note = "Armadura de pele mínima para h > 40 cm"
    else:
        As_torsion_top = 0.0
        As_torsion_bot = 0.0
        As_skin_face_req = 0.0
        skin_note = "Não aplicável: h <= 40 cm sem torção relevante"

    skin_choice = _choose_skin_rebar_detailing(As_skin_face_req)
    As_skin_face_prov = float(skin_choice.get("skin_area_prov_face_mm2", 0.0) or 0.0)

    As_req_bot = max(As_min_bot, flex_pos["As_req"] + As_torsion_bot)
    As_req_top = max(As_min_top, flex_neg["As_req"] + flex_pos.get("As_comp_req", 0.0) + As_torsion_top)

    bot = self.choose_longitudinal_bars(As_req_bot, bw, max_layers=MAX_MAIN_LAYERS_Detailing)
    top = self.choose_longitudinal_bars(As_req_top, bw, max_layers=MAX_MAIN_LAYERS_Detailing)

    d_bot = h - bot.centroid_from_edge_mm if bot.n_bars else d_bot_initial
    d_top = h - top.centroid_from_edge_mm if top.n_bars else d_top_initial
    d2_top = top.centroid_from_edge_mm if top.n_bars else d2_initial
    d2_bot = bot.centroid_from_edge_mm if bot.n_bars else d2_initial

    Mrd_pos = self.flexural_capacity(bot.area_mm2, top.area_mm2, section, d_bot, d2_top, fcd, fyd)
    Mrd_neg = self.flexural_capacity(top.area_mm2, bot.area_mm2, section_neg, d_top, d2_bot, fcd, fyd)
    eta_m_pos = Mpos / max(Mrd_pos.get("MRd_kNm") or 0.0, 1e-9) if Mpos > 0 else 0.0
    eta_m_neg = Mneg / max(Mrd_neg.get("MRd_kNm") or 0.0, 1e-9) if Mneg > 0 else 0.0

    shear = self.shear_requirements(VEd, bw, min(d_bot, d_top), max(bot.area_mm2, top.area_mm2), fck, fcd, fyd)
    Asw_s_total = max(float(shear["Asw_s_min_mm2_per_mm"]), float(shear["Asw_s_shear_req_mm2_per_mm"])) + float(tors["Asw_s_torsion_req_mm2_per_mm"] or 0.0)
    stir = self.choose_stirrups(Asw_s_total, bw, h, min(d_bot, d_top), torsion=torsion_active)

    els = self.serviceability(row, bot.area_mm2, top.area_mm2, d_bot, d_top, section, cp)
    det = self.detailing_check(section, bot, top, stir, As_skin_face_req)

    failure_reasons = []
    if bot.status != "OK" or top.status != "OK":
        failure_reasons.append("armadura longitudinal não cabe com os diâmetros e espaçamentos adoptados")
    if eta_m_pos > 1.0 + 1e-6:
        failure_reasons.append("flexão positiva não verifica")
    if eta_m_neg > 1.0 + 1e-6:
        failure_reasons.append("flexão negativa não verifica")
    if "Não conforme" in str(shear["shear_status"]):
        failure_reasons.append("esforço transverso excede VRd,max")
    if "Não conforme" in str(tors["torsion_status"]):
        failure_reasons.append("torção excede TRd,max")
    if stir.get("stirrup_status") != "OK":
        failure_reasons.append("armadura transversal não cabe")
    if det.get("detailing_status") != "OK":
        failure_reasons.append("pormenorização não conforme")
    status = "OK" if not failure_reasons else "Falha"
    failure_reason = "; ".join(dict.fromkeys(failure_reasons))
    failure_type = ""
    if failure_reasons:
        txt = failure_reason.lower()
        if "flexão" in txt:
            failure_type = "flexao"
        elif "transverso" in txt:
            failure_type = "esforco_transverso"
        elif "torção" in txt:
            failure_type = "torcao"
        elif "cabe" in txt or "pormenorização" in txt:
            failure_type = "pormenorizacao"
        else:
            failure_type = "outra"

    stirrup_label = "Estribos fechados" if torsion_active else "Estribos"
    sol = f"Inf.: {bot.label}; Sup.: {top.label}; {stirrup_label} Ø{int(stir['phi_st_mm'])}/{int(stir['stirrup_legs'])}r // {float(stir['s_st_mm'])/10:.1f} cm"
    if As_skin_face_req > 1e-6:
        sol += f"; pele/alma: {skin_choice['skin_rebar']}"

    eta_shear_max = VEd / max(finite(shear.get("VRd_max_kN")), 1e-9) if VEd > 0 else 0.0
    eta_torsion_max = TEd / max(finite(tors.get("TRd_max_kNm")), 1e-9) if TEd > 0 else 0.0
    eta_flexure_max = max(eta_m_pos, eta_m_neg)
    eta_global = max(eta_flexure_max, eta_shear_max, eta_torsion_max)

    return {
        "member": row.get("member", ""),
        "case": row.get("case", ""),
        "combination_number": row.get("combination_number", extract_combination_number(row.get("case", ""))),
        "limit_state": row.get("limit_state", classify_limit_state(row.get("case", ""))),
        "name": row.get("name", ""),
        "story": row.get("story", ""),
        "node_i": row.get("node_i", ""),
        "node_j": row.get("node_j", ""),
        "n_points_found": row.get("n_points_found", None),
        "length_m": finite(row.get("length"), 0.0),
        "material": material,
        "material_source": row.get("material_source", "tabela"),
        "section_type": section.section_type,
        "bw_cm": bw / 10.0,
        "h_cm": h / 10.0,
        "bf_cm": bf / 10.0 if bf > 0 else None,
        "hf_cm": hf / 10.0 if hf > 0 else None,
        "cover_mm": self.cover_mm,
        "fck_MPa": fck,
        "fcd_MPa": fcd,
        "fyk_MPa": self.fyk,
        "fyd_MPa": fyd,
        "moment_axis": row.get("moment_axis", "MY"),
        "shear_axis": row.get("shear_axis", "FZ"),
        "torsion_axis": row.get("torsion_axis", "MX"),
        "m_pos_ed_kNm": Mpos,
        "m_pos_at": row.get("m_pos_at", ""),
        "m_neg_ed_kNm": Mneg,
        "m_neg_at": row.get("m_neg_at", ""),
        "v_ed_kN": VEd,
        "v_at": row.get("v_at", ""),
        "t_ed_kNm": TEd,
        "t_at": row.get("t_at", ""),
        "as_min_bot_mm2": As_min_bot,
        "as_min_top_mm2": As_min_top,
        "as_req_bot_mm2": As_req_bot,
        "as_req_top_mm2": As_req_top,
        "as_prov_bot_mm2": bot.area_mm2,
        "as_prov_top_mm2": top.area_mm2,
        "bot_rebar": bot.label,
        "top_rebar": top.label,
        "bot_layers": bot.layers,
        "top_layers": top.layers,
        "bot_bars_per_layer": bot.bars_per_layer,
        "top_bars_per_layer": top.bars_per_layer,
        "bot_clear_spacing_mm": bot.clear_spacing_mm,
        "top_clear_spacing_mm": top.clear_spacing_mm,
        "d_bot_mm": d_bot,
        "d_top_mm": d_top,
        "mrd_pos_kNm": Mrd_pos.get("MRd_kNm"),
        "mrd_neg_kNm": Mrd_neg.get("MRd_kNm"),
        "eta_m_pos": eta_m_pos,
        "eta_m_neg": eta_m_neg,
        "x_pos_mm": Mrd_pos.get("x_mm"),
        "x_neg_mm": Mrd_neg.get("x_mm"),
        "ductility_pos": flex_pos.get("ductility_status"),
        "ductility_neg": flex_neg.get("ductility_status"),
        "VRd_c_kN": shear.get("VRd_c_kN"),
        "VRd_max_kN": shear.get("VRd_max_kN"),
        "Asw_s_shear_req_mm2_per_m": float(shear.get("Asw_s_shear_req_mm2_per_mm", 0.0)) * 1000.0,
        "Asw_s_min_mm2_per_m": float(shear.get("Asw_s_min_mm2_per_mm", 0.0)) * 1000.0,
        "TRd_max_kNm": tors.get("TRd_max_kNm"),
        "Asw_s_torsion_req_mm2_per_m": float(tors.get("Asw_s_torsion_req_mm2_per_mm", 0.0) or 0.0) * 1000.0,
        "Asl_torsion_req_mm2": tors.get("Asl_torsion_req_mm2"),
        "torsion_considered": tors.get("torsion_considered"),
        "eta_torsion_design": tors.get("eta_torsion_design"),
        "torsion_relevance_limit": TORSION_RELEVANCE_ETA_Detailing,
        "Asw_s_total_req_mm2_per_m": Asw_s_total * 1000.0,
        "phi_st_mm": stir.get("phi_st_mm"),
        "stirrup_legs": stir.get("stirrup_legs"),
        "s_st_mm": stir.get("s_st_mm"),
        "Asw_s_prov_mm2_per_m": float(stir.get("Asw_s_prov_mm2_per_mm", 0.0)) * 1000.0,
        "shear_status": shear.get("shear_status"),
        "torsion_status": tors.get("torsion_status"),
        "stirrup_status": stir.get("stirrup_status"),
        "skin_reinf_face_mm2": As_skin_face_req,
        "skin_reinf_face_prov_mm2": As_skin_face_prov,
        "skin_rebar": skin_choice.get("skin_rebar"),
        "skin_n_per_face": skin_choice.get("skin_n_per_face"),
        "skin_phi_mm": skin_choice.get("skin_phi_mm"),
        "skin_reinf_threshold_cm": SKIN_REINF_MIN_HEIGHT_MM / 10.0,
        "skin_reinf_note": skin_note,
        "norma": NORMATIVE_SUPPORT,
        "design_basis": "Suporte normativo fixo para Portugal",
        "calc_version": APP_VERSION,
        "flexure_method": "NP EN 1992-1-1:2010, Secção 6.1; MRd com armadura adoptada",
        "shear_method": "NP EN 1992-1-1:2010, Secção 6.2; VRd,c, VRd,max e Asw/s",
        "torsion_method": "NP EN 1992-1-1:2010, Secção 6.3; torção dimensionada apenas quando relevante",
        "sls_method": "NP EN 1992-1-1:2010, Secção 7; ELS expedito",
        "detailing_method": "NP EN 1992-1-1:2010, Secções 8 e 9.2; espaçamentos mínimos e pormenorização construtiva",
        "eta_flexure_max": eta_flexure_max,
        "eta_shear_max": eta_shear_max,
        "eta_torsion_max": eta_torsion_max,
        "eta_global": eta_global,
        **els,
        **det,
        "status": status,
        "failure_reason": failure_reason,
        "failure_type": failure_type,
        "recommendations": "",
        "solution": sol,
        "shortlist_text": f"Bot {bot.label}: As={bot.area_mm2:.0f} mm², camadas={bot.layers}, clear={bot.clear_spacing_mm:.0f} mm; Top {top.label}: As={top.area_mm2:.0f} mm², camadas={top.layers}, clear={top.clear_spacing_mm:.0f} mm; Estribos Asw/s={float(stir.get('Asw_s_prov_mm2_per_mm',0))*1000:.0f} mm²/m; Pele/alma={skin_choice.get('skin_rebar')}",
    }


BeamDesigner.design_one = _design_one_detailing


def _detailing_audit_df_detailing(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","story","bot_rebar","top_rebar","bot_layers","top_layers","bot_bars_per_layer","top_bars_per_layer","bot_clear_spacing_mm","top_clear_spacing_mm","phi_st_mm","stirrup_legs","s_st_mm","skin_rebar","skin_reinf_face_mm2","skin_reinf_face_prov_mm2","detailing_status","detailing_issues","solution"]
    return results[[c for c in cols if c in results.columns]].copy() if results is not None and not results.empty else pd.DataFrame()


# substituir a função usada nos relatórios e memória de cálculo
_detaling_old_name = globals().get('_detailing_audit_df_report_base')
_detaling_old_name = _detaling_old_name
_detaling_old_name
_detaling_old_name = None
globals()['_detailing_audit_df_report_base'] = _detailing_audit_df_detailing


def _update_report_detailing(self):
    self.report_txt.delete("1.0", "end")
    if self.df_results is None or self.df_results.empty:
        self.report_txt.insert("1.0", "Sem resultados. Importe a tabela e execute o cálculo.")
        return
    source = self.df_summary if self.df_summary is not None and not self.df_summary.empty else self.df_results
    n_total = len(self.df_results)
    n_ok = int((self.df_results["status"] == "OK").sum()) if "status" in self.df_results.columns else 0
    n_fail = int((self.df_results["status"] == "Falha").sum()) if "status" in self.df_results.columns else 0
    lines = [
        f"{APP_NAME} {APP_VERSION}\n",
        f"Relatório técnico de dimensionamento de vigas - {NORMATIVE_SUPPORT}\n\n",
        f"Envelopes analisados: {n_total} | OK: {n_ok} | Falhas: {n_fail}\n",
        f"Recobrimento: {self.var_cover.get()} mm | fyk: {self.var_fyk.get()} MPa | cotθ: {self.var_cot_theta.get()}\n",
        f"PDF seleccionado: {(self.var_pdf_scope.get() if hasattr(self, 'var_pdf_scope') else 'Completo')}\n\n",
    ]
    for _, r in source.head(120).iterrows():
        sec = f"{finite(r.get('bw_cm')):.0f} x {finite(r.get('h_cm')):.0f} cm"
        story = str(r.get('story','') or '')
        lines.append(f"Viga {r.get('member','')} | Caso {r.get('case','')} | Piso {story}\n")
        lines.append(f"  Secção: {sec} | {r.get('section_type','')} | Material: {r.get('material','')}\n")
        lines.append(f"  Esforços: M+Ed={_fmt_report_base(r.get('m_pos_ed_kNm'),2)} kNm; M-Ed={_fmt_report_base(r.get('m_neg_ed_kNm'),2)} kNm; VEd={_fmt_report_base(r.get('v_ed_kN'),2)} kN; TEd={_fmt_report_base(r.get('t_ed_kNm'),2)} kNm\n")
        lines.append(f"  Flexão: MRd+={_fmt_report_base(r.get('mrd_pos_kNm'),2)} kNm (η={_fmt_report_base(r.get('eta_m_pos'),3)}); MRd-={_fmt_report_base(r.get('mrd_neg_kNm'),2)} kNm (η={_fmt_report_base(r.get('eta_m_neg'),3)})\n")
        torsion_status = str(r.get('torsion_status','') or '')
        if torsion_status:
            lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'),2)} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'),2)} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'),2)} kNm; Torção: {torsion_status}\n")
        else:
            lines.append(f"  V/T: VRd,c={_fmt_report_base(r.get('VRd_c_kN'),2)} kN; VRd,max={_fmt_report_base(r.get('VRd_max_kN'),2)} kN; TRd,max={_fmt_report_base(r.get('TRd_max_kNm'),2)} kNm\n")
        skin = str(r.get('skin_rebar','') or 'n/a')
        skin_txt = f" | Pele/alma: {skin}" if skin and skin.lower() != 'n/a' else ""
        stirrup_prefix = "Estribos fechados" if str(r.get('torsion_considered','Não')) == 'Sim' else "Estribos"
        lines.append(f"  Armaduras: Inf. {r.get('bot_rebar','')} | Sup. {r.get('top_rebar','')} | {stirrup_prefix} Ø{_fmt_report_base(r.get('phi_st_mm'),0)}/{_fmt_report_base(r.get('stirrup_legs'),0)}r // {_fmt_report_base(finite(r.get('s_st_mm'))/10,1)} cm{skin_txt}\n")
        lines.append(f"  Pormenorização: clear inf.={_fmt_report_base(r.get('bot_clear_spacing_mm'),0)} mm; clear sup.={_fmt_report_base(r.get('top_clear_spacing_mm'),0)} mm; Estado porm.: {r.get('detailing_status','')}\n")
        lines.append(f"  ELS: {r.get('service_status','')} | Estado: {r.get('status','')}\n")
        motivo = str(r.get("failure_reason", "") or "").strip()
        if motivo:
            lines.append(f"  Motivo: {motivo}\n")
        lines.append("\n")
    self.report_txt.insert("1.0", "".join(lines))


BeamsEC2App.update_report = _update_report_detailing


def _metadata_df_detailing(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Programa", APP_NAME], ["Versão", APP_VERSION], ["Autor", APP_AUTHOR], ["Repositório", GITHUB_URL],
        ["Data de exportação", datetime.now().strftime("%Y-%m-%d %H:%M")], ["Ficheiro de origem", self.input_file_path or "-"],
        ["Suporte normativo", NORMATIVE_SUPPORT], ["Âmbito", "Dimensionamento/verificação ELU e ELS expedito de vigas de betão armado"],
        ["Critério de torção", f"Armadura específica apenas se TEd/TRd,max >= {TORSION_RELEVANCE_ETA_Detailing:.2f}"],
        ["Diâmetros principais", "Ø12, Ø16, Ø20, Ø25"],
        ["Diâmetros de estribos", "Ø6, Ø8, Ø10; mínimo 2 ramos"],
        ["Descrição", APP_TABLE_DESCRIPTION],
    ], columns=["Campo", "Valor"])


BeamsEC2App._metadata_df = _metadata_df_detailing


def _parameters_df_detailing(self) -> pd.DataFrame:
    return pd.DataFrame([
        ["Suporte normativo", NORMATIVE_SUPPORT],
        ["Recobrimento [mm]", self.var_cover.get()], ["Agregado dg [mm]", self.var_agg.get()], ["Aço fyk [MPa]", self.var_fyk.get()],
        ["γc", self.var_gamma_c.get()], ["γs", self.var_gamma_s.get()], ["cotθ", self.var_cot_theta.get()],
        ["wk,lim [mm]", self.var_crack_limit.get()], ["L/d limite", self.var_ld_limit.get()],
        ["Momento principal", self.var_moment_axis.get()], ["Corte vertical", self.var_shear_axis.get()], ["Torção", self.var_torsion_axis.get()],
        ["Diâmetros principais adoptados", "Ø12, Ø16, Ø20, Ø25"],
        ["Diâmetros de estribos adoptados", "Ø6, Ø8, Ø10; mínimo 2 ramos"],
        ["Máximo de camadas principais", MAX_MAIN_LAYERS_Detailing],
        ["Limiar de torção", f"TEd/TRd,max >= {TORSION_RELEVANCE_ETA_Detailing:.2f}"],
        ["Armadura de pele/alma", "reportada em formato nØ/face quando aplicável"],
        ["Redução para casos governantes", "Sim" if self.var_reduce_cases.get() else "Não"],
        ["Tipo de relatório PDF", self.var_pdf_scope.get() if hasattr(self, 'var_pdf_scope') else "Completo"],
    ], columns=["Parâmetro", "Valor"])


BeamsEC2App._parameters_df = _parameters_df_detailing


def _vt_audit_df_detailing(results: pd.DataFrame) -> pd.DataFrame:
    cols = ["member","case","story","v_ed_kN","v_at","VRd_c_kN","VRd_max_kN","eta_shear_max","Asw_s_shear_req_mm2_per_m","Asw_s_min_mm2_per_m","t_ed_kNm","t_at","TRd_max_kNm","eta_torsion_max","eta_torsion_design","torsion_relevance_limit","torsion_considered","Asw_s_torsion_req_mm2_per_m","Asl_torsion_req_mm2","Asw_s_total_req_mm2_per_m","phi_st_mm","stirrup_legs","s_st_mm","Asw_s_prov_mm2_per_m","shear_status","torsion_status","stirrup_status"]
    return results[[c for c in cols if c in results.columns]].copy() if results is not None and not results.empty else pd.DataFrame()

globals()['_vt_audit_df_report_base'] = _vt_audit_df_detailing


def _write_pdf_detailing(self, path: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
    styles = _pdf_styles_report_base()
    scope = self.var_pdf_scope.get() if hasattr(self, "var_pdf_scope") else "Completo"
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    doc.title = APP_NAME; doc.author = APP_AUTHOR; doc.subject = APP_SUBJECT
    story = []
    results = self.df_results
    summary = self.df_summary if self.df_summary is not None and not self.df_summary.empty else results
    failures = self.df_failures if self.df_failures is not None else pd.DataFrame()
    n_total = len(results); n_ok = int((results["status"]=="OK").sum()) if "status" in results.columns else 0; n_fail = int((results["status"]=="Falha").sum()) if "status" in results.columns else 0
    story.append(Paragraph(APP_NAME, styles["ReportTitle"])); story.append(Paragraph(f"Dimensionamento de vigas — {NORMATIVE_SUPPORT}", styles["ReportSubtitle"]))
    meta = [["Programa", f"{APP_NAME} {APP_VERSION}", "Autor", APP_AUTHOR], ["Data", datetime.now().strftime("%Y-%m-%d %H:%M"), "Suporte normativo", NORMATIVE_SUPPORT], ["Envelopes", str(n_total), "OK/Falhas", f"{n_ok}/{n_fail}"], ["Diâmetros principais", "Ø12, Ø16, Ø20, Ø25", "Torção", f"ηT ≥ {TORSION_RELEVANCE_ETA_Detailing:.2f}"]]
    t = Table(meta, colWidths=[38*mm,90*mm,38*mm,105*mm]); t.setStyle(self._pdf_table_style(header=False)); story += [t, Spacer(1,5*mm)]
    if scope in ["Completo", "Resumo executivo"]:
        story.extend([Paragraph("Resumo executivo", styles["Section"]), self._pdf_df_table(_executive_summary_df_report_base(results, summary), ["Indicador","Valor"], max_rows=25, widths=[90,180]), Spacer(1,5*mm), Paragraph("Resumo por viga", styles["BodyCourier"]), self._pdf_df_table(summary, ["member","story","case","section_type","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm","bot_rebar","top_rebar","skin_rebar","solution","status"], max_rows=30)])
    if scope in ["Completo", "Relatório técnico"]:
        if story and scope == "Completo": story.append(PageBreak())
        story.extend([Paragraph("Relatório técnico", styles["Section"]), Paragraph("Critérios de cálculo", styles["BodyCourier"]), self._pdf_df_table(self._parameters_df(), ["Parâmetro","Valor"], max_rows=45, widths=[90,180]), Spacer(1,5*mm), Paragraph("Envelopes de esforços", styles["BodyCourier"]), self._pdf_df_table(self.df_env, ["member","story","case","n_points_found","length","material","hy","hz","bf","hf","m_pos_ed_kNm","m_neg_ed_kNm","v_ed_kN","t_ed_kNm"], max_rows=38), Spacer(1,5*mm), Paragraph("Flexão", styles["BodyCourier"]), self._pdf_df_table(_flexure_audit_df_report_base(summary), ["member","case","m_pos_ed_kNm","mrd_pos_kNm","eta_m_pos","bot_rebar","m_neg_ed_kNm","mrd_neg_kNm","eta_m_neg","top_rebar","ductility_pos","ductility_neg"], max_rows=34), Spacer(1,5*mm), Paragraph("Esforço transverso e torção", styles["BodyCourier"]), self._pdf_df_table(_vt_audit_df_detailing(summary), ["member","case","v_ed_kN","VRd_c_kN","VRd_max_kN","t_ed_kNm","TRd_max_kNm","eta_torsion_design","torsion_considered","Asw_s_total_req_mm2_per_m","Asw_s_prov_mm2_per_m","shear_status","torsion_status"], max_rows=34)])
    if scope in ["Completo", "Memória de cálculo"]:
        if story and scope == "Completo": story.append(PageBreak())
        mem = _calc_memory_df_report_base(summary)
        story.extend([Paragraph("Memória de cálculo", styles["Section"]), self._pdf_df_table(mem, ["Viga","Caso","Piso","Secção","Etapa","Item","Valor","Unidade","Critério/Referência","Estado/Nota"], max_rows=80)])
    if scope in ["Completo", "Relatório técnico"]:
        story.extend([PageBreak(), Paragraph("ELS e pormenorização", styles["Section"]), Paragraph("Verificações em serviço", styles["BodyCourier"]), self._pdf_df_table(_sls_audit_df_report_base(summary), ["member","case","service_sigma_s_MPa","service_wk_est_mm","service_wk_lim_mm","service_L_over_d","service_status","service_note"], max_rows=30), Spacer(1,5*mm), Paragraph("Pormenorização", styles["BodyCourier"]), self._pdf_df_table(_detailing_audit_df_detailing(summary), ["member","case","bot_rebar","top_rebar","bot_clear_spacing_mm","top_clear_spacing_mm","phi_st_mm","stirrup_legs","s_st_mm","skin_rebar","detailing_status","detailing_issues"], max_rows=30)])
        if failures is not None and not failures.empty:
            story.extend([PageBreak(), Paragraph("Falhas", styles["Section"]), self._pdf_df_table(failures, ["member","story","case","failure_type","failure_reason"], max_rows=45)])
    story.append(Spacer(1,5*mm)); story.append(Paragraph("Nota: workbook memória de cálculo contém a auditoria completa por envelope, incluindo espaçamentos livres e critérios de torção.", styles["Small"]))
    def footer(canvas, doc_obj):
        canvas.saveState(); canvas.setAuthor(APP_AUTHOR); canvas.setTitle(APP_NAME); canvas.setSubject(APP_SUBJECT); canvas.setFont("Courier",7); canvas.setFillColor(colors.grey); canvas.drawString(12*mm,7*mm,f"{APP_NAME} {APP_VERSION} | {APP_AUTHOR}"); canvas.drawRightString(285*mm,7*mm,f"Página {doc_obj.page}"); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


BeamsEC2App._write_pdf = _write_pdf_detailing



